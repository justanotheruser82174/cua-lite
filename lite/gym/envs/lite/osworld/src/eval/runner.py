"""OSWorld evaluator for lite_osworld tasks.

Uses Flask server endpoints for all operations that metrics depend on
(matching OSWorld's getter behavior exactly). Uses computer.interface
only for file downloads (base64 transfer).

Usage:
    from lite.gym.envs.lite.osworld.src.eval.runner import evaluate_osworld_task
"""

from __future__ import annotations

import asyncio
import base64
import hashlib
import functools
import json
import logging
import os
from pathlib import Path
import shlex
import tempfile
import urllib.request

from lite.gym.envs.lite.osworld.src.utils.dispatch import (
    dispatch_action,
    CHROME_DATA_DIR as _CHROME_DATA_DIR,
    _replace_templates,
)
logger = logging.getLogger(__name__)

_SERVER = "http://localhost:5000"

# Bare shell operator tokens that appear as their OWN element in an argv-list getter
# command (e.g. ["code","--list-extensions","|","grep",X]); left UNQUOTED when the
# list is joined to a `bash -c` string so pipes/redirects/and-or keep working, while
# every non-operator token is shlex.quote'd (so a `python3 -c '<script>'` arg keeps
# its quoting). See _normalize_osworld_command_config.
_SHELL_OP_TOKENS = frozenset({
    "|", "||", "&&", ";", "&", ">", ">>", "<", "2>", "2>>", "2>&1", "|&", "&>",
})


def _cache_output_path(cache_dir: str, name: str, *, label: str = "output") -> str:
    """Resolve a host-side evaluator artifact under ``cache_dir``."""
    if not isinstance(name, str) or not name:
        raise ValueError(f"{label} must be a non-empty relative path")
    raw = Path(name)
    if not raw.parts:
        raise ValueError(f"{label} must be a non-empty relative path")
    if raw.is_absolute() or ".." in raw.parts:
        raise ValueError(f"{label} must stay inside cache_dir: {name!r}")
    cache_root = Path(cache_dir).resolve()
    path = (cache_root / raw).resolve()
    if path != cache_root and cache_root not in path.parents:
        raise ValueError(f"{label} must stay inside cache_dir: {name!r}")
    path.parent.mkdir(parents=True, exist_ok=True)
    return str(path)


# ---------------------------------------------------------------------------
# Flask server helpers
# ---------------------------------------------------------------------------

async def _flask_get(computer, endpoint: str) -> str:
    """Call Flask server GET endpoint."""
    r = await computer.interface.run_command(f"curl -s {_SERVER}{endpoint}")
    return r.stdout


async def _flask_post(computer, endpoint: str, data: dict) -> str:
    """Call Flask server POST endpoint."""
    payload = shlex.quote(json.dumps(data))
    r = await computer.interface.run_command(
        f"curl -s -X POST {_SERVER}{endpoint} -H 'Content-Type: application/json' -d {payload}"
    )
    return r.stdout


async def _flask_post_json(computer, endpoint: str, data: dict) -> dict:
    """Call Flask server POST endpoint, return parsed JSON."""
    raw = await _flask_post(computer, endpoint, data)
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        return {}


# ---------------------------------------------------------------------------
# Main evaluator
# ---------------------------------------------------------------------------

def _summarize(data: object, limit: int = 2000) -> object:
    """Truncate large strings for debug output; pass through other types."""
    if isinstance(data, str) and len(data) > limit:
        return data[:limit] + f"... ({len(data)} chars)"
    return data


async def evaluate_osworld_task(
    computer, evaluator: dict, cache_dir: str | None = None, *, debug: bool = False,
) -> float | tuple[float, dict]:
    """Run OSWorld evaluator: postconfig → getter(result) → getter(expected) → metric."""
    if cache_dir is None:
        cache_dir = tempfile.mkdtemp(prefix="osworld_eval_")
    os.makedirs(cache_dir, exist_ok=True)

    func = evaluator.get("func", "")
    if func == "infeasible":
        return (0.0, {"func": "infeasible"}) if debug else 0.0

    # --- Postconfig (uses shared dispatch) ---
    if not evaluator.get("_postconfig_done"):
        for step in evaluator.get("postconfig", []):
            await dispatch_action(computer, step, cache_dir=cache_dir)

    # --- Resolve metrics ---
    func_list = func if isinstance(func, list) else [func]
    result_list = evaluator.get("result", {})
    if not isinstance(result_list, list):
        result_list = [result_list]
    expected_list = evaluator.get("expected", {})
    if not isinstance(expected_list, list):
        expected_list = [expected_list]
    options_list = evaluator.get("options", {})
    if not isinstance(options_list, list):
        options_list = [options_list]

    n = len(func_list)
    result_list += [{}] * (n - len(result_list))
    expected_list += [{}] * (n - len(expected_list))
    options_list += [{}] * (n - len(options_list))

    conj = evaluator.get("conj", "and")
    scores = []
    details = [] if debug else None

    # desktop_env presence is enforced at lite.osworld/main.py module-load
    # time (raises EnvDepsMissingError if missing → env removed from registry).
    # By the time eval runs it's guaranteed importable — no fallback needed.
    from desktop_env.evaluators import metrics

    # Local lite_osworld metrics override / extend the upstream OSWorld set.
    # See custom_metrics.py for the rationale (compare_docx_files is text-only,
    # so all format-only writer tasks trivially pass without it).
    from lite.gym.envs.lite.osworld.src.eval import metrics as custom_metrics

    for i, fn_name in enumerate(func_list):
        try:
            result_data = await _get_result(computer, result_list[i] or {}, cache_dir)
            # DIAG dump (set CUALITE_EVAL_DUMP=/path to enable)
            _dump = os.environ.get("CUALITE_EVAL_DUMP")
            if _dump and isinstance(result_data, str) and result_data.lstrip().startswith("<"):
                try:
                    with open(_dump, "w") as _f:
                        _f.write(result_data)
                except Exception:
                    pass
            # When expected_list has more entries than func_list (multi-expected
            # pattern, e.g. compare_unique_train_records which needs both a gold
            # file and an initial-state reference), fetch all remaining entries
            # and pass them as a list to the metric.
            if i == len(func_list) - 1 and len(expected_list) > len(func_list):
                _remaining = expected_list[i:]
                if len(_remaining) > 1:
                    _items = []
                    for _cfg in _remaining:
                        _items.append(await _get_expected(computer, _cfg or {}, cache_dir))
                    expected_data = _items
                else:
                    expected_data = await _get_expected(computer, expected_list[i] or {}, cache_dir)
            else:
                expected_data = await _get_expected(computer, expected_list[i] or {}, cache_dir)
            metric_fn = getattr(custom_metrics, fn_name, None)
            if metric_fn is None:
                metric_fn = getattr(metrics, fn_name, None)
            if metric_fn is None:
                logger.warning("Metric not found: %s", fn_name)
                scores.append(0.0)
                if details is not None:
                    details.append({"func": fn_name, "error": "metric not found", "score": 0.0})
                continue
            # Match OSWorld's calling pattern:
            # - with expected: metric(result, expected, **options)
            # - without expected: metric(result, **options)
            opts = options_list[i] or {}
            # Run the metric in a thread so CPU-heavy evaluators (SSIM on
            # large images, PDF parsing, audio DTW, etc.) don't block the
            # async event loop and stall other concurrent tasks.
            import sys as _sys

            def _run_metric(
                _fn=metric_fn, _rd=result_data, _ed=expected_data,
                _opts=opts, _cache=cache_dir,
            ):
                # Snapshot sys.modules so we can evict task-specific modules
                # (e.g. settings, tetris, block) imported from cache_dir.
                _modules_before = set(_sys.modules.keys())
                if _ed is not None:
                    result = _fn(_rd, _ed, **_opts)
                else:
                    result = _fn(_rd, **_opts)
                # Remove modules imported from the temp cache_dir
                for _mod in list(_sys.modules.keys()):
                    if _mod not in _modules_before:
                        _mod_obj = _sys.modules.get(_mod)
                        _mod_file = getattr(_mod_obj, "__file__", None) or ""
                        if _cache and _mod_file.startswith(_cache):
                            del _sys.modules[_mod]
                return result

            loop = asyncio.get_running_loop()
            score = await loop.run_in_executor(None, _run_metric)
            score_f = float(score) if isinstance(score, (int, float)) else (1.0 if score else 0.0)
            scores.append(score_f)
            if details is not None:
                details.append({
                    "func": fn_name,
                    "result_config": result_list[i],
                    "result_data": _summarize(result_data),
                    "expected_data": _summarize(expected_data),
                    "score": score_f,
                })
        except Exception as e:
            logger.warning("Metric %s[%d] failed: %s", fn_name, i, e)
            scores.append(0.0)
            if details is not None:
                details.append({"func": fn_name, "error": str(e), "score": 0.0})

    # Faithful to upstream OSWorld DesktopEnv.evaluate(): return the RAW aggregate
    # score (partial credit preserved) — NEVER binarize at 0.5. The old
    # `1.0 if all(s >= 0.5)` promoted a partial (e.g. compare_references 0.855) to a
    # full 1.0 pass, making the eval strictly EASIER than upstream and inflating
    # results. Upstream aggregation: "and" → 0.0 if any sub-metric is exactly 0.0
    # else the mean; "or" → max. For a single metric both reduce to the raw
    # metric value. (Mirrors lite.scalecua
    # verify._combine_scores, which is already faithful.)
    if not scores:
        final = 0.0
    elif conj == "and":
        final = 0.0 if any(s == 0.0 for s in scores) else sum(scores) / len(scores)
    else:
        final = max(scores)
    if debug:
        return final, {"conj": conj, "scores": scores, "details": details}
    return final


# ---------------------------------------------------------------------------
# AT / CDP helpers
# ---------------------------------------------------------------------------

async def _get_at_xml(computer) -> str | None:
    """Fetch accessibility tree XML from the OSWorld Flask server."""
    raw = await _flask_get(computer, "/accessibility")
    try:
        return json.loads(raw).get("AT", "")
    except (json.JSONDecodeError, TypeError):
        return None


def _extract_address_bar_url(at_xml: str) -> str | None:
    """Extract the Chrome address bar text from AT XML."""
    try:
        import lxml.etree
        from lxml.cssselect import CSSSelector

        tree = lxml.etree.fromstring(at_xml.encode() if isinstance(at_xml, str) else at_xml)

        for app_name in ("Google\\ Chrome", "Chromium"):
            sel_str = f"application[name={app_name}] entry[name=Address\\ and\\ search\\ bar]"
            try:
                sel = CSSSelector(sel_str, namespaces={"": "https://accessibility.ubuntu.example.org/ns/component"})
                elems = sel(tree)
                if elems and elems[0].text:
                    return elems[0].text
            except Exception:
                pass

        for elem in tree.iter():
            name = elem.get("name", "")
            if "Address" in name and "search" in name.lower() and elem.text:
                return elem.text

    except Exception as e:
        logger.warning("Failed to extract URL from AT: %s", e)
    return None


async def _get_active_url(computer, config: dict) -> str | None:
    """Get active Chrome URL. CDP first, AT-SPI fallback."""
    url = await _get_active_url_via_cdp(computer)
    if url:
        return url
    at_xml = await _get_at_xml(computer)
    if at_xml:
        raw_url = await asyncio.to_thread(_extract_address_bar_url, at_xml)
        if raw_url:
            goto_prefix = config.get("goto_prefix", "https://")
            return f"{goto_prefix}{raw_url}"
    return None


async def _get_active_url_via_cdp(computer) -> str | None:
    """Get the active tab URL via Chrome DevTools Protocol."""
    r = await computer.interface.run_command("curl -s http://localhost:1337/json")
    try:
        tabs = json.loads(r.stdout)
        pages = [t for t in tabs if t.get("type") == "page"]
        if pages:
            return pages[0].get("url", "")
    except (json.JSONDecodeError, TypeError):
        pass
    return None


async def _get_open_tabs_via_cdp(computer) -> list[dict]:
    """Get list of open Chrome tabs via CDP."""
    r = await computer.interface.run_command("curl -s http://localhost:1337/json")
    try:
        tabs = json.loads(r.stdout)
        # Filter out Chrome internal pages (chrome:// scheme) — e.g. the omnibox
        # popup (chrome://omnibox-popup.top-chrome/) which appears as a CDP "page"
        # but is never a real user tab.
        return [{"title": t.get("title", ""), "url": t.get("url", "")}
                for t in tabs if t.get("type") == "page"
                and not t.get("url", "").startswith("chrome://")]
    except (json.JSONDecodeError, TypeError):
        return []


async def _get_html_parse_via_cdp(computer, url: str, config: dict) -> dict | None:
    """Use python + CDP inside container to get HTML and parse it."""
    config_json = json.dumps(config).replace('"', '\\"').replace("'", "\\'")
    script = f'''
import json, requests, time
from urllib.parse import unquote
try:
    from bs4 import BeautifulSoup
except ImportError:
    print(json.dumps({{"error": "bs4 not installed"}}))
    exit()

config = json.loads('{config_json}')
url = "{url}"

pages = requests.get("http://localhost:1337/json").json()
target = None
for p in pages:
    if p.get("type") == "page":
        def norm(u): return unquote(u).rstrip("/")
        if norm(p["url"]) == norm(url):
            target = p
            break
if not target:
    for p in pages:
        if p.get("type") == "page":
            target = p
            break
if not target:
    print(json.dumps({{"error": "no page found"}}))
    exit()

ws_url = target["webSocketDebuggerUrl"]
import websocket
ws = websocket.create_connection(ws_url)
ws.send(json.dumps({{"id": 1, "method": "Runtime.evaluate", "params": {{"expression": "document.documentElement.outerHTML"}}}}))
result = json.loads(ws.recv())
ws.close()
html = result.get("result", {{}}).get("result", {{}}).get("value", "")
if not html:
    print(json.dumps({{"error": "empty html"}}))
    exit()

soup = BeautifulSoup(html, "html.parser")
extracted = {{}}
category = config.get("category", "")

def _texts_of(cls):
    return [el.get_text(strip=True) for el in soup.find_all(class_=cls)]

def _direct_text_nodes_of(cls):
    # OSWorld safely_get_direct_text_nodes_playwright: for the first matching element,
    # return its direct text node children (text content not inside child tags).
    elems = soup.find_all(class_=cls)
    if not elems:
        return []
    out = []
    for child in elems[0].children:
        if getattr(child, "name", None) is None:
            txt = str(child).strip()
            if txt:
                out.append(txt)
    return out

def _only_child_h3_of(cls):
    # OSWorld safely_get_only_child_text_content: for each element with class,
    # return the text of its <h3> descendant.
    out = []
    for el in soup.find_all(class_=cls):
        h3 = el.find("h3")
        if h3:
            out.append(h3.get_text(strip=True))
    return out

def _direct_li_of(cls):
    # OSWorld safely_get_direct_li_playwright: for each element with class,
    # find child 'li.catAllProducts' and return its span text.
    out = []
    for el in soup.find_all(class_=cls):
        for li in el.find_all("li", class_="catAllProducts"):
            span = li.find("span")
            if span:
                out.append(span.get_text(strip=True))
    return out

if category == "class":
    class_multi = config.get("class_multiObject", {{}})
    for cls, rank_map in class_multi.items():
        texts = _texts_of(cls)
        for order_key, key in rank_map.items():
            idx = int(order_key)
            extracted[key] = texts[idx] if idx < len(texts) else ""

    class_multi_child = config.get("class_multiObject_child", {{}})
    for cls, rank_map in class_multi_child.items():
        texts = _direct_text_nodes_of(cls)
        for order_key, key in rank_map.items():
            idx = int(order_key)
            extracted[key] = texts[idx] if idx < len(texts) else ""

    class_multi_only_child = config.get("class_multiObject_only_child", {{}})
    for cls, rank_map in class_multi_only_child.items():
        texts = _only_child_h3_of(cls)
        for order_key, key in rank_map.items():
            idx = int(order_key)
            extracted[key] = texts[idx] if idx < len(texts) else ""

    class_multi_search = config.get("class_multiObject_search_exist", {{}})
    for cls, object_list in class_multi_search.items():
        texts = _texts_of(cls)
        for each in object_list:
            if each == "is_other_exist":
                continue
            extracted[each] = each in texts
        if "is_other_exist" in object_list:
            extras = [t for t in texts if t not in object_list]
            if "is_other_exist" not in extracted:
                extracted["is_other_exist"] = bool(extras)

    class_single = config.get("class_singleObject", {{}})
    for cls, key in class_single.items():
        texts = _texts_of(cls)
        extracted[key] = texts[0] if texts else ""

elif category == "label":
    label_obj = config.get("labelObject", {{}})
    for selector, key in label_obj.items():
        label = soup.find("label", string=lambda t: t and selector in t)
        if label:
            extracted[key] = label.parent.get_text(strip=True) if label.parent else ""
        else:
            extracted[key] = ""

elif category == "xpath":
    import lxml.etree
    tree = lxml.etree.HTML(html)
    xpath_obj = config.get("xpathObject", {{}})
    for xpath, key in xpath_obj.items():
        elems = tree.xpath(xpath) if tree is not None else []
        if elems:
            el = elems[0]
            if hasattr(el, "text_content"):
                txt = el.text_content() or ""
            elif hasattr(el, "text"):
                txt = el.text or ""
            else:
                txt = str(el)
            extracted[key] = txt.strip()
        else:
            extracted[key] = ""

elif category == "input":
    import lxml.etree
    tree = lxml.etree.HTML(html)
    input_obj = config.get("inputObject", {{}})
    for xpath, key in input_obj.items():
        elems = tree.xpath(xpath) if tree is not None else []
        if elems:
            extracted[key] = elems[0].get("value", "") if hasattr(elems[0], "get") else ""
        else:
            extracted[key] = ""

elif category == "class&url":
    class_multi = config.get("class_multiObject", {{}})
    for cls, object_list in class_multi.items():
        texts = _texts_of(cls)
        lower_texts = [t.lower() for t in texts]
        for each in object_list:
            if any(each.lower() == t for t in lower_texts):
                extracted[each.lower()] = True
        # is_other_exist: any text that doesn't substring-match any expected
        other = False
        for t in texts:
            if all(t.lower() not in item.lower() for item in object_list):
                other = True
                break
        if "is_other_exist" not in extracted:
            extracted["is_other_exist"] = other

    class_multi_li = config.get("class_multiObject_li", {{}})
    for cls, object_list in class_multi_li.items():
        texts = _direct_li_of(cls)
        lower_texts = [t.lower() for t in texts]
        for each in object_list:
            if any(each.lower() == t for t in lower_texts):
                extracted[each.lower()] = True
        other = False
        for t in texts:
            if all(t.lower() not in item.lower() for item in object_list):
                other = True
                break
        if "is_other_exist" not in extracted:
            extracted["is_other_exist"] = other

    page_url_lower = (target.get("url","") or "").lower()
    url_include_expected = config.get("url_include_expected", [])
    for key in url_include_expected:
        k = key.lower()
        if k not in extracted:
            extracted[k] = k in page_url_lower

    url_include_expected_multichoice = config.get("url_include_expected_multichoice", {{}})
    for key, value in url_include_expected_multichoice.items():
        v = value.lower()
        if v not in extracted:
            extracted[v] = key.lower() in page_url_lower

print(json.dumps(extracted))
'''
    # bare python3 resolves to the env venv (requests/bs4/lxml/websocket) via the
    # harness PATH.
    r = await computer.interface.run_command(
        f"python3 -c {shlex.quote(script)}")
    try:
        return json.loads(r.stdout)
    except (json.JSONDecodeError, TypeError):
        logger.warning("active_tab_html_parse failed: %s", r.stdout[:200])
        return None


async def _get_chrome_appearance_mode(computer) -> str | None:
    """Get Chrome appearance mode via CDP JS evaluation."""
    script = '''
import json, requests, time
try:
    import websocket
except ImportError:
    print("system")
    exit()

pages = requests.get("http://localhost:1337/json").json()
ws_url = None
for p in pages:
    if p.get("type") == "page":
        ws_url = p["webSocketDebuggerUrl"]
        break
if not ws_url:
    print("system")
    exit()

ws = websocket.create_connection(ws_url)
ws.send(json.dumps({"id": 1, "method": "Page.navigate", "params": {"url": "chrome://settings/appearance"}}))
ws.recv()
time.sleep(3)

js = """
(() => {
  try {
    const ui = document.querySelector('settings-ui');
    if (!ui || !ui.shadowRoot) return 'system';
    const prefs = ui.prefs;
    if (prefs && prefs.browser && prefs.browser.theme) {
      const cs = prefs.browser.theme.color_scheme;
      if (cs && cs.value !== undefined) {
        if (cs.value === 1) return 'light';
        if (cs.value === 2) return 'dark';
        return 'system';
      }
    }
    return 'system';
  } catch(e) { return 'system'; }
})()
"""
ws.send(json.dumps({"id": 2, "method": "Runtime.evaluate", "params": {"expression": js}}))
result = json.loads(ws.recv())
ws.close()
mode = result.get("result", {}).get("result", {}).get("value", "system")
print(mode)
'''
    # bare python3 resolves to the env venv (requests/websocket) via the harness PATH.
    r = await computer.interface.run_command(
        f"python3 -c {shlex.quote(script)}")
    raw = r.stdout.strip()
    return raw if raw in ("light", "dark", "system") else "system"


async def _get_page_info_via_cdp(computer, target_url: str) -> dict:
    """Get page info (title, url, content) via CDP inside container."""
    script = '''
import json, requests, time, sys
target_url = sys.argv[1]
try:
    import websocket
except ImportError:
    print(json.dumps({"title": "", "url": target_url, "content": ""}))
    exit()

pages = requests.get("http://localhost:1337/json").json()
ws_url = None
for p in pages:
    if p.get("type") == "page":
        ws_url = p["webSocketDebuggerUrl"]
        break
if not ws_url:
    print(json.dumps({"title": "", "url": target_url, "content": ""}))
    exit()

ws = websocket.create_connection(ws_url)
ws.send(json.dumps({"id": 1, "method": "Page.navigate", "params": {"url": target_url}}))
ws.recv()
time.sleep(3)
ws.send(json.dumps({"id": 2, "method": "Runtime.evaluate", "params": {"expression": "document.title"}}))
title = json.loads(ws.recv()).get("result", {}).get("result", {}).get("value", "")
ws.send(json.dumps({"id": 3, "method": "Runtime.evaluate", "params": {"expression": "window.location.href"}}))
final_url = json.loads(ws.recv()).get("result", {}).get("result", {}).get("value", "")
ws.send(json.dumps({"id": 4, "method": "Runtime.evaluate", "params": {"expression": "document.documentElement.outerHTML"}}))
content = json.loads(ws.recv()).get("result", {}).get("result", {}).get("value", "")
ws.close()
print(json.dumps({"title": title, "url": final_url, "content": content}))
'''
    # bare python3 resolves to the env venv (requests/websocket) via the harness PATH.
    r = await computer.interface.run_command(
        f"python3 -c {shlex.quote(script)} {shlex.quote(target_url)}")
    try:
        return json.loads(r.stdout)
    except (json.JSONDecodeError, TypeError):
        return {"title": "", "url": target_url, "content": ""}


async def _get_media_from_slide(computer, config: dict, cache_dir: str, media_type: str = "audio"):
    """Extract audio or background image from a PowerPoint slide."""
    import xml.etree.ElementTree as ET
    import zipfile

    ppt_path = config.get("ppt_file_path", "")
    slide_index = int(config.get("slide_index", 0))

    local_pptx = await _download_from_container(computer, ppt_path, cache_dir)
    if not local_pptx:
        return None

    with zipfile.ZipFile(local_pptx, "r") as zf:
        if media_type == "image":
            slide_xml = f"ppt/slides/slide{slide_index + 1}.xml"
            if slide_xml not in zf.namelist():
                return None
            with zf.open(slide_xml) as f:
                tree = ET.parse(f)
                root = tree.getroot()
            bg_tag = "{http://schemas.openxmlformats.org/presentationml/2006/main}bgPr"
            blip_tag = "{http://schemas.openxmlformats.org/drawingml/2006/main}blip"
            embed_attr = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
            image_id = None
            for child in root.iter(bg_tag):
                for elem in child.iter(blip_tag):
                    image_id = elem.attrib.get(embed_attr)
                    if image_id:
                        break
                if image_id:
                    break
            if not image_id:
                return None
            rel_type_match = "image"
            target_id = image_id
        else:
            rel_type_match = "audio"
            target_id = None

        rels_xml = f"ppt/slides/_rels/slide{slide_index + 1}.xml.rels"
        if rels_xml not in zf.namelist():
            return None
        with zf.open(rels_xml) as f:
            tree = ET.parse(f)
            root = tree.getroot()
        ns = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
        for rel in root.findall("r:Relationship", ns):
            if rel_type_match not in rel.attrib.get("Type", ""):
                continue
            if target_id and rel.attrib.get("Id") != target_id:
                continue
            target = rel.attrib["Target"]
            if target.startswith(".."):
                resolved = os.path.normpath(os.path.join("ppt/slides", target)).replace("\\", "/")
                tmpdir = os.path.dirname(local_pptx)
                zf.extract(resolved, tmpdir)
                return os.path.join(tmpdir, resolved)
            elif target.startswith("file://"):
                return await _download_from_container(computer, target[7:], cache_dir)

    return None


# ---------------------------------------------------------------------------
# Chrome Preferences extractors
# ---------------------------------------------------------------------------

def _extract_chrome_pref(result_type: str, prefs: dict):
    """Extract a specific value from Chrome Preferences.

    Each extractor matches the return value of the corresponding OSWorld getter
    in desktop_env/evaluators/getters/chrome.py.
    """
    # --- get_enable_do_not_track ---
    if result_type == "enable_do_not_track":
        return "true" if prefs.get("enable_do_not_track", {}) else "false"

    # --- get_check_dns_prefetch (not in OSWorld, keep simple) ---
    if result_type == "check_dns_prefetch":
        return "true" if prefs.get("dns_prefetching", {}).get("enabled", True) else "false"

    # --- get_enable_safe_browsing ---
    if result_type == "enable_safe_browsing":
        sb = prefs.get("safebrowsing", {})
        is_enhanced = bool(sb.get("enhanced", False))
        is_enabled = bool(sb.get("enabled", False))
        return "true" if (is_enhanced or is_enabled) else "false"

    # --- get_data_delete_automacally ---
    # The real Chrome UI control ("Delete data sites have saved to your device
    # when you close all windows", chrome://settings/content/siteData) writes
    # profile.default_content_setting_values.cookies = 4 (clear-on-exit). LIVE-
    # verified on Chrome 149: toggling that radio flips the key None→4, and
    # browser.clear_data.browsing_data_lifetime is NEVER written by the UI (it's
    # an enterprise-policy key). The old getter checked browsing_data_lifetime
    # (co-changed with the oracle), so a real agent driving the UI scored 0
    # on chrome_99146c54. Upstream's "default_content_setting_values is not
    # None" check trivially passed — wrong the other way. We check the specific
    # cookies==4 the UI writes: positive proof, no trivial pass.
    if result_type == "data_delete_automacally":
        cookies = (prefs.get("profile", {})
                   .get("default_content_setting_values", {})
                   .get("cookies"))
        return "true" if cookies == 4 else "false"

    # --- get_new_startup_page ---
    # OSWorld: requires session.restore_on_startup == 5 (positive proof).
    # Old behavior returned "true" when "session" was missing — that auto-passed
    # any composite Chrome+X task where Chrome was never launched (no Preferences
    # file → empty prefs dict). Now require the actual key to be present.
    if result_type == "new_startup_page":
        restore = prefs.get("session", {}).get("restore_on_startup", {})
        return "true" if restore == 5 else "false"

    # --- get_default_search_engine ---
    if result_type == "default_search_engine":
        tud = prefs.get("default_search_provider_data", {}).get("template_url_data", {})
        name = tud.get("short_name", tud.get("keyword", "Google"))
        # Chrome localizes built-in provider short names (e.g. "Yahoo! Hong
        # Kong", "Bing (UK)"). Eval alias lists use the canonical brand, so
        # collapse a trailing locale qualifier for known providers.
        n = name.strip()
        for brand in ("Yahoo!", "Yahoo", "Bing", "Google", "DuckDuckGo", "Baidu", "Ecosia"):
            if n == brand or n.startswith(brand + " "):
                return brand
        return n

    # --- get_chrome_font_size ---
    # OSWorld returns full webkit.webprefs dict with defaults
    if result_type == "chrome_font_size":
        return prefs.get("webkit", {}).get("webprefs", {
            "default_fixed_font_size": 13,
            "default_font_size": 16,
            "minimum_font_size": 13,
        })

    # --- get_profile_name ---
    if result_type == "profile_name":
        return prefs.get("profile", {}).get("name", "")

    # --- get_chrome_color_scheme ---
    # Canonical-first: Chrome's Settings UI writes
    # browser.theme.color_scheme; color_scheme2 is a stale mirror left by
    # perturbation setup. Prefer the canonical key (agrees with the JS
    # chrome_appearance_mode_ui getter at ~L548); fall back to the mirror only
    # when canonical is absent (older-style writes) so no state regresses.
    if result_type == "chrome_color_scheme":
        def _normalize(raw):
            if isinstance(raw, str):
                n = raw.strip().lower()
                if n in ("system", "default", "0"): return "system"
                if n in ("light", "1"): return "light"
                if n in ("dark", "2"): return "dark"
            if isinstance(raw, (int, float)):
                return {0: "system", 1: "light", 2: "dark"}.get(int(raw))
            return None
        theme = prefs.get("browser", {}).get("theme", {})
        raw = theme.get("color_scheme", theme.get("color_scheme2"))
        mode = _normalize(raw)
        # Fallback: extensions.theme.system_theme
        st = prefs.get("extensions", {}).get("theme", {}).get("system_theme")
        if st in (0, "0", False):
            return "light"
        return mode if mode else "system"

    return prefs


# ---------------------------------------------------------------------------
# Result getters
# ---------------------------------------------------------------------------

def _redirect_chrome_profile(cmd):
    """Rewrite the upstream default Chrome profile path to the ``--user-data-dir``
    Chrome actually launches with (``/home/user/chrome-data``) — symmetric with
    dispatch.py's setup-side redirect. str or argv-list; no-op otherwise."""
    if isinstance(cmd, str):
        return cmd.replace("/home/user/.config/google-chrome", _CHROME_DATA_DIR)
    if isinstance(cmd, (list, tuple)):
        return type(cmd)(
            _redirect_chrome_profile(c) if isinstance(c, str) else c for c in cmd
        )
    return cmd


def _normalize_osworld_command_config(cmd, shell):
    """Value-rewrites symmetric with the setup/oracle dispatch path (dispatch.py
    applies these to config/oracle ``execute`` actions): OSWorld template vars
    ({CLIENT_PASSWORD}/{SCREEN_WIDTH}/…) and the chrome-profile redirect
    (/home/user/.config/google-chrome → the --user-data-dir Chrome actually uses).
    Applied to EVERY getter; each rewrite is a no-op on a command that lacks its
    tokens. Every getter runs as the desktop user (the exec-stdio server's identity)
    with the desktop-session env (DISPLAY/DBUS/XDG) already available, so no manual
    session prefix or path rewrite here.
    """
    cmd = _replace_templates(cmd)
    cmd = _redirect_chrome_profile(cmd)
    # Both getters (osworld + scalecua, which delegates here) send `cmd` straight to
    # run_command, which requires a STRING (the in-container server's _bracket_pkill
    # does `re.sub` on it → TypeError on a list). OSWorld/ScaleCUA config rows supply
    # argv-LIST getter commands too (the pre-collapse Flask path accepted + joined
    # them). Join OPERATOR-AWARE so the string is correct for EVERY list shape the
    # corpus uses: `shlex.quote` each token so a `bash -c '<script>'` / `python3 -c
    # 'import …; …'` arg keeps its quoting (a plain `" ".join` would spill the script's
    # `;`/spaces into the shell → wrong output → silent reward-0), but leave a bare
    # shell OPERATOR token (`|`/`&&`/`;`/`>`…) UNQUOTED so a
    # ["code","--list-extensions","|","grep",X] pipe stays a pipe under `bash -c`.
    if isinstance(cmd, (list, tuple)):
        if shell:
            # shell=True argv (rare): join OPERATOR-AWARE so a `|`/`&&`/`;`/`>` token
            # stays a shell operator under `bash -c` (a pipe stays a pipe), while
            # every non-operator token is `shlex.quote`d to keep its own quoting.
            cmd = " ".join(
                s if (s := str(x)) in _SHELL_OP_TOKENS else shlex.quote(s) for x in cmd
            )
        else:
            # shell=False argv: OSWorld runs it via `subprocess.run(cmd, shell=False)`
            # — execve, NO shell interpretation, so `|`/`grep`/… are LITERAL operands
            # to the first program, NOT a pipe. run_command needs a string, so emulate
            # that under `bash -c` by `shlex.quote`ing EVERY token (operators included)
            # → they become literal args, matching upstream byte-for-byte. Fixes the
            # `ls DIR | grep FILE` getters (vs_code_5e2d93d8 / _57242fad): upstream's
            # literal read surfaces FILE from the getter's cwd (home) and passes, while
            # lite's spurious real pipe filtered against DIR and false-failed (reward 0
            # vs VM 1). The 4 `code --list-extensions | grep X` getters stay correct —
            # `code` still dumps the full extension list and the `contain X` rule
            # matches, exactly as on the VM (all shell=False, verified via eval.jsonl).
            cmd = " ".join(shlex.quote(str(x)) for x in cmd)
    return cmd, shell


# ---------------------------------------------------------------------------
# LibreOffice registry config-flush (read-before-flush, LO family)
# ---------------------------------------------------------------------------
# LibreOffice commits its user registry (registrymodifications.xcu) to disk ONLY
# on a clean shutdown, so a bare `vm_file` download of that path while the
# agent's soffice.bin is still running reads a STALE/empty xcu -> the setting the
# agent just changed is absent -> false-negative. Mirrors scalecua verify.py's
# gimp/thunderbird/vlc flush infra: send ctrl+q so the app's own quit->save path
# runs, then SIGTERM soffice.bin (whose handler flushes the user registry) --
# NEVER -9, which skips it -- then sync. The env-vendored `xdotool` (resolved via
# the server's PATH) drives the user X session on DISPLAY=:1 as the desktop user
# (see dispatch.py); the guaranteed flush is the SIGTERM, which fires even if the
# window search is empty.
#
# Oracle/no-op replay: the eval fixtures already `killall -9 soffice.bin`
# before the getter, so `pgrep -x soffice.bin` is empty, the SIGTERM loop breaks
# immediately, and the body reduces to a bare `sync`. Already-green fixtures are
# undisturbed; the 0->1 recovery this enables is rollout-only.

_LIBREOFFICE_CONFIG_FILE_MARKERS = (
    "registrymodifications.xcu",
    "/.config/libreoffice/",
    "~/.config/libreoffice/",
    ".config/libreoffice/",
)


def _mentions_libreoffice_config_file(value) -> bool:
    """True if any string in `value` (recursively) names the LibreOffice user
    registry / config dir. File-marker gate (no-overfit): fires on ANY vm_file
    read of registrymodifications.xcu regardless of the task's metric func, and
    excludes `.docx`/`.odt`/`.xlsx` document reads (which read the saved file,
    not the registry -- a separate document-save concern, not this flush)."""
    if isinstance(value, str):
        lowered = value.lower()
        return any(marker in lowered for marker in _LIBREOFFICE_CONFIG_FILE_MARKERS)
    if isinstance(value, dict):
        return any(_mentions_libreoffice_config_file(item) for item in value.values())
    if isinstance(value, (list, tuple)):
        return any(_mentions_libreoffice_config_file(item) for item in value)
    return False


def _best_effort_flush(fn):
    """Make a config-flush helper best-effort — a flush persists in-memory app
    state before a getter reads it, but it is an OPTIMIZATION, not a correctness
    requirement: a flush that hangs/times out under load must NEVER fail the
    reward (mirror of scalecua verify.py's `_flush_vlc` fix: a TimeoutExpired
    there must never become reward 0.0). Swallow + log any Exception. Worst case the getter reads
    un-flushed (stale) state, exactly as if the flush had not run — score can only
    go DOWN, never fabricate reward (no new FP). `Exception` only, so
    asyncio.CancelledError still propagates.
    """
    @functools.wraps(fn)
    async def _wrapped(*args, **kwargs):
        try:
            return await fn(*args, **kwargs)
        except Exception as exc:  # noqa: BLE001 — best-effort by design
            logger.warning(
                "config-flush %s failed (best-effort, ignored): %r", fn.__name__, exc
            )
    return _wrapped


@_best_effort_flush
async def _flush_libreoffice_config(computer) -> None:
    """Graceful-quit LibreOffice so it commits registrymodifications.xcu before
    an eval download reads it. ctrl+q (best-effort) -> SIGTERM soffice.bin
    (guaranteed registry flush; NEVER -9) -> sync. No-op when no soffice runs."""
    command = r"""
sent_quit=0
for display in "${DISPLAY:-}" :1 :0; do
  [ -z "$display" ] && continue
  DISPLAY="$display"; export DISPLAY
  for wid in $(xdotool search --class 'libreoffice' 2>/dev/null; xdotool search --class 'soffice' 2>/dev/null); do
    xdotool windowactivate --sync "$wid" 2>/dev/null || true
    xdotool key --window "$wid" ctrl+q 2>/dev/null || true
    sent_quit=1
    break 2
  done
done
if [ "$sent_quit" -eq 1 ]; then
  sleep 2
  xdotool key Return 2>/dev/null || true
fi
# Guaranteed flush: SIGTERM soffice.bin runs its user-registry save handler
# (NEVER -9, which skips it). No-op when no soffice is running during replay.
for _ in $(seq 1 30); do
  pgrep -x soffice.bin >/dev/null 2>&1 || break
  pkill -TERM soffice.bin 2>/dev/null || true
  pkill -TERM oosplash 2>/dev/null || true
  sleep 1
done
sync
"""
    await computer.interface.run_command(
        "bash -lc " + shlex.quote(command),
        timeout=45,
    )


# ---------------------------------------------------------------------------
# Chrome profile config-flush (read-before-flush, Chrome family)
# ---------------------------------------------------------------------------
# Chrome persists Preferences / Bookmarks (ImportantFileWriter, batched ~commit
# timer + write-on-clean-shutdown), so a bare on-disk read while chrome is still
# running reads a STALE profile -> FN. Proven in-container: a UI font-size change
# (webkit.webprefs.default_font_size 16->22) is held in memory; the on-disk read
# 0.5s later still returns 16 (SCORE 0); a graceful quit flushes it to 22
# (SCORE 1). accept_languages / safebrowsing.enabled stay stale 30s+.
#
# Mirror of scalecua verify.py::_flush_chrome_profile + _needs_chrome_profile_flush
# and of the LO flush above. Whitelist gate on the ON-DISK chrome result_types
# ONLY -- the live CDP/AT readers (active_url*, open_tabs_info, active_tab_*,
# chrome_appearance_mode_ui) are excluded, since killing chrome would break them.
# SAFE by corpus audit: across all 300 osworld chrome tasks, NO evaluator mixes an
# on-disk reader with a live reader, so a per-getter flush can never break a live
# read. Kill+relaunch tasks (postconfig `pkill chrome`) already flush via SIGTERM
# before relaunch; re-flushing their on-disk read is harmless (no live reader).
#
# Oracle/no-op replay: when no chrome runs the wait loop breaks immediately and
# the body reduces to a bare `sync`. Flush only PERSISTS the agent's true state --
# it cannot fabricate reward (negative control: no change -> disk stays at default
# 16 -> a task expecting 22 still SCORE 0).

_CHROME_ONDISK_RESULT_TYPES = frozenset({
    "enable_do_not_track", "check_dns_prefetch", "enable_safe_browsing",
    "data_delete_automacally", "new_startup_page", "default_search_engine",
    "chrome_font_size", "profile_name", "chrome_color_scheme",
    "bookmarks", "cookie_data", "history",
    "find_unpacked_extension_path", "find_installed_extension_name",
})

_CHROME_PROFILE_FILE_MARKERS = (
    "/home/user/chrome-data/",
    "/home/user/.config/google-chrome/",
    "/home/user/.config/chromium/",
)


def _mentions_chrome_profile_file(value) -> bool:
    """True if any string in `value` (recursively) names the Chrome profile dir.
    Future-proofing marker gate (parallels `_mentions_libreoffice_config_file`);
    the result-type whitelist is the primary gate."""
    if isinstance(value, str):
        lowered = value.lower()
        return any(marker in lowered for marker in _CHROME_PROFILE_FILE_MARKERS)
    if isinstance(value, dict):
        return any(_mentions_chrome_profile_file(item) for item in value.values())
    if isinstance(value, (list, tuple)):
        return any(_mentions_chrome_profile_file(item) for item in value)
    return False


def _needs_chrome_profile_flush(result_type: str, config: dict) -> bool:
    base = str(result_type or "").split("__", 1)[0]
    return base in _CHROME_ONDISK_RESULT_TYPES or _mentions_chrome_profile_file(config)


@_best_effort_flush
async def _flush_chrome_profile(computer) -> None:
    """Graceful-quit Chrome so it commits Preferences / Bookmarks / History before
    an eval reads the on-disk profile. Alt+F4 (best-effort) -> wait-for-exit ->
    SIGTERM chrome (guaranteed flush handler; NEVER -9, which skips it) -> sync.
    No-op when no chrome runs (oracle / no-kill fixtures)."""
    command = r"""
for display in "${DISPLAY:-}" :1 :0; do
  [ -z "$display" ] && continue
  DISPLAY="$display"; export DISPLAY
  for wid in $(xdotool search --class 'google-chrome' 2>/dev/null; xdotool search --class 'chromium' 2>/dev/null); do
    xdotool windowactivate --sync "$wid" 2>/dev/null || true
    xdotool key --window "$wid" alt+F4 2>/dev/null || true
  done
done
# Guaranteed flush: SIGTERM chrome runs its on-shutdown pref-commit handler
# (NEVER -9, which skips it). No-op when no chrome is running.
for _ in $(seq 1 30); do
  pgrep -x chrome >/dev/null 2>&1 || pgrep -x chromium >/dev/null 2>&1 || break
  pkill -TERM chrome 2>/dev/null || true
  pkill -TERM chromium 2>/dev/null || true
  sleep 0.5
done
# WAL checkpoint (AFTER chrome has exited, so nothing re-opens the db). SIGTERM
# alone is not enough for the sqlite-backed profile stores: a UI-driven change --
# Settings -> "Delete browsing data" -> Delete -- is committed to the WAL
# sidecar, while the getters (`is_cookie_deleted`, history readers) download and
# read the MAIN .db file. Without this the eval sees the pre-delete rows and a
# visibly-completed deletion scores 0 -- a false negative, observed on
# osworld_chrome_7b6c7e24 (agent emptied every amazon cookie + history entry on
# screen, reward 0). The synth chrome generator already documents and applies the
# same checkpoint for the tasks it emits (src/gen/train/synth/chrome.py); this
# puts it on the shared flush so eval rows get it too.
# The checkpoint must never CREATE a store. `sqlite3.connect(path)` opens
# rw-create, so a bare connect on an absent store writes a 0-byte file whenever
# the parent profile dir exists -- and /home/user/chrome-data/Default always
# does. That fabricates `Cookies` / `History` / `Web Data` / `Login Data` out of
# nothing, so an existence-checking getter (or a cookie-deletion evaluator) reads
# an empty table instead of seeing "absent" and can pass for the wrong reason.
# `mode=rw` (no `c`) makes the open FAIL on a missing file, atomically -- no
# exists()/connect() race, and it cannot be defeated by a store that appears
# between the two calls.
python3 - <<'PYEOF' 2>/dev/null || true
import sqlite3
for profile in ("/home/user/chrome-data/Default",
                "/home/user/.config/google-chrome/Default",
                "/home/user/.config/chromium/Default"):
    for store in ("Cookies", "History", "Web Data", "Login Data"):
        try:
            conn = sqlite3.connect(f"file:{profile}/{store}?mode=rw", uri=True)
            conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            conn.commit()
            conn.close()
        except Exception:
            pass
PYEOF
sync
"""
    await computer.interface.run_command(
        "bash -lc " + shlex.quote(command),
        timeout=45,
    )


async def _get_result(computer, config: dict, cache_dir: str):
    result_type = config.get("type", "")
    if not result_type:
        return None

    # Read-before-flush (Chrome family): Chrome persists Preferences /
    # Bookmarks / History only lazily (commit timer) or on clean shutdown, so an
    # on-disk read while chrome runs returns stale state -> FN. Flush BEFORE the
    # read. Whitelist gate excludes the live CDP/AT readers (no corpus task mixes
    # on-disk + live, so this never breaks a live getter).
    if _needs_chrome_profile_flush(result_type, config):
        await _flush_chrome_profile(computer)

    # --- File from container ---
    if result_type == "vm_file":
        # Read-before-flush (LO family): LibreOffice writes
        # registrymodifications.xcu only on clean shutdown, so a live download
        # while soffice.bin runs reads a stale xcu -> FN. Flush BEFORE the read.
        if _mentions_libreoffice_config_file(config):
            await _flush_libreoffice_config(computer)
        path = config.get("path", "")
        if isinstance(path, list):
            dests = config.get("dest", [])
            gives = set(config.get("gives", [0]))
            results = []
            for i, p in enumerate(path):
                dest_name = dests[i] if i < len(dests) else ""
                local = await _download_from_container(computer, p, cache_dir)
                if local and dest_name and os.path.basename(local) != dest_name:
                    dest_path = _cache_output_path(cache_dir, dest_name, label="vm_file dest")
                    import shutil as _shutil
                    # File copy can be large (GB xlsx/pptx); offload to thread.
                    await asyncio.to_thread(_shutil.copy2, local, dest_path)
                    local = dest_path
                if local and local.endswith((".xlsx", ".xls")):
                    # openpyxl load_workbook + iter_rows is CPU-heavy — offload.
                    await asyncio.to_thread(_ensure_csv_exports, local, cache_dir)
                # Validation: mirror `_get_expected` — normalize docx/pptx for
                # symmetric strict comparison. See note below in single-path
                # branch for rationale.
                if local and isinstance(local, str) and local.endswith(".docx") and os.path.isfile(local):
                    await _lo_normalize_docx(computer, local)
                elif local and isinstance(local, str) and local.endswith(".pptx") and os.path.isfile(local):
                    await _lo_normalize_pptx(computer, local)
                if i in gives:
                    results.append(local)
            # Match OSWorld: return string if single result, list if multiple
            if len(results) == 1:
                return results[0]
            return results if results else None
        local = await _download_from_container(computer, path, cache_dir)
        if local and local.endswith((".xlsx", ".xls")):
            await asyncio.to_thread(_ensure_csv_exports, local, cache_dir)
        # Validation: mirror `_get_expected`'s LO-normalize for symmetric
        # comparison. Without this, agent's interactive Ctrl+S can produce
        # `style.name='normal'` (preserved) while `_get_expected` headless
        # convert produces `'LO-Normal'` (canonicalized) — strict's
        # string-equal style.name check then false-fails on documents
        # with simple style refs (e.g. HK_train_record). Normalize is
        # idempotent so already-normalized files pass through unchanged.
        if local and isinstance(local, str) and local.endswith(".docx") and os.path.isfile(local):
            await _lo_normalize_docx(computer, local)
        elif local and isinstance(local, str) and local.endswith(".pptx") and os.path.isfile(local):
            await _lo_normalize_pptx(computer, local)
        return local

    # --- Command execution (via Flask to preserve shell param) ---
    if result_type == "vm_command_line":
        cmd = config.get("command", "")
        shell = config.get("shell", False)
        cmd, _shell = _normalize_osworld_command_config(cmd, shell)
        # Runs as the desktop user (the server's uid) — reads the user's live state
        # (dconf/Gio, AT-SPI) directly, and a `su -`-driven password check (upstream's
        # `check_password.sh`) prompts correctly. No channel/classifier.
        r = await computer.interface.run_command(cmd)
        return r.stdout

    if result_type == "vm_command_error":
        cmd = config.get("command", "")
        shell = config.get("shell", False)
        cmd, _shell = _normalize_osworld_command_config(cmd, shell)
        r = await computer.interface.run_command(cmd)
        return r.stderr

    # --- Terminal output (requires Flask server) ---
    if result_type == "vm_terminal_output":
        raw = await _flask_get(computer, "/terminal")
        try:
            return json.loads(raw).get("output", "")
        except (json.JSONDecodeError, TypeError):
            return raw

    # --- Chrome URL getters (CDP primary, AT-SPI fallback) ---
    if result_type == "active_url_from_accessTree":
        return await _get_active_url(computer, config)

    if result_type == "active_tab_info":
        url = await _get_active_url(computer, config)
        return {"url": url, "title": "", "content": ""} if url else None

    if result_type == "active_tab_url_parse":
        url = await _get_active_url(computer, config)
        if url is None:
            return None
        from urllib.parse import urlparse, parse_qs
        parsed = urlparse(url)
        query_params = parse_qs(parsed.query)
        keys_of_interest = config.get("parse_keys", [])
        extracted = {k: query_params.get(k, [""])[0] for k in keys_of_interest}
        if config.get("replace"):
            for old_key, new_key in config["replace"].items():
                extracted[new_key] = extracted.pop(old_key, "")
        if config.get("split_list"):
            extracted = {k: v.split(",") for k, v in extracted.items()}
        return extracted

    if result_type == "active_tab_html_parse":
        url = await _get_active_url(computer, config)
        if url is None:
            return None
        return await _get_html_parse_via_cdp(computer, url, config)

    if result_type == "open_tabs_info":
        return await _get_open_tabs_via_cdp(computer)

    if result_type == "chrome_appearance_mode_ui":
        return await _get_chrome_appearance_mode(computer)

    # --- Chrome Preferences ---
    chrome_pref_types = {
        "enable_do_not_track", "check_dns_prefetch", "enable_safe_browsing",
        "data_delete_automacally", "new_startup_page", "default_search_engine",
        "chrome_font_size", "profile_name", "chrome_color_scheme",
    }
    if result_type in chrome_pref_types:
        r = await computer.interface.run_command(
            "cat /home/user/chrome-data/Default/Preferences 2>/dev/null || "
            "cat /home/user/.config/google-chrome/Default/Preferences 2>/dev/null || echo '{}'"
        )
        try:
            prefs = json.loads(r.stdout)
        except (json.JSONDecodeError, TypeError):
            prefs = {}
        return _extract_chrome_pref(result_type, prefs)

    if result_type == "bookmarks":
        r = await computer.interface.run_command(
            "cat /home/user/chrome-data/Default/Bookmarks 2>/dev/null || "
            "cat /home/user/.config/google-chrome/Default/Bookmarks 2>/dev/null || echo '{}'"
        )
        try:
            data = json.loads(r.stdout)
            return data.get("roots", data)
        except (json.JSONDecodeError, TypeError):
            return {}

    if result_type == "cookie_data":
        local = await _download_from_container(computer, "/home/user/chrome-data/Default/Cookies", cache_dir)
        if not local:
            return None
        import sqlite3
        try:
            conn = sqlite3.connect(local)
            cookies = conn.cursor().execute("SELECT * FROM cookies").fetchall()
            conn.close()
            return cookies
        except Exception as e:
            logger.warning("cookie_data query failed: %s", e)
            return None

    if result_type == "history":
        local = await _download_from_container(computer, "/home/user/chrome-data/Default/History", cache_dir)
        if not local:
            return None
        import sqlite3
        try:
            conn = sqlite3.connect(local)
            rows = conn.cursor().execute("SELECT url, title, last_visit_time FROM urls").fetchall()
            conn.close()
            return rows
        except Exception as e:
            logger.warning("history query failed: %s", e)
            return None

    if result_type == "find_unpacked_extension_path":
        r = await computer.interface.run_command(
            "python3 -c \"import json,os; "
            "p='/home/user/chrome-data/Default/Preferences'; "
            "d=json.load(open(p)) if os.path.exists(p) else {}; "
            "exts=d.get('extensions',{}).get('settings',{}); "
            "print(json.dumps([exts[k]['path'] for k in exts]))\""
        )
        try:
            return json.loads(r.stdout)
        except (json.JSONDecodeError, TypeError):
            return []

    if result_type == "find_installed_extension_name":
        r = await computer.interface.run_command(
            "python3 -c \"import json,os; "
            "p='/home/user/chrome-data/Default/Preferences'; "
            "d=json.load(open(p)) if os.path.exists(p) else {}; "
            "exts=d.get('extensions',{}).get('settings',{}); "
            "names=[exts[k].get('manifest',{}).get('name','') for k in exts if 'manifest' in exts[k]]; "
            "print(json.dumps(names))\""
        )
        try:
            return json.loads(r.stdout)
        except (json.JSONDecodeError, TypeError):
            return []

    if result_type == "shortcuts_on_desktop":
        r = await computer.interface.run_command(
            "python3 -c \"import os,json; "
            "d='/home/user/Desktop'; "
            "res={}; "
            "[res.__setitem__(f, open(os.path.join(d,f)).read()) "
            "for f in os.listdir(d) if f.endswith('.desktop')]; "
            "print(json.dumps(res))\""
        )
        try:
            return json.loads(r.stdout)
        except (json.JSONDecodeError, TypeError):
            return {}

    # --- VLC ---
    if result_type == "vlc_config":
        return await _download_from_container(computer, "/home/user/.config/vlc/vlcrc", cache_dir)

    if result_type == "vlc_playing_info":
        dest = config.get("dest", "vlc_status.xml")
        local_path = _cache_output_path(cache_dir, dest, label="vlc status dest")
        # The only vlcrc writer (dispatch.py) always sets http-password=password,
        # so a single :password attempt is sufficient.
        for curl_cmd in [
            "curl -s --user :password http://localhost:8080/requests/status.xml",
        ]:
            r = await computer.interface.run_command(curl_cmd)
            if r.stdout.strip() and "<root>" in r.stdout:
                with open(local_path, "w") as f:
                    f.write(r.stdout)
                return local_path
        # Fallback: write whatever we got
        if r.stdout.strip():
            with open(local_path, "w") as f:
                f.write(r.stdout)
            return local_path
        return None

    if result_type == "default_video_player":
        from collections import Counter
        extensions = ["mp4", "x-matroska", "webm", "x-msvideo", "quicktime", "x-flv", "mpeg", "ogg"]
        apps = []
        for ext in extensions:
            r = await computer.interface.run_command(f"xdg-mime query default video/{ext}")
            app = r.stdout.strip()
            if app:
                apps.append(app)
        return Counter(apps).most_common(1)[0][0] if apps else "unknown"

    # --- GIMP ---
    if result_type == "gimp_config_file":
        filename = config.get("file_name", config.get("config_file", "gimprc"))
        return await _download_from_container(computer, f"/home/user/.config/GIMP/2.10/{filename}", cache_dir)

    # --- Screen / window / directory (via Flask for correct format) ---
    if result_type == "vm_window_size":
        # OSWorld: env.controller.get_vm_window_size(app_class_name=...)
        # Flask server endpoint: /window_size
        acn = config.get("app_class_name", "")
        r = await computer.interface.run_command(
            f"curl -s -X POST {_SERVER}/window_size -d 'app_class_name={acn}'"
        )
        try:
            data = json.loads(r.stdout)
            return data if data else None
        except (json.JSONDecodeError, TypeError):
            return None

    if result_type == "vm_screen_size":
        # OSWorld: env.controller.get_vm_screen_size() → Flask POST /screen_size
        r = await computer.interface.run_command(
            f"curl -s -X POST {_SERVER}/screen_size"
        )
        try:
            data = json.loads(r.stdout)
            return data if data else None
        except (json.JSONDecodeError, TypeError):
            return None

    if result_type == "list_directory":
        # OSWorld: env.controller.get_vm_directory_tree(path) → Flask POST /list_directory
        # Returns response.json()["directory_tree"]
        path = config.get("path", "/home/user")
        resp = await _flask_post_json(computer, "/list_directory", {"path": path})
        return resp.get("directory_tree", resp)

    if result_type == "vm_wallpaper":
        return await _download_from_container(computer, "/tmp/wallpaper.png", cache_dir,
                                              pre_cmd="curl -s -X POST http://localhost:5000/wallpaper -o /tmp/wallpaper.png")

    # --- VS Code ---
    if result_type == "vscode_config":
        # The eval-helper VS Code extension is not baked into the image.
        # Instead of driving the command palette to make the extension write the
        # answer file, reproduce that file by reading VS Code's own on-disk state
        # (settings.json / workspace.json / state.vscdb). No extension, no palette,
        # no image rebuild. The existing download of `path` + the existing metric
        # then run unchanged. Shared with scalecua verify.py (hashed vscode_* types).
        ext_cmd = config.get("vscode_extension_command", "")
        if ext_cmd:
            await _vscode_write_answer_file(computer, config)
        path = config.get("path", "")
        if path:
            return await _download_from_container(computer, path, cache_dir)
        return None

    # --- Slide media ---
    if result_type == "audio_in_slide":
        return await _get_media_from_slide(computer, config, cache_dir, media_type="audio")

    if result_type == "background_image_in_slide":
        return await _get_media_from_slide(computer, config, cache_dir, media_type="image")

    # --- Accessibility tree (requires Flask server) ---
    if result_type == "accessibility_tree":
        return await _get_at_xml(computer)

    # --- Content from VM file ---
    if result_type == "content_from_vm_file":
        path = config.get("path", "")
        if not path:
            return None
        local = await _download_from_container(computer, path, cache_dir)
        if not local:
            return None
        file_type = config.get("file_type", "")
        file_content = config.get("file_content", "")
        if file_type == "xlsx" and file_content == "last_row":
            import pandas as pd
            df = pd.read_excel(local)
            return df.iloc[-1].astype(str).tolist()
        return local

    # --- Pass-through types ---
    if result_type == "rule":
        return config.get("rules", config)

    if result_type == "cache_file":
        rel_path = config.get("path", "")
        local = _cache_output_path(cache_dir, rel_path, label="cache_file path")
        if os.path.exists(local):
            return local
        if rel_path:
            return await _download_from_container(computer, rel_path, cache_dir)
        return None

    # --- Cloud file (as result) ---
    if result_type == "cloud_file":
        url = config.get("path", "")
        dest = config.get("dest", "")
        if isinstance(url, list):
            gives = set(config.get("gives", [0]))
            dests = config.get("dest", [])
            paths = []
            for i, (u, d) in enumerate(zip(url, dests if isinstance(dests, list) else [dests])):
                local = _download_url(u, cache_dir, d)
                if i in gives:
                    paths.append(local)
            return paths[0] if len(paths) == 1 else paths
        return _download_url(url, cache_dir, dest)

    # --- URL extraction ---
    if result_type == "url_dashPart":
        url = await _get_active_url(computer, config)
        if url is None:
            return None
        try:
            part_index = int(config.get("partIndex", 0))
        except (ValueError, TypeError):
            return None
        parts = url.split("/")
        if part_index >= len(parts):
            return None
        dash_part = parts[part_index]
        if config.get("needDeleteId", False):
            dash_part = dash_part.split("?")[0]
        if config.get("returnType", "string") == "json":
            return {config.get("key", "value"): dash_part}
        return dash_part

    if result_type == "url_path_parse":
        url = await _get_active_url(computer, config)
        if url is None:
            return None
        from urllib.parse import urlparse, unquote
        parsed = urlparse(url)
        path = unquote(parsed.path)
        result = {"mens_clothing": True if "mens-clothing" in path else None}
        path_parts = path.strip("/").split("/")
        key_value_json = {}
        shirts_flag = "shirts" in path
        for i in range(len(path_parts) - 1):
            if "," in path_parts[i] and "," in path_parts[i + 1]:
                keys = [k.strip() for k in path_parts[i].split(",")]
                values = [v.strip() for v in path_parts[i + 1].split(",")]
                for k, v in zip(keys, values):
                    if k == "Price_discount_range":
                        if v and "|" in v:
                            key_value_json[k] = [x.strip() for x in v.split("|")]
                        else:
                            key_value_json[k] = v if v else None
                    else:
                        key_value_json[k] = v if v else None
                    if k == "Product_department" and v and v.lower().startswith("shirt"):
                        shirts_flag = True
                    if k == "Sleeve_length" and v and "short" in v.lower():
                        key_value_json["short_sleeve"] = True
                break
        result["shirts"] = shirts_flag if shirts_flag else None
        for key in config.get("parse_keys", []):
            if key in key_value_json:
                result[key] = key_value_json[key]
        return result

    if result_type == "page_info":
        return await _get_page_info_via_cdp(computer, config.get("url", ""))

    # --- Infeasible ---
    if result_type == "googledrive_file":
        logger.warning("googledrive_file requires Google Drive auth — infeasible")
        return None

    # --- Fallback ---
    logger.warning("Unknown result type '%s'", result_type)
    cmd = config.get("command", "")
    if isinstance(cmd, list):
        cmd = " ".join(cmd)
    if cmd:
        r = await computer.interface.run_command(cmd)
        return r.stdout
    return None


# ---------------------------------------------------------------------------
# Expected getter helpers
# ---------------------------------------------------------------------------

def _resolve_relative_time(config: dict) -> dict:
    """Direct passthrough to OSWorld's get_rule_relativeTime.

    The pip-installed desktop_env package's get_rule_relativeTime
    handles env=None gracefully via _get_vm_now_datetime fallback,
    so we can call it directly. We inject the host's local timezone
    so the resolved 'today' matches what the test runner sees, instead
    of letting OSWorld call get_timezone_from_ip() (slow + can return
    a different tz than expected).

    Upstream bug (folded in locally, upstream is NOT mutated): for
    'first monday four/eight months later', upstream computes
    `next_month = (now.month + N) % 12`, which yields 0 when
    now.month + N is a multiple of 12 (e.g. August + 4 = 12 → 0,
    April + 8 = 12 → 0), so datetime() raises "month must be in 1..12".
    We catch that ValueError and recompute with correct wrap arithmetic
    `((now.month - 1 + N) % 12) + 1` (year carry via `// 12`).
    """
    from desktop_env.evaluators.getters.misc import get_rule_relativeTime
    import copy
    cfg = copy.deepcopy(config)
    cfg.setdefault("rules", {})
    if "timezone" not in cfg["rules"]:
        # Use host's local timezone so absoluteDay matches test runner's
        # "now". Read /etc/timezone, falling back to UTC.
        try:
            with open("/etc/timezone") as f:
                cfg["rules"]["timezone"] = f.read().strip() or "UTC"
        except OSError:
            cfg["rules"]["timezone"] = "UTC"
    try:
        return get_rule_relativeTime(env=None, config=cfg)
    except ValueError as e:
        if "month must be in 1..12" not in str(e):
            raise
        # Fold of the former upstream monkey-patch: recompute the broken
        # 'first monday four/eight months later' cases with correct wrap
        # arithmetic instead of upstream's `(now.month + N) % 12` (which
        # yields month 0). Upstream is left untouched.
        from datetime import datetime as _dt, timedelta as _td
        from desktop_env.evaluators.getters.misc import apply_rules_to_timeFormat
        from_time = cfg["rules"]["relativeTime"].get("from", "")
        if "first monday" not in from_time:
            raise
        now = _dt.now()
        shift = 4 if "four months" in from_time else 8
        next_month = ((now.month - 1 + shift) % 12) + 1
        next_year = now.year + ((now.month - 1 + shift) // 12)
        temp = _dt(next_year, next_month, 1)
        days_to_monday = ((6 - temp.weekday()) + 1) % 7
        absolute_day = temp + _td(days=days_to_monday)
        time_value = cfg["rules"]["expected"]["time"]
        if isinstance(time_value, list):
            cfg["rules"]["expected"]["time"] = [
                apply_rules_to_timeFormat(t, absolute_day) for t in time_value
            ]
        else:
            cfg["rules"]["expected"]["time"] = apply_rules_to_timeFormat(
                time_value, absolute_day
            )
        return cfg["rules"]


async def _get_pdf_from_url(computer, config: dict, cache_dir: str) -> str | None:
    """Print a webpage to PDF via Chrome CDP (matching OSWorld's get_pdf_from_url)."""
    url = config.get("path", "")
    dest = config.get("dest", "expected.pdf")

    script = f'''
import json, urllib.request, socket, struct, os, base64, time

def ws_connect(url):
    url = url.replace("ws://", "")
    host_port, path = url.split("/", 1)
    host, port = host_port.split(":")
    port = int(port)
    sock = socket.create_connection((host, port), timeout=30)
    key = base64.b64encode(os.urandom(16)).decode()
    handshake = f"GET /{{path}} HTTP/1.1\\r\\nHost: {{host}}:{{port}}\\r\\nUpgrade: websocket\\r\\nConnection: Upgrade\\r\\nSec-WebSocket-Key: {{key}}\\r\\nSec-WebSocket-Version: 13\\r\\n\\r\\n"
    sock.sendall(handshake.encode())
    response = b""
    while b"\\r\\n\\r\\n" not in response:
        response += sock.recv(4096)
    return sock

def ws_send(sock, message):
    data = message.encode()
    frame = bytearray([0x81])
    mask_key = os.urandom(4)
    length = len(data)
    if length < 126:
        frame.append(0x80 | length)
    elif length < 65536:
        frame.append(0x80 | 126)
        frame.extend(struct.pack(">H", length))
    else:
        frame.append(0x80 | 127)
        frame.extend(struct.pack(">Q", length))
    frame.extend(mask_key)
    frame.extend(bytearray(b ^ mask_key[i % 4] for i, b in enumerate(data)))
    sock.sendall(frame)

def ws_recv(sock, timeout=60):
    sock.settimeout(timeout)
    header = sock.recv(2)
    if len(header) < 2:
        return ""
    length = header[1] & 0x7f
    if length == 126:
        length = struct.unpack(">H", sock.recv(2))[0]
    elif length == 127:
        length = struct.unpack(">Q", sock.recv(8))[0]
    data = b""
    while len(data) < length:
        chunk = sock.recv(min(65536, length - len(data)))
        if not chunk:
            break
        data += chunk
    return data.decode(errors="replace")

pages = json.loads(urllib.request.urlopen("http://localhost:1337/json").read())
ws_url = None
for p in pages:
    if p.get("type") == "page":
        ws_url = p["webSocketDebuggerUrl"]
        break

if not ws_url:
    print("NO_PAGE")
    exit(1)

sock = ws_connect(ws_url)
ws_send(sock, json.dumps({{"id": 1, "method": "Page.navigate", "params": {{"url": "{url}"}}}}) )
ws_recv(sock)
time.sleep(15)

ws_send(sock, json.dumps({{"id": 2, "method": "Page.printToPDF", "params": {{"marginTop": 0, "marginBottom": 0, "marginLeft": 0, "marginRight": 0, "printBackground": True}}}}))
result = ""
deadline = time.time() + 90
while time.time() < deadline:
    chunk = ws_recv(sock, timeout=60)
    if not chunk:  # socket EOF/closed: ws_recv returns "" -> old `while True` busy-spun at 100% CPU
        break
    result += chunk
    try:
        data = json.loads(result)
        break
    except:
        continue

pdf_data = json.loads(result).get("result", {{}}).get("data", "")
if pdf_data:
    with open("/tmp/expected_pdf.pdf", "wb") as f:
        f.write(base64.b64decode(pdf_data))
    print("OK")
else:
    print("NO_DATA")
sock.close()
'''
    # CDP printToPDF getter. The script is stdlib-only; it writes
    # /tmp/expected_pdf.pdf as the desktop user, which read_bytes then pulls back.
    r = await computer.interface.run_command(
        f"python3 -c {shlex.quote(script)}")
    if "OK" in r.stdout:
        local = await _download_from_container(computer, "/tmp/expected_pdf.pdf", cache_dir)
        # Reap the getter's in-container artifact so it does not linger on the
        # agent-visible filesystem (VM parity: the VM getter leaves nothing behind).
        await computer.interface.run_command(
            "rm -f /tmp/expected_pdf.pdf"
        )
        if local:
            final_path = _cache_output_path(cache_dir, dest, label="pdf dest")
            if local != final_path:
                import shutil
                await asyncio.to_thread(shutil.move, local, final_path)
            return final_path
    return None


async def _get_recreation_page_html(computer, config: dict, cache_dir: str) -> dict | None:
    """Parse HTML content from recreation.gov page via CDP (matching OSWorld getter).

    Returns a dict with "expected" key wrapping the parsed values, because
    check_direct_json_object expects rules["expected"].
    """
    url = await _get_active_url(computer, config)
    if not url:
        return None

    class_name = config.get("class", "")
    order = int(config.get("order", 0))

    script = f'''
import json, requests, time
try:
    from bs4 import BeautifulSoup
except ImportError:
    print(json.dumps({{"error": "bs4_not_installed"}}))
    exit()

pages = requests.get("http://localhost:1337/json").json()
target = None
for p in pages:
    if p.get("type") == "page":
        target = p
        break
if not target:
    print(json.dumps({{"error": "no_page"}}))
    exit()

import websocket
ws = websocket.create_connection(target["webSocketDebuggerUrl"])
ws.send(json.dumps({{"id": 1, "method": "Runtime.evaluate", "params": {{"expression": "document.documentElement.outerHTML"}}}}))
result = json.loads(ws.recv())
ws.close()
html = result.get("result", {{}}).get("result", {{}}).get("value", "")
if not html:
    print(json.dumps({{"error": "empty_html"}}))
    exit()

soup = BeautifulSoup(html, "html.parser")
elems = soup.find_all(class_="{class_name}")
if len(elems) > {order}:
    text = elems[{order}].get_text(strip=True)
    print(json.dumps({{"camp-sortable-column-header": text}}))
else:
    print(json.dumps({{"camp-sortable-column-header": ""}}))
'''
    # bare python3 resolves to the env venv (requests/bs4/websocket) via the harness PATH.
    r = await computer.interface.run_command(
        f"python3 -c {shlex.quote(script)}")
    try:
        parsed = json.loads(r.stdout)
        # Wrap in {"expected": ...} for check_direct_json_object
        return {"expected": parsed}
    except (json.JSONDecodeError, TypeError):
        return None


# ---------------------------------------------------------------------------
# Expected getters
# ---------------------------------------------------------------------------

async def _get_expected(computer, config: dict, cache_dir: str):
    if not config:
        return None
    t = config.get("type", "")
    if t == "cloud_file":
        # _download_url is sync (requests.get); _ensure_csv_exports is sync
        # openpyxl. Both must run in a thread to avoid stalling the 32-way
        # concurrent rollout event loop.
        if config.get("multi", False):
            urls = config.get("path", [])
            dests = config.get("dest", [])
            gives = set(config.get("gives", [0]))
            # Two passes. Pass 1: download EVERY explicit cloud dest first — incl.
            # any authoritative ``*-Sheet1.csv`` gold rendered cloud-side WITH
            # number formats applied. Pass 2: only then synthesize per-sheet CSVs
            # for xlsx golds; _ensure_csv_exports skips a CSV that already exists,
            # so it doesn't clobber a cloud CSV with RAW (format-less) values.
            # (A single interleaved pass false-failed 5 calc number-format
            #  sheet_print tasks — _ensure_csv_exports ran on the xlsx BEFORE the
            #  csv dest downloaded, and _download_url then early-returned on the
            #  pre-created raw file.)
            locals_ = []
            for url, dest in zip(urls, dests):
                locals_.append(await asyncio.to_thread(_download_url, url, cache_dir, dest))
            for local in locals_:
                if local and local.endswith((".xlsx", ".xls")):
                    await asyncio.to_thread(_ensure_csv_exports, local, cache_dir)
            paths = [locals_[i] for i in range(len(locals_)) if i in gives]
            return paths[0] if len(paths) == 1 else paths
        url = config.get("path", "")
        dest = config.get("dest", "")
        if isinstance(url, list):
            results = []
            for u in url:
                results.append(await asyncio.to_thread(_download_url, u, cache_dir))
            return results
        local = await asyncio.to_thread(_download_url, url, cache_dir, dest)
        if local and local.endswith((".xlsx", ".xls")):
            await asyncio.to_thread(_ensure_csv_exports, local, cache_dir)
        # Symmetric LO normalize: agent-saved result files (vm_file) go through
        # `_lo_normalize_pptx`/`_lo_normalize_docx` so XML rewrites done by LO
        # on save (style.name='Normal' → 'LO-Normal', indent promotion, etc.)
        # don't false-fail compare_pptx_files / compare_docx_strict. Cloud-file
        # gold pptx/docx have to receive the same treatment or the result side
        # diverges semantically after normalize while expected stays raw —
        # symptomatic on the GNOME base, where LO's read+write cycle reshuffles
        # slideLayout/slideMaster XML content.
        if local and isinstance(local, str) and local.endswith(".docx") and os.path.isfile(local):
            await _lo_normalize_docx(computer, local)
        elif local and isinstance(local, str) and local.endswith(".pptx") and os.path.isfile(local):
            await _lo_normalize_pptx(computer, local)
        return local
    if t == "rule":
        return config.get("rules", config)
    if t == "list":
        return config.get("list", [])
    # Support various result types in expected (some tasks use this)
    if t in ("vm_command_line", "vm_command_error", "vm_file", "accessibility_tree",
             "content_from_vm_file", "vm_window_size", "vm_screen_size"):
        # `_get_result` already handles OOXML normalization for vm_file-like
        # getters. Do not normalize again here: LO pptx conversion is not
        # perfectly idempotent and a second pass can move untouched placeholders
        # differently from the result side.
        return await _get_result(computer, config, cache_dir)

    # --- rule_relativeTime: resolve time-dependent expected values ---
    if t == "rule_relativeTime":
        return _resolve_relative_time(config)

    # --- pdf_from_url: download a PDF by printing from Chrome via CDP ---
    if t == "pdf_from_url":
        return await _get_pdf_from_url(computer, config, cache_dir)

    # --- gotoRecreationPage_and_get_html_content: parse HTML from active tab ---
    if t == "gotoRecreationPage_and_get_html_content":
        return await _get_recreation_page_html(computer, config, cache_dir)

    # --- info_from_website: use backup values (web scraping not supported) ---
    if t == "info_from_website":
        backups = config.get("backups", [])
        if backups:
            return backups
        return config

    return config


async def _lo_normalize_ooxml(computer, local_path: str, fmt: str) -> None:
    """Run LO headless ``--convert-to <fmt>`` on *local_path* inside the container.

    The file is at *local_path* on the HOST (downloaded from the container into
    a temp cache dir). We push it back into the container, normalize in-place
    with LO headless, then pull it back so the local file reflects LO's XML
    structure (style names, spacing, indent promotion, run merging, etc.).

    Runs AFTER the task (LO interactive already open). LO can run multiple
    headless instances in parallel with a running interactive session as long as
    the input file path differs from any already-open document.  If LO headless
    fails (non-zero exit or output file absent), the original file is left
    unchanged — safe fallback.

    *fmt* is the LO format filter name, e.g. ``"docx"`` or ``"pptx"``.
    """
    import os as _os
    remote_tmp = f"/tmp/_lo_norm_{_os.path.basename(local_path)}"
    with open(local_path, "rb") as fh:
        payload = fh.read()
    await computer.interface.write_bytes(remote_tmp, payload)
    await computer.interface.run_command(
        f"tmpd=$(mktemp -d) && "
        f"DISPLAY=:1 soffice --headless --norestore --nofirststartwizard "
        f"--convert-to {fmt} --outdir \"$tmpd\" '{remote_tmp}' 2>/dev/null && "
        f"[ -f \"$tmpd/$(basename '{remote_tmp}')\" ] && "
        f"cp \"$tmpd/$(basename '{remote_tmp}')\" '{remote_tmp}'; "
        f"rm -rf \"$tmpd\"; true"
    )
    try:
        data = await computer.interface.read_bytes(remote_tmp)
        if data:
            with open(local_path, "wb") as fh:
                fh.write(data)
    except Exception:
        pass
    # §K: reap the temp in /tmp (produced by the desktop-user command above, so a
    # plain rm removes it). Class: applies to BOTH docx and pptx (shared
    # _lo_normalize_ooxml); xlsx is host-side openpyxl so it's exempt.
    await computer.interface.run_command(f"rm -f '{remote_tmp}'")


async def _lo_normalize_docx(computer, local_path: str) -> None:
    await _lo_normalize_ooxml(computer, local_path, "docx")


async def _lo_normalize_pptx(computer, local_path: str) -> None:
    await _lo_normalize_ooxml(computer, local_path, "pptx")


# ---------------------------------------------------------------------------
# File helpers
# ---------------------------------------------------------------------------

def _ensure_csv_exports(xlsx_path: str, cache_dir: str):
    """Generate per-sheet CSV exports from xlsx if they don't exist.

    compare_table's sheet_print rules expect ``{stem}-{SheetName}.csv``.
    """
    try:
        import csv as csv_mod
        import openpyxl
        stem = os.path.splitext(xlsx_path)[0]
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        for sheet_name in wb.sheetnames:
            csv_path = f"{stem}-{sheet_name}.csv"
            if os.path.exists(csv_path):
                continue
            ws = wb[sheet_name]
            with open(csv_path, "w", newline="") as f:
                writer = csv_mod.writer(f)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(row)
            logger.info("Generated CSV export: %s", csv_path)
        wb.close()
    except Exception as e:
        logger.warning("CSV export failed for %s: %s", xlsx_path, e)


# --- VS Code on-disk state reader --------------------------------------------
# One general family keyed on `vscode_extension_command`, reproducing the answer
# file an eval-helper extension would write to config["path"] by instead reading
# VS Code's persisted state. Read logic mirrors the in-repo cuagym reward.py refs
# (state.vscdb ItemTable, workspace.json, settings.json). No extension, no
# palette, no image rebuild. Reused by scalecua/src/osworld/verify.py.
_VSCODE_ONDISK_READER = r'''
import os, sys, json, sqlite3, glob, time, subprocess, urllib.parse

CMD = os.environ["VSCODE_CMD"]
OUT = os.environ["VSCODE_OUT"]
CODE_USER = os.path.expanduser("~/.config/Code/User")
WS_STORAGE = os.path.join(CODE_USER, "workspaceStorage")

# Lazy mementos: flush by gracefully terminating VS Code (SIGTERM, never KILL),
# which writes state.vscdb + storage.json on exit. Folder/settings/explorer are
# written at open-time / are plain config, so they skip the flush.
LAZY = {"GetOpenFile", "OpenFile", "GetOpenTabs", "GetTerminalStatus", "GetSearchResults"}
if CMD in LAZY:
    subprocess.call(["pkill", "-TERM", "-f", "code .*--no-sandbox"])
    for _ in range(20):
        if subprocess.call(["pgrep", "-f", "code .*--no-sandbox"],
                           stdout=subprocess.DEVNULL) != 0:
            break
        time.sleep(0.5)
    time.sleep(1.0)

def uri_to_path(uri):
    if not isinstance(uri, str) or not uri.startswith("file://"):
        return None
    return urllib.parse.unquote(urllib.parse.urlparse(uri).path)

def ws_dirs():
    ds = [d for d in glob.glob(os.path.join(WS_STORAGE, "*")) if os.path.isdir(d)]
    ds.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return ds

def resolve_folder():
    for d in ws_dirs():
        wj = os.path.join(d, "workspace.json")
        if os.path.exists(wj):
            try:
                p = uri_to_path(json.load(open(wj)).get("folder"))
            except Exception:
                p = None
            if p:
                return p
    sj = os.path.join(CODE_USER, "globalStorage", "storage.json")
    if os.path.exists(sj):
        try:
            data = json.load(open(sj))
        except Exception:
            data = {}
        p = uri_to_path((data.get("windowsState", {}).get("lastActiveWindow", {}) or {}).get("folder"))
        if p:
            return p
        hist = data.get("history", {})
        entries = hist.get("recentlyOpenedPathsList", {}).get("entries", []) if isinstance(hist, dict) else []
        for e in entries:
            p = uri_to_path(e.get("folderUri"))
            if p:
                return p
    return None

def state_db():
    for d in ws_dirs():
        db = os.path.join(d, "state.vscdb")
        if os.path.exists(db):
            return db
    return None

def db_item(key):
    db = state_db()
    if not db:
        return None
    try:
        conn = sqlite3.connect(db)
        row = conn.execute("SELECT value FROM ItemTable WHERE key = ?", (key,)).fetchone()
        conn.close()
        return json.loads(row[0]) if row else None
    except Exception:
        return None

def editor_leaves():
    st = db_item("memento/workbench.parts.editor") or {}
    root = st.get("editorpart.state", {}).get("serializedGrid", {}).get("root", {})
    out = []
    def walk(n):
        if n.get("type") == "leaf":
            out.append(n.get("data", {}))
        elif n.get("type") == "branch":
            for c in n.get("data", []):
                walk(c)
    walk(root)
    return out

def editor_paths():
    paths = []
    for grp in editor_leaves():
        for ed in grp.get("editors", []):
            try:
                val = json.loads(ed.get("value", "{}"))
            except Exception:
                val = {}
            res = val.get("resourceJSON", {}) or {}
            p = res.get("fsPath") or res.get("path") or uri_to_path(val.get("resource", ""))
            if p:
                paths.append(p)
    return paths

def write(text):
    with open(OUT, "w") as f:
        f.write(text)

if CMD in ("OpenProject", "GetWorkspace"):
    write(resolve_folder() or "")
elif CMD == "GetExplorer":
    folder = resolve_folder()
    write("\n".join(sorted(os.listdir(folder))) if folder and os.path.isdir(folder) else "")
elif CMD == "WorkspaceFiles":
    folder = resolve_folder()
    files = []
    if folder and os.path.isdir(folder):
        for r, _d, fs in os.walk(folder):
            files.extend(os.path.join(r, fn) for fn in fs)
    write("\n".join(files))
elif CMD in ("GetOpenFile", "OpenFile"):
    ps = editor_paths()
    write(ps[0] if ps else "")            # single-file tasks: first == active
elif CMD == "GetOpenTabs":
    write("\n".join(editor_paths()))
elif CMD == "GetSettings":
    settings = {}
    uj = os.path.join(CODE_USER, "settings.json")
    if os.path.exists(uj):
        try: settings.update(json.load(open(uj)))
        except Exception: pass
    folder = resolve_folder()
    if folder:
        wj = os.path.join(folder, ".vscode", "settings.json")
        if os.path.exists(wj):
            try: settings.update(json.load(open(wj)))   # workspace overrides user
            except Exception: pass
    write(json.dumps(settings))
elif CMD == "GetTerminalStatus":
    layout = db_item("terminal.integrated.layoutInfo") or {}
    write("open" if layout.get("tabs") else "closed")
elif CMD == "GetSearchResults":
    mem = db_item("memento/workbench.view.search") or {}
    q = ""
    for k in ("query", "contentPattern", "searchQuery"):
        v = mem.get(k)
        if isinstance(v, dict):
            v = v.get("pattern") or v.get("value")
        if v:
            q = v; break
    lines = [q]
    folder = resolve_folder()
    if q and folder and os.path.isdir(folder):
        try:
            r = subprocess.run(["grep", "-rn", q, folder],
                               capture_output=True, text=True, timeout=15)
            lines.append(r.stdout)
        except Exception:
            pass
    write("\n".join(lines))
# else: e.g. "Preferences: Open Settings (JSON)" — path is the settings file
# itself; it already exists on disk, so write nothing and let the caller download.
'''


async def _vscode_write_answer_file(computer, config: dict) -> None:
    """Reproduce the eval-helper extension's answer file at ``config["path"]``
    by reading VS Code's on-disk state. Dispatch is on
    ``vscode_extension_command`` (the 10 getter families), never per-task.
    Idempotent; safe to call more than once.
    """
    ext_cmd = config.get("vscode_extension_command", "")
    out_path = config.get("path", "")
    if not ext_cmd or not out_path:
        return
    r = await computer.interface.run_command(
        f"VSCODE_CMD={shlex.quote(ext_cmd)} VSCODE_OUT={shlex.quote(out_path)} "
        f"python3 -c {shlex.quote(_VSCODE_ONDISK_READER)}"
    )
    if getattr(r, "returncode", 0) != 0:
        logger.warning("vscode on-disk reader (%s) failed: %s", ext_cmd, (r.stderr or "")[:200])


async def _download_from_container(computer, remote_path: str, cache_dir: str, pre_cmd: str = "") -> str | None:
    """Download file from container via the exec-stdio ``read_bytes`` op.

    Symmetric with host_push (host→container) which uses
    ``interface.write_bytes`` over the SAME exec-stdio session — no host
    port published, works in rootless docker and isolated container
    networks where bridge-IP routing is unavailable.

    Single-line base64 payload, bounded by the host-side ``_STREAM_LIMIT``
    (64 MiB) in ``lite/gym/sandbox/exec_stdio/client.py``. Falls back to
    ``run_command("base64 -w0 <path>")`` (also exec-stdio) if the
    primary op raises a transient error; final return is ``None`` if
    both fail.
    """
    if not remote_path:
        return None
    filename = os.path.basename(remote_path)
    local_path = _cache_output_path(cache_dir, filename, label="download filename")

    if pre_cmd:
        await computer.interface.run_command(pre_cmd)

    # Primary: read_bytes RPC (symmetric with host_push's write_bytes).
    try:
        data = await computer.interface.read_bytes(remote_path)
        if data:
            with open(local_path, "wb") as f:
                f.write(data)
            return local_path
    except Exception as e:
        logger.warning("read_bytes failed for %s: %s", remote_path, e)

    # Fallback: base64-encode inside container, decode on host. Useful when
    # the primary read_bytes op hits a transient session error — the
    # base64 path rides the same exec-stdio session through run_command,
    # which has a different retry surface (one extra round-trip is cheap).
    import shlex as _shlex
    try:
        r = await computer.interface.run_command(
            f"base64 -w0 {_shlex.quote(remote_path)}"
        )
        b64 = r.stdout.strip()
        if b64:
            data = base64.b64decode(b64)
            if data:
                with open(local_path, "wb") as f:
                    f.write(data)
                return local_path
    except Exception as e:
        logger.warning("base64 download failed for %s: %s", remote_path, e)

    return None


def _download_url(url: str, cache_dir: str, dest: str = "") -> str | None:
    if not url:
        return None
    if dest:
        local = _cache_output_path(cache_dir, dest, label="download dest")
    else:
        h = hashlib.md5(url.encode()).hexdigest()[:12]
        local = _cache_output_path(
            cache_dir,
            f"{h}_{os.path.basename(url)}",
            label="download filename",
        )
    if os.path.exists(local) and os.path.getsize(local) > 0:
        return local
    import time
    delay = 3
    for attempt in range(1, 6):
        try:
            tmp = local + ".tmp"
            urllib.request.urlretrieve(url, tmp)
            if os.path.getsize(tmp) > 0:
                os.replace(tmp, local)
                return local
            os.unlink(tmp)
        except Exception as e:
            logger.warning("Download attempt %d/5 failed %s: %s", attempt, url, e)
        if attempt < 5:
            time.sleep(delay)
            delay = min(delay * 2, 30)
    logger.warning("All 5 download attempts failed: %s", url)
    return None
