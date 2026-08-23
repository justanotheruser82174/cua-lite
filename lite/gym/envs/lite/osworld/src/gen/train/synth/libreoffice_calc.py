"""LibreOffice Calc synth generator (Track A — host-heredoc design).

Per AGENTS.md / libreoffice_calc.md: each row encodes BOTH the source xlsx
AND the gold xlsx as openpyxl heredocs in `pre_config_steps`. Every gold-py
ENDS WITH `_LO_NORMALIZE_TAIL` (Hard Constraint #13: re-save via soffice
headless to populate `<v>` cached values + normalize datetime serialization
— calc-only requirement). Oracle = `cp <expected> <src>`. Postconfig =
`LO_SAVE_POSTCONFIG`. `oracle_after_postconfig=True` (so LO_SAVE_POSTCONFIG runs
BEFORE the cp-oracle's effect would be flushed by LO's own save).
`open_command=["libreoffice","--calc",<src>]` per row (calc has no
DOMAIN_DEFAULT_OPEN entry).

    Cartesian generator layout:
        The file-end section is the only template factory: file × tasks × params
        via `File` / `Param` / `FileTask` dataclasses. 90 File instances drive
        180 FileTask entries; cap-2×2 emits ≤180 templates. Other domains mirror
        this pattern.

Usage:
    uv run python -m lite.gym.envs.lite.osworld.src.gen.train \\
        --track synth --domain libreoffice_calc
"""

from __future__ import annotations

import random
import textwrap
from dataclasses import dataclass, field
from typing import Callable

from lite.gym.envs.lite.osworld.src.gen.common import (
    LO_SAVE_POSTCONFIG,
)
from lite.gym.envs.lite.osworld.src.gen.train.synth._utils import (
    SynthTemplate,
    _stage_asset,
    _stable_hash,
)


# ---------------------------------------------------------------------------
# LO post-launch settle steps (validation — Trigger F race fix). Owned by the
# LO domain per the per-domain setup convention; previously lived in
# common.py and was applied via a `template.domain == "libreoffice_*"`
# conditional.
# ---------------------------------------------------------------------------
_LO_POSTLAUNCH_SETTLE: list[dict] = [
    {"type": "sleep", "parameters": {"seconds": 6}},
    {"type": "activate_window", "parameters": {"window_name": "LibreOffice"}},
]


# ---------------------------------------------------------------------------
# Mandatory tail (constraint #13). Verbatim copy from
# perturb/libreoffice_calc.py:_LO_NORMALIZE_TAIL.
# ---------------------------------------------------------------------------
_LO_NORMALIZE_TAIL = """
import os as _os, subprocess as _sp, tempfile as _tf, shutil as _sh
_td = _tf.mkdtemp()
try:
    _sp.run(["soffice", "--headless", "--norestore",
             "--nofirststartwizard", "--convert-to", "xlsx",
             "--outdir", _td, {expected_path!r}],
            capture_output=True, env={{**_os.environ, "DISPLAY": ":1"}}, timeout=120)
    _conv = _os.path.join(_td, _os.path.basename({expected_path!r}))
    if _os.path.exists(_conv):
        _sh.copy(_conv, {expected_path!r})
finally:
    _sh.rmtree(_td, ignore_errors=True)
"""


# ---------------------------------------------------------------------------
# Common heredoc helpers
# ---------------------------------------------------------------------------


def _py_step(py_code: str) -> dict:
    """Wrap a python code string in an `execute` config step using a heredoc."""
    return {
        "type": "execute",
        "parameters": {
            "command": f"python3 << 'PYEOF'\n{py_code}\nPYEOF",
            "shell": True,
        },
    }


def _cp_oracle(expected_path: str, sink_path: str) -> list[dict]:
    return [{
        "type": "execute",
        "parameters": {
            "command": f"cp '{expected_path}' '{sink_path}'",
            "shell": True,
        },
    }]


def _pick_compare_func(rules: list[dict]) -> str:
    """Route a rule set to the right compare_table variant. Chart rules need the
    series-ref-agnostic `compare_calc_chart_type`; `number_format`-style rules
    need the locale-tolerant `compare_table_numfmt_tolerant`; all else flows
    through upstream `compare_table`."""
    if any(r.get("type") == "chart" for r in rules):
        return "compare_calc_chart_type"
    if any(r.get("type") == "style" and (r.get("props") or []) == ["number_format"]
           for r in rules):
        return "compare_table_numfmt_tolerant"
    return "compare_table"


def _eval_compare_table(sink: str, expected: str, rules: list[dict]) -> dict:
    # deferred queue ( calc chart cluster:
    # f_calc_55__chart_box_office_bar + f_calc_77__chart_ticket_price). When
    # ANY rule in the rule list has type=='chart', route through the local
    # `compare_calc_chart_type` override that compares chart info dicts
    # WITHOUT the series-reference key (which LO normalizes — e.g. drops
    # the sheet prefix or expands the range — between gold and result even
    # when chart_props=['type'] match). Non-chart rules still flow through
    # upstream compare_table verbatim.
    #
    # validation Bug 2: freeze rules are also routed out, but to the
    # local `check_xlsx_freeze_pane` helper. Upstream's `freeze` rule
    # compares result vs expected `ws.freeze_panes`, but
    # `_LO_NORMALIZE_TAIL` strips `<pane state="frozen"/>` from the gold
    # — so both sides read None and the rule trivially passes before the
    # agent acts. The helper probes the agent's xlsx directly against
    # the literal expected cell (carried on the rule dict as
    # `freeze_cell`).
    sink_result = {"type": "vm_file", "path": sink,
                   "dest": sink.split("/")[-1]}
    expected_result = {"type": "vm_file", "path": expected,
                       "dest": "expected_file"}

    freeze_rules = [r for r in rules if r.get("type") == "freeze"]
    remaining = [r for r in rules if r.get("type") != "freeze"]

    if freeze_rules:
        # Build a multi-func evaluator: each freeze rule gets a
        # `check_xlsx_freeze_pane` slot keyed off the result file, plus
        # one upstream compare_table slot for the remaining rules. All
        # slots use `conj="and"` (the default).
        funcs: list[str] = []
        result_list: list[dict] = []
        expected_list: list[dict] = []
        options_list: list[dict] = []
        for r in freeze_rules:
            cell = r.get("freeze_cell", "A2")
            col = "".join(c for c in cell if c.isalpha()) or "A"
            row = "".join(c for c in cell if c.isdigit()) or "2"
            funcs.append("check_xlsx_freeze_pane")
            result_list.append(sink_result)
            expected_list.append({})  # helper takes result_path only
            options_list.append({
                "expected_first_row": int(row),
                "expected_first_col_letter": col,
            })
        if remaining:
            funcs.append(_pick_compare_func(remaining))
            result_list.append(sink_result)
            expected_list.append(expected_result)
            options_list.append({"rules": remaining})
        return {
            "func": funcs,
            "result": result_list,
            "expected": expected_list,
            "options": options_list,
            "postconfig": LO_SAVE_POSTCONFIG,
        }

    func = _pick_compare_func(rules)
    return {
        "func": func,
        "result": sink_result,
        "expected": expected_result,
        "options": {"rules": rules},
        "postconfig": LO_SAVE_POSTCONFIG,
    }


# ---------------------------------------------------------------------------
# Source-xlsx builders (each returns a python source-code string that
# writes the source xlsx when executed via `python3 << PYEOF`).
#
# All builders accept (src_path, rng_seed) — rng_seed lets us add minor
# scalar variations per-seed without changing structural shape.
# ---------------------------------------------------------------------------


def _src_movies(src_path: str, seed: int) -> str:
    """30×5 movie list: Title / Year / Director / Rating / Genre.

    Curated mix of 30 widely-recognized films across 6 decades (1940s-2020s)
    spanning drama, sci-fi, action, animation, romance, thriller. Real-world
    titles + directors so the source xlsx looks like a genuine film catalog
    rather than a placeholder shape.

    validation (2026-05-10): rows are deterministically scrambled (seed+template
    hash) so the source is NOT already Year-ascending — otherwise the
    `sort_movies_year_asc` task trivially passes its `compare_table sheet_data`
    eval before the agent acts. The scramble keeps the asset realistic (still
    1 film per year, no value-tampering) but breaks the canonical Year order.
    """
    return textwrap.dedent(f"""\
        import openpyxl, random
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Movies"
        ws.append(["Title", "Year", "Director", "Rating", "Genre"])
        rows = [
            ("Casablanca", 1942, "Michael Curtiz", 8.5, "Drama"),
            ("Singin' in the Rain", 1952, "Stanley Donen", 8.3, "Musical"),
            ("12 Angry Men", 1957, "Sidney Lumet", 9.0, "Drama"),
            ("Psycho", 1960, "Alfred Hitchcock", 8.5, "Thriller"),
            ("2001: A Space Odyssey", 1968, "Stanley Kubrick", 8.3, "Sci-Fi"),
            ("The Godfather", 1972, "Francis Ford Coppola", 9.2, "Crime"),
            ("Jaws", 1975, "Steven Spielberg", 8.1, "Thriller"),
            ("Star Wars: A New Hope", 1977, "George Lucas", 8.6, "Sci-Fi"),
            ("Alien", 1979, "Ridley Scott", 8.5, "Sci-Fi"),
            ("Blade Runner", 1982, "Ridley Scott", 8.1, "Sci-Fi"),
            ("Back to the Future", 1985, "Robert Zemeckis", 8.5, "Adventure"),
            ("Die Hard", 1988, "John McTiernan", 8.2, "Action"),
            ("Goodfellas", 1990, "Martin Scorsese", 8.7, "Crime"),
            ("Jurassic Park", 1993, "Steven Spielberg", 8.2, "Adventure"),
            ("Pulp Fiction", 1994, "Quentin Tarantino", 8.9, "Crime"),
            ("Toy Story", 1995, "John Lasseter", 8.3, "Animation"),
            ("Titanic", 1997, "James Cameron", 7.9, "Romance"),
            ("The Matrix", 1999, "Lana Wachowski", 8.7, "Sci-Fi"),
            ("Spirited Away", 2001, "Hayao Miyazaki", 8.6, "Animation"),
            ("The Dark Knight", 2008, "Christopher Nolan", 9.0, "Action"),
            ("Up", 2009, "Pete Docter", 8.3, "Animation"),
            ("Inception", 2010, "Christopher Nolan", 8.8, "Sci-Fi"),
            ("The Social Network", 2010, "David Fincher", 7.8, "Drama"),
            ("Mad Max: Fury Road", 2015, "George Miller", 8.1, "Action"),
            ("Spotlight", 2015, "Tom McCarthy", 8.1, "Drama"),
            ("La La Land", 2016, "Damien Chazelle", 8.0, "Romance"),
            ("Get Out", 2017, "Jordan Peele", 7.7, "Thriller"),
            ("Parasite", 2019, "Bong Joon-ho", 8.6, "Drama"),
            ("Dune", 2021, "Denis Villeneuve", 8.0, "Sci-Fi"),
            ("Everything Everywhere All at Once", 2022, "Daniel Kwan", 7.8, "Adventure"),
        ]
        random.Random({seed} ^ 0xC0FFEE).shuffle(rows)
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_expenses(src_path: str, seed: int) -> str:
    """24×4 household expense ledger: Month / Category / Amount / Notes.

    Realistic monthly household budget across 6 categories
    (Rent / Utilities / Groceries / Transportation / Insurance / Healthcare)
    over 4 months Jan-Apr. Dollar amounts modeled on a typical US single-
    earner household: rent ~2400, utilities ~140-220 (winter spike),
    groceries ~520-640, etc.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"
        ws.append(["Month", "Category", "Amount", "Notes"])
        rows = [
            ("January", "Rent",           2400.00, "monthly lease"),
            ("January", "Utilities",       218.45, "heating spike"),
            ("January", "Groceries",       612.30, "Whole Foods + Trader Joe's"),
            ("January", "Transportation",  142.80, "gas + transit pass"),
            ("January", "Insurance",       185.00, "auto premium"),
            ("January", "Healthcare",       45.00, "copay urgent care"),
            ("February", "Rent",          2400.00, "monthly lease"),
            ("February", "Utilities",      196.20, "heating tapering"),
            ("February", "Groceries",      548.95, "shorter month"),
            ("February", "Transportation", 168.40, "valentines weekend trip"),
            ("February", "Insurance",      185.00, "auto premium"),
            ("February", "Healthcare",     120.00, "dental cleaning"),
            ("March", "Rent",             2400.00, "monthly lease"),
            ("March", "Utilities",         152.10, "milder weather"),
            ("March", "Groceries",         634.20, "spring restock"),
            ("March", "Transportation",    156.95, "regular commute"),
            ("March", "Insurance",         185.00, "auto premium"),
            ("March", "Healthcare",         85.00, "annual physical copay"),
            ("April", "Rent",             2400.00, "monthly lease"),
            ("April", "Utilities",         138.75, "AC not yet on"),
            ("April", "Groceries",         586.40, "Easter + farmers market"),
            ("April", "Transportation",    175.20, "weekend road trip"),
            ("April", "Insurance",         185.00, "auto premium"),
            ("April", "Healthcare",         60.00, "prescription refill"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_sales(src_path: str, seed: int) -> str:
    """12-month retail sales: Month / Sales / COGS / Profit / Region / Channel.

    Realistic seasonality: low Q1 trough (post-holiday slump ~12-18K),
    Q2 mid-range, Q3 back-to-school bump, Q4 holiday spike (Nov/Dec
    >60K). COGS = 42-48% of Sales (typical retail margin). Region
    rotates Northeast/Midwest/South/West; channel rotates
    Online/Retail/Wholesale.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append(["Month", "Sales", "COGS", "Profit", "Region", "Channel"])
        # Seasonal pattern: Q1 trough (12-18K), Q2 mid (18-26K),
        # Q3 back-to-school (24-32K), Q4 holiday spike (38-72K).
        rows = [
            ("January",   13420, 6307, 7113, "Northeast", "Online"),
            ("February",  11865, 5577, 6288, "Midwest",   "Retail"),
            ("March",     16240, 7470, 8770, "South",     "Wholesale"),
            ("April",     19880, 9145, 10735, "West",     "Online"),
            ("May",       22130, 10182, 11948, "Northeast", "Retail"),
            ("June",      24560, 11553, 13007, "Midwest",  "Wholesale"),
            ("July",      21980, 10331, 11649, "South",    "Online"),
            ("August",    27140, 12756, 14384, "West",     "Retail"),
            ("September", 31420, 14767, 16653, "Northeast", "Wholesale"),
            ("October",   38760, 18217, 20543, "Midwest",  "Online"),
            ("November",  56480, 26545, 29935, "South",    "Retail"),
            ("December",  72340, 33999, 38341, "West",     "Online"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_inventory(src_path: str, seed: int) -> str:
    """8×3 inventory P&L: Product / Revenue / Margin.

    Eight named SKUs with realistic year-revenue ints (45K-220K range)
    and realistic gross-margin decimals (0.18-0.41). Product names
    follow a typical industrial-component catalog (Widget A, Sprocket B,
    etc.) so the source xlsx reads as a real per-line P&L extract.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"
        ws.append(["Product", "Revenue", "Margin"])
        rows = [
            ("Widget A",      45000, 0.32),
            ("Sprocket B",    78000, 0.41),
            ("Bolt C",       128400, 0.18),
            ("Hinge D",       62300, 0.27),
            ("Bearing E",    219800, 0.36),
            ("Gasket F",      89500, 0.24),
            ("Filter G",     156200, 0.31),
            ("Coupler H",    103700, 0.29),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_phonics_titles(src_path: str, seed: int) -> str:
    """22×4 mangled film titles: Idx / Raw / Year / Clean (empty)."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PhonicsTitles"
        ws.append(["Idx", "Raw", "Year", "Clean"])
        rows = [
            (1,  "  the QUIET garden  ", 1998, None),
            (2,  "echoes of  TOMORROW",  2014, None),
            (3,  "  Salt And BREAD ",     2007, None),
            (4,  "river OF  glass",       2001, None),
            (5,  "  blue HOUR ",           2019, None),
            (6,  "the   crimson SAIL",     1995, None),
            (7,  "WINDOW gardens",         2003, None),
            (8,  "moonless  NIGHT",        2011, None),
            (9,  " Lighthouse YEAR  ",     2008, None),
            (10, "stone  LANTERN ",        1996, None),
            (11, "  ASHES of october",     2017, None),
            (12, "the   gold COAST",       2009, None),
            (13, "saltwood  mile",         2023, None),
            (14, " quiet MILE ",           1994, None),
            (15, "  brass  WINGS",         2016, None),
            (16, "TIDE of storms",         2012, None),
            (17, "the  unwritten ",         2002, None),
            (18, "Glass DOORS",             1997, None),
            (19, " halcyon  drift ",        2006, None),
            (20, "  AURORA tide ",          2000, None),
            (21, "voltage ",                 2021, None),
            (22, " quiver",                  2004, None),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_pnl_2col(src_path: str, seed: int) -> str:
    """1 sheet 24×3: Month / Revenue / Expenses (for sheet2 Forecast copy)."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PnL"
        ws.append(["Month", "Revenue", "Expenses"])
        months = ["2024-01","2024-02","2024-03","2024-04","2024-05","2024-06",
                  "2024-07","2024-08","2024-09","2024-10","2024-11","2024-12",
                  "2025-01","2025-02","2025-03","2025-04","2025-05","2025-06",
                  "2025-07","2025-08","2025-09","2025-10","2025-11","2025-12"]
        rev = [82400, 91200, 105800, 88300, 116400, 122800, 109500, 113700, 134600,
               142800, 158400, 171300, 92100, 99500, 116900, 102300, 128400, 137200,
               118800, 126400, 148700, 159200, 171800, 188400]
        exp = [54200, 58100, 64900, 56400, 71500, 75200, 69100, 70600, 81200,
               85100, 93600, 102200, 60800, 64300, 71700, 65500, 78600, 84900,
               74200, 79100, 89600, 96200, 102400, 113800]
        for m, r, e in zip(months, rev, exp):
            ws.append([m, r, e])
        wb.save({src_path!r})
        """)


def _src_sales_rep_quarter(src_path: str, seed: int) -> str:
    """1 sheet 7-col: Rep / Q1..Q6 (10 rows). For Sheet2 'Summary' rollup."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reps"
        ws.append(["Rep", "Q1", "Q2", "Q3", "Q4", "Q5", "Q6"])
        rows = [
            ("Alvarez",  18400, 19200, 21000, 22300, 24100, 25600),
            ("Brennan",  15200, 16700, 17800, 18400, 19200, 20100),
            ("Caldwell", 22800, 24500, 26100, 27300, 28800, 30200),
            ("Diop",     19500, 20800, 22300, 23400, 25100, 26700),
            ("Eklund",   17200, 18400, 19600, 20800, 22100, 23400),
            ("Fontaine", 21300, 22800, 24500, 25700, 27300, 28900),
            ("Gupta",    20100, 21500, 22900, 24100, 25600, 27100),
            ("Hartmann", 18800, 20100, 21500, 22700, 24100, 25600),
            ("Inoue",    16400, 17600, 18800, 19900, 21100, 22300),
            ("Janos",    23700, 25400, 27200, 28500, 30300, 32000),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_orders_north(src_path: str, seed: int) -> str:
    """1 sheet 24×4 customer order log; ~10 North rows + decoys."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append(["OrderID", "Customer", "Region", "Amount"])
        rows = [
            ("ORD-1001", "Acme Co",         "North", 1240.50),
            ("ORD-1002", "Bayshore Ltd",    "South",  812.00),
            ("ORD-1003", "Cypress Group",   "North", 1980.25),
            ("ORD-1004", "Driftwood Inc",   "West",   442.75),
            ("ORD-1005", "Evergreen Sys",   "North",  690.00),
            ("ORD-1006", "Forge & Co",      "East",  1124.40),
            ("ORD-1007", "Granite Hold",    "North", 1488.30),
            ("ORD-1008", "Harbor Freight",  "South",  376.10),
            ("ORD-1009", "Ironwood Bros",   "North", 2104.80),
            ("ORD-1010", "Juniper Stk",     "West",   915.55),
            ("ORD-1011", "Kestrel Inc",     "North", 1342.65),
            ("ORD-1012", "Linden Mfg",      "East",   780.20),
            ("ORD-1013", "Mariner Co",      "North",  588.40),
            ("ORD-1014", "Northpoint Ltd",  "South",  995.70),
            ("ORD-1015", "Oakridge Sys",    "North", 1772.10),
            ("ORD-1016", "Pinecrest LLC",   "West",   449.85),
            ("ORD-1017", "Quartz Hold",     "North",  834.95),
            ("ORD-1018", "Riverbend Corp",  "East",  1245.00),
            ("ORD-1019", "Saltwood Inc",    "North",  697.60),
            ("ORD-1020", "Tidewater Co",    "South",  528.30),
            ("ORD-1021", "Underwood Ltd",   "North", 1925.45),
            ("ORD-1022", "Vermillion LLC",  "West",   620.10),
            ("ORD-1023", "Westbrook Sys",   "East",   876.80),
            ("ORD-1024", "Yarrow Co",       "South",  704.25),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_loan_5col(src_path: str, seed: int) -> str:
    """5-col loan amortization (Loan# / Principal / Term / Rate / Periods) 4 rows."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Loans"
        ws.append(["Loan#", "Principal", "Term", "Rate", "Periods"])
        rows = [
            ("L-2401", 250000, 30, 0.0625, 360),
            ("L-2402", 180000, 15, 0.0550, 180),
            ("L-2403", 425000, 30, 0.0700, 360),
            ("L-2404",  92000, 10, 0.0475, 120),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_questionnaire(src_path: str, seed: int) -> str:
    """4-col questionnaire (Respondents / Sex / Civil Status / Highest Educ)."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Survey"
        ws.append(["Respondents", "Sex", "Civil Status", "Highest Educ Attainment"])
        sx = ["F","M"]
        cs = ["Single","Married","Widowed","Separated"]
        ed = ["High School","College","Bachelor","Master","Doctorate"]
        rows = []
        for i in range(1, 41):
            rows.append((f"R-{{i:03d}}", sx[i % 2], cs[i % 4], ed[i % 5]))
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_bus_schedule(src_path: str, seed: int) -> str:
    """6-col bus schedule."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Schedule"
        ws.append(["Route","Origin","Dest","Depart","Arrive","Duration"])
        rows = [
            ("R-12","Downtown","Airport",   "06:15","07:05","00:50"),
            ("R-12","Downtown","Airport",   "07:00","07:50","00:50"),
            ("R-22","Westside","Mainplaza", "06:30","07:10","00:40"),
            ("R-22","Westside","Mainplaza", "07:15","07:55","00:40"),
            ("R-37","Eastpark","Med Center","06:45","07:25","00:40"),
            ("R-37","Eastpark","Med Center","07:30","08:10","00:40"),
            ("R-44","Hilltop", "Library",   "07:00","07:35","00:35"),
            ("R-44","Hilltop", "Library",   "07:50","08:25","00:35"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_tournament(src_path: str, seed: int) -> str:
    """4-col tournament (Match / Team A / Team B / Score) 8 matches."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bracket"
        ws.append(["Match","Team A","Team B","Score"])
        rows = [
            ("QF1","Pumas",   "Eagles",   "2-1"),
            ("QF2","Foxes",   "Wolves",   "0-3"),
            ("QF3","Hawks",   "Bears",    "1-1"),
            ("QF4","Tigers",  "Lions",    "2-2"),
            ("QF5","Stags",   "Otters",   "4-0"),
            ("QF6","Marlins", "Sharks",   "1-2"),
            ("QF7","Bulls",   "Stallions","3-3"),
            ("QF8","Coyotes", "Lynx",     "2-0"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_gradebook(src_path: str, seed: int) -> str:
    """2-col gradebook Name / Score — 12 students with realistic full names.

    Score distribution roughly mean ~70 stdev ~15 with a mix of pass
    (>=60) and fail (<60), so the eval's pass/fail binary fill applies
    non-trivially to both categories. Names are drawn from a culturally
    diverse roster (Wong / Chen / Patel / Garcia / Okafor / Kim / etc.).
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grades"
        ws.append(["Name","Score"])
        rows = [
            ("Alice Wong",       88),
            ("Marcus Chen",      72),
            ("Priya Patel",      94),
            ("Diego Garcia",     58),
            ("Aisha Okafor",     81),
            ("Jordan Kim",       45),
            ("Sofia Hernandez",  76),
            ("Liam O'Connor",    52),
            ("Yuki Tanaka",      68),
            ("Noah Williams",    91),
            ("Zara Ahmed",       55),
            ("Ethan Brooks",     79),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_safety_inspection(src_path: str, seed: int) -> str:
    """2-col safety-inspection (Item / Pass-Fail) 15 rows."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inspection"
        ws.append(["Item","Status"])
        rows = [
            ("Fire extinguisher A1", "Pass"),
            ("Fire extinguisher B2", "Fail"),
            ("Emergency exit east",  "Pass"),
            ("Emergency exit west",  "Pass"),
            ("Smoke detector L1",    "Fail"),
            ("Smoke detector L2",    "Pass"),
            ("First-aid kit floor 1","Pass"),
            ("First-aid kit floor 2","Fail"),
            ("Sprinkler zone A",     "Pass"),
            ("Sprinkler zone B",     "Pass"),
            ("Eyewash station",      "Fail"),
            ("Hazmat cabinet",       "Pass"),
            ("Spill kit lobby",      "Pass"),
            ("Stairwell lighting",   "Pass"),
            ("Roof access alarm",    "Fail"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_user_emails(src_path: str, seed: int) -> str:
    """2-col user-list (UserId / Email mixed-case+spaces, 30 rows). Empty 'Clean' col.

    Rows are stored in swap-adjacent-pairs order (U-002,U-001,U-004,U-003,…) so
    neither ascending nor descending sort of UserId is a no-op (trivial_pass guard).
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Users"
        ws.append(["UserId","Email","Clean"])
        rows = [
            ("U-002"," BOB_smith@MAIL.org",      None),
            ("U-001","  Alice.Doe@Example.COM ", None),
            ("U-004"," Dan.Reed@example.COM",     None),
            ("U-003","Carol-Jones@Site.NET ",     None),
            ("U-006","FRANK.b@SITE.NET",            None),
            ("U-005","   Erin@ Mail.org ",         None),
            ("U-008","Henry.Q @MAIL.ORG ",          None),
            ("U-007","  Gail@Example.COM",          None),
            ("U-010"," JIM@example.COM ",            None),
            ("U-009","Iris.K@Site.Net",              None),
            ("U-012","Liam@Site.NET ",                None),
            ("U-011","Kara.s@MAIL.org",              None),
            ("U-014","NICO@MAIL.ORG ",                 None),
            ("U-013"," Mia.X@example.COM",            None),
            ("U-016"," Pia@example.COM",                None),
            ("U-015","Owen@Site.net ",                  None),
            ("U-018","Ray.S@Site.NET",                    None),
            ("U-017","Quinn@MAIL.org ",                  None),
            ("U-020","TIM@MAIL.org ",                       None),
            ("U-019","Sara.t@example.COM ",                None),
            ("U-022"," Vik@example.COM",                     None),
            ("U-021","Uma.K@Site.NET",                       None),
            ("U-024","XAN@Site.NET",                           None),
            ("U-023","Wren@MAIL.org ",                        None),
            ("U-026","Zara.M@MAIL.org",                         None),
            ("U-025","Yara.L@example.COM ",                    None),
            ("U-028","Beth.J@example.COM",                       None),
            ("U-027","Aiden@Site.NET ",                          None),
            ("U-030","Dev.K@Site.NET",                             None),
            ("U-029","Cleo@MAIL.org ",                            None),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_product_codes(src_path: str, seed: int) -> str:
    """2-col product codes ALL CAPS + empty 'Clean' col.

    Rows stored in swap-adjacent-pairs order (S-002,S-001,S-004,S-003,…) so
    neither ascending nor descending sort of SKU is a no-op (trivial_pass guard).
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["SKU","Code","Clean"])
        rows = [
            ("S-002","SPROCKET-B2",None),
            ("S-001","WIDGET-A1",  None),
            ("S-004","NUT-D4",     None),
            ("S-003","BOLT-C3",    None),
            ("S-006","BRACKET-F6", None),
            ("S-005","WASHER-E5",  None),
            ("S-008","LATCH-H8",   None),
            ("S-007","HINGE-G7",   None),
            ("S-010","STUD-J10",   None),
            ("S-009","RIVET-I9",   None),
            ("S-012","SHIM-L12",   None),
            ("S-011","GASKET-K11", None),
            ("S-014","PIN-N14",    None),
            ("S-013","CLIP-M13",   None),
            ("S-016","BEARING-P16",None),
            ("S-015","SPRING-O15", None),
            ("S-018","RAIL-R18",   None),
            ("S-017","COIL-Q17",   None),
            ("S-020","ROLLER-T20", None),
            ("S-019","TRACK-S19",  None),
            ("S-022","CAP-V22",    None),
            ("S-021","VALVE-U21",  None),
            ("S-024","FAN-X24",    None),
            ("S-023","DUCT-W23",   None),
            ("S-025","FILTER-Y25", None),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


# ---------------------------------------------------------------------------
# Gold-py builders — one per op family. Every builder ends with
# `_LO_NORMALIZE_TAIL.format(expected_path=...)`.
# ---------------------------------------------------------------------------


def _gold_sort(src: str, exp: str, col_idx: int, reverse: bool) -> str:
    """Sort-with-formula-row-rewrite (matches LO Calc Sort behavior).

    Verbatim port of perturb._build_calc_sort_gold_py.
    """
    return textwrap.dedent(f"""\
        import re
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        if max_r >= 2 and max_c >= 1:
            rows = []
            for r in range(2, max_r + 1):
                row_cells = [ws.cell(r, c).value for c in range(1, max_c + 1)]
                rows.append((r, row_cells))
            def _sk(item, ci={col_idx}):
                v = item[1][ci]
                if v is None: return (1, 0, 0.0, "")
                if isinstance(v, (int, float)): return (0, 0, float(v), "")
                return (0, 1, 0.0, str(v))
            rows.sort(key=_sk, reverse={reverse!r})
            row_map = {{}}
            for new_r, (old_r, _) in enumerate(rows, 2):
                row_map[old_r] = new_r
            ref_re = re.compile(r"(\\$?)([A-Z]+)(\\$?)(\\d+)")
            def _rewrite(formula, rmap):
                def repl(m):
                    col_dollar, col_letters, row_dollar, row_digits = m.groups()
                    old = int(row_digits)
                    if old in rmap and not row_dollar:
                        return f"{{col_dollar}}{{col_letters}}{{row_dollar}}{{rmap[old]}}"
                    return m.group(0)
                return ref_re.sub(repl, formula)
            for new_r, (old_r, row_cells) in enumerate(rows, 2):
                for j, val in enumerate(row_cells, 1):
                    if isinstance(val, str) and val.startswith("="):
                        val = "=" + _rewrite(val[1:], row_map)
                    ws.cell(new_r, j).value = val  # #155 #2: .value= sets None (cell(...,None) no-ops -> stale labels)
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_total_row(src: str, exp: str, sum_col_idxs: list[int],
                    label: str, label_col_idx: int) -> str:
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_c = ws.max_column or 0
        max_r = ws.max_row or 0
        while max_r > 1 and all(ws.cell(max_r, c).value is None for c in range(1, max_c + 1)):
            max_r -= 1
        new_r = max_r + 1
        for ci in {sum_col_idxs!r}:
            total = 0.0
            any_num = False
            for r in range(2, max_r + 1):
                v = ws.cell(r, ci + 1).value
                if isinstance(v, (int, float)):
                    total += float(v)
                    any_num = True
            ws.cell(new_r, ci + 1, total if any_num else None)
        ws.cell(new_r, {label_col_idx + 1}, {label!r})
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_derived_col(src: str, exp: str, new_header: str,
                      src_col_idx: int, expr: str,
                      number_format: str | None = None) -> str:
    nf_lines = ""
    if number_format:
        nf_lines = (
            f"for _r in range(2, max_r + 1):\n"
            f"    _cc = ws.cell(_r, new_c)\n"
            f"    if _cc.value is not None:\n"
            f"        _cc.number_format = {number_format!r}\n"
        )
    body = (
        f"import openpyxl\n"
        # The derived exprs call round(); Python's builtin round() uses
        # banker's rounding AND inherits float-repr error (round(1412.175,2)->1412.17),
        # but the agent's natural LibreOffice =ROUND() rounds the DECIMAL value
        # half-up (->1412.18). Override round() here with a half-up version so the
        # gold matches a correct LO-rounded result (~78 derived tasks were false-
        # failing on this). Non-numeric / log exprs are unaffected.
        #
        # `%.15g` (NOT str(x)) is what makes the two agree at a rounding boundary.
        # LO's ROUND() is rtl::math::round(..., Corrected): the IEEE double is first
        # snapped to its 15-SIGNIFICANT-DIGIT decimal value (rtl_math_approxValue),
        # THEN rounded half-away-from-zero. `str(x)` keeps the full 17-digit repr, so
        # a product that is decimally exact but binary-inexact rounds the wrong way:
        # 65*1.085 -> 70.52499999999999, str() -> 70.52 but LO ROUND() -> 70.53.
        # Verified against real headless LibreOffice Calc in the lite container:
        # 5/5 boundary cells across f_calc_40/44/76 now match `=ROUND(...)`.
        # Values whose repr already fits 15 digits are untouched (1412.175 -> 1412.18
        # either way), so this only changes the boundary class.
        f"from decimal import Decimal as _D, ROUND_HALF_UP as _HU\n"
        f"def round(x, n=0):\n"
        f"    if not isinstance(x, (int, float)) or isinstance(x, bool):\n"
        f"        return x\n"
        f"    try:\n"
        f"        _d = _D('%.15g' % x) if isinstance(x, float) else _D(x)\n"
        f"        return float(_d.quantize(_D(1).scaleb(-n), rounding=_HU))\n"
        f"    except Exception:\n"
        f"        return x\n"
        f"wb = openpyxl.load_workbook({src!r})\n"
        f"ws = wb.worksheets[0]\n"
        f"max_r = ws.max_row or 0\n"
        f"max_c = ws.max_column or 0\n"
        f"new_c = max_c + 1\n"
        f"ws.cell(1, new_c, {new_header!r})\n"
        f"for r in range(2, max_r + 1):\n"
        f"    v = ws.cell(r, {src_col_idx + 1}).value\n"
        f"    try:\n"
        f"        ws.cell(r, new_c, ({expr}))\n"
        f"    except (TypeError, ValueError, AttributeError):\n"
        f"        ws.cell(r, new_c, None)\n"
        f"{nf_lines}"
        f"wb.save({exp!r})\n"
    )
    return body + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_numfmt(src: str, exp: str, col_fmt: list[tuple[int, str]]) -> str:
    """Apply a per-column number_format to every non-null data cell."""
    loops = ""
    for ci, fmt in col_fmt:
        loops += (
            f"for r in range(2, max_r + 1):\n"
            f"    cc = ws.cell(r, {ci + 1})\n"
            f"    if cc.value is not None:\n"
            f"        cc.number_format = {fmt!r}\n"
        )
    body = (
        f"import openpyxl\n"
        f"wb = openpyxl.load_workbook({src!r})\n"
        f"ws = wb.worksheets[0]\n"
        f"max_r = ws.max_row or 0\n"
        f"{loops}"
        f"wb.save({exp!r})\n"
    )
    return body + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_string_clean(src: str, exp: str, src_col_idx: int,
                       dst_col_idx: int, op: str) -> str:
    """Write cleaned strings into dst column.

    op ∈ {'lower', 'upper', 'title', 'strip', 'proper_strip'}
    """
    op_expr = {
        "lower": "v.lower()",
        "upper": "v.upper()",
        "title": "v.title()",
        "strip": "' '.join(v.split())",
        "proper_strip": "' '.join(v.split()).title()",
    }[op]
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_r = ws.max_row or 0
        for r in range(2, max_r + 1):
            v = ws.cell(r, {src_col_idx + 1}).value
            if isinstance(v, str):
                ws.cell(r, {dst_col_idx + 1}, {op_expr})
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_sheet2_copy_col(src: str, exp: str, src_col_idx: int,
                          new_sheet_name: str, new_header: str) -> str:
    """Create a new sheet and copy a single column from sheet 0 (with header)."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws1 = wb.worksheets[0]
        max_r = ws1.max_row or 0
        if {new_sheet_name!r} in wb.sheetnames:
            ws2 = wb[{new_sheet_name!r}]
        else:
            ws2 = wb.create_sheet({new_sheet_name!r})
        ws2.cell(1, 1, {new_header!r})
        for r in range(2, max_r + 1):
            ws2.cell(r, 1, ws1.cell(r, {src_col_idx + 1}).value)
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_sheet2_aggregate(src: str, exp: str,
                           sum_cols: list[tuple[int, str]],
                           new_sheet_name: str,
                           agg_label_header: str,
                           agg_value_header: str) -> str:
    """Create a new sheet with one row per (label, sum) pair from source columns.

    sum_cols: list of (col_idx_in_source, label) — source column to sum, label
    written in the new sheet.
    """
    rows_lit = repr(sum_cols)
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws1 = wb.worksheets[0]
        max_r = ws1.max_row or 0
        if {new_sheet_name!r} in wb.sheetnames:
            ws2 = wb[{new_sheet_name!r}]
        else:
            ws2 = wb.create_sheet({new_sheet_name!r})
        ws2.cell(1, 1, {agg_label_header!r})
        ws2.cell(1, 2, {agg_value_header!r})
        out_r = 2
        for ci, label in {rows_lit}:
            tot = 0.0
            for r in range(2, max_r + 1):
                v = ws1.cell(r, ci + 1).value
                if isinstance(v, (int, float)):
                    tot += float(v)
            ws2.cell(out_r, 1, label)
            ws2.cell(out_r, 2, tot)
            out_r += 1
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_sheet2_filter(src: str, exp: str,
                        new_sheet_name: str,
                        filter_col_idx: int,
                        filter_value: str) -> str:
    """Copy header + rows whose filter_col == filter_value to a new sheet."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws1 = wb.worksheets[0]
        max_r = ws1.max_row or 0
        max_c = ws1.max_column or 0
        if {new_sheet_name!r} in wb.sheetnames:
            ws2 = wb[{new_sheet_name!r}]
        else:
            ws2 = wb.create_sheet({new_sheet_name!r})
        for c in range(1, max_c + 1):
            ws2.cell(1, c, ws1.cell(1, c).value)
        out_r = 2
        for r in range(2, max_r + 1):
            v = ws1.cell(r, {filter_col_idx + 1}).value
            if v == {filter_value!r}:
                for c in range(1, max_c + 1):
                    ws2.cell(out_r, c, ws1.cell(r, c).value)
                out_r += 1
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_merge_header(src: str, exp: str,
                       merge_range: str,
                       merge_text: str,
                       *,
                       bold: bool = False,
                       italic: bool = False,
                       underline: bool = False,
                       fill_argb: str | None = None) -> str:
    """Merge cells across `merge_range` and write `merge_text` with optional style.

    fill_argb: 8-char ARGB hex, e.g. 'FFADD8E6' for light blue.
    """
    font_kwargs = []
    if bold:
        font_kwargs.append("bold=True")
    if italic:
        font_kwargs.append("italic=True")
    if underline:
        font_kwargs.append("underline='single'")
    style_block = ""
    if font_kwargs:
        style_block += (
            f"from openpyxl.styles import Font\n"
            f"first_cell.font = Font({', '.join(font_kwargs)})\n"
        )
    if fill_argb:
        style_block += (
            f"from openpyxl.styles import PatternFill\n"
            f"first_cell.fill = PatternFill(start_color={fill_argb!r}, "
            f"end_color={fill_argb!r}, fill_type='solid')\n"
        )
    body = (
        f"import openpyxl\n"
        f"wb = openpyxl.load_workbook({src!r})\n"
        f"ws = wb.worksheets[0]\n"
        f"# Shift existing rows down by 1 to make room for merged header at row 1.\n"
        f"max_r = ws.max_row or 0\n"
        f"max_c = ws.max_column or 0\n"
        f"for r in range(max_r, 0, -1):\n"
        f"    for c in range(1, max_c + 1):\n"
        f"        src_cell = ws.cell(r, c)\n"
        f"        ws.cell(r + 1, c, src_cell.value)\n"
        f"        ws.cell(r, c, None)\n"
        f"ws.merge_cells({merge_range!r})\n"
        f"first_coord = {merge_range!r}.split(':')[0]\n"
        f"first_cell = ws[first_coord]\n"
        f"first_cell.value = {merge_text!r}\n"
        f"{style_block}"
        f"wb.save({exp!r})\n"
    )
    return body + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_cell_color_by_predicate(src: str, exp: str,
                                  rules_py: str,
                                  apply_kind: str,
                                  apply_argb: str) -> str:
    """Iterate data rows and apply font/fill color when `rules_py` returns True.

    `rules_py` is a Python expression in scope of locals: r (1-based row),
    row (list of cell values for that row, 0-indexed). Returns True/False.

    apply_kind ∈ {'fill', 'font'}. apply_argb: 8-char ARGB hex.
    """
    if apply_kind == "fill":
        style_decl = (
            f"from openpyxl.styles import PatternFill\n"
            f"_style = PatternFill(start_color={apply_argb!r}, "
            f"end_color={apply_argb!r}, fill_type='solid')\n"
        )
        apply_attr = "fill"
    elif apply_kind == "font":
        style_decl = (
            f"from openpyxl.styles import Font\n"
            f"_style = Font(color={apply_argb!r})\n"
        )
        apply_attr = "font"
    else:
        raise ValueError(f"unknown apply_kind: {apply_kind!r}")
    body = (
        f"import openpyxl\n"
        f"{style_decl}"
        f"wb = openpyxl.load_workbook({src!r})\n"
        f"ws = wb.worksheets[0]\n"
        f"max_r = ws.max_row or 0\n"
        f"max_c = ws.max_column or 0\n"
        f"for r in range(2, max_r + 1):\n"
        f"    row = [ws.cell(r, c).value for c in range(1, max_c + 1)]\n"
        f"    if {rules_py}:\n"
        f"        for c in range(1, max_c + 1):\n"
        f"            ws.cell(r, c).{apply_attr} = _style\n"
        f"wb.save({exp!r})\n"
    )
    return body + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_two_color_by_predicate(src: str, exp: str,
                                 pred_a_py: str, argb_a: str,
                                 pred_b_py: str, argb_b: str,
                                 apply_kind: str = "fill") -> str:
    """Apply colour A when pred_a is True, else B when pred_b is True (font/fill)."""
    if apply_kind == "fill":
        style_decl = (
            f"from openpyxl.styles import PatternFill\n"
            f"_sa = PatternFill(start_color={argb_a!r}, end_color={argb_a!r}, fill_type='solid')\n"
            f"_sb = PatternFill(start_color={argb_b!r}, end_color={argb_b!r}, fill_type='solid')\n"
        )
        apply_attr = "fill"
    elif apply_kind == "font":
        style_decl = (
            f"from openpyxl.styles import Font\n"
            f"_sa = Font(color={argb_a!r})\n"
            f"_sb = Font(color={argb_b!r})\n"
        )
        apply_attr = "font"
    else:
        raise ValueError(f"unknown apply_kind: {apply_kind!r}")
    body = (
        f"import openpyxl\n"
        f"{style_decl}"
        f"wb = openpyxl.load_workbook({src!r})\n"
        f"ws = wb.worksheets[0]\n"
        f"max_r = ws.max_row or 0\n"
        f"max_c = ws.max_column or 0\n"
        f"for r in range(2, max_r + 1):\n"
        f"    row = [ws.cell(r, c).value for c in range(1, max_c + 1)]\n"
        f"    if {pred_a_py}:\n"
        f"        for c in range(1, max_c + 1):\n"
        f"            ws.cell(r, c).{apply_attr} = _sa\n"
        f"    elif {pred_b_py}:\n"
        f"        for c in range(1, max_c + 1):\n"
        f"            ws.cell(r, c).{apply_attr} = _sb\n"
        f"wb.save({exp!r})\n"
    )
    return body + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_sheet2_filter_score_threshold(src: str, exp: str, *,
                                        new_sheet_name: str,
                                        score_col_idx: int,
                                        predicate: str,
                                        threshold: float) -> str:
    """Copy header + rows whose `score_col` op `threshold` to a new sheet.

    `predicate` ∈ {"lt","le","gt","ge","eq"}. Used by F-CALC-2 gradebook
    Failing / Honor-Roll filter task.
    """
    cmp = {"lt": "<", "le": "<=", "gt": ">", "ge": ">=", "eq": "=="}[predicate]
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws1 = wb.worksheets[0]
        max_r = ws1.max_row or 0
        max_c = ws1.max_column or 0
        if {new_sheet_name!r} in wb.sheetnames:
            ws2 = wb[{new_sheet_name!r}]
        else:
            ws2 = wb.create_sheet({new_sheet_name!r})
        for c in range(1, max_c + 1):
            ws2.cell(1, c, ws1.cell(1, c).value)
        out_r = 2
        for r in range(2, max_r + 1):
            v = ws1.cell(r, {score_col_idx + 1}).value
            if isinstance(v, (int, float)) and v {cmp} {threshold}:
                for c in range(1, max_c + 1):
                    ws2.cell(out_r, c, ws1.cell(r, c).value)
                out_r += 1
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_sheet2_groupby_sum(src: str, exp: str, *,
                             new_sheet_name: str,
                             key_col_idx: int,
                             value_col_idx: int,
                             agg: str,
                             key_header: str,
                             value_header: str) -> str:
    """Group-by + aggregate `value_col` per distinct `key_col` value → new sheet.

    `agg` ∈ {"sum","count","avg"}. Row order = first-seen order of the key
    in the source sheet (deterministic). Used by F-CALC-3 orders
    group-by-Region task.
    """
    if agg not in {"sum", "count", "avg"}:
        raise ValueError(f"unknown agg: {agg!r}")
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws1 = wb.worksheets[0]
        max_r = ws1.max_row or 0
        if {new_sheet_name!r} in wb.sheetnames:
            ws2 = wb[{new_sheet_name!r}]
        else:
            ws2 = wb.create_sheet({new_sheet_name!r})
        ws2.cell(1, 1, {key_header!r})
        ws2.cell(1, 2, {value_header!r})
        order = []
        tot = {{}}
        cnt = {{}}
        for r in range(2, max_r + 1):
            k = ws1.cell(r, {key_col_idx + 1}).value
            v = ws1.cell(r, {value_col_idx + 1}).value
            if k is None:
                continue
            if k not in tot:
                tot[k] = 0.0
                cnt[k] = 0
                order.append(k)
            if isinstance(v, (int, float)):
                tot[k] += float(v)
            cnt[k] += 1
        out_r = 2
        for k in order:
            if {agg!r} == "sum":
                out_v = tot[k]
            elif {agg!r} == "count":
                out_v = cnt[k]
            else:  # avg
                out_v = tot[k] / cnt[k] if cnt[k] else 0.0
            ws2.cell(out_r, 1, k)
            ws2.cell(out_r, 2, out_v)
            out_r += 1
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


# ---------------------------------------------------------------------------
# Rule constants — referenced by §I FILE_TASKS Param.rules.
# ---------------------------------------------------------------------------

# Standard "compare_table sheet_data" rule on the active sheet.
_RULE_SHEET_DATA = [{"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"}]
_RULE_SHEET_NAME_AND_DATA_NAMED = lambda sheet: [
    # validation P2 calc rebalance: eval's Sheet2-creating tasks (filter / groupby /
    # aggregate / copy_col) test ONLY the sheet_data rule indexed by `RNSheet2`/
    # `ENSheet2` — they do NOT add a separate `sheet_name` rule. The named-sheet
    # lookup already fails if the agent never created the sheet, so the extra
    # `sheet_name` rule is redundant. Removing it shifts the synth `rule_combo`
    # distribution from 24% `sheet_data+sheet_name` (vs eval 2.2%) toward
    # `sheet_data` (eval 43.5%). Kept the named function so downstream Param.rules
    # entries don't need rewriting.
    {"type": "sheet_data",
     "sheet_idx0": f"RN{sheet}", "sheet_idx1": f"EN{sheet}"},
]
_RULE_SHEET_DATA_AND_STYLE = [
    {"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
    # `_read_cell_style` does not accept bare "fill" — it dispatches on
    # specific style names. For background fills, "bgcolor" reads
    # `cell.fill.fgColor.rgb` which is what `PatternFill(start_color=..,
    # end_color=.., fill_type='solid')` actually populates.
    {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0", "props": ["bgcolor"]},
]
_RULE_SHEET_DATA_AND_FONT = [
    {"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
    # `_read_cell_style` does not accept bare "font" — it dispatches on
    # specific style names. For bold/italic header tasks we compare
    # `font_bold` and `font_italic` directly.
    {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
     "props": ["font_bold", "font_italic"]},
]
# Batch — eval-aligned rule shapes for new skill families.
# `chart` rule with chart_props=['type'] only — validation retained
# `_RULE_CHART_NO_TITLE` semantics (title round-trips imperfectly through LO).
_RULE_CHART_TYPE_ACTIVE = [
    {"type": "chart", "sheet_idx0": 0, "sheet_idx1": "EI0",
     "chart_props": ["type"]},
]
# freeze rule carries `freeze_cell` so `_eval_compare_table` can route
# the freeze probe to the local `check_xlsx_freeze_pane` helper. Upstream
# `compare_table`'s `freeze` rule compares `ws.freeze_panes` between
# result and expected, which is defeated when `_LO_NORMALIZE_TAIL`'s
# `soffice --convert-to xlsx` strips `<pane state="frozen"/>` from the
# openpyxl-written gold (validation Bug 2: both sides become None,
# matches trivially). The new helper checks the result directly against
# the literal expected cell, bypassing the normalize round-trip.
_RULE_FREEZE_AND_DATA = lambda freeze_cell="A2": [
    {"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
    {"type": "freeze", "sheet_idx0": 0, "sheet_idx1": "EI0",
     "freeze_cell": freeze_cell},
]
_RULE_ZOOM_AND_DATA = lambda scale: [
    {"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
    {"type": "zoom", "sheet_idx": 0, "method": "eq", "ref": scale},
]
_RULE_DATA_VALIDATION_AND_DATA = lambda allowed: [
    {"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
    {"type": "data_validation", "sheet_idx": 0,
     "dv_props": [{"type": {"method": "eq", "ref": "list"},
                   "formula1": {"method": "str_set_eq", "ref": list(allowed)}}]},
]
_RULE_ROW_PROPS_HIDDEN = [
    {"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
    {"type": "row_props", "sheet_idx0": 0, "sheet_idx1": "EI0",
     "props": ["hidden"]},
]


def _RULE_CHECK_CELL_NAMED_SHEET(sheet: str, coord: str, value) -> list[dict]:
    """validation calc P1 — eval-anchored `check_cell` skill_class.

    Eval has 2 `check_cell` rows (`osworld_libreoffice_calc_4172478b` and the
    investment-summary merged-header task) — synth had zero. A `check_cell` rule
    pinned at a specific coordinate verifies the exact text the agent typed
    independently of `sheet_data`'s range scan. Combined with the existing
    `sheet_data` rule, this also shifts `rule_combo` toward eval's
    `check_cell+sheet_data` shape.
    """
    return [
        {"type": "check_cell",
         "sheet_idx": f"RN{sheet}",
         "coordinate": coord,
         "props": {"value": {"method": "eq", "ref": value}}},
        {"type": "sheet_data",
         "sheet_idx0": f"RN{sheet}", "sheet_idx1": f"EN{sheet}"},
    ]


TEMPLATES: list[SynthTemplate] = []  # Populated below by §I.f _emit_templates(FILE_TASKS)


# ===========================================================================
# Real-CSV builders — read assets/synth/data/csv/*.csv into a one-sheet xlsx.
# Each `_csv_src_*` returns a python heredoc; the matching `_make_csv_file`
# call in §I.c stages the CSV via `_stage_asset` before the heredoc runs.
# ===========================================================================


def _csv_src_us_population_states(csv_in: str, src_path: str, _seed: int) -> str:
    """Read us-population-states.csv → first 30×3 rows (State / Population2020 / StateFIPS)."""
    return textwrap.dedent(f"""\
        import csv, openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Population"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            n = 0
            for row in rd:
                if not row:
                    continue
                if n >= 30:
                    break
                # State, Population2020 (int), StateFIPS (str — preserve leading zero)
                ws.append([row[0], int(row[1]), row[2]])
                n += 1
        wb.save({src_path!r})
        """)


def _csv_src_us_gdp(csv_in: str, src_path: str, _seed: int) -> str:
    """Read us-gdp.csv → first 32 quarterly rows (date / GDP)."""
    return textwrap.dedent(f"""\
        import csv, openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "GDP"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            n = 0
            for row in rd:
                if not row or n >= 32:
                    break
                ws.append([row[0], float(row[1])])
                n += 1
        wb.save({src_path!r})
        """)


def _csv_src_us_unemployment(csv_in: str, src_path: str, _seed: int) -> str:
    """Read us-unemployment.csv → first 36 monthly rows (date / UNRATE)."""
    return textwrap.dedent(f"""\
        import csv, openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Unemployment"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            n = 0
            for row in rd:
                if not row or n >= 36:
                    break
                ws.append([row[0], float(row[1])])
                n += 1
        wb.save({src_path!r})
        """)


def _csv_src_world_gdp_2022(csv_in: str, src_path: str, _seed: int) -> str:
    """Read world-gdp-2022.csv → TOP-30 COUNTRIES by GDP (CountryCode / CountryName / Year / GDP_USD).

    validation B2 fix: (1) filter out World-Bank AGGREGATE codes (AFE/AFW/
    EUU/WLD/HIC/LIC/LMC/...) and income-group rows so the slice contains real
    ISO countries; (2) sort countries by GDP_USD DESCENDING and take the top
    30 so the slice is "top economies" (USA, CHN, JPN, DEU, ...) per the task
    instruction wording (e.g. "color top economies", "sort by GDP desc").
    Pre-fix the first-30 was an alphabetical-by-CountryCode walk that
    happened to land entirely on WB aggregates — neither real countries NOR
    top economies.
    """
    return textwrap.dedent(f"""\
        import csv, openpyxl
        _WB_AGG = {{
            'AFE','AFW','ARB','CSS','CEB','EAR','EAS','EAP','TEA','EMU',
            'ECS','ECA','TEC','EUU','FCS','HPC','IBD','IBT','IDB','IDX',
            'IDA','LTE','LCN','LAC','TLA','LDC','LMY','MEA','MNA','TMN',
            'MIC','NAC','INX','OED','OSS','PSS','PST','SST','SAS','TSS',
            'SSF','SSA','TSA','UMC','WLD','HIC','LIC','LMC','LME','PRE',
        }}
        _INCOME_NAMES = {{
            'High income','Low income','Lower middle income','Upper middle income',
        }}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WorldGDP"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            countries = []
            for row in rd:
                if not row:
                    continue
                code = (row[0] or '').strip()
                name = (row[1] or '').strip()
                # Skip aggregates / income groups.
                if not code or code in _WB_AGG or name in _INCOME_NAMES:
                    continue
                try:
                    gdp = float(row[3])
                except (ValueError, IndexError):
                    continue
                countries.append((code, name, int(row[2]), gdp))
            # Top-30 by GDP descending (largest economies first).
            countries.sort(key=lambda r: r[3], reverse=True)
            for rec in countries[:30]:
                ws.append(list(rec))
        wb.save({src_path!r})
        """)


def _csv_src_oil_wti_daily(csv_in: str, src_path: str, _seed: int) -> str:
    """Read oil-wti-daily.csv → first 30 daily price rows (date / DCOILWTICO)."""
    return textwrap.dedent(f"""\
        import csv, openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WTIPrice"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            n = 0
            for row in rd:
                if not row or n >= 30:
                    break
                try:
                    p = float(row[1])
                except (ValueError, IndexError):
                    continue
                ws.append([row[0], p])
                n += 1
        wb.save({src_path!r})
        """)


def _csv_src_us_fed_funds(csv_in: str, src_path: str, _seed: int) -> str:
    """Read us-fed-funds-rate.csv → first 36 monthly rows (date / FEDFUNDS)."""
    return textwrap.dedent(f"""\
        import csv, openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FedFunds"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            n = 0
            for row in rd:
                if not row or n >= 36:
                    break
                try:
                    r = float(row[1])
                except (ValueError, IndexError):
                    continue
                ws.append([row[0], r])
                n += 1
        wb.save({src_path!r})
        """)


def _csv_src_us_housing_starts(csv_in: str, src_path: str, _seed: int) -> str:
    """Read us-housing-starts.csv → first 36 monthly rows (date / HOUST)."""
    return textwrap.dedent(f"""\
        import csv, openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Housing"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            n = 0
            for row in rd:
                if not row or n >= 36:
                    break
                try:
                    h = float(row[1])
                except (ValueError, IndexError):
                    continue
                ws.append([row[0], h])
                n += 1
        wb.save({src_path!r})
        """)


def _csv_src_us_inflation_cpi(csv_in: str, src_path: str, _seed: int) -> str:
    """Read us-inflation-cpi.csv → first 36 monthly rows (date / CPIAUCSL)."""
    return textwrap.dedent(f"""\
        import csv, openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CPI"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            n = 0
            for row in rd:
                if not row or n >= 36:
                    break
                try:
                    v = float(row[1])
                except (ValueError, IndexError):
                    continue
                ws.append([row[0], v])
                n += 1
        wb.save({src_path!r})
        """)


def _csv_src_us_mortgage_30yr(csv_in: str, src_path: str, _seed: int) -> str:
    """Read us-mortgage-30yr.csv → first 26 weekly rows (date / MORTGAGE30US)."""
    return textwrap.dedent(f"""\
        import csv, openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Mortgage"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            n = 0
            for row in rd:
                if not row or n >= 26:
                    break
                try:
                    r = float(row[1])
                except (ValueError, IndexError):
                    continue
                ws.append([row[0], r])
                n += 1
        wb.save({src_path!r})
        """)


def _csv_src_us_state_median_income(csv_in: str, src_path: str, _seed: int) -> str:
    """Read us-state-median-income.csv → first 30 rows (State / MedianHouseholdIncome_USD / StateFIPS).

    Capped at 30 rows to land inside the eval-row Q-scale band (eval median 22,
    p75 30, max ~45) per the Q-scale guard. Was 53 (uncapped).
    """
    return textwrap.dedent(f"""\
        import csv, openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "StateIncome"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            n = 0
            for row in rd:
                if not row or n >= 30:
                    break
                # State, MedianHouseholdIncome_USD (int), StateFIPS (str — keep leading zero)
                try:
                    inc = int(row[1])
                except (ValueError, IndexError):
                    continue
                ws.append([row[0], inc, row[2]])
                n += 1
        wb.save({src_path!r})
        """)


def _csv_src_world_population_2022(csv_in: str, src_path: str, _seed: int) -> str:
    """Read world-population-2022.csv → TOP-30 COUNTRIES by Population (CountryCode / CountryName / Year / Population).

    Capped at 30 rows to land inside the eval Q-scale band (median 22, p75 30,
    max ~45). Was 50.

    validation B2 fix: (1) filter out WB aggregate codes + income-group
    rows; (2) sort by Population DESCENDING and take top 30 so the slice is
    "most-populous countries" (CHN, IND, USA, IDN, PAK, ...) per the
    task instruction wording. Pre-fix the first 30 entries were all
    aggregates (AFE/EUU/WLD/...) so "top-N countries by population" tasks
    had zero valid rows.
    """
    return textwrap.dedent(f"""\
        import csv, openpyxl
        _WB_AGG = {{
            'AFE','AFW','ARB','CSS','CEB','EAR','EAS','EAP','TEA','EMU',
            'ECS','ECA','TEC','EUU','FCS','HPC','IBD','IBT','IDB','IDX',
            'IDA','LTE','LCN','LAC','TLA','LDC','LMY','MEA','MNA','TMN',
            'MIC','NAC','INX','OED','OSS','PSS','PST','SST','SAS','TSS',
            'SSF','SSA','TSA','UMC','WLD','HIC','LIC','LMC','LME','PRE',
        }}
        _INCOME_NAMES = {{
            'High income','Low income','Lower middle income','Upper middle income',
        }}
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "WorldPop"
        with open({csv_in!r}, newline='') as _f:
            rd = csv.reader(_f)
            header = next(rd)
            ws.append(header)
            countries = []
            for row in rd:
                if not row:
                    continue
                code = (row[0] or '').strip()
                name = (row[1] or '').strip()
                if not code or code in _WB_AGG or name in _INCOME_NAMES:
                    continue
                try:
                    pop = int(row[3])
                except (ValueError, IndexError):
                    continue
                countries.append((code, name, int(row[2]), pop))
            countries.sort(key=lambda r: r[3], reverse=True)
            for rec in countries[:30]:
                ws.append(list(rec))
        wb.save({src_path!r})
        """)



# ===========================================================================
# Additional `_src_*` xlsx builders used by §I.c File instances.
# ===========================================================================


def _src_attendance_school(src_path: str, _seed: int) -> str:
    """24×5 elementary-school attendance log."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"
        ws.append(["Date", "Student", "Grade", "Status", "Tardy_Min"])
        rows = [
            ("2026-04-01", "Aiden Park",       3, "Present",  0),
            ("2026-04-01", "Beatriz Silva",    3, "Tardy",   12),
            ("2026-04-01", "Caleb Singh",      4, "Present",  0),
            ("2026-04-01", "Diana Lopez",      4, "Absent",   0),
            ("2026-04-02", "Aiden Park",       3, "Present",  0),
            ("2026-04-02", "Beatriz Silva",    3, "Present",  0),
            ("2026-04-02", "Caleb Singh",      4, "Tardy",    7),
            ("2026-04-02", "Diana Lopez",      4, "Present",  0),
            ("2026-04-03", "Aiden Park",       3, "Absent",   0),
            ("2026-04-03", "Beatriz Silva",    3, "Present",  0),
            ("2026-04-03", "Caleb Singh",      4, "Present",  0),
            ("2026-04-03", "Diana Lopez",      4, "Present",  0),
            ("2026-04-06", "Aiden Park",       3, "Present",  0),
            ("2026-04-06", "Beatriz Silva",    3, "Tardy",    5),
            ("2026-04-06", "Caleb Singh",      4, "Present",  0),
            ("2026-04-06", "Diana Lopez",      4, "Tardy",   18),
            ("2026-04-07", "Aiden Park",       3, "Tardy",    3),
            ("2026-04-07", "Beatriz Silva",    3, "Present",  0),
            ("2026-04-07", "Caleb Singh",      4, "Absent",   0),
            ("2026-04-07", "Diana Lopez",      4, "Present",  0),
            ("2026-04-08", "Aiden Park",       3, "Present",  0),
            ("2026-04-08", "Beatriz Silva",    3, "Present",  0),
            ("2026-04-08", "Caleb Singh",      4, "Present",  0),
            ("2026-04-08", "Diana Lopez",      4, "Present",  0),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_inventory_warehouse(src_path: str, _seed: int) -> str:
    """21×5 warehouse SKU inventory.

    validation fix (vacuous-predicate audit): added an out-of-stock row
    ("SKU-1021", "Widget-X Connector", "H-01", 0, 100) so the Param[1]
    `row[3] == 0` predicate on F_CALC_18.color_low_stock has ≥1 match
    (was 0/20 with previous min OnHand=17).
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Warehouse"
        ws.append(["SKU", "Item", "Bin", "OnHand", "ReorderPoint"])
        rows = [
            ("SKU-1002", "Galvanized Bolt M8",  "A-13",   38, 80),
            ("SKU-1001", "Steel Bracket 4in",   "A-12",  142, 50),
            ("SKU-1004", "Lock Washer M8",      "A-15",  640, 150),
            ("SKU-1003", "Hex Nut M8",          "A-14",  920, 200),
            ("SKU-1006", "PVC Coupler 3/4",     "B-03",  118, 30),
            ("SKU-1005", "Rubber Grommet 1/2",  "B-02",   72, 40),
            ("SKU-1008", "Wire Connector AWG14","C-01",  315, 100),
            ("SKU-1007", "Copper Pipe 1ft",     "B-04",   24, 25),
            ("SKU-1010", "Power Strip 6-Outlet","C-03",   17, 10),
            ("SKU-1009", "Heat-Shrink Tubing",  "C-02",   84, 50),
            ("SKU-1012", "Smoke Detector 9V",   "D-09",   46, 25),
            ("SKU-1011", "LED Bulb 60W-eq",     "D-08",  205, 80),
            ("SKU-1014", "Hose Clamp 1in",      "E-05",  198, 75),
            ("SKU-1013", "Air Filter 16x20",    "E-04",   58, 40),
            ("SKU-1016", "O-Ring AS568-014",    "F-12",  812, 250),
            ("SKU-1015", "Ball Bearing 608",    "F-11",  264, 120),
            ("SKU-1018", "Cable Tie 8in",       "G-01",  605, 300),
            ("SKU-1017", "Stainless Screw M4",  "F-13",  476, 200),
            ("SKU-1020", "Anti-static Bag 6x9", "G-03",  342, 100),
            ("SKU-1019", "Velcro Strap 12in",   "G-02",   89, 50),
            ("SKU-1021", "Widget-X Connector",  "H-01",    0, 100),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_quarterly_sales(src_path: str, _seed: int) -> str:
    """16×4 quarterly sales by region."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "QuarterlySales"
        ws.append(["Quarter", "Region", "Revenue", "UnitsSold"])
        rows = [
            ("Q1 2025", "Northeast",  286400, 1240),
            ("Q1 2025", "Midwest",    312800, 1410),
            ("Q1 2025", "South",      415900, 1880),
            ("Q1 2025", "West",       378200, 1720),
            ("Q2 2025", "Northeast",  324500, 1420),
            ("Q2 2025", "Midwest",    348100, 1560),
            ("Q2 2025", "South",      462300, 2050),
            ("Q2 2025", "West",       418600, 1890),
            ("Q3 2025", "Northeast",  368900, 1610),
            ("Q3 2025", "Midwest",    391200, 1740),
            ("Q3 2025", "South",      508400, 2280),
            ("Q3 2025", "West",       472100, 2110),
            ("Q4 2025", "Northeast",  548700, 2450),
            ("Q4 2025", "Midwest",    582300, 2620),
            ("Q4 2025", "South",      721500, 3290),
            ("Q4 2025", "West",       689400, 3140),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)



def _src_fitness_steps(src_path: str, _seed: int) -> str:
    """30×3 daily fitness log: Date / Steps / Calories.

    validation fix (vacuous-predicate audit): lowered two recovery days
    (2026-04-06 → 3920, 2026-04-24 → 4180) so the Param[1] `row[1] < 5000`
    predicate on F_CALC_20.color_high_step_days has ≥1 match (was 0/30
    with previous min Steps=6120).
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fitness"
        ws.append(["Date", "Steps", "Calories"])
        # Realistic 30-day arc: weekday avg ~7-9K, weekend higher (~12-15K),
        # with two sub-5K recovery / rest days.
        rows = [
            ("2026-04-01",  6840,  295),
            ("2026-04-02",  7920,  342),
            ("2026-04-03",  8410,  368),
            ("2026-04-04", 14260,  612),
            ("2026-04-05", 13180,  564),
            ("2026-04-06",  3920,  169),
            ("2026-04-07",  7480,  321),
            ("2026-04-08",  8260,  357),
            ("2026-04-09",  9140,  395),
            ("2026-04-10",  7860,  339),
            ("2026-04-11", 12740,  549),
            ("2026-04-12", 15310,  658),
            ("2026-04-13",  8120,  351),
            ("2026-04-14",  7240,  313),
            ("2026-04-15",  6920,  299),
            ("2026-04-16",  9580,  413),
            ("2026-04-17",  8470,  366),
            ("2026-04-18", 13850,  596),
            ("2026-04-19", 11620,  500),
            ("2026-04-20",  7340,  317),
            ("2026-04-21",  8280,  358),
            ("2026-04-22",  9120,  394),
            ("2026-04-23",  7910,  342),
            ("2026-04-24",  4180,  180),
            ("2026-04-25", 14710,  632),
            ("2026-04-26", 12930,  556),
            ("2026-04-27",  7160,  309),
            ("2026-04-28",  8640,  373),
            ("2026-04-29",  9210,  398),
            ("2026-04-30",  8910,  385),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_concert_revenue(src_path: str, _seed: int) -> str:
    """12×3 concert tour revenue: Show / Venue / Revenue."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tour"
        ws.append(["Show", "Venue", "Revenue"])
        rows = [
            ("2026-05-02", "Madison Square Garden, NYC",   1245000),
            ("2026-05-05", "TD Garden, Boston",             682000),
            ("2026-05-08", "Wells Fargo Center, Philly",    715000),
            ("2026-05-11", "Capital One Arena, DC",         598000),
            ("2026-05-14", "State Farm Arena, Atlanta",     541000),
            ("2026-05-17", "American Airlines Center, DAL", 627000),
            ("2026-05-20", "Toyota Center, Houston",        584000),
            ("2026-05-23", "Ball Arena, Denver",            462000),
            ("2026-05-26", "Chase Center, SF",              738000),
            ("2026-05-29", "Crypto.com Arena, LA",         1126000),
            ("2026-06-01", "Footprint Center, Phoenix",     489000),
            ("2026-06-04", "T-Mobile Arena, Las Vegas",     815000),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_market_share(src_path: str, _seed: int) -> str:
    """6×2 market share: Brand / Share.

    validation fix (vacuous-predicate audit): added a low-share brand ("Nano", 0.04)
    so the Param[1] `row[1] < 0.05` predicate on F_CALC_22.color_top_share has
    ≥1 match (was 0/5 with previous min=0.12).
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MarketShare"
        ws.append(["Brand", "Share"])
        rows = [
            ("Acme",        0.34),
            ("Globex",      0.21),
            ("Initech",     0.18),
            ("Soylent",     0.15),
            ("Umbrella",    0.12),
            ("Nano",        0.04),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)






def _src_invoices(src_path: str, _seed: int) -> str:
    """12×4 invoice log with mix of OVERDUE / PAID / PENDING statuses."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"
        ws.append(["InvoiceID", "Customer", "Amount", "Status"])
        rows = [
            ("INV-3001", "Pinecrest Mfg",     2480.50, "PAID"),
            ("INV-3002", "Bayside Grocers",   1842.00, "OVERDUE"),
            ("INV-3003", "Northwood Dental",   648.75, "PAID"),
            ("INV-3004", "Lakeshore Realty",  3920.40, "PENDING"),
            ("INV-3005", "Cypress Logistics", 5175.00, "OVERDUE"),
            ("INV-3006", "Driftwood Cafe",     394.20, "PAID"),
            ("INV-3007", "Evergreen Clinic",  2150.30, "PENDING"),
            ("INV-3008", "Forge & Co",        1820.85, "OVERDUE"),
            ("INV-3009", "Granite Holdings",  4580.00, "PAID"),
            ("INV-3010", "Harbor Freight",     895.65, "PENDING"),
            ("INV-3011", "Ironwood Bros",     2740.10, "OVERDUE"),
            ("INV-3012", "Juniper Studios",   1295.40, "PAID"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_op_metrics(src_path: str, _seed: int) -> str:
    """10×3 ops metrics: Server / UptimePct / IncidentCount."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OpsMetrics"
        ws.append(["Server", "UptimePct", "IncidentCount"])
        rows = [
            ("api-east-01",  99.92, 1),
            ("api-east-02",  99.71, 2),
            ("api-west-01",  98.45, 7),
            ("db-primary",   99.99, 0),
            ("db-replica-1", 99.34, 4),
            ("db-replica-2", 97.82, 9),
            ("cache-01",     99.88, 1),
            ("cache-02",     98.12, 6),
            ("queue-01",     99.61, 3),
            ("worker-pool",  96.74, 12),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)




def _src_attendance_lookup(src_path: str, _seed: int) -> str:
    """Two-sheet xlsx: Sheet1 = attendance (Name / LateMin / Fee=empty),
    Sheet2 = LateFeeScale (LateMin lookup → Fee).
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Attendance"
        ws1.append(["Name", "LateMin", "Fee"])
        rows = [
            ("Alice Wong",       0,  None),
            ("Marcus Chen",      8,  None),
            ("Priya Patel",     22,  None),
            ("Diego Garcia",     3,  None),
            ("Aisha Okafor",    35,  None),
            ("Jordan Kim",      14,  None),
            ("Sofia Hernandez",  0,  None),
            ("Liam O'Connor",   45,  None),
            ("Yuki Tanaka",      6,  None),
            ("Noah Williams",   28,  None),
        ]
        for r in rows:
            ws1.append(list(r))
        ws2 = wb.create_sheet("LateFeeScale")
        ws2.append(["LateMin", "Fee"])
        # Stepwise scale: 0 -> $0, 5+ -> $5, 15+ -> $15, 30+ -> $30, 60+ -> $60.
        for lm, fee in [(0, 0), (5, 5), (15, 15), (30, 30), (60, 60)]:
            ws2.append([lm, fee])
        wb.save({src_path!r})
        """)


def _gold_vlookup_late_fee(src: str, exp: str) -> str:
    """Fill Fee column on Sheet1 by stepped-VLOOKUP against LateFeeScale.

    Resolves the lookup in python (max breakpoint <= LateMin → Fee).
    Equivalent to VLOOKUP(LateMin, LateFeeScale!A:B, 2, TRUE).
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws1 = wb["Attendance"]
        ws2 = wb["LateFeeScale"]
        # Build the breakpoint→fee table (sorted ascending by breakpoint).
        scale = []
        for r in range(2, (ws2.max_row or 0) + 1):
            lm = ws2.cell(r, 1).value
            fe = ws2.cell(r, 2).value
            if isinstance(lm, (int, float)) and isinstance(fe, (int, float)):
                scale.append((float(lm), float(fe)))
        scale.sort(key=lambda t: t[0])
        max_r = ws1.max_row or 0
        for r in range(2, max_r + 1):
            v = ws1.cell(r, 2).value
            if not isinstance(v, (int, float)):
                continue
            fee = 0.0
            for bp, fe in scale:
                if float(v) >= bp:
                    fee = fe
                else:
                    break
            ws1.cell(r, 3, fee)
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _src_quarterly_3sheets(src_path: str, _seed: int) -> str:
    """Three quarter sheets (Q1, Q2, Q3) — same shape, different numbers."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        # Q1
        ws1 = wb.active
        ws1.title = "Q1"
        ws1.append(["Region", "Revenue", "Expenses"])
        for row in [("Northeast",  286400, 192300),
                    ("Midwest",    312800, 215600),
                    ("South",      415900, 278400),
                    ("West",       378200, 251800)]:
            ws1.append(list(row))
        # Q2
        ws2 = wb.create_sheet("Q2")
        ws2.append(["Region", "Revenue", "Expenses"])
        for row in [("Northeast",  324500, 218600),
                    ("Midwest",    348100, 231900),
                    ("South",      462300, 309800),
                    ("West",       418600, 281200)]:
            ws2.append(list(row))
        # Q3
        ws3 = wb.create_sheet("Q3")
        ws3.append(["Region", "Revenue", "Expenses"])
        for row in [("Northeast",  368900, 246800),
                    ("Midwest",    391200, 263400),
                    ("South",      508400, 339200),
                    ("West",       472100, 318900)]:
            ws3.append(list(row))
        wb.save({src_path!r})
        """)


def _gold_quarterly_summary(src: str, exp: str) -> str:
    """Build a 4th 'Summary' sheet: Region / TotalRevenue / TotalExpenses
    rolling up Q1+Q2+Q3 by Region.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        sums = {{}}  # region -> [rev, exp]
        for sn in ("Q1", "Q2", "Q3"):
            ws = wb[sn]
            mr = ws.max_row or 0
            for r in range(2, mr + 1):
                region = ws.cell(r, 1).value
                rev = ws.cell(r, 2).value
                exp = ws.cell(r, 3).value
                if region is None:
                    continue
                cur = sums.setdefault(region, [0.0, 0.0])
                if isinstance(rev, (int, float)):
                    cur[0] += float(rev)
                if isinstance(exp, (int, float)):
                    cur[1] += float(exp)
        if "Summary" in wb.sheetnames:
            ws_s = wb["Summary"]
        else:
            ws_s = wb.create_sheet("Summary")
        ws_s.cell(1, 1, "Region")
        ws_s.cell(1, 2, "TotalRevenue")
        ws_s.cell(1, 3, "TotalExpenses")
        # Preserve insertion order (Northeast / Midwest / South / West).
        out_r = 2
        for region in ["Northeast", "Midwest", "South", "West"]:
            if region in sums:
                rev, exp_v = sums[region]
                ws_s.cell(out_r, 1, region)
                ws_s.cell(out_r, 2, rev)
                ws_s.cell(out_r, 3, exp_v)
                out_r += 1
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


# ---------------------------------------------------------------------------
# New templates (instances).
# ---------------------------------------------------------------------------

# Batch fix: chart rules dropped the `title` prop — openpyxl chart
# title XML round-trips imperfectly through LO save (numRef.f format
# mismatch between agent-built and LO-normalized charts), so we keep


# ===========================================================================
# §I-prelude. Additional authored xlsx src builders (Batch file growth).
# Each is a small, structurally distinct file affording a specific eval skill.
# ===========================================================================


def _src_bank_transactions(src_path: str, _seed: int) -> str:
    """20×4 bank transactions: Date / Description / Amount / Balance (Balance empty).

    Target eval skill: running balance derive; conditional fmt overdrafts.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"
        ws.append(["Date", "Description", "Amount", "Balance"])
        rows = [
            ("2026-03-01", "Opening balance",       1500.00, None),
            ("2026-03-02", "Grocery — Whole Foods",  -82.45, None),
            ("2026-03-03", "ATM withdrawal",        -100.00, None),
            ("2026-03-04", "Payroll deposit",       2350.00, None),
            ("2026-03-05", "Rent — autopay",       -1400.00, None),
            ("2026-03-07", "Electric utility",       -78.20, None),
            ("2026-03-08", "Gas station",            -45.60, None),
            ("2026-03-10", "Refund — Amazon",         32.10, None),
            ("2026-03-12", "Restaurant — Tartine",   -56.80, None),
            ("2026-03-14", "Phone bill",             -42.00, None),
            ("2026-03-15", "Side gig payout",        320.00, None),
            ("2026-03-17", "Pharmacy",               -28.50, None),
            ("2026-03-18", "Subscription — Netflix", -15.99, None),
            ("2026-03-20", "Grocery — Trader Joes",  -64.30, None),
            ("2026-03-22", "Cash deposit",           150.00, None),
            ("2026-03-24", "Insurance — auto",      -185.00, None),
            ("2026-03-26", "Restaurant — Saigon",    -34.20, None),
            ("2026-03-28", "Online purchase — Etsy", -48.75, None),
            ("2026-03-30", "Payroll deposit",       2350.00, None),
            ("2026-03-31", "Bank fee",                -5.00, None),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_event_schedule(src_path: str, _seed: int) -> str:
    """14×4 event schedule: EventID / Title / Date / Status.

    Target: dedup events; conditional fmt past dates.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Events"
        ws.append(["EventID", "Title", "Date", "Status"])
        rows = [
            ("E-001", "Spring Hackathon",     "2026-03-12", "Confirmed"),
            ("E-002", "Quarterly Townhall",   "2026-04-05", "Confirmed"),
            ("E-003", "Designer Meetup",      "2026-04-18", "Cancelled"),
            ("E-004", "Product Launch",       "2026-05-02", "Confirmed"),
            ("E-005", "Customer Workshop",    "2026-05-15", "Pending"),
            ("E-006", "Spring Hackathon",     "2026-03-12", "Confirmed"),
            ("E-007", "Engineering Offsite",  "2026-06-08", "Confirmed"),
            ("E-008", "Open House",           "2026-02-22", "Confirmed"),
            ("E-009", "Recruiting Fair",      "2026-04-26", "Pending"),
            ("E-010", "Designer Meetup",      "2026-04-18", "Cancelled"),
            ("E-011", "Annual Dinner",        "2026-07-14", "Pending"),
            ("E-012", "Customer Workshop",    "2026-05-15", "Pending"),
            ("E-013", "Code Review Bootcamp", "2026-06-22", "Confirmed"),
            ("E-014", "Open House",           "2026-02-22", "Confirmed"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_product_catalog(src_path: str, _seed: int) -> str:
    """18×4 product catalog: SKU / Product / Price / Category.

    Target: multi-col sort, filter by category.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalog"
        ws.append(["SKU", "Product", "Price", "Category"])
        rows = [
            ("P-00012", "Hex Driver Set",      24.95, "Tools"),
            ("P-00027", "Cordless Drill",     189.00, "Tools"),
            ("P-00031", "Halogen Bulb 4-pk",   18.40, "Lighting"),
            ("P-00045", "LED Strip 5m",        32.25, "Lighting"),
            ("P-00058", "Garden Trowel",       12.80, "Garden"),
            ("P-00064", "Pruning Shears",      27.60, "Garden"),
            ("P-00073", "Wheelbarrow",         98.50, "Garden"),
            ("P-00081", "Bluetooth Speaker",   65.00, "Electronics"),
            ("P-00092", "USB-C Hub",           42.15, "Electronics"),
            ("P-00104", "Mechanical Keyboard",129.95, "Electronics"),
            ("P-00118", "Yoga Mat",            38.00, "Fitness"),
            ("P-00125", "Resistance Band Set", 24.50, "Fitness"),
            ("P-00139", "Dumbbell 10kg",       58.20, "Fitness"),
            ("P-00148", "Cast Iron Skillet",   54.80, "Kitchen"),
            ("P-00156", "Chef Knife 8in",      89.95, "Kitchen"),
            ("P-00162", "Wooden Cutting Board",36.40, "Kitchen"),
            ("P-00174", "Coffee Grinder",      72.30, "Kitchen"),
            ("P-00188", "Tea Infuser Pot",     28.95, "Kitchen"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_clinic_visits(src_path: str, _seed: int) -> str:
    """20×5 clinic visits: PatientID / Age / VisitType / Department / Cost.

    Target: filter by visit type; group by department.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Visits"
        ws.append(["PatientID", "Age", "VisitType", "Department", "Cost"])
        rows = [
            ("PT-1001", 34, "Routine",   "Cardiology",  240.00),
            ("PT-1002", 58, "Emergency", "Cardiology", 1280.50),
            ("PT-1003", 12, "Routine",   "Pediatrics",  180.00),
            ("PT-1004", 45, "Routine",   "Dermatology", 195.00),
            ("PT-1005", 71, "Emergency", "Cardiology", 1450.00),
            ("PT-1006", 28, "Followup",  "Orthopedics", 320.00),
            ("PT-1007",  8, "Routine",   "Pediatrics",  180.00),
            ("PT-1008", 62, "Followup",  "Cardiology",  290.00),
            ("PT-1009", 41, "Routine",   "Dermatology", 195.00),
            ("PT-1010", 17, "Emergency", "Orthopedics",  920.00),
            ("PT-1011", 33, "Routine",   "Cardiology",  240.00),
            ("PT-1012", 55, "Emergency", "Cardiology", 1380.00),
            ("PT-1013", 22, "Followup",  "Dermatology", 220.00),
            ("PT-1014",  5, "Routine",   "Pediatrics",  180.00),
            ("PT-1015", 67, "Routine",   "Cardiology",  240.00),
            ("PT-1016", 38, "Followup",  "Orthopedics", 320.00),
            ("PT-1017", 51, "Emergency", "Cardiology", 1620.00),
            ("PT-1018", 26, "Routine",   "Dermatology", 195.00),
            ("PT-1019", 14, "Followup",  "Pediatrics",  210.00),
            ("PT-1020", 49, "Routine",   "Orthopedics", 280.00),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_survey_responses(src_path: str, _seed: int) -> str:
    """22×4 survey responses: RespondentID / Region / Rating / Complete.

    Includes deliberate duplicates (RespondentID) for dedup tasks.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Survey"
        ws.append(["RespondentID", "Region", "Rating", "Complete"])
        rows = [
            ("R-101", "Northeast", 4, "Yes"),
            ("R-102", "South",     5, "Yes"),
            ("R-103", "West",      3, "No"),
            ("R-104", "Midwest",   4, "Yes"),
            ("R-105", "South",     2, "No"),
            ("R-106", "Northeast", 5, "Yes"),
            ("R-101", "Northeast", 4, "Yes"),
            ("R-107", "West",      4, "Yes"),
            ("R-108", "Midwest",   3, "No"),
            ("R-109", "South",     5, "Yes"),
            ("R-103", "West",      3, "No"),
            ("R-110", "Northeast", 4, "Yes"),
            ("R-111", "Midwest",   1, "No"),
            ("R-112", "South",     5, "Yes"),
            ("R-113", "West",      4, "Yes"),
            ("R-114", "Northeast", 3, "No"),
            ("R-105", "South",     2, "No"),
            ("R-115", "Midwest",   5, "Yes"),
            ("R-116", "West",      4, "Yes"),
            ("R-117", "South",     3, "No"),
            ("R-118", "Northeast", 5, "Yes"),
            ("R-119", "Midwest",   4, "Yes"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_temperature_log(src_path: str, _seed: int) -> str:
    """24×3 hourly temperature log: Timestamp / Station / TempC.

    Target: hour extraction, decimal formatting.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Temperature"
        ws.append(["Timestamp", "Station", "TempC"])
        rows = [
            ("2026-04-15 00:00", "NS-1",  9.4),
            ("2026-04-15 01:00", "NS-1",  8.9),
            ("2026-04-15 02:00", "NS-1",  8.5),
            ("2026-04-15 03:00", "NS-1",  8.1),
            ("2026-04-15 04:00", "NS-1",  7.8),
            ("2026-04-15 05:00", "NS-1",  7.6),
            ("2026-04-15 06:00", "NS-1",  8.2),
            ("2026-04-15 07:00", "NS-1",  9.7),
            ("2026-04-15 08:00", "NS-1", 11.4),
            ("2026-04-15 09:00", "NS-1", 13.6),
            ("2026-04-15 10:00", "NS-1", 15.8),
            ("2026-04-15 11:00", "NS-1", 17.5),
            ("2026-04-15 12:00", "NS-1", 19.1),
            ("2026-04-15 13:00", "NS-1", 20.4),
            ("2026-04-15 14:00", "NS-1", 21.2),
            ("2026-04-15 15:00", "NS-1", 21.6),
            ("2026-04-15 16:00", "NS-1", 21.0),
            ("2026-04-15 17:00", "NS-1", 19.8),
            ("2026-04-15 18:00", "NS-1", 18.1),
            ("2026-04-15 19:00", "NS-1", 16.3),
            ("2026-04-15 20:00", "NS-1", 14.5),
            ("2026-04-15 21:00", "NS-1", 12.8),
            ("2026-04-15 22:00", "NS-1", 11.2),
            ("2026-04-15 23:00", "NS-1", 10.1),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_commute_log(src_path: str, _seed: int) -> str:
    """20×4 commute log: Date / Mode / Miles / Minutes.

    Target: running total miles; sort by minutes.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Commute"
        ws.append(["Date", "Mode", "Miles", "Minutes"])
        rows = [
            ("2026-04-01", "Car",   12.4, 28),
            ("2026-04-02", "Bike",   8.2, 35),
            ("2026-04-03", "Car",   12.4, 32),
            ("2026-04-04", "Transit",10.5, 48),
            ("2026-04-05", "Car",   12.4, 26),
            ("2026-04-08", "Bike",   8.2, 33),
            ("2026-04-09", "Car",   12.4, 30),
            ("2026-04-10", "Transit",10.5, 45),
            ("2026-04-11", "Car",   12.4, 27),
            ("2026-04-12", "Bike",   8.2, 36),
            ("2026-04-15", "Car",   12.4, 29),
            ("2026-04-16", "Transit",10.5, 50),
            ("2026-04-17", "Bike",   8.2, 34),
            ("2026-04-18", "Car",   12.4, 31),
            ("2026-04-19", "Car",   12.4, 28),
            ("2026-04-22", "Bike",   8.2, 32),
            ("2026-04-23", "Transit",10.5, 47),
            ("2026-04-24", "Car",   12.4, 26),
            ("2026-04-25", "Car",   12.4, 33),
            ("2026-04-26", "Bike",   8.2, 35),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_subscription_data(src_path: str, _seed: int) -> str:
    """16×4 subscription data: SubID / StartDate / Plan / Status.

    Target: filter active; conditional fmt cancelled.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subscriptions"
        ws.append(["SubID", "StartDate", "Plan", "Status"])
        rows = [
            ("SUB-2001", "2025-09-12", "Basic",   "Active"),
            ("SUB-2002", "2025-10-04", "Pro",     "Active"),
            ("SUB-2003", "2025-08-22", "Pro",     "Cancelled"),
            ("SUB-2004", "2026-01-18", "Basic",   "Active"),
            ("SUB-2005", "2025-11-30", "Premium", "Active"),
            ("SUB-2006", "2025-07-15", "Basic",   "Cancelled"),
            ("SUB-2007", "2026-02-10", "Premium", "Active"),
            ("SUB-2008", "2025-12-04", "Pro",     "Paused"),
            ("SUB-2009", "2026-01-25", "Basic",   "Active"),
            ("SUB-2010", "2025-10-19", "Premium", "Cancelled"),
            ("SUB-2011", "2026-02-28", "Pro",     "Active"),
            ("SUB-2012", "2025-09-07", "Basic",   "Paused"),
            ("SUB-2013", "2026-03-14", "Premium", "Active"),
            ("SUB-2014", "2025-11-12", "Pro",     "Active"),
            ("SUB-2015", "2025-08-30", "Basic",   "Cancelled"),
            ("SUB-2016", "2026-03-22", "Premium", "Active"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_recipe_ingredients(src_path: str, _seed: int) -> str:
    """18×4 recipe ingredients: Recipe / Ingredient / Qty / UnitCost.

    Target: dedup ingredients; derived line cost.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recipes"
        ws.append(["Recipe", "Ingredient", "Qty", "UnitCost"])
        rows = [
            ("Pasta Primavera", "Pasta",      2.0, 1.40),
            ("Pasta Primavera", "Olive Oil",  0.1, 6.50),
            ("Pasta Primavera", "Garlic",     0.05,3.20),
            ("Pasta Primavera", "Zucchini",   1.0, 1.80),
            ("Pasta Primavera", "Parmesan",   0.2, 12.40),
            ("Veggie Stir Fry", "Tofu",       1.0, 3.40),
            ("Veggie Stir Fry", "Soy Sauce",  0.05,4.20),
            ("Veggie Stir Fry", "Bell Pepper",1.0, 1.60),
            ("Veggie Stir Fry", "Garlic",     0.05,3.20),
            ("Veggie Stir Fry", "Rice",       1.5, 2.10),
            ("Tomato Soup",     "Tomato",     2.0, 1.90),
            ("Tomato Soup",     "Onion",      0.5, 1.10),
            ("Tomato Soup",     "Olive Oil",  0.05,6.50),
            ("Tomato Soup",     "Basil",      0.02,8.80),
            ("Apple Pie",       "Apple",      3.0, 0.90),
            ("Apple Pie",       "Flour",      1.0, 1.20),
            ("Apple Pie",       "Sugar",      0.5, 0.80),
            ("Apple Pie",       "Cinnamon",   0.02,9.50),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_student_grades_multi(src_path: str, _seed: int) -> str:
    """24×5 multi-subject grades: StudentID / Name / Math / Science / English.

    Target: column reorder; derived average; conditional fmt low scores.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grades"
        ws.append(["StudentID", "Name", "Math", "Science", "English"])
        rows = [
            ("S-001", "Aiden Park",        88, 91, 84),
            ("S-002", "Beatriz Silva",     72, 78, 81),
            ("S-003", "Caleb Singh",       95, 89, 92),
            ("S-004", "Diana Lopez",       58, 65, 72),
            ("S-005", "Ethan Brooks",      81, 84, 78),
            ("S-006", "Fiona Park",        94, 96, 90),
            ("S-007", "Gavin Cho",         63, 70, 68),
            ("S-008", "Hana Tanaka",       87, 82, 89),
            ("S-009", "Ian Murphy",        55, 60, 64),
            ("S-010", "Julia Romero",      76, 82, 80),
            ("S-011", "Kenji Nakamura",    91, 88, 86),
            ("S-012", "Lila Khan",         68, 71, 74),
            ("S-013", "Marcus Chen",       83, 79, 88),
            ("S-014", "Nora Patel",        92, 94, 91),
            ("S-015", "Owen Brooks",       49, 54, 60),
            ("S-016", "Priya Mehta",       86, 81, 87),
            ("S-017", "Quinn Adams",       77, 73, 79),
            ("S-018", "Rosa Garcia",       93, 90, 95),
            ("S-019", "Sam Lee",           70, 67, 72),
            ("S-020", "Tomas Reyes",       65, 72, 70),
            ("S-021", "Uma Singh",         88, 85, 91),
            ("S-022", "Vince Romano",      52, 58, 61),
            ("S-023", "Willow Park",       79, 81, 76),
            ("S-024", "Xenia Volkov",      96, 93, 94),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_library_loans(src_path: str, _seed: int) -> str:
    """16×5 library loans: LoanID / Book / Patron / DueDate / Returned.

    Target: filter overdue; conditional fmt by Returned status.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Loans"
        ws.append(["LoanID", "Book", "Patron", "DueDate", "Returned"])
        rows = [
            ("LN-301", "The Quiet Garden",      "A. Park",     "2026-03-04", "Yes"),
            ("LN-302", "Echoes of Tomorrow",    "B. Silva",    "2026-03-06", "No"),
            ("LN-303", "Salt and Bread",        "C. Singh",    "2026-03-10", "Yes"),
            ("LN-304", "River of Glass",        "D. Lopez",    "2026-03-12", "No"),
            ("LN-305", "Blue Hour",             "E. Brooks",   "2026-03-18", "Yes"),
            ("LN-306", "The Crimson Sail",      "F. Park",     "2026-03-20", "No"),
            ("LN-307", "Window Gardens",        "G. Cho",      "2026-03-22", "Yes"),
            ("LN-308", "Moonless Night",        "H. Tanaka",   "2026-03-25", "Yes"),
            ("LN-309", "Lighthouse Year",       "I. Murphy",   "2026-03-28", "No"),
            ("LN-310", "Stone Lantern",         "J. Romero",   "2026-04-01", "Yes"),
            ("LN-311", "Ashes of October",      "K. Nakamura", "2026-04-04", "No"),
            ("LN-312", "The Gold Coast",        "L. Khan",     "2026-04-06", "Yes"),
            ("LN-313", "Saltwood Mile",         "M. Chen",     "2026-04-08", "Yes"),
            ("LN-314", "Quiet Mile",            "N. Patel",    "2026-04-10", "No"),
            ("LN-315", "Brass Wings",           "O. Brooks",   "2026-04-12", "Yes"),
            ("LN-316", "Tide of Storms",        "P. Mehta",    "2026-04-15", "Yes"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_fleet_vehicles(src_path: str, _seed: int) -> str:
    """14×5 fleet vehicles: Plate / Make / Year / Mileage / Status.

    Target: sort by mileage; filter by status.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fleet"
        ws.append(["Plate", "Make", "Year", "Mileage", "Status"])
        rows = [
            ("FLT-001", "Ford Transit",      2019, 124800, "Active"),
            ("FLT-002", "Toyota HiAce",      2021,  68400, "Active"),
            ("FLT-003", "Ford Transit",      2017, 198500, "Maintenance"),
            ("FLT-004", "Mercedes Sprinter", 2022,  42300, "Active"),
            ("FLT-005", "Toyota HiAce",      2018, 156900, "Retired"),
            ("FLT-006", "Ram ProMaster",     2020,  98700, "Active"),
            ("FLT-007", "Ford Transit",      2016, 224100, "Retired"),
            ("FLT-008", "Mercedes Sprinter", 2023,  18600, "Active"),
            ("FLT-009", "Ram ProMaster",     2021,  76200, "Maintenance"),
            ("FLT-010", "Toyota HiAce",      2020, 108300, "Active"),
            ("FLT-011", "Ford Transit",      2018, 168700, "Maintenance"),
            ("FLT-012", "Mercedes Sprinter", 2021,  84500, "Active"),
            ("FLT-013", "Ram ProMaster",     2019, 132600, "Active"),
            ("FLT-014", "Toyota HiAce",      2022,  31900, "Active"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_employees_payroll(src_path: str, _seed: int) -> str:
    """18×5 payroll: EmpID / Name / Dept / Salary / Bonus.

    Target: group by dept totals; sort by salary.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payroll"
        ws.append(["EmpID", "Name", "Dept", "Salary", "Bonus"])
        rows = [
            ("E-2001", "Alice Wong",       "Engineering", 142000,  8500),
            ("E-2002", "Bruno Mata",       "Sales",        92000,  6200),
            ("E-2003", "Cara Lin",         "Engineering", 128500,  7400),
            ("E-2004", "Diego Reyes",      "Marketing",    96400,  5100),
            ("E-2005", "Elena Park",       "Engineering", 156800,  9600),
            ("E-2006", "Faisal Khan",      "Sales",        88500,  5800),
            ("E-2007", "Grace Okafor",     "HR",           76200,  3400),
            ("E-2008", "Henry Brooks",     "Marketing",   102300,  5600),
            ("E-2009", "Iris Tanaka",      "Engineering", 134900,  7800),
            ("E-2010", "Jaden Cho",        "Sales",       104600,  7200),
            ("E-2011", "Kira Patel",       "HR",           82100,  3700),
            ("E-2012", "Liam Walsh",       "Engineering", 148200,  8900),
            ("E-2013", "Maya Reddy",       "Marketing",    94700,  5300),
            ("E-2014", "Nico Almeida",     "Sales",        86400,  5500),
            ("E-2015", "Owen Park",        "Engineering", 138600,  7600),
            ("E-2016", "Priya Mehta",      "HR",           78900,  3500),
            ("E-2017", "Quinn Adams",      "Marketing",   108400,  6100),
            ("E-2018", "Rosa Garcia",      "Engineering", 162500, 10200),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_weather_log(src_path: str, _seed: int) -> str:
    """20×4 daily weather: Date / HighC / LowC / Precip_mm.

    Target: derived range; conditional fmt cold/hot days.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weather"
        ws.append(["Date", "HighC", "LowC", "Precip_mm"])
        rows = [
            ("2026-04-01", 14.2,  6.8, 0.0),
            ("2026-04-02", 16.5,  8.1, 0.0),
            ("2026-04-03", 12.4,  5.6, 3.2),
            ("2026-04-04",  9.8,  3.4, 8.6),
            ("2026-04-05", 11.6,  4.9, 1.4),
            ("2026-04-06", 15.3,  7.2, 0.0),
            ("2026-04-07", 18.7, 10.4, 0.0),
            ("2026-04-08", 21.4, 11.8, 0.0),
            ("2026-04-09", 19.6, 12.3, 2.8),
            ("2026-04-10", 17.2,  9.7, 5.1),
            ("2026-04-11", 13.8,  6.4, 12.4),
            ("2026-04-12",  8.5,  2.1, 6.8),
            ("2026-04-13", 10.4,  3.8, 0.0),
            ("2026-04-14", 14.9,  6.2, 0.0),
            ("2026-04-15", 19.1, 10.3, 0.0),
            ("2026-04-16", 22.6, 12.9, 0.0),
            ("2026-04-17", 24.3, 14.5, 0.0),
            ("2026-04-18", 20.8, 12.6, 4.4),
            ("2026-04-19", 16.4,  9.1, 9.3),
            ("2026-04-20", 12.7,  6.0, 1.6),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_restaurant_menu(src_path: str, _seed: int) -> str:
    """16×4 restaurant menu: Item / Category / Price / SoldCount.

    Target: filter by category; conditional fmt top sellers.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Menu"
        ws.append(["Item", "Category", "Price", "SoldCount"])
        rows = [
            ("Caesar Salad",         "Starter",  9.50,  142),
            ("Tomato Soup",          "Starter",  7.80,   98),
            ("Bruschetta",           "Starter",  8.40,  126),
            ("Margherita Pizza",     "Main",    14.50,  308),
            ("Spaghetti Carbonara",  "Main",    16.80,  214),
            ("Grilled Salmon",       "Main",    22.40,  186),
            ("Mushroom Risotto",     "Main",    17.60,  152),
            ("Ribeye Steak",         "Main",    32.50,  124),
            ("Tiramisu",             "Dessert",  8.20,  168),
            ("Chocolate Cake",       "Dessert",  7.80,  204),
            ("Gelato",               "Dessert",  6.40,  186),
            ("Panna Cotta",          "Dessert",  7.20,   92),
            ("Espresso",             "Beverage", 3.50,  412),
            ("Cappuccino",           "Beverage", 4.20,  348),
            ("Sparkling Water",      "Beverage", 3.80,  286),
            ("House Wine Glass",     "Beverage", 7.50,  238),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_donations_log(src_path: str, _seed: int) -> str:
    """18×4 donation log: Donor / Date / Amount / Campaign.

    Target: group by campaign; sort by amount.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Donations"
        ws.append(["Donor", "Date", "Amount", "Campaign"])
        rows = [
            ("Avery N.",     "2026-02-04",  120.00, "SpringFund"),
            ("Brooke T.",    "2026-02-08",   55.00, "EmergencyAid"),
            ("Cory L.",      "2026-02-12",  300.00, "ScholarshipDrive"),
            ("Davis K.",     "2026-02-18",   85.00, "SpringFund"),
            ("Elena P.",     "2026-02-22",  500.00, "ScholarshipDrive"),
            ("Felipe G.",    "2026-02-27",   40.00, "EmergencyAid"),
            ("Grace M.",     "2026-03-02",  200.00, "SpringFund"),
            ("Henry O.",     "2026-03-05",   75.00, "EmergencyAid"),
            ("Iris L.",      "2026-03-08",  150.00, "ScholarshipDrive"),
            ("Joaquin R.",   "2026-03-12",   90.00, "SpringFund"),
            ("Kara S.",      "2026-03-16",  250.00, "ScholarshipDrive"),
            ("Liam D.",      "2026-03-20",   60.00, "EmergencyAid"),
            ("Maya V.",      "2026-03-24",  180.00, "SpringFund"),
            ("Nico A.",      "2026-03-28",  400.00, "ScholarshipDrive"),
            ("Owen B.",      "2026-04-02",   45.00, "EmergencyAid"),
            ("Priya K.",     "2026-04-06",  100.00, "SpringFund"),
            ("Quinn W.",     "2026-04-10",  220.00, "ScholarshipDrive"),
            ("Rosa H.",      "2026-04-14",   65.00, "EmergencyAid"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_books_catalog(src_path: str, _seed: int) -> str:
    """20×5 books: ISBN / Title / Author / Year / Pages.

    Target: sort by year; sort alpha by author.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Books"
        ws.append(["ISBN", "Title", "Author", "Year", "Pages"])
        rows = [
            ("978-0-001", "Quiet Streams",          "Adler, J.",    2014, 312),
            ("978-0-002", "Borrowed Light",         "Beckett, M.",  2008, 248),
            ("978-0-003", "Calderwood",             "Chen, P.",     2019, 416),
            ("978-0-004", "Dune Maps",              "Davies, R.",   2002, 284),
            ("978-0-005", "Echo Plains",            "Eklund, A.",   2016, 352),
            ("978-0-006", "Falling Forward",        "Fontana, S.",  2010, 192),
            ("978-0-007", "Glasswright",            "Garcia, L.",   2021, 408),
            ("978-0-008", "Hollow Years",           "Hayashi, K.",  2005, 276),
            ("978-0-009", "Indigo Salt",            "Iverson, T.",  2018, 328),
            ("978-0-010", "Junipers",               "Jordan, F.",   2011, 240),
            ("978-0-011", "Kestrel Days",           "Khan, A.",     2023, 384),
            ("978-0-012", "Lighthouse Trails",      "Liu, Y.",      2007, 296),
            ("978-0-013", "Maplecourt",             "Mahmoud, R.",  2015, 360),
            ("978-0-014", "Narrow Bridges",         "Navarro, E.",  2020, 432),
            ("978-0-015", "Of Stone and Wind",      "O'Hara, P.",   2003, 304),
            ("978-0-016", "Petals & Print",         "Park, M.",     2017, 268),
            ("978-0-017", "Quarry Notes",           "Quinn, S.",    2012, 256),
            ("978-0-018", "River Glass",            "Reyes, D.",    2022, 392),
            ("978-0-019", "Sundial",                "Singh, V.",    2009, 336),
            ("978-0-020", "Tinder & Tide",          "Tanaka, H.",   2019, 376),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_movie_box_office(src_path: str, _seed: int) -> str:
    """16×4 movie box office: Title / OpeningWeek / DomesticTotal / IntlTotal.

    Target: derived global total; sort.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BoxOffice"
        ws.append(["Title", "OpeningWeek", "DomesticTotal", "IntlTotal"])
        rows = [
            ("Stellar Drift",        45_200_000, 184_500_000, 213_800_000),
            ("Brass Garden",         18_600_000,  72_400_000,  48_100_000),
            ("Echo Plains",          62_800_000, 248_700_000, 312_400_000),
            ("Hollow Years",         12_400_000,  56_200_000,  29_800_000),
            ("Tomorrow's Glass",     38_900_000, 162_300_000, 187_600_000),
            ("Salt and Bread",       28_400_000, 121_500_000, 134_200_000),
            ("Crimson Sail",         54_100_000, 218_400_000, 246_500_000),
            ("Lighthouse Year",       9_800_000,  41_200_000,  22_300_000),
            ("Window Gardens",       21_600_000,  92_700_000,  74_800_000),
            ("Moonless Night",       42_500_000, 178_600_000, 198_400_000),
            ("Gold Coast",           16_300_000,  68_900_000,  43_500_000),
            ("Saltwood Mile",         7_200_000,  31_400_000,  19_600_000),
            ("Quiet Mile",           24_800_000, 104_200_000,  89_400_000),
            ("Brass Wings",          49_700_000, 196_800_000, 224_300_000),
            ("Tide of Storms",       33_200_000, 138_600_000, 156_900_000),
            ("Unwritten",            14_100_000,  62_400_000,  38_700_000),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_real_estate_listings(src_path: str, _seed: int) -> str:
    """18×5 real estate listings: ListingID / City / Beds / SqFt / Price.

    Target: sort by price; filter by city; derived price-per-sqft.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Listings"
        ws.append(["ListingID", "City", "Beds", "SqFt", "Price"])
        rows = [
            ("RE-101", "Portland",  3, 1840,  618000),
            ("RE-102", "Seattle",   2, 1240,  745000),
            ("RE-103", "Portland",  4, 2340,  825000),
            ("RE-104", "Boise",     3, 1680,  412000),
            ("RE-105", "Seattle",   3, 1920,  892000),
            ("RE-106", "Boise",     2, 1080,  328000),
            ("RE-107", "Portland",  2, 1320,  548000),
            ("RE-108", "Spokane",   3, 1740,  362000),
            ("RE-109", "Seattle",   4, 2780, 1240000),
            ("RE-110", "Spokane",   2, 1180,  264000),
            ("RE-111", "Boise",     4, 2480,  582000),
            ("RE-112", "Portland",  3, 1960,  698000),
            ("RE-113", "Spokane",   3, 1820,  398000),
            ("RE-114", "Seattle",   2, 1380,  812000),
            ("RE-115", "Boise",     3, 1620,  386000),
            ("RE-116", "Portland",  4, 2640,  912000),
            ("RE-117", "Spokane",   4, 2360,  468000),
            ("RE-118", "Seattle",   3, 2080, 1080000),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_purchase_orders(src_path: str, _seed: int) -> str:
    """18×4 purchase orders: PO# / Vendor / Amount / Status.

    Target: group by vendor; filter by status.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "POs"
        ws.append(["PO#", "Vendor", "Amount", "Status"])
        rows = [
            ("PO-9001", "Acme Supplies",     2480.50, "Approved"),
            ("PO-9002", "Bayside Materials", 1845.00, "Pending"),
            ("PO-9003", "Cypress Goods",     3920.40, "Approved"),
            ("PO-9004", "Driftwood Trade",   1124.40, "Rejected"),
            ("PO-9005", "Evergreen Mill",    2104.80, "Approved"),
            ("PO-9006", "Forge & Co",         580.20, "Pending"),
            ("PO-9007", "Granite Hold",      4580.00, "Approved"),
            ("PO-9008", "Harbor Freight",     895.65, "Pending"),
            ("PO-9009", "Acme Supplies",     1740.10, "Rejected"),
            ("PO-9010", "Juniper Studios",   1295.40, "Approved"),
            ("PO-9011", "Bayside Materials", 2580.00, "Approved"),
            ("PO-9012", "Cypress Goods",     1490.00, "Pending"),
            ("PO-9013", "Forge & Co",        2840.30, "Approved"),
            ("PO-9014", "Granite Hold",      3260.50, "Pending"),
            ("PO-9015", "Acme Supplies",      675.20, "Approved"),
            ("PO-9016", "Harbor Freight",    1080.40, "Rejected"),
            ("PO-9017", "Juniper Studios",   2240.80, "Approved"),
            ("PO-9018", "Evergreen Mill",     940.50, "Pending"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_employee_skills(src_path: str, _seed: int) -> str:
    """16×4 employee skills matrix: EmpID / Name / Skill / Level."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Skills"
        ws.append(["EmpID", "Name", "Skill", "Level"])
        rows = [
            ("E-101", "Alice Wong",   "Python",     "Expert"),
            ("E-102", "Bruno Mata",   "JavaScript", "Intermediate"),
            ("E-103", "Cara Lin",     "Python",     "Intermediate"),
            ("E-104", "Diego Reyes",  "SQL",        "Expert"),
            ("E-105", "Elena Park",   "Java",       "Beginner"),
            ("E-106", "Faisal Khan",  "Python",     "Beginner"),
            ("E-107", "Grace Okafor", "SQL",        "Intermediate"),
            ("E-108", "Henry Brooks", "JavaScript", "Expert"),
            ("E-109", "Iris Tanaka",  "Java",       "Expert"),
            ("E-110", "Jaden Cho",    "Python",     "Expert"),
            ("E-111", "Kira Patel",   "SQL",        "Beginner"),
            ("E-112", "Liam Walsh",   "Java",       "Intermediate"),
            ("E-113", "Maya Reddy",   "JavaScript", "Beginner"),
            ("E-114", "Nico Almeida", "Python",     "Intermediate"),
            ("E-115", "Owen Park",    "SQL",        "Expert"),
            ("E-116", "Priya Mehta",  "JavaScript", "Intermediate"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_warehouse_orders(src_path: str, _seed: int) -> str:
    """20×4 warehouse outbound orders: OrderID / Warehouse / Items / Weight_kg."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Outbound"
        ws.append(["OrderID", "Warehouse", "Items", "Weight_kg"])
        rows = [
            ("WO-2001", "WH-East",   12,  82.4),
            ("WO-2002", "WH-East",    8,  56.2),
            ("WO-2003", "WH-West",   18, 124.6),
            ("WO-2004", "WH-South",   5,  31.8),
            ("WO-2005", "WH-East",   14,  92.7),
            ("WO-2006", "WH-West",   10,  68.1),
            ("WO-2007", "WH-South",  22, 154.3),
            ("WO-2008", "WH-North",  16, 108.9),
            ("WO-2009", "WH-East",    7,  48.5),
            ("WO-2010", "WH-West",   25, 178.4),
            ("WO-2011", "WH-South",  11,  74.2),
            ("WO-2012", "WH-North",   9,  62.7),
            ("WO-2013", "WH-East",   13,  86.1),
            ("WO-2014", "WH-West",   19, 132.8),
            ("WO-2015", "WH-South",   6,  39.5),
            ("WO-2016", "WH-North",  20, 138.6),
            ("WO-2017", "WH-East",   15, 102.4),
            ("WO-2018", "WH-West",   17, 118.9),
            ("WO-2019", "WH-South",  23, 162.5),
            ("WO-2020", "WH-North",  12,  84.7),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_taxi_fares(src_path: str, _seed: int) -> str:
    """18×4 taxi fares: TripID / Distance_km / Duration_min / Fare_USD."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Fares"
        ws.append(["TripID", "Distance_km", "Duration_min", "Fare_USD"])
        rows = [
            ("T-3001",  4.2, 12, 12.40),
            ("T-3002",  8.7, 23, 21.80),
            ("T-3003",  2.1,  8,  7.60),
            ("T-3004", 12.5, 35, 32.50),
            ("T-3005",  6.3, 18, 16.20),
            ("T-3006",  3.4, 11,  9.80),
            ("T-3007", 18.2, 48, 46.40),
            ("T-3008",  5.6, 14, 14.30),
            ("T-3009",  9.4, 26, 24.60),
            ("T-3010", 14.7, 41, 38.20),
            ("T-3011",  2.8, 10,  8.40),
            ("T-3012",  7.1, 19, 18.50),
            ("T-3013", 11.3, 32, 28.40),
            ("T-3014",  4.5, 13, 12.90),
            ("T-3015", 16.8, 44, 42.60),
            ("T-3016",  6.9, 20, 17.80),
            ("T-3017",  3.2, 11,  9.40),
            ("T-3018", 10.4, 28, 26.50),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_workout_log(src_path: str, _seed: int) -> str:
    """20×4 workout log: Date / Activity / Duration_min / CaloriesBurned."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Workouts"
        ws.append(["Date", "Activity", "Duration_min", "CaloriesBurned"])
        rows = [
            ("2026-04-01", "Run",      35, 412),
            ("2026-04-02", "Yoga",     45, 198),
            ("2026-04-03", "Cycle",    50, 528),
            ("2026-04-04", "Swim",     40, 462),
            ("2026-04-05", "Run",      30, 348),
            ("2026-04-06", "Strength", 60, 386),
            ("2026-04-07", "Yoga",     45, 198),
            ("2026-04-08", "Cycle",    55, 580),
            ("2026-04-09", "Run",      40, 472),
            ("2026-04-10", "Swim",     35, 405),
            ("2026-04-11", "Strength", 55, 358),
            ("2026-04-12", "Run",      45, 528),
            ("2026-04-13", "Yoga",     50, 220),
            ("2026-04-14", "Cycle",    60, 632),
            ("2026-04-15", "Swim",     45, 522),
            ("2026-04-16", "Strength", 50, 326),
            ("2026-04-17", "Run",      40, 472),
            ("2026-04-18", "Cycle",    65, 684),
            ("2026-04-19", "Yoga",     40, 178),
            ("2026-04-20", "Swim",     50, 580),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_apartment_rents(src_path: str, _seed: int) -> str:
    """16×4 apartment rents: Building / Unit / Beds / MonthlyRent."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rents"
        ws.append(["Building", "Unit", "Beds", "MonthlyRent"])
        rows = [
            ("A", "101", 1, 1850),
            ("A", "102", 1, 1875),
            ("A", "201", 2, 2480),
            ("A", "202", 2, 2520),
            ("B", "101", 1, 1920),
            ("B", "102", 2, 2640),
            ("B", "201", 3, 3280),
            ("B", "202", 2, 2680),
            ("C", "101", 1, 1780),
            ("C", "102", 1, 1820),
            ("C", "201", 2, 2380),
            ("C", "202", 3, 3120),
            ("D", "101", 2, 2540),
            ("D", "102", 1, 1780),
            ("D", "201", 3, 3320),
            ("D", "202", 2, 2580),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_blog_posts(src_path: str, _seed: int) -> str:
    """16×4 blog posts: PostID / Title / Author / ViewCount."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Posts"
        ws.append(["PostID", "Title", "Author", "ViewCount"])
        rows = [
            ("BP-001", "Getting Started with Rust",        "A. Park",     12480),
            ("BP-002", "A Tour of the Type System",        "B. Silva",     8240),
            ("BP-003", "Async Patterns Explained",         "C. Singh",    21380),
            ("BP-004", "Memory Safety Without GC",         "D. Lopez",    15720),
            ("BP-005", "Macros: A Practical Guide",        "E. Brooks",    6450),
            ("BP-006", "Building CLIs with clap",          "F. Park",     18620),
            ("BP-007", "Lifetimes Demystified",            "G. Cho",      24100),
            ("BP-008", "Performance Tuning Tips",          "H. Tanaka",   11280),
            ("BP-009", "Embedded Rust Basics",             "I. Murphy",    5860),
            ("BP-010", "Iterator Combinators",             "J. Romero",   14580),
            ("BP-011", "Error Handling Patterns",          "K. Nakamura", 19340),
            ("BP-012", "WebAssembly with Rust",            "L. Khan",     16240),
            ("BP-013", "Testing Strategies",               "M. Chen",      9420),
            ("BP-014", "Refactoring Legacy Code",          "N. Patel",    13680),
            ("BP-015", "Building a REST API",              "O. Brooks",   22480),
            ("BP-016", "Concurrency Without Fear",         "P. Mehta",    17320),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_classroom_quiz(src_path: str, _seed: int) -> str:
    """20×4 classroom quiz: QuestionID / Topic / Difficulty / CorrectPct."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quiz"
        ws.append(["QuestionID", "Topic", "Difficulty", "CorrectPct"])
        rows = [
            ("Q-101", "Algebra",   "Easy",   84.2),
            ("Q-102", "Algebra",   "Medium", 62.5),
            ("Q-103", "Geometry",  "Easy",   78.6),
            ("Q-104", "Geometry",  "Hard",   34.8),
            ("Q-105", "Calculus",  "Medium", 56.4),
            ("Q-106", "Calculus",  "Hard",   28.2),
            ("Q-107", "Algebra",   "Hard",   42.7),
            ("Q-108", "Statistics","Easy",   76.3),
            ("Q-109", "Statistics","Medium", 58.1),
            ("Q-110", "Geometry",  "Medium", 64.5),
            ("Q-111", "Calculus",  "Easy",   72.8),
            ("Q-112", "Statistics","Hard",   38.4),
            ("Q-113", "Algebra",   "Easy",   86.4),
            ("Q-114", "Geometry",  "Easy",   80.2),
            ("Q-115", "Calculus",  "Medium", 54.7),
            ("Q-116", "Statistics","Easy",   78.9),
            ("Q-117", "Algebra",   "Medium", 64.3),
            ("Q-118", "Geometry",  "Hard",   32.6),
            ("Q-119", "Calculus",  "Hard",   26.8),
            ("Q-120", "Statistics","Medium", 60.5),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_film_festival(src_path: str, _seed: int) -> str:
    """16×4 film festival: Title / Country / RuntimeMin / Rating."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Festival"
        ws.append(["Title", "Country", "RuntimeMin", "Rating"])
        rows = [
            ("Stellar Drift",        "USA",     128, 8.4),
            ("Brass Garden",         "Germany",  98, 7.6),
            ("Echo Plains",          "Japan",   142, 8.8),
            ("Hollow Years",         "France",  104, 7.2),
            ("Tomorrow's Glass",     "Korea",   116, 8.1),
            ("Salt and Bread",       "Spain",   132, 8.3),
            ("Crimson Sail",         "USA",     124, 7.9),
            ("Lighthouse Year",      "Norway",   86, 6.8),
            ("Window Gardens",       "Italy",    96, 7.4),
            ("Moonless Night",       "Japan",   118, 8.2),
            ("Gold Coast",           "Brazil",  108, 7.6),
            ("Saltwood Mile",        "Mexico",   92, 7.0),
            ("Quiet Mile",           "France",  102, 7.8),
            ("Brass Wings",          "USA",     136, 8.5),
            ("Tide of Storms",       "Iceland", 114, 7.7),
            ("Unwritten",            "UK",       88, 7.1),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_charity_pledges(src_path: str, _seed: int) -> str:
    """16×4 charity pledges: PledgeID / Donor / Amount / Type."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pledges"
        ws.append(["PledgeID", "Donor", "Amount", "Type"])
        rows = [
            ("PL-401", "A. Park",     250.00, "OneTime"),
            ("PL-402", "B. Silva",     50.00, "Monthly"),
            ("PL-403", "C. Singh",    500.00, "OneTime"),
            ("PL-404", "D. Lopez",    100.00, "Monthly"),
            ("PL-405", "E. Brooks",   125.00, "OneTime"),
            ("PL-406", "F. Park",      75.00, "Monthly"),
            ("PL-407", "G. Cho",     1000.00, "OneTime"),
            ("PL-408", "H. Tanaka",    40.00, "Monthly"),
            ("PL-409", "I. Murphy",    80.00, "OneTime"),
            ("PL-410", "J. Romero",   150.00, "Monthly"),
            ("PL-411", "K. Nakamura", 300.00, "OneTime"),
            ("PL-412", "L. Khan",      60.00, "Monthly"),
            ("PL-413", "M. Chen",     200.00, "OneTime"),
            ("PL-414", "N. Patel",     90.00, "Monthly"),
            ("PL-415", "O. Brooks",   175.00, "OneTime"),
            ("PL-416", "P. Mehta",     45.00, "Monthly"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_software_bugs(src_path: str, _seed: int) -> str:
    """20×4 software bug log: BugID / Severity / Status / DaysOpen."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bugs"
        ws.append(["BugID", "Severity", "Status", "DaysOpen"])
        rows = [
            ("BUG-001", "Critical", "Open",       12),
            ("BUG-002", "Major",    "Resolved",   28),
            ("BUG-003", "Minor",    "Open",        5),
            ("BUG-004", "Critical", "Resolved",    8),
            ("BUG-005", "Major",    "Open",       18),
            ("BUG-006", "Minor",    "Resolved",   42),
            ("BUG-007", "Major",    "Open",        4),
            ("BUG-008", "Critical", "Open",        2),
            ("BUG-009", "Minor",    "Open",       16),
            ("BUG-010", "Major",    "Resolved",   34),
            ("BUG-011", "Minor",    "Resolved",   38),
            ("BUG-012", "Critical", "Open",        7),
            ("BUG-013", "Major",    "Open",       22),
            ("BUG-014", "Minor",    "Resolved",   48),
            ("BUG-015", "Critical", "Resolved",   14),
            ("BUG-016", "Major",    "Open",       11),
            ("BUG-017", "Minor",    "Open",       26),
            ("BUG-018", "Critical", "Open",        3),
            ("BUG-019", "Major",    "Resolved",   30),
            ("BUG-020", "Minor",    "Resolved",   52),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_hotel_bookings(src_path: str, _seed: int) -> str:
    """16×5 hotel bookings: BookingID / Guest / Room / Nights / TotalUSD."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bookings"
        ws.append(["BookingID", "Guest", "Room", "Nights", "TotalUSD"])
        rows = [
            ("HB-501", "A. PARK",     "Standard",  3,  348.00),
            ("HB-502", "B. SILVA",    "Suite",     2,  580.00),
            ("HB-503", "C. SINGH",    "Standard",  4,  464.00),
            ("HB-504", "D. LOPEZ",    "Deluxe",    5,  780.00),
            ("HB-505", "E. BROOKS",   "Suite",     3,  870.00),
            ("HB-506", "F. PARK",     "Standard",  2,  232.00),
            ("HB-507", "G. CHO",      "Deluxe",    4,  624.00),
            ("HB-508", "H. TANAKA",   "Suite",     5, 1450.00),
            ("HB-509", "I. MURPHY",   "Standard",  3,  348.00),
            ("HB-510", "J. ROMERO",   "Deluxe",    6,  936.00),
            ("HB-511", "K. NAKAMURA", "Suite",     4, 1160.00),
            ("HB-512", "L. KHAN",     "Standard",  2,  232.00),
            ("HB-513", "M. CHEN",     "Deluxe",    3,  468.00),
            ("HB-514", "N. PATEL",    "Suite",     7, 2030.00),
            ("HB-515", "O. BROOKS",   "Standard",  4,  464.00),
            ("HB-516", "P. MEHTA",    "Deluxe",    5,  780.00),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_app_downloads(src_path: str, _seed: int) -> str:
    """16×4 app downloads: AppID / Platform / DownloadCount / RatingAvg."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AppStats"
        ws.append(["AppID", "Platform", "DownloadCount", "RatingAvg"])
        rows = [
            ("APP-001", "iOS",     124800, 4.6),
            ("APP-002", "Android", 218400, 4.4),
            ("APP-003", "iOS",      48200, 4.2),
            ("APP-004", "Android",  92400, 4.0),
            ("APP-005", "iOS",     186300, 4.7),
            ("APP-006", "Android", 312500, 4.3),
            ("APP-007", "iOS",      72400, 4.1),
            ("APP-008", "Android", 148600, 4.5),
            ("APP-009", "iOS",      36500, 3.9),
            ("APP-010", "Android", 248700, 4.4),
            ("APP-011", "iOS",      96400, 4.5),
            ("APP-012", "Android", 178200, 4.2),
            ("APP-013", "iOS",     142800, 4.6),
            ("APP-014", "Android", 268500, 4.3),
            ("APP-015", "iOS",      58200, 4.0),
            ("APP-016", "Android", 198400, 4.4),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_country_capitals(src_path: str, _seed: int) -> str:
    """20×4 country capitals: Country / Capital / Continent / Pop_M."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Capitals"
        ws.append(["Country", "Capital", "Continent", "Pop_M"])
        rows = [
            ("Argentina",  "Buenos Aires",   "South America",   3.1),
            ("Brazil",     "Brasilia",       "South America",   3.1),
            ("Canada",     "Ottawa",         "North America",   1.0),
            ("Egypt",      "Cairo",          "Africa",         20.4),
            ("France",     "Paris",          "Europe",          2.1),
            ("Germany",    "Berlin",         "Europe",          3.6),
            ("India",      "New Delhi",      "Asia",           28.5),
            ("Japan",      "Tokyo",          "Asia",           37.4),
            ("Kenya",      "Nairobi",        "Africa",          4.4),
            ("Mexico",     "Mexico City",    "North America",  21.7),
            ("Netherlands","Amsterdam",      "Europe",          1.1),
            ("Nigeria",    "Abuja",          "Africa",          3.6),
            ("Peru",       "Lima",           "South America",  10.7),
            ("Russia",     "Moscow",         "Europe",         12.5),
            ("South Korea","Seoul",          "Asia",            9.7),
            ("Spain",      "Madrid",         "Europe",          3.3),
            ("Thailand",   "Bangkok",        "Asia",           10.5),
            ("Turkey",     "Ankara",         "Europe",          5.7),
            ("UK",         "London",         "Europe",          9.0),
            ("USA",        "Washington DC",  "North America",   0.7),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_class_attendance(src_path: str, _seed: int) -> str:
    """20×4 class attendance: Date / Course / Section / Headcount.

    Target: group by course; sort by headcount.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ClassAttendance"
        ws.append(["Date", "Course", "Section", "Headcount"])
        rows = [
            ("2026-03-02", "CS101",  "A", 28),
            ("2026-03-02", "CS101",  "B", 24),
            ("2026-03-02", "MATH210","A", 32),
            ("2026-03-02", "MATH210","B", 30),
            ("2026-03-03", "BIO150", "A", 26),
            ("2026-03-03", "BIO150", "B", 25),
            ("2026-03-04", "CS101",  "A", 27),
            ("2026-03-04", "CS101",  "B", 23),
            ("2026-03-04", "MATH210","A", 31),
            ("2026-03-04", "MATH210","B", 29),
            ("2026-03-05", "BIO150", "A", 26),
            ("2026-03-05", "BIO150", "B", 24),
            ("2026-03-09", "CS101",  "A", 28),
            ("2026-03-09", "CS101",  "B", 22),
            ("2026-03-09", "MATH210","A", 30),
            ("2026-03-09", "MATH210","B", 28),
            ("2026-03-10", "BIO150", "A", 25),
            ("2026-03-10", "BIO150", "B", 23),
            ("2026-03-11", "CS101",  "A", 27),
            ("2026-03-11", "MATH210","A", 32),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


# ---------------------------------------------------------------------------
# Final batch: 19 new authored xlsx files (F-CALC-72..F-CALC-90).
# Each file is structurally distinct (different column shape / value
# distribution) so each affords its own task-pair without op-axis cloning.
# ---------------------------------------------------------------------------


def _src_gym_membership(src_path: str, _seed: int) -> str:
    """16×4 gym membership: MemberID / Tier / MonthlyFee / Status."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Members"
        ws.append(["MemberID", "Tier", "MonthlyFee", "Status"])
        rows = [
            ("GM-001", "Basic",   29.00, "Active"),
            ("GM-002", "Premium", 79.00, "Active"),
            ("GM-003", "Basic",   29.00, "Cancelled"),
            ("GM-004", "Pro",     49.00, "Active"),
            ("GM-005", "Premium", 79.00, "Paused"),
            ("GM-006", "Basic",   29.00, "Active"),
            ("GM-007", "Pro",     49.00, "Cancelled"),
            ("GM-008", "Premium", 79.00, "Active"),
            ("GM-009", "Basic",   29.00, "Active"),
            ("GM-010", "Pro",     49.00, "Active"),
            ("GM-011", "Premium", 79.00, "Cancelled"),
            ("GM-012", "Basic",   29.00, "Paused"),
            ("GM-013", "Pro",     49.00, "Active"),
            ("GM-014", "Premium", 79.00, "Active"),
            ("GM-015", "Basic",   29.00, "Active"),
            ("GM-016", "Pro",     49.00, "Cancelled"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_garden_plants(src_path: str, _seed: int) -> str:
    """16×4 garden plants: PlantID / Species / SunNeeded / WaterNeeded."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plants"
        ws.append(["PlantID", "Species", "SunNeeded", "WaterNeeded"])
        rows = [
            ("PL-101", "TOMATO",     "Full",    "High"),
            ("PL-102", "BASIL",      "Full",    "Medium"),
            ("PL-103", "FERN",       "Shade",   "High"),
            ("PL-104", "HOSTA",      "Shade",   "Medium"),
            ("PL-105", "LAVENDER",   "Full",    "Low"),
            ("PL-106", "MINT",       "Partial", "High"),
            ("PL-107", "CACTUS",     "Full",    "Low"),
            ("PL-108", "HYDRANGEA",  "Partial", "High"),
            ("PL-109", "ROSEMARY",   "Full",    "Low"),
            ("PL-110", "IMPATIENS",  "Shade",   "Medium"),
            ("PL-111", "PEPPER",     "Full",    "Medium"),
            ("PL-112", "STRAWBERRY", "Full",    "High"),
            ("PL-113", "COLEUS",     "Shade",   "Medium"),
            ("PL-114", "SAGE",       "Full",    "Low"),
            ("PL-115", "BEGONIA",    "Partial", "Medium"),
            ("PL-116", "SUNFLOWER",  "Full",    "Medium"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_warehouse_skus(src_path: str, _seed: int) -> str:
    """18×4 warehouse SKUs with leading-zero-padded codes: SKU / Item / OnHand / Bin."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SKUs"
        ws.append(["SKU", "Item", "OnHand", "Bin"])
        rows = [
            ("00045", "Hex Bolt M6",     420, "A-01"),
            ("00112", "Wing Nut M5",     180, "A-02"),
            ("00203", "Lock Washer M8",  640, "A-03"),
            ("00318", "Cable Tie 8in",   840, "B-01"),
            ("00427", "Heat Shrink",     220, "B-02"),
            ("00534", "Cable Gland 1/2", 110, "B-03"),
            ("00641", "PVC Coupler 3/4", 198, "C-01"),
            ("00758", "Copper Pipe 1ft",  74, "C-02"),
            ("00865", "Solder Wire 1lb",  48, "C-03"),
            ("00972", "Wire Brush 4in",  120, "D-01"),
            ("01084", "Steel Cable 10ft",105, "D-02"),
            ("01196", "Spring Pin 3mm",  580, "D-03"),
            ("01278", "O-Ring 1in",      720, "E-01"),
            ("01385", "Felt Pad 2in",    320, "E-02"),
            ("01492", "Velcro Strap",    150, "E-03"),
            ("01504", "Foam Tape 1in",    96, "F-01"),
            ("01617", "Anti-Slip Mat",    62, "F-02"),
            ("01729", "Steel Mesh 1ft",   88, "F-03"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_streaming_subs(src_path: str, _seed: int) -> str:
    """16×4 streaming subscribers: SubID / Plan / DeviceCount / MonthlyFee."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Subs"
        ws.append(["SubID", "Plan", "DeviceCount", "MonthlyFee"])
        rows = [
            ("SS-001", "Basic",    2,  9.99),
            ("SS-002", "Standard", 4, 14.99),
            ("SS-003", "Premium",  8, 24.99),
            ("SS-004", "Basic",    2,  9.99),
            ("SS-005", "Standard", 4, 14.99),
            ("SS-006", "Premium",  8, 24.99),
            ("SS-007", "Standard", 4, 14.99),
            ("SS-008", "Basic",    2,  9.99),
            ("SS-009", "Premium",  8, 24.99),
            ("SS-010", "Standard", 4, 14.99),
            ("SS-011", "Basic",    2,  9.99),
            ("SS-012", "Premium",  8, 24.99),
            ("SS-013", "Standard", 4, 14.99),
            ("SS-014", "Basic",    2,  9.99),
            ("SS-015", "Premium",  8, 24.99),
            ("SS-016", "Standard", 4, 14.99),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_pet_clinic(src_path: str, _seed: int) -> str:
    """16×4 pet clinic visits: VisitID / Pet / Species / Cost."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PetVisits"
        ws.append(["VisitID", "Pet", "Species", "Cost"])
        rows = [
            ("PV-201", "Bella",    "Dog",    140.00),
            ("PV-202", "Whiskers", "Cat",     95.00),
            ("PV-203", "Max",      "Dog",    220.00),
            ("PV-204", "Luna",     "Cat",     80.00),
            ("PV-205", "Charlie",  "Dog",    175.00),
            ("PV-206", "Coco",     "Rabbit",  65.00),
            ("PV-207", "Rocky",    "Dog",    310.00),
            ("PV-208", "Mittens",  "Cat",    120.00),
            ("PV-209", "Buddy",    "Dog",    160.00),
            ("PV-210", "Shadow",   "Cat",     90.00),
            ("PV-211", "Daisy",    "Dog",    240.00),
            ("PV-212", "Pepper",   "Rabbit",  55.00),
            ("PV-213", "Oliver",   "Cat",    105.00),
            ("PV-214", "Cooper",   "Dog",    195.00),
            ("PV-215", "Tiger",    "Cat",     85.00),
            ("PV-216", "Zeus",     "Dog",    280.00),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_concert_tickets(src_path: str, _seed: int) -> str:
    """18×4 concert tickets: TicketID / Section / Price / Sold."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets"
        ws.append(["TicketID", "Section", "Price", "Sold"])
        rows = [
            ("TK-501", "VIP",       250.00, "Yes"),
            ("TK-502", "Floor",     180.00, "Yes"),
            ("TK-503", "Mezzanine", 120.00, "Yes"),
            ("TK-504", "Balcony",    75.00, "No"),
            ("TK-505", "VIP",       250.00, "Yes"),
            ("TK-506", "Floor",     180.00, "Yes"),
            ("TK-507", "Mezzanine", 120.00, "No"),
            ("TK-508", "Balcony",    75.00, "Yes"),
            ("TK-509", "VIP",       250.00, "Yes"),
            ("TK-510", "Floor",     180.00, "Yes"),
            ("TK-511", "Mezzanine", 120.00, "Yes"),
            ("TK-512", "Balcony",    75.00, "Yes"),
            ("TK-513", "VIP",       250.00, "No"),
            ("TK-514", "Floor",     180.00, "Yes"),
            ("TK-515", "Mezzanine", 120.00, "Yes"),
            ("TK-516", "Balcony",    75.00, "No"),
            ("TK-517", "VIP",       250.00, "Yes"),
            ("TK-518", "Floor",     180.00, "Yes"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_runners_log(src_path: str, _seed: int) -> str:
    """20×4 race results: RunnerID / Name / Distance_km / TimeMin."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Runners"
        ws.append(["RunnerID", "Name", "Distance_km", "TimeMin"])
        rows = [
            ("R-301", "Adam Park",      5,  22.4),
            ("R-302", "Beth Silva",    10,  48.2),
            ("R-303", "Carl Singh",     5,  19.8),
            ("R-304", "Dora Lopez",    21,  98.6),
            ("R-305", "Erik Brooks",   10,  44.5),
            ("R-306", "Fay Park",       5,  24.1),
            ("R-307", "Gus Cho",       21, 105.3),
            ("R-308", "Hua Tanaka",    10,  46.8),
            ("R-309", "Ian Murphy",     5,  21.6),
            ("R-310", "Jane Romero",   21, 102.4),
            ("R-311", "Ken Nakamura",  10,  43.9),
            ("R-312", "Lila Khan",      5,  23.5),
            ("R-313", "Marc Chen",     10,  47.2),
            ("R-314", "Nora Patel",    21,  96.8),
            ("R-315", "Owen Brooks",    5,  20.4),
            ("R-316", "Priya Mehta",   10,  45.6),
            ("R-317", "Quinn Adams",   21, 108.2),
            ("R-318", "Rosa Garcia",    5,  18.9),
            ("R-319", "Sam Lee",       10,  49.4),
            ("R-320", "Tom Reyes",     21, 104.5),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_recipe_ratings(src_path: str, _seed: int) -> str:
    """16×4 recipe ratings: RecipeID / Cuisine / Rating / ReviewCount."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recipes"
        ws.append(["RecipeID", "Cuisine", "Rating", "ReviewCount"])
        rows = [
            ("RC-101", "Italian",  4.6, 248),
            ("RC-102", "Japanese", 4.8, 412),
            ("RC-103", "Mexican",  4.4, 168),
            ("RC-104", "Indian",   4.7, 326),
            ("RC-105", "Italian",  4.2,  92),
            ("RC-106", "French",   4.5, 184),
            ("RC-107", "Japanese", 4.9, 528),
            ("RC-108", "Mexican",  4.3, 142),
            ("RC-109", "Indian",   4.6, 286),
            ("RC-110", "French",   4.4, 156),
            ("RC-111", "Italian",  4.7, 314),
            ("RC-112", "Japanese", 4.5, 198),
            ("RC-113", "Mexican",  4.6, 268),
            ("RC-114", "Indian",   4.8, 392),
            ("RC-115", "French",   4.3, 124),
            ("RC-116", "Italian",  4.5, 218),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_invoice_aging(src_path: str, _seed: int) -> str:
    """18×4 invoice aging: InvoiceID / Customer / DaysOverdue / Amount."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Aging"
        ws.append(["InvoiceID", "Customer", "DaysOverdue", "Amount"])
        rows = [
            ("INV-7001", "Acme Co",       12, 1240.50),
            ("INV-7002", "Bayshore Ltd",  45,  812.00),
            ("INV-7003", "Cypress Group",  3, 1980.25),
            ("INV-7004", "Driftwood Inc", 78,  442.75),
            ("INV-7005", "Evergreen Sys", 21,  690.00),
            ("INV-7006", "Forge & Co",     7, 1124.40),
            ("INV-7007", "Granite Hold",  62, 1488.30),
            ("INV-7008", "Harbor Freight",  0,  376.10),
            ("INV-7009", "Ironwood Bros", 35, 2104.80),
            ("INV-7010", "Juniper Stk",   18,  915.55),
            ("INV-7011", "Kestrel Inc",    5, 1342.65),
            ("INV-7012", "Linden Mfg",    52,  780.20),
            ("INV-7013", "Mariner Co",     0,  588.40),
            ("INV-7014", "Northpoint Ltd",26,  995.70),
            ("INV-7015", "Oakridge Sys",  91, 1772.10),
            ("INV-7016", "Pinecrest LLC", 14,  449.85),
            ("INV-7017", "Quartz Hold",    8,  834.95),
            ("INV-7018", "Riverbend Corp", 41, 1245.00),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_university_courses(src_path: str, _seed: int) -> str:
    """16×4 university courses: CourseCode / Title / Credits / Enrolled."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Courses"
        ws.append(["CourseCode", "Title", "Credits", "Enrolled"])
        rows = [
            ("CS101",  "Intro to Programming",         3, 124),
            ("CS201",  "Data Structures",              4,  98),
            ("CS301",  "Algorithms",                   4,  72),
            ("MATH101","Calculus I",                   4, 156),
            ("MATH201","Linear Algebra",               3,  84),
            ("MATH301","Real Analysis",                4,  42),
            ("PHYS101","Mechanics",                    4, 102),
            ("PHYS201","Electromagnetism",             4,  68),
            ("BIO101", "Cell Biology",                 3, 138),
            ("BIO201", "Genetics",                     4,  86),
            ("CHEM101","General Chemistry",            4, 112),
            ("CHEM201","Organic Chemistry",            4,  78),
            ("ENG101", "Composition",                  3, 168),
            ("HIST101","World History",                3, 124),
            ("PSYC101","Intro to Psychology",          3, 184),
            ("ECON101","Microeconomics",               3,  96),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_smartphone_models(src_path: str, _seed: int) -> str:
    """16×4 smartphones: ModelID / Brand / StorageGB / Price."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Phones"
        ws.append(["ModelID", "Brand", "StorageGB", "Price"])
        rows = [
            ("SP-001", "Apple",   128,  799),
            ("SP-002", "Apple",   256,  899),
            ("SP-003", "Apple",   512, 1099),
            ("SP-004", "Samsung", 128,  749),
            ("SP-005", "Samsung", 256,  849),
            ("SP-006", "Samsung", 512, 1049),
            ("SP-007", "Google",  128,  599),
            ("SP-008", "Google",  256,  699),
            ("SP-009", "OnePlus", 128,  649),
            ("SP-010", "OnePlus", 256,  749),
            ("SP-011", "Xiaomi",  128,  449),
            ("SP-012", "Xiaomi",  256,  549),
            ("SP-013", "Apple",   128,  799),
            ("SP-014", "Samsung", 256,  849),
            ("SP-015", "Google",  256,  699),
            ("SP-016", "Xiaomi",  512,  749),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_lab_results(src_path: str, _seed: int) -> str:
    """18×4 lab results: SampleID / TestType / Result / FlagStatus."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lab"
        ws.append(["SampleID", "TestType", "Result", "FlagStatus"])
        rows = [
            ("LAB-101", "Glucose",      92.0, "Normal"),
            ("LAB-102", "Cholesterol", 245.0, "High"),
            ("LAB-103", "Hemoglobin",   13.8, "Normal"),
            ("LAB-104", "Glucose",     186.0, "High"),
            ("LAB-105", "Cholesterol", 198.0, "Normal"),
            ("LAB-106", "Hemoglobin",   10.4, "Low"),
            ("LAB-107", "Glucose",      88.0, "Normal"),
            ("LAB-108", "Cholesterol", 165.0, "Normal"),
            ("LAB-109", "Hemoglobin",   15.6, "Normal"),
            ("LAB-110", "Glucose",     142.0, "High"),
            ("LAB-111", "Cholesterol", 280.0, "High"),
            ("LAB-112", "Hemoglobin",    9.8, "Low"),
            ("LAB-113", "Glucose",      98.0, "Normal"),
            ("LAB-114", "Cholesterol", 175.0, "Normal"),
            ("LAB-115", "Hemoglobin",   12.4, "Normal"),
            ("LAB-116", "Glucose",     112.0, "Normal"),
            ("LAB-117", "Cholesterol", 256.0, "High"),
            ("LAB-118", "Hemoglobin",   11.2, "Low"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_freight_routes(src_path: str, _seed: int) -> str:
    """16×4 freight routes: RouteID / Origin / Destination / Distance_km."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Routes"
        ws.append(["RouteID", "Origin", "Destination", "Distance_km"])
        rows = [
            ("FR-101", "Chicago",     "Detroit",      450),
            ("FR-102", "Los Angeles", "San Diego",    195),
            ("FR-103", "New York",    "Boston",       350),
            ("FR-104", "Houston",     "Dallas",       385),
            ("FR-105", "Seattle",     "Portland",     280),
            ("FR-106", "Atlanta",     "Charlotte",    400),
            ("FR-107", "Denver",      "Salt Lake City",660),
            ("FR-108", "Phoenix",     "Las Vegas",    480),
            ("FR-109", "Miami",       "Orlando",      380),
            ("FR-110", "Chicago",     "Cincinnati",   480),
            ("FR-111", "Boston",      "Philadelphia", 480),
            ("FR-112", "Dallas",      "Oklahoma City",330),
            ("FR-113", "Portland",    "Boise",        680),
            ("FR-114", "Atlanta",     "Nashville",    400),
            ("FR-115", "Denver",      "Albuquerque",  720),
            ("FR-116", "Las Vegas",   "Reno",         700),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_internet_speeds(src_path: str, _seed: int) -> str:
    """16×4 internet speed tests: TestID / ISP / DownloadMbps / UploadMbps."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SpeedTests"
        ws.append(["TestID", "ISP", "DownloadMbps", "UploadMbps"])
        rows = [
            ("ST-001", "Comcast",  248.4,  12.6),
            ("ST-002", "Verizon",  812.7, 745.2),
            ("ST-003", "ATT",      542.1, 524.8),
            ("ST-004", "Spectrum", 312.5,  18.4),
            ("ST-005", "Comcast",  186.3,   8.9),
            ("ST-006", "Verizon",  892.4, 824.6),
            ("ST-007", "ATT",      438.7, 412.3),
            ("ST-008", "Spectrum", 264.8,  15.2),
            ("ST-009", "Comcast",  198.6,  10.4),
            ("ST-010", "Verizon",  748.2, 692.5),
            ("ST-011", "ATT",      592.3, 568.1),
            ("ST-012", "Spectrum", 348.5,  22.7),
            ("ST-013", "Comcast",  216.4,  11.8),
            ("ST-014", "Verizon",  834.7, 798.4),
            ("ST-015", "ATT",      512.6, 484.2),
            ("ST-016", "Spectrum", 282.1,  17.5),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_volunteer_hours(src_path: str, _seed: int) -> str:
    """18×4 volunteer hours: VolunteerID / Project / HoursLogged / Status."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Volunteers"
        ws.append(["VolunteerID", "Project", "HoursLogged", "Status"])
        rows = [
            ("V-101", "FoodBank",     24, "Approved"),
            ("V-102", "TreePlanting", 12, "Approved"),
            ("V-103", "Tutoring",     36, "Pending"),
            ("V-104", "FoodBank",     18, "Approved"),
            ("V-105", "Cleanup",      10, "Approved"),
            ("V-106", "TreePlanting",  8, "Pending"),
            ("V-107", "Tutoring",     42, "Approved"),
            ("V-108", "FoodBank",     20, "Approved"),
            ("V-109", "Cleanup",      14, "Pending"),
            ("V-110", "Tutoring",     28, "Approved"),
            ("V-111", "TreePlanting", 16, "Approved"),
            ("V-112", "FoodBank",     22, "Pending"),
            ("V-113", "Cleanup",       6, "Approved"),
            ("V-114", "Tutoring",     32, "Approved"),
            ("V-115", "FoodBank",     26, "Approved"),
            ("V-116", "TreePlanting", 10, "Pending"),
            ("V-117", "Cleanup",      18, "Approved"),
            ("V-118", "Tutoring",     40, "Approved"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_solar_panels(src_path: str, _seed: int) -> str:
    """16×4 solar panel installs: InstallID / Panels / KwCapacity / CostUSD."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Solar"
        ws.append(["InstallID", "Panels", "KwCapacity", "CostUSD"])
        rows = [
            ("SOL-101", 24,  9.6, 18400),
            ("SOL-102", 18,  7.2, 14200),
            ("SOL-103", 32, 12.8, 24800),
            ("SOL-104", 12,  4.8,  9600),
            ("SOL-105", 28, 11.2, 21500),
            ("SOL-106", 20,  8.0, 15800),
            ("SOL-107", 16,  6.4, 12700),
            ("SOL-108", 36, 14.4, 27800),
            ("SOL-109", 22,  8.8, 17200),
            ("SOL-110", 14,  5.6, 11200),
            ("SOL-111", 30, 12.0, 23200),
            ("SOL-112", 26, 10.4, 20400),
            ("SOL-113", 19,  7.6, 14900),
            ("SOL-114", 33, 13.2, 25600),
            ("SOL-115", 21,  8.4, 16500),
            ("SOL-116", 25, 10.0, 19400),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_wine_inventory(src_path: str, _seed: int) -> str:
    """18×4 wine inventory: BottleID / Varietal / Vintage / BottleCount."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Wines"
        ws.append(["BottleID", "Varietal", "Vintage", "BottleCount"])
        rows = [
            ("WN-101", "Cabernet",   2018, 24),
            ("WN-102", "Chardonnay", 2020, 36),
            ("WN-103", "Pinot Noir", 2017, 18),
            ("WN-104", "Merlot",     2019, 28),
            ("WN-105", "Sauvignon",  2021, 42),
            ("WN-106", "Cabernet",   2016, 12),
            ("WN-107", "Chardonnay", 2019, 30),
            ("WN-108", "Pinot Noir", 2020, 22),
            ("WN-109", "Merlot",     2018, 26),
            ("WN-110", "Sauvignon",  2020, 38),
            ("WN-111", "Cabernet",   2019, 20),
            ("WN-112", "Chardonnay", 2021, 44),
            ("WN-113", "Pinot Noir", 2018, 16),
            ("WN-114", "Merlot",     2020, 32),
            ("WN-115", "Sauvignon",  2019, 36),
            ("WN-116", "Cabernet",   2020, 18),
            ("WN-117", "Chardonnay", 2018, 26),
            ("WN-118", "Pinot Noir", 2019, 24),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_pottery_orders(src_path: str, _seed: int) -> str:
    """16×4 pottery orders: OrderID / Item / Quantity / UnitPrice."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pottery"
        ws.append(["OrderID", "Item", "Quantity", "UnitPrice"])
        rows = [
            ("PO-101", "Mug",        12, 18.50),
            ("PO-102", "Bowl",        6, 32.00),
            ("PO-103", "Vase",        3, 65.00),
            ("PO-104", "Plate",      18, 24.00),
            ("PO-105", "Mug",        24, 18.50),
            ("PO-106", "Teapot",      4, 85.00),
            ("PO-107", "Bowl",       12, 32.00),
            ("PO-108", "Plate",       9, 24.00),
            ("PO-109", "Vase",        2, 65.00),
            ("PO-110", "Mug",         8, 18.50),
            ("PO-111", "Teapot",      6, 85.00),
            ("PO-112", "Bowl",       15, 32.00),
            ("PO-113", "Plate",      12, 24.00),
            ("PO-114", "Vase",        4, 65.00),
            ("PO-115", "Teapot",      3, 85.00),
            ("PO-116", "Mug",        16, 18.50),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_helpdesk_tickets(src_path: str, _seed: int) -> str:
    """20×4 helpdesk tickets: TicketID / Priority / Category / Status."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets"
        ws.append(["TicketID", "Priority", "Category", "Status"])
        rows = [
            ("HD-1001", "High",   "Network",   "Open"),
            ("HD-1002", "Medium", "Hardware",  "Resolved"),
            ("HD-1003", "Low",    "Software",  "Open"),
            ("HD-1004", "High",   "Network",   "InProgress"),
            ("HD-1005", "Low",    "Account",   "Resolved"),
            ("HD-1006", "Medium", "Software",  "Open"),
            ("HD-1007", "High",   "Hardware",  "Resolved"),
            ("HD-1008", "Low",    "Account",   "Open"),
            ("HD-1009", "Medium", "Network",   "InProgress"),
            ("HD-1010", "High",   "Software",  "Open"),
            ("HD-1011", "Low",    "Hardware",  "Resolved"),
            ("HD-1012", "Medium", "Account",   "Open"),
            ("HD-1013", "High",   "Network",   "Resolved"),
            ("HD-1014", "Low",    "Software",  "InProgress"),
            ("HD-1015", "Medium", "Hardware",  "Open"),
            ("HD-1016", "High",   "Account",   "Resolved"),
            ("HD-1017", "Low",    "Network",   "Open"),
            ("HD-1018", "Medium", "Software",  "Resolved"),
            ("HD-1019", "High",   "Hardware",  "InProgress"),
            ("HD-1020", "Low",    "Account",   "Open"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


# ---------------------------------------------------------------------------
# Batch: eval-alignment src + gold builders (P2-P6).
#
# P2: clustered-column chart on sales data (compare_table verifies source-data
#     tab; chart presence verification is NOT supported by compare_table — see
#     SKIP note in validation comment below).
# P3: column reorder + sheet rename/copy/delete (multi-sheet management).
# P4: date arithmetic (age from birthday, duration from start/end) + pad-zeros.
# P5: custom number-format M / B (cell.number_format = '0.0,,"M"').
# P6: total row + month-on-month growth column + filter+aggregate compound.
# ---------------------------------------------------------------------------


def _src_sales_chart_data(src_path: str, _seed: int) -> str:
    """12-row Week / Sales / COGS for clustered-column chart task."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Week", "Sales", "COGS"])
        rows = [
            ("W01", 12400, 5820),
            ("W02", 13150, 6160),
            ("W03", 14820, 6940),
            ("W04", 16380, 7670),
            ("W05", 15200, 7130),
            ("W06", 17640, 8260),
            ("W07", 19320, 9050),
            ("W08", 21180, 9920),
            ("W09", 22840, 10700),
            ("W10", 24560, 11500),
            ("W11", 26230, 12290),
            ("W12", 28100, 13160),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_expense_chart_data(src_path: str, _seed: int) -> str:
    """12-row Month / Revenue / Expenses for chart variant."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Month", "Revenue", "Expenses"])
        rows = [
            ("Jan",  82400, 54200),
            ("Feb",  91200, 58100),
            ("Mar", 105800, 64900),
            ("Apr",  88300, 56400),
            ("May", 116400, 71500),
            ("Jun", 122800, 75200),
            ("Jul", 109500, 69100),
            ("Aug", 113700, 70600),
            ("Sep", 134600, 81200),
            ("Oct", 142800, 85100),
            ("Nov", 158400, 93600),
            ("Dec", 171300, 102200),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_customer_orders_5col(src_path: str, _seed: int) -> str:
    """5-col table (Order ID / Sales / First Name / Last Name / Date) for column-reorder."""
    return textwrap.dedent(f"""\
        import openpyxl, datetime
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Orders"
        ws.append(["Order ID", "Sales", "First Name", "Last Name", "Date"])
        rows = [
            ("ORD-2001",  1240.50, "Alice",   "Wong",      datetime.date(2026, 1, 12)),
            ("ORD-2002",   812.00, "Marcus",  "Chen",      datetime.date(2026, 1, 18)),
            ("ORD-2003",  1980.25, "Priya",   "Patel",     datetime.date(2026, 2,  3)),
            ("ORD-2004",   442.75, "Diego",   "Garcia",    datetime.date(2026, 2,  9)),
            ("ORD-2005",   690.00, "Aisha",   "Okafor",    datetime.date(2026, 2, 21)),
            ("ORD-2006",  1124.40, "Jordan",  "Kim",       datetime.date(2026, 3,  2)),
            ("ORD-2007",  1488.30, "Sofia",   "Hernandez", datetime.date(2026, 3, 14)),
            ("ORD-2008",   376.10, "Liam",    "O'Brien",   datetime.date(2026, 3, 25)),
            ("ORD-2009",  2104.80, "Ava",     "Singh",     datetime.date(2026, 4,  6)),
            ("ORD-2010",   915.55, "Noah",    "Adebayo",   datetime.date(2026, 4, 18)),
            ("ORD-2011",  1342.65, "Mia",     "Rossi",     datetime.date(2026, 4, 28)),
            ("ORD-2012",   780.20, "Ethan",   "Park",      datetime.date(2026, 5,  9)),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_lars_two_sheet(src_path: str, _seed: int) -> str:
    """2-sheet workbook (Sheet1 with resources, Sheet2 placeholder)."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Resource", "Type", "Hours"])
        rows = [
            ("Adobe Reader",     "Software", 40),
            ("Visio Pro",        "Software", 25),
            ("Engineering Team", "People",   120),
            ("Testing Lab",      "Facility", 80),
            ("Quality Control",  "People",   95),
            ("Build Server",     "Hardware", 30),
        ]
        for r in rows:
            ws1.append(list(r))
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Note"])
        ws2.append(["placeholder"])
        wb.save({src_path!r})
        """)


def _src_employee_birthday(src_path: str, _seed: int) -> str:
    """5-col employee table with Birthday column for age computation."""
    return textwrap.dedent(f"""\
        import openpyxl, datetime
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Employees"
        ws.append(["EmpID", "Name", "Dept", "Birthday", "Salary"])
        rows = [
            ("E-101", "Alice Wong",    "Eng",   datetime.date(1987,  3, 12),  92000),
            ("E-102", "Marcus Chen",   "Sales", datetime.date(1992,  7, 25),  78000),
            ("E-103", "Priya Patel",   "Eng",   datetime.date(1985, 11,  8), 105000),
            ("E-104", "Diego Garcia",  "Ops",   datetime.date(1990,  1, 30),  72000),
            ("E-105", "Aisha Okafor",  "Eng",   datetime.date(1995,  9, 14),  68000),
            ("E-106", "Jordan Kim",    "Sales", datetime.date(1988,  6,  3),  85000),
            ("E-107", "Sofia Hernandez","HR",   datetime.date(1993,  4, 21),  74000),
            ("E-108", "Liam O'Brien",  "Eng",   datetime.date(1986, 12,  7),  98000),
            ("E-109", "Ava Singh",     "Mkt",   datetime.date(1991, 10, 18),  81000),
            ("E-110", "Noah Adebayo",  "Eng",   datetime.date(1989,  2, 28),  94000),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_old_id_padding(src_path: str, _seed: int) -> str:
    """2-col table with Old ID (numeric, <7 digits) and empty New 7 Digit Id column."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customers"
        ws.append(["Old ID", "New 7 Digit Id"])
        rows = [
            (12,      None),
            (4815,    None),
            (732,     None),
            (90210,   None),
            (1,       None),
            (6543,    None),
            (8,       None),
            (12345,   None),
            (567,     None),
            (88,      None),
            (3141,    None),
            (4096,    None),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_project_dates(src_path: str, _seed: int) -> str:
    """4-col project table with StartDate / EndDate for duration computation."""
    return textwrap.dedent(f"""\
        import openpyxl, datetime
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Projects"
        ws.append(["ProjectID", "Name", "StartDate", "EndDate"])
        rows = [
            ("P-001", "Atlas",     datetime.date(2025,  1, 10), datetime.date(2025,  4, 25)),
            ("P-002", "Beacon",    datetime.date(2025,  2,  5), datetime.date(2025,  6, 18)),
            ("P-003", "Catalyst",  datetime.date(2025,  3, 12), datetime.date(2025,  8, 30)),
            ("P-004", "Delta",     datetime.date(2025,  4,  1), datetime.date(2025,  7, 15)),
            ("P-005", "Eclipse",   datetime.date(2025,  5, 20), datetime.date(2025, 11,  9)),
            ("P-006", "Falcon",    datetime.date(2025,  6,  8), datetime.date(2025, 10, 22)),
            ("P-007", "Gemini",    datetime.date(2025,  7, 14), datetime.date(2026,  1, 30)),
            ("P-008", "Helios",    datetime.date(2025,  8,  3), datetime.date(2025, 12, 18)),
            ("P-009", "Iris",      datetime.date(2025,  9, 27), datetime.date(2026,  3, 12)),
            ("P-010", "Juno",      datetime.date(2025, 10, 14), datetime.date(2026,  4,  6)),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_company_financials(src_path: str, _seed: int) -> str:
    """3-col company table with large numbers in millions / billions ranges."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financials"
        ws.append(["Company", "Revenue", "MarketCap"])
        rows = [
            ("Apex Corp",       2_400_000_000,   42_800_000_000),
            ("Borealis Ltd",      850_000_000,   12_400_000_000),
            ("Catalyst Inc",    5_600_000_000,   98_700_000_000),
            ("Drake Hold",        320_000_000,    4_100_000_000),
            ("Evergreen Sys",   1_180_000_000,   18_900_000_000),
            ("Forge Industries",  720_000_000,    9_300_000_000),
            ("Granite Pro",     3_840_000_000,   65_200_000_000),
            ("Harbour & Co",      490_000_000,    6_700_000_000),
            ("Indigo Ltd",      2_950_000_000,   38_500_000_000),
            ("Juniper Sys",     1_620_000_000,   22_100_000_000),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_monthly_sales_growth(src_path: str, _seed: int) -> str:
    """3-col table Month / Sales / Costs — used for total-row + MoM growth col."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MonthlySales"
        ws.append(["Month", "Sales", "Costs"])
        rows = [
            ("Jan",  82400, 54200),
            ("Feb",  91200, 58100),
            ("Mar", 105800, 64900),
            ("Apr",  88300, 56400),
            ("May", 116400, 71500),
            ("Jun", 122800, 75200),
            ("Jul", 109500, 69100),
            ("Aug", 113700, 70600),
            ("Sep", 134600, 81200),
            ("Oct", 142800, 85100),
            ("Nov", 158400, 93600),
            ("Dec", 171300, 102200),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_region_sales(src_path: str, _seed: int) -> str:
    """3-col Region / Product / Sales table for filter+aggregate compound."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales"
        ws.append(["Region", "Product", "Sales"])
        rows = [
            ("North",  "Widget A", 1240),
            ("South",  "Widget A",  812),
            ("North",  "Widget B", 1980),
            ("West",   "Widget A",  442),
            ("North",  "Widget C",  690),
            ("East",   "Widget B", 1124),
            ("North",  "Widget B", 1488),
            ("South",  "Widget A",  376),
            ("North",  "Widget A", 2104),
            ("West",   "Widget B",  915),
            ("North",  "Widget C", 1342),
            ("East",   "Widget A",  780),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


# ---------------------------------------------------------------------------
# Batch: gold builders for new task families.
# ---------------------------------------------------------------------------


def _gold_chart_data_passthrough(src: str, exp: str, *,
                                  chart_sheet_name: str) -> str:
    """Gold for "create chart in a new sheet" task — keeps Sheet1 data intact,
    adds `chart_sheet_name`, AND inserts a real BarChart (clustered-column
    family — openpyxl emits `<barChart>` for both bar and clustered-column,
    which is the tagname compare_calc_chart_type reads when chart_props=['type']).

    validation B4 fix: prior version added only an EMPTY target sheet.
    compare_table's chart-rule (`chart_props=['type']`) read 0 charts in the
    expected sheet, so the eval was vacuous on chart presence — `--max-turns 0`
    scored 1.0 because the agent's "no-op" workbook also has 0 charts there.
    Now the gold places a real chart referencing Sheet1!$B$2:$C$13, so any
    agent run without a chart in `chart_sheet_name` fails the chart rule.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        from openpyxl.chart import BarChart
        from openpyxl.chart.series import Series
        from openpyxl.chart.data_source import NumDataSource, NumRef, AxDataSource
        wb = openpyxl.load_workbook({src!r})
        # Add the target chart sheet so the agent's "create chart in Sheet2"
        # produces a workbook with the same sheet topology.
        if {chart_sheet_name!r} not in wb.sheetnames:
            wb.create_sheet({chart_sheet_name!r})
        ws_target = wb[{chart_sheet_name!r}]
        # Insert a real clustered-column (BarChart) so the chart-rule eval
        # (chart_props=['type']) sees a `barChart` tagname here. The actual
        # series ranges don't matter because compare_calc_chart_type drops
        # the series-ref key before comparing.
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"
        ser = Series()
        ser.val = NumDataSource(numRef=NumRef(f="Sheet1!$B$2:$B$13"))
        chart.series.append(ser)
        ws_target.add_chart(chart, "E2")
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_reorder_columns(src: str, exp: str, *,
                           new_order: list[int]) -> str:
    """Reorder columns on sheet 0 per `new_order` (list of 0-based source
    column indices). All rows are rewritten in the new column order.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        new_order = {new_order!r}
        # Snapshot all rows.
        snap = []
        for r in range(1, max_r + 1):
            snap.append([ws.cell(r, c).value for c in range(1, max_c + 1)])
        # Clear and rewrite in new order.
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                ws.cell(r, c, None)
        for r, row in enumerate(snap, 1):
            for new_c, src_idx in enumerate(new_order, 1):
                if src_idx < len(row):
                    ws.cell(r, new_c, row[src_idx])
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_rename_and_copy_sheet(src: str, exp: str, *,
                                 old_name: str, new_name: str,
                                 copy_suffix: str) -> str:
    """Rename sheet `old_name` to `new_name`, then duplicate as
    `new_name + copy_suffix` placed BEFORE Sheet2.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        from copy import copy
        wb = openpyxl.load_workbook({src!r})
        # Rename first sheet.
        ws_old = wb[{old_name!r}]
        ws_old.title = {new_name!r}
        # Create copy by replicating cells.
        ws_copy = wb.copy_worksheet(ws_old)
        ws_copy.title = {new_name!r} + {copy_suffix!r}
        # Move the copy to position 1 (before Sheet2 which is at position 1
        # originally, but Sheet2 stays where it is; copy goes before it).
        wb.move_sheet(ws_copy, offset=-(len(wb.sheetnames) - 2))
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_rename_and_delete_sheet(src: str, exp: str, *,
                                   old_name: str, new_name: str,
                                   delete_name: str) -> str:
    """Rename sheet `old_name` -> `new_name`, then delete `delete_name`."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        wb[{old_name!r}].title = {new_name!r}
        if {delete_name!r} in wb.sheetnames:
            del wb[{delete_name!r}]
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_age_from_birthday(src: str, exp: str, *,
                             birthday_col_idx: int,
                             age_header: str,
                             reference_year: int) -> str:
    """Append an Age column computed as `reference_year - year(Birthday)`."""
    return textwrap.dedent(f"""\
        import openpyxl, datetime
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        new_c = max_c + 1
        ws.cell(1, new_c, {age_header!r})
        for r in range(2, max_r + 1):
            v = ws.cell(r, {birthday_col_idx + 1}).value
            yr = None
            if isinstance(v, datetime.date):
                yr = v.year
            elif isinstance(v, datetime.datetime):
                yr = v.year
            if yr is not None:
                ws.cell(r, new_c, {reference_year} - yr)
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_pad_zeros(src: str, exp: str, *,
                     src_col_idx: int, dst_col_idx: int,
                     width: int) -> str:
    """Copy numeric values from `src_col_idx` into `dst_col_idx` as text
    padded with leading zeros to `width` digits (e.g. width=7 -> "0000012").
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_r = ws.max_row or 0
        for r in range(2, max_r + 1):
            v = ws.cell(r, {src_col_idx + 1}).value
            if isinstance(v, (int, float)):
                ws.cell(r, {dst_col_idx + 1}, str(int(v)).zfill({width}))
            elif isinstance(v, str) and v.isdigit():
                ws.cell(r, {dst_col_idx + 1}, v.zfill({width}))
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_date_duration(src: str, exp: str, *,
                         start_col_idx: int, end_col_idx: int,
                         duration_header: str) -> str:
    """Append a duration-in-days column = EndDate - StartDate."""
    return textwrap.dedent(f"""\
        import openpyxl, datetime
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        new_c = max_c + 1
        ws.cell(1, new_c, {duration_header!r})
        for r in range(2, max_r + 1):
            s = ws.cell(r, {start_col_idx + 1}).value
            e = ws.cell(r, {end_col_idx + 1}).value
            sd = s.date() if isinstance(s, datetime.datetime) else s
            ed = e.date() if isinstance(e, datetime.datetime) else e
            if isinstance(sd, datetime.date) and isinstance(ed, datetime.date):
                ws.cell(r, new_c, (ed - sd).days)
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_numfmt_scaled(src: str, exp: str, *,
                         col_idx: int, number_format: str) -> str:
    """Apply a custom number_format (e.g. '0.0,,"M"' / '0.0,,,"B"') to every
    non-null data cell in `col_idx`. Values themselves are untouched —
    custom format does the display scaling.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_r = ws.max_row or 0
        for r in range(2, max_r + 1):
            cc = ws.cell(r, {col_idx + 1})
            if cc.value is not None:
                cc.number_format = {number_format!r}
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_total_and_growth(src: str, exp: str, *,
                            value_col_idx: int,
                            label_col_idx: int,
                            total_label: str,
                            growth_header: str) -> str:
    """Append a Total row labelled `total_label` AND a Growth% column =
    month-on-month percentage growth on the value column.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        # 1) Append growth column.
        new_c = max_c + 1
        ws.cell(1, new_c, {growth_header!r})
        prev = None
        for r in range(2, max_r + 1):
            v = ws.cell(r, {value_col_idx + 1}).value
            if isinstance(v, (int, float)) and isinstance(prev, (int, float)) and prev != 0:
                ws.cell(r, new_c, round((float(v) - float(prev)) / float(prev), 4))
            prev = v
        # 2) Append Total row.
        total_r = max_r + 1
        ws.cell(total_r, {label_col_idx + 1}, {total_label!r})
        tot = 0.0
        for r in range(2, max_r + 1):
            v = ws.cell(r, {value_col_idx + 1}).value
            if isinstance(v, (int, float)):
                tot += float(v)
        ws.cell(total_r, {value_col_idx + 1}, tot)
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_filter_to_sheet_with_total(src: str, exp: str, *,
                                       new_sheet_name: str,
                                       filter_col_idx: int,
                                       filter_value: str,
                                       sum_col_idx: int,
                                       total_label_col_idx: int,
                                       total_label: str) -> str:
    """Filter rows where filter_col == filter_value to a new sheet, then
    append a Total row summing `sum_col_idx`.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws1 = wb.worksheets[0]
        max_r = ws1.max_row or 0
        max_c = ws1.max_column or 0
        if {new_sheet_name!r} in wb.sheetnames:
            ws2 = wb[{new_sheet_name!r}]
        else:
            ws2 = wb.create_sheet({new_sheet_name!r})
        # Copy header.
        for c in range(1, max_c + 1):
            ws2.cell(1, c, ws1.cell(1, c).value)
        out_r = 2
        tot = 0.0
        for r in range(2, max_r + 1):
            v = ws1.cell(r, {filter_col_idx + 1}).value
            if v == {filter_value!r}:
                for c in range(1, max_c + 1):
                    ws2.cell(out_r, c, ws1.cell(r, c).value)
                sv = ws1.cell(r, {sum_col_idx + 1}).value
                if isinstance(sv, (int, float)):
                    tot += float(sv)
                out_r += 1
        # Append Total row in the new sheet.
        ws2.cell(out_r, {total_label_col_idx + 1}, {total_label!r})
        ws2.cell(out_r, {sum_col_idx + 1}, tot)
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


# ---------------------------------------------------------------------------
# Batch — gold builders for the new eval-aligned skills introduced
# post validation cut. Each is intentionally minimal; cap-2x2 + validation scaler
# constraints keep the per-helper footprint small.
# ---------------------------------------------------------------------------


def _gold_real_chart(src: str, exp: str, *,
                     chart_type: str,
                     series_ref: str,
                     cat_ref: str | None = None,
                     target_sheet: str = "active") -> str:
    """Insert a real openpyxl LineChart / BarChart referencing the given column.

    Mirrors perturb/libreoffice_calc.py:_build_calc_chart_gold_py. The chart
    is anchored at E2 on the active sheet (or a named target_sheet which is
    created if missing). Eval rule = chart with `chart_props=['type']` per
    validation note (title round-trips imperfectly through LO save).
    """
    cls = {"line": "LineChart", "bar": "BarChart"}[chart_type]
    return textwrap.dedent(f"""\
        import openpyxl
        from openpyxl.chart import {cls}
        from openpyxl.chart.series import Series
        from openpyxl.chart.data_source import NumDataSource, NumRef, AxDataSource
        wb = openpyxl.load_workbook({src!r})
        target_sheet = {target_sheet!r}
        if target_sheet == "active":
            ws_target = wb.worksheets[0]
        else:
            if target_sheet in wb.sheetnames:
                ws_target = wb[target_sheet]
            else:
                ws_target = wb.create_sheet(target_sheet)
        chart = {cls}()
        ser = Series()
        ser.val = NumDataSource(numRef=NumRef(f={series_ref!r}))
        _cat_ref = {cat_ref!r}
        if _cat_ref:
            ser.cat = AxDataSource(numRef=NumRef(f=_cat_ref))
        chart.series.append(ser)
        ws_target.add_chart(chart, "E2")
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_freeze_panes(src: str, exp: str, *, freeze_cell: str) -> str:
    """Set `ws.freeze_panes = freeze_cell` on sheet 0 (e.g. 'A2' for header freeze).

    NB: this gold INTENTIONALLY skips `_LO_NORMALIZE_TAIL`. The
    `soffice --headless --convert-to xlsx` pass strips the
    `<sheetView><pane state="frozen"/></sheetView>` element that openpyxl
    writes, so a normalized gold has `freeze_panes=None` and the oracle
    (`cp expected → result`) propagates the strip into the result — which
    then fails the `check_xlsx_freeze_pane` probe in
    `_eval_compare_table`. Keeping the gold openpyxl-native preserves the
    frozen pane so both oracle replay (cp-then-probe) and agent paths
    (LO-interactive Ctrl+S, which DOES retain the pane element) score 1.0.
    sheet_data is unaffected because the openpyxl save round-trips the
    same cell values.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        ws.freeze_panes = {freeze_cell!r}
        wb.save({exp!r})
        """)


def _gold_zoom(src: str, exp: str, *, zoom_scale: int) -> str:
    """Set the sheet-view zoom scale on sheet 0."""
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        ws.sheet_view.zoomScale = {zoom_scale!r}
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_data_validation_list(src: str, exp: str, *,
                                col_letter: str,
                                last_row: int,
                                allowed: list[str]) -> str:
    """Attach a list-type data validation to <col_letter>2:<col_letter><last_row>."""
    formula = '"' + ",".join(allowed) + '"'
    return textwrap.dedent(f"""\
        import openpyxl
        from openpyxl.worksheet.datavalidation import DataValidation
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        dv = DataValidation(type="list", formula1={formula!r}, allow_blank=True)
        dv.add({col_letter!r} + "2:" + {col_letter!r} + str({last_row}))
        ws.data_validations.append(dv)
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_hide_rows(src: str, exp: str, *,
                     predicate_col_idx: int,
                     predicate_value: str) -> str:
    """Hide every data row where ws.cell(r, predicate_col_idx+1) == predicate_value.

    Sets `ws.row_dimensions[r].hidden = True` on matching rows.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_r = ws.max_row or 0
        for r in range(2, max_r + 1):
            if ws.cell(r, {predicate_col_idx + 1}).value == {predicate_value!r}:
                ws.row_dimensions[r].hidden = True
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_transpose_block(src: str, exp: str, *,
                            src_range_top_left: str,
                            src_range_bottom_right: str,
                            dst_top_left: str) -> str:
    """Copy a rectangular cell range and write it transposed at dst_top_left.

    Source cells are NOT cleared; only the destination block is populated.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        from openpyxl.utils import coordinate_to_tuple
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        r0, c0 = coordinate_to_tuple({src_range_top_left!r})
        r1, c1 = coordinate_to_tuple({src_range_bottom_right!r})
        dr0, dc0 = coordinate_to_tuple({dst_top_left!r})
        for ri, r in enumerate(range(r0, r1 + 1)):
            for ci, c in enumerate(range(c0, c1 + 1)):
                v = ws.cell(r, c).value
                # transpose: src(ri, ci) -> dst(ci, ri)
                ws.cell(dr0 + ci, dc0 + ri, v)
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


# ---------------------------------------------------------------------------
# Cycle-N (this subagent) — eval-essence emulation builders.
#
# Targeted gaps (per measure_gap calc bucket):
#   - eval_fn=compare_csv 0% synth vs 2.2% eval (emulate
#     osworld_libreoffice_calc_3aaa4e37: "export current sheet to a csv file")
#   - eval_fn=check_pdf_pages+compare_pdfs 0% vs 2.2% (emulate
#     osworld_libreoffice_calc_aa3a8974: "resize cells to fit onto one page
#     and export to PDF")
#   - weekend-highlight (date-aware conditional fmt; emulates
#     osworld_libreoffice_calc_8b1ce5f2)
#   - TEXT() decimal display (emulates osworld_libreoffice_calc_4f07fbe9)
# ---------------------------------------------------------------------------


def _src_weekend_calendar(src_path: str, _seed: int) -> str:
    """26-row event calendar: Date / Event / Owner. Real `datetime.date`
    cells span ~4 weeks so Sat/Sun rows appear naturally for the weekend-
    highlight task (emulates osworld_libreoffice_calc_8b1ce5f2).
    """
    return textwrap.dedent(f"""\
        import openpyxl, datetime
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Calendar"
        ws.append(["Date", "Event", "Owner"])
        rows = [
            (datetime.date(2026, 3,  2), "Sprint Planning",        "Alice"),
            (datetime.date(2026, 3,  3), "Design Review",          "Marcus"),
            (datetime.date(2026, 3,  4), "Standup",                "Priya"),
            (datetime.date(2026, 3,  5), "Customer Demo",          "Diego"),
            (datetime.date(2026, 3,  6), "Retro",                  "Aisha"),
            (datetime.date(2026, 3,  7), "Hackathon Saturday",     "Team"),
            (datetime.date(2026, 3,  8), "Office Closed",          "—"),
            (datetime.date(2026, 3,  9), "Architecture Sync",      "Jordan"),
            (datetime.date(2026, 3, 10), "1:1 Reviews",            "Sofia"),
            (datetime.date(2026, 3, 11), "Roadmap Workshop",       "Liam"),
            (datetime.date(2026, 3, 12), "Backlog Grooming",       "Ava"),
            (datetime.date(2026, 3, 13), "All-Hands",              "Noah"),
            (datetime.date(2026, 3, 14), "Pi Day Social",          "Team"),
            (datetime.date(2026, 3, 15), "On-call Handoff",        "Mia"),
            (datetime.date(2026, 3, 16), "Stakeholder Review",     "Ethan"),
            (datetime.date(2026, 3, 17), "Release Cut",            "Alice"),
            (datetime.date(2026, 3, 18), "Postmortem",             "Marcus"),
            (datetime.date(2026, 3, 19), "Bug Bash",               "Priya"),
            (datetime.date(2026, 3, 20), "Demo Day Prep",          "Diego"),
            (datetime.date(2026, 3, 21), "Saturday Workshop",      "Aisha"),
            (datetime.date(2026, 3, 22), "Quiet Sunday",           "—"),
            (datetime.date(2026, 3, 23), "Quarterly Kickoff",      "Jordan"),
            (datetime.date(2026, 3, 24), "Vendor Onsite",          "Sofia"),
            (datetime.date(2026, 3, 25), "Code Freeze",            "Liam"),
            (datetime.date(2026, 3, 26), "Release Notes",          "Ava"),
            (datetime.date(2026, 3, 27), "Launch Rehearsal",       "Noah"),
        ]
        for r in rows:
            ws.append(list(r))
        wb.save({src_path!r})
        """)


def _src_hourly_value(src_path: str, _seed: int) -> str:
    """12-row reading log: Reading / Value (raw float). Source for the
    TEXT()-decimal task (emulates osworld_libreoffice_calc_4f07fbe9).
    The Value cells are written as raw floats (e.g. 12.3456) so the
    agent's job is to render them with 2 decimals via TEXT() in column 3.
    """
    return textwrap.dedent(f"""\
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Readings"
        ws.append(["Reading", "Value", "Display"])
        rows = [
            ("R-001", 12.3456),
            ("R-002",  8.0500),
            ("R-003", 19.7891),
            ("R-004",  4.2345),
            ("R-005", 25.9999),
            ("R-006",  7.1000),
            ("R-007", 14.5678),
            ("R-008", 22.4444),
            ("R-009",  3.3333),
            ("R-010", 18.0001),
            ("R-011",  6.6666),
            ("R-012", 11.5),
        ]
        for r in rows:
            ws.append([r[0], r[1], None])
        wb.save({src_path!r})
        """)


def _gold_color_weekends(src: str, exp: str, *,
                          date_col_idx: int,
                          argb: str = "FFFF0000") -> str:
    """Set every row's bgcolor to `argb` when the cell at `date_col_idx`
    is a Saturday or Sunday. Mirrors weekday-aware conditional formatting
    (osworld_libreoffice_calc_8b1ce5f2).
    """
    return textwrap.dedent(f"""\
        import openpyxl, datetime
        from openpyxl.styles import PatternFill
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        _fill = PatternFill(start_color={argb!r}, end_color={argb!r}, fill_type='solid')
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        for r in range(2, max_r + 1):
            v = ws.cell(r, {date_col_idx + 1}).value
            if isinstance(v, (datetime.date, datetime.datetime)) and v.weekday() >= 5:
                for c in range(1, max_c + 1):
                    ws.cell(r, c).fill = _fill
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


def _gold_text_decimal_display(src: str, exp: str, *,
                                 value_col_idx: int,
                                 dst_col_idx: int,
                                 decimals: int = 2) -> str:
    """Write a TEXT-formatted display string into dst_col_idx for every data
    row whose value cell is numeric. Stores both the formula and the
    pre-computed cached value, since `_LO_NORMALIZE_TAIL` will repopulate
    `<v>` via soffice headless. Emulates osworld_libreoffice_calc_4f07fbe9.
    """
    # Same class as the round() override above: the agent's
    # natural LibreOffice =TEXT(v,"0.000") rounds the DECIMAL value half-up
    # (4.2345 -> "4.235"), but Python's format(float(v), '.3f') applies banker's
    # rounding on the float-repr (float(4.2345)==4.23449… -> "4.234"), so the gold
    # false-failed a correct agent. Emit a half-up Decimal quantize to match a
    # correct LO-rendered TEXT display. Guards bool + falls back to the old
    # format on any Decimal edge case. `%.15g` (not str(v)) for the same reason
    # as the round() shim at L830-855: LO snaps the double to 15 significant
    # decimal digits before rounding, so a 16/17-digit repr must not decide the
    # rounding direction.
    return textwrap.dedent(f"""\
        import openpyxl
        from decimal import Decimal as _D, ROUND_HALF_UP as _HU
        wb = openpyxl.load_workbook({src!r})
        ws = wb.worksheets[0]
        max_r = ws.max_row or 0
        for r in range(2, max_r + 1):
            v = ws.cell(r, {value_col_idx + 1}).value
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                try:
                    _dv = _D('%.15g' % v) if isinstance(v, float) else _D(v)
                    _disp = str(_dv.quantize(_D(1).scaleb(-{decimals}), rounding=_HU))
                except Exception:
                    _disp = format(float(v), '.{decimals}f')
                ws.cell(r, {dst_col_idx + 1}, _disp)
        wb.save({exp!r})
        """) + _LO_NORMALIZE_TAIL.format(expected_path=exp)


# ---------------------------------------------------------------------------
# Cycle-N — Custom-evaluator templates (CSV export + PDF export).
#
# These do NOT fit the FileTask / compare_table pattern — eval funcs are
# compare_csv (3aaa4e37) and check_pdf_pages+compare_pdfs (aa3a8974).
# They build their own SynthTemplate directly and append to TEMPLATES at
# the bottom of the file. Verified upstream funcs exist:
#   .venv/lib/python3.12/site-packages/desktop_env/evaluators/metrics/table.py
#   .venv/lib/python3.12/site-packages/desktop_env/evaluators/metrics/pdf.py
#   .venv/lib/python3.12/site-packages/desktop_env/evaluators/metrics/chrome.py
# ---------------------------------------------------------------------------


def _gold_export_csv(src_xlsx: str, exp_csv: str) -> str:
    """Convert the source xlsx to CSV via soffice headless, producing
    the gold csv at `exp_csv`. The csv is what `compare_csv` matches.
    """
    exp_dir = exp_csv.rsplit("/", 1)[0]
    exp_name = exp_csv.rsplit("/", 1)[1]
    return textwrap.dedent(f"""\
        import os as _os, subprocess as _sp, shutil as _sh, tempfile as _tf
        _td = _tf.mkdtemp()
        try:
            _sp.run(["soffice", "--headless", "--norestore",
                     "--nofirststartwizard", "--convert-to", "csv",
                     "--outdir", _td, {src_xlsx!r}],
                    capture_output=True,
                    env={{**_os.environ, "DISPLAY": ":1"}}, timeout=120)
            _base = _os.path.splitext(_os.path.basename({src_xlsx!r}))[0]
            _conv = _os.path.join(_td, _base + ".csv")
            _os.makedirs({exp_dir!r}, exist_ok=True)
            if _os.path.exists(_conv):
                _sh.copy(_conv, {exp_csv!r})
        finally:
            _sh.rmtree(_td, ignore_errors=True)
        """)


def _gold_export_pdf_fit_one_page(src_xlsx: str, exp_pdf: str) -> str:
    """Apply fit-to-one-page page setup on the xlsx and convert to PDF
    via soffice headless. Mirrors osworld_libreoffice_calc_aa3a8974's
    expected single-page PDF output.
    """
    exp_dir = exp_pdf.rsplit("/", 1)[0]
    return textwrap.dedent(f"""\
        import os as _os, subprocess as _sp, shutil as _sh, tempfile as _tf
        import openpyxl as _ox
        _td = _tf.mkdtemp()
        try:
            # Fit to 1 page (width × height) before PDF conversion.
            _wb = _ox.load_workbook({src_xlsx!r})
            _ws = _wb.worksheets[0]
            _ws.page_setup.fitToWidth = 1
            _ws.page_setup.fitToHeight = 1
            _ws.sheet_properties.pageSetUpPr.fitToPage = True
            _staged = _os.path.join(_td, _os.path.basename({src_xlsx!r}))
            _wb.save(_staged)
            _sp.run(["soffice", "--headless", "--norestore",
                     "--nofirststartwizard", "--convert-to", "pdf",
                     "--outdir", _td, _staged],
                    capture_output=True,
                    env={{**_os.environ, "DISPLAY": ":1"}}, timeout=180)
            _base = _os.path.splitext(_os.path.basename({src_xlsx!r}))[0]
            _conv = _os.path.join(_td, _base + ".pdf")
            _os.makedirs({exp_dir!r}, exist_ok=True)
            if _os.path.exists(_conv):
                _sh.copy(_conv, {exp_pdf!r})
        finally:
            _sh.rmtree(_td, ignore_errors=True)
        """)


def _eval_compare_csv(result_csv: str, expected_csv: str) -> dict:
    """compare_csv evaluator dict (eval_fn=compare_csv, eval_class=compare_csv).
    Matches osworld_libreoffice_calc_3aaa4e37's evaluator shape (default csv).
    """
    return {
        "func": "compare_csv",
        "result": {"type": "vm_file", "path": result_csv,
                   "dest": result_csv.split("/")[-1]},
        "expected": {"type": "vm_file", "path": expected_csv,
                     "dest": "expected_" + expected_csv.split("/")[-1]},
        "postconfig": LO_SAVE_POSTCONFIG,
    }


def _eval_pdf_pages_and_compare(result_pdf: str, expected_pdf: str,
                                 nb_pages: int = 1) -> dict:
    """check_pdf_pages + compare_pdfs combined evaluator
    (osworld_libreoffice_calc_aa3a8974 shape).
    """
    return {
        "func": ["check_pdf_pages", "compare_pdfs"],
        "result": [
            {"type": "vm_file", "path": result_pdf,
             "dest": result_pdf.split("/")[-1]},
            {"type": "vm_file", "path": result_pdf,
             "dest": result_pdf.split("/")[-1]},
        ],
        "expected": [
            {"type": "rule",
             "rules": {"relation": "eq", "ref_value": nb_pages}},
            {"type": "vm_file", "path": expected_pdf,
             "dest": "expected_" + expected_pdf.split("/")[-1]},
        ],
        "postconfig": LO_SAVE_POSTCONFIG,
    }


def _make_csv_export_template(*, template_id: str, setup_class: str,
                                basename_xlsx: str,
                                src_builder: Callable[[str, int], str],
                                instructions: list[str]) -> SynthTemplate:
    """Build a CSV-export SynthTemplate. Each seed picks one of `instructions`.

    Oracle: cp the gold csv to the agent's expected csv output path
    (sibling of the xlsx, same stem). The agent's path is /home/user/<stem>.csv
    matching the eval task's "share the file name" requirement.
    """
    stem = basename_xlsx.rsplit(".", 1)[0]
    n = len(instructions)

    def _params(seed: int) -> dict:
        src_path = f"/home/user/{basename_xlsx}"
        result_csv = f"/home/user/{stem}.csv"
        expected_csv = f"/tmp/expected_csv_{template_id}_{seed:04d}.csv"
        pre: list[dict] = []
        pre.append(_py_step(src_builder(src_path, seed)))
        pre.append(_py_step(_gold_export_csv(src_path, expected_csv)))
        pre.append({"type": "open", "parameters": {"path": src_path}})
        return {
            "instruction": instructions[seed % n],
            "out_path": result_csv,
            "expected_path": expected_csv,
            "pre_config_steps": pre,
            "post_open_config_steps": list(_LO_POSTLAUNCH_SETTLE),
            "oracle_after_postconfig": True,
        }

    return SynthTemplate(
        template_id=template_id,
        domain="libreoffice_calc",
        instruction_fn=lambda p: p["instruction"],
        evaluator_fn=lambda p: _eval_compare_csv(p["out_path"], p["expected_path"]),
        oracle_fn=lambda p: _cp_oracle(p["expected_path"], p["out_path"]),
        postconfig_fn=lambda _p: None,
        param_fn=_params,
        n_rows=n,
        setup_class=setup_class,
        eval_class="compare_csv",
    )


def _make_pdf_fit_one_page_template(*, template_id: str, setup_class: str,
                                       basename_xlsx: str,
                                       src_builder: Callable[[str, int], str],
                                       instructions: list[str]) -> SynthTemplate:
    """Build a fit-to-one-page → PDF-export SynthTemplate."""
    stem = basename_xlsx.rsplit(".", 1)[0]
    n = len(instructions)

    def _params(seed: int) -> dict:
        src_path = f"/home/user/{basename_xlsx}"
        result_pdf = f"/home/user/{stem}.pdf"
        expected_pdf = f"/tmp/expected_pdf_{template_id}_{seed:04d}.pdf"
        pre: list[dict] = []
        pre.append(_py_step(src_builder(src_path, seed)))
        pre.append(_py_step(_gold_export_pdf_fit_one_page(src_path, expected_pdf)))
        pre.append({"type": "open", "parameters": {"path": src_path}})
        return {
            "instruction": instructions[seed % n],
            "out_path": result_pdf,
            "expected_path": expected_pdf,
            "pre_config_steps": pre,
            "post_open_config_steps": list(_LO_POSTLAUNCH_SETTLE),
            "oracle_after_postconfig": True,
        }

    return SynthTemplate(
        template_id=template_id,
        domain="libreoffice_calc",
        instruction_fn=lambda p: p["instruction"],
        evaluator_fn=lambda p: _eval_pdf_pages_and_compare(
            p["out_path"], p["expected_path"], nb_pages=1
        ),
        oracle_fn=lambda p: _cp_oracle(p["expected_path"], p["out_path"]),
        postconfig_fn=lambda _p: None,
        param_fn=_params,
        n_rows=n,
        setup_class=setup_class,
        eval_class="check_pdf_pages+compare_pdfs",
    )


# ===========================================================================
# §I. File-task templates (Batch, dataclass form)
#
# Symmetric across all synth/*.py — every other
# domain mirrors this exact section layout (§I.a–§I.f). Only the field
# *types* on Param vary per domain (calc uses `rules`; impress uses
# `gold_mutate`+`examine_field`; gimp uses image-compare opts; ...).
#
# Unit of design: ONE `FileTask` = ONE (file × task) pair. Each FileTask
# emits ONE SynthTemplate with `n_rows = len(params)` ≤ cap. Topic family
# is resolved INSIDE the source builder per seed (free visual augmentation,
# NOT an axis). cartesian enumeration lives in FILE_TASKS expansion:
#
#     N files × M tasks/file = total templates
#                ↓ cap
#     emit ≤ N × SYNTH_CAP_TASKS_PER_FILE × SYNTH_CAP_PARAMS_PER_TASK
#
# Layout:
#   §I.a  Caps
#   §I.b  Dataclasses (File / Param / FileTask)
#   §I.c  File instances (define each File ONCE)
#   §I.d  Factory  (_to_synth_template + _emit_templates)
#   §I.e  FILE_TASKS — flat list, each entry is one (file, task) pair
#   §I.f  Emission — TEMPLATES.extend(_emit_templates(FILE_TASKS))
# ===========================================================================


# §I.a — caps. Hard upper bound; scale volume by adding more File entries.
SYNTH_CAP_TASKS_PER_FILE: int = 2
SYNTH_CAP_PARAMS_PER_TASK: int = 2


# §I.b — Dataclasses.

@dataclass(frozen=True)
class File:
    """One structurally distinct source xlsx.

    The `src` callable receives (path, seed) and returns a python heredoc
    string (passed to `_py_step`). Topic family is picked inside `src` via
    `random.Random(seed ^ _stable_hash(...))` — different seeds may pick
    different topics → free visual augmentation.

    `pre_steps`: tuple of config steps to run BEFORE the src heredoc (used
    by CSV-staged files for `_stage_asset`). Default empty for inline files.
    """
    id: str
    setup_class: str
    basename: str
    src: Callable[[str, int], str]
    pre_steps: tuple[dict, ...] = ()


@dataclass(frozen=True)
class Param:
    """One concrete parameterization of a task.

    Each Param's three fields rotate together — change one and you change
    a real operation, not just paraphrase. `gold_args` feeds the gold
    builder; `rules` is the compare_table rule list; `instr` is the
    rendered (NOT paraphrased) instruction string.
    """
    gold_args: dict
    rules: list
    instr: str


@dataclass(frozen=True)
class FileTask:
    """One (file, task) pair → one SynthTemplate at emit time.

    Adding a new task on an existing file = add one FileTask entry.
    Adding a new file with N tasks = define one File + N FileTask entries.
    Adding more params to an existing task = grow the `params` list.
    """
    file: File
    task_id: str
    eval_class: str
    gold: Callable[..., str]   # (src_path, exp_path, **gold_args) -> py heredoc
    params: list[Param] = field(default_factory=list)


# §I.c — File instances. Each is defined ONCE; FileTask entries reference
# it. Adding a new file = one new instance below.

F_CALC_1 = File(
    id="F-CALC-1", setup_class="pnl_table",
    basename="pnl.xlsx", src=_src_pnl_2col,
)
F_CALC_2 = File(
    id="F-CALC-2", setup_class="gradebook_table",
    basename="gradebook.xlsx", src=_src_gradebook,
)
F_CALC_3 = File(
    id="F-CALC-3", setup_class="orders_table",
    basename="orders.xlsx", src=_src_orders_north,
)
F_CALC_4 = File(
    id="F-CALC-4", setup_class="movies_table",
    basename="movies.xlsx", src=_src_movies,
)
F_CALC_5 = File(
    id="F-CALC-5", setup_class="expense_table",
    basename="expenses.xlsx", src=_src_expenses,
)
F_CALC_6 = File(
    id="F-CALC-6", setup_class="sales_table",
    basename="sales.xlsx", src=_src_sales,
)
F_CALC_7 = File(
    id="F-CALC-7", setup_class="inventory_table",
    basename="inventory.xlsx", src=_src_inventory,
)
F_CALC_8 = File(
    id="F-CALC-8", setup_class="reps_quarter_table",
    basename="sales_reps.xlsx", src=_src_sales_rep_quarter,
)
F_CALC_9 = File(
    id="F-CALC-9", setup_class="loan_table",
    basename="loans.xlsx", src=_src_loan_5col,
)
F_CALC_10 = File(
    id="F-CALC-10", setup_class="safety_inspection_table",
    basename="safety.xlsx", src=_src_safety_inspection,
)

# Additional File instances reusing existing _src_* builders.
F_CALC_11 = File(
    id="F-CALC-11", setup_class="user_emails_table",
    basename="user_emails.xlsx", src=_src_user_emails,
)
F_CALC_12 = File(
    id="F-CALC-12", setup_class="product_codes_table",
    basename="product_codes.xlsx", src=_src_product_codes,
)
F_CALC_13 = File(
    id="F-CALC-13", setup_class="phonics_titles_table",
    basename="phonics_titles.xlsx", src=_src_phonics_titles,
)
F_CALC_14 = File(
    id="F-CALC-14", setup_class="questionnaire_table",
    basename="questionnaire.xlsx", src=_src_questionnaire,
)
F_CALC_15 = File(
    id="F-CALC-15", setup_class="bus_schedule_table",
    basename="bus_schedule.xlsx", src=_src_bus_schedule,
)
F_CALC_16 = File(
    id="F-CALC-16", setup_class="tournament_table",
    basename="tournament.xlsx", src=_src_tournament,
)


# ---------------------------------------------------------------------------
# CSV-backed File factory: stage real CSV via host_push, then run a
# `_csv_src_*` builder that reads from that staged path. The deterministic
# /tmp/_synth_csv_<id>.csv path means the src closure can hardcode it.
# ---------------------------------------------------------------------------

def _make_csv_file(*, file_id: str, setup_class: str, basename: str,
                   csv_rel: str,
                   csv_builder: Callable[[str, str, int], str]) -> File:
    """Build a `File` that stages a real CSV before running its builder.

    `csv_builder(csv_in, src_path, seed) -> str` returns the python heredoc
    that reads the staged CSV at `csv_in` and writes the source xlsx at
    `src_path`. The CSV is staged at a deterministic /tmp path so the
    builder closure can capture it without per-seed leakage.
    """
    csv_dst = f"/tmp/_synth_csv_{file_id.lower().replace('-', '_')}.csv"

    def _src(src_path: str, seed: int) -> str:
        return csv_builder(csv_dst, src_path, seed)

    return File(
        id=file_id, setup_class=setup_class, basename=basename, src=_src,
        pre_steps=(_stage_asset(csv_rel, csv_dst),),
    )


# Additional File instances using existing `_src_*` builders in this file.
F_CALC_17 = File(
    id="F-CALC-17", setup_class="attendance_school_table",
    basename="attendance-school.xlsx", src=_src_attendance_school,
)
F_CALC_18 = File(
    id="F-CALC-18", setup_class="warehouse_inventory_table",
    basename="warehouse-inventory.xlsx", src=_src_inventory_warehouse,
)
F_CALC_19 = File(
    id="F-CALC-19", setup_class="quarterly_sales_table",
    basename="quarterly-sales.xlsx", src=_src_quarterly_sales,
)
F_CALC_20 = File(
    id="F-CALC-20", setup_class="fitness_log_table",
    basename="fitness-steps.xlsx", src=_src_fitness_steps,
)
F_CALC_21 = File(
    id="F-CALC-21", setup_class="concert_tour_table",
    basename="concert-tour.xlsx", src=_src_concert_revenue,
)
F_CALC_22 = File(
    id="F-CALC-22", setup_class="market_share_table",
    basename="market-share.xlsx", src=_src_market_share,
)
F_CALC_23 = File(
    id="F-CALC-23", setup_class="invoices_table",
    basename="invoices.xlsx", src=_src_invoices,
)
F_CALC_24 = File(
    id="F-CALC-24", setup_class="op_metrics_table",
    basename="ops-metrics.xlsx", src=_src_op_metrics,
)
F_CALC_25 = File(
    id="F-CALC-25", setup_class="attendance_lookup_table",
    basename="attendance-fees.xlsx", src=_src_attendance_lookup,
)
F_CALC_26 = File(
    id="F-CALC-26", setup_class="quarterly_3sheets_table",
    basename="quarterly-rollup.xlsx", src=_src_quarterly_3sheets,
)


# Real CSV-backed files (file-as-topic, real economic data).
F_CALC_27 = _make_csv_file(
    file_id="F-CALC-27", setup_class="us_gdp_table",
    basename="us-gdp.xlsx", csv_rel="data/csv/us-gdp.csv",
    csv_builder=_csv_src_us_gdp,
)
F_CALC_28 = _make_csv_file(
    file_id="F-CALC-28", setup_class="us_population_states_table",
    basename="us-population.xlsx", csv_rel="data/csv/us-population-states.csv",
    csv_builder=_csv_src_us_population_states,
)
F_CALC_29 = _make_csv_file(
    file_id="F-CALC-29", setup_class="us_unemployment_table",
    basename="us-unemployment.xlsx", csv_rel="data/csv/us-unemployment.csv",
    csv_builder=_csv_src_us_unemployment,
)
F_CALC_30 = _make_csv_file(
    file_id="F-CALC-30", setup_class="world_gdp_table",
    basename="world-gdp-2022.xlsx", csv_rel="data/csv/world-gdp-2022.csv",
    csv_builder=_csv_src_world_gdp_2022,
)
F_CALC_31 = _make_csv_file(
    file_id="F-CALC-31", setup_class="oil_wti_table",
    basename="oil-wti-daily.xlsx", csv_rel="data/csv/oil-wti-daily.csv",
    csv_builder=_csv_src_oil_wti_daily,
)
F_CALC_32 = _make_csv_file(
    file_id="F-CALC-32", setup_class="us_fed_funds_table",
    basename="us-fed-funds-rate.xlsx", csv_rel="data/csv/us-fed-funds-rate.csv",
    csv_builder=_csv_src_us_fed_funds,
)
F_CALC_33 = _make_csv_file(
    file_id="F-CALC-33", setup_class="us_housing_starts_table",
    basename="us-housing-starts.xlsx", csv_rel="data/csv/us-housing-starts.csv",
    csv_builder=_csv_src_us_housing_starts,
)
F_CALC_34 = _make_csv_file(
    file_id="F-CALC-34", setup_class="us_inflation_cpi_table",
    basename="us-inflation-cpi.xlsx", csv_rel="data/csv/us-inflation-cpi.csv",
    csv_builder=_csv_src_us_inflation_cpi,
)
F_CALC_35 = _make_csv_file(
    file_id="F-CALC-35", setup_class="us_mortgage_30yr_table",
    basename="us-mortgage-30yr.xlsx", csv_rel="data/csv/us-mortgage-30yr.csv",
    csv_builder=_csv_src_us_mortgage_30yr,
)
F_CALC_36 = _make_csv_file(
    file_id="F-CALC-36", setup_class="us_state_median_income_table",
    basename="us-state-median-income.xlsx",
    csv_rel="data/csv/us-state-median-income.csv",
    csv_builder=_csv_src_us_state_median_income,
)
F_CALC_37 = _make_csv_file(
    file_id="F-CALC-37", setup_class="world_population_table",
    basename="world-population-2022.xlsx",
    csv_rel="data/csv/world-population-2022.csv",
    csv_builder=_csv_src_world_population_2022,
)


# Newly authored xlsx files.
F_CALC_38 = File(
    id="F-CALC-38", setup_class="bank_transactions_table",
    basename="bank-transactions.xlsx", src=_src_bank_transactions,
)
F_CALC_39 = File(
    id="F-CALC-39", setup_class="event_schedule_table",
    basename="event-schedule.xlsx", src=_src_event_schedule,
)
F_CALC_40 = File(
    id="F-CALC-40", setup_class="product_catalog_table",
    basename="product-catalog.xlsx", src=_src_product_catalog,
)
F_CALC_41 = File(
    id="F-CALC-41", setup_class="clinic_visits_table",
    basename="clinic-visits.xlsx", src=_src_clinic_visits,
)
F_CALC_42 = File(
    id="F-CALC-42", setup_class="survey_responses_table",
    basename="survey-responses.xlsx", src=_src_survey_responses,
)
F_CALC_43 = File(
    id="F-CALC-43", setup_class="temperature_log_table",
    basename="temperature-log.xlsx", src=_src_temperature_log,
)
F_CALC_44 = File(
    id="F-CALC-44", setup_class="commute_log_table",
    basename="commute-log.xlsx", src=_src_commute_log,
)
F_CALC_45 = File(
    id="F-CALC-45", setup_class="subscription_data_table",
    basename="subscription-data.xlsx", src=_src_subscription_data,
)
F_CALC_46 = File(
    id="F-CALC-46", setup_class="recipe_ingredients_table",
    basename="recipe-ingredients.xlsx", src=_src_recipe_ingredients,
)
F_CALC_47 = File(
    id="F-CALC-47", setup_class="student_grades_multi_table",
    basename="student-grades-multi.xlsx", src=_src_student_grades_multi,
)
F_CALC_48 = File(
    id="F-CALC-48", setup_class="library_loans_table",
    basename="library-loans.xlsx", src=_src_library_loans,
)
F_CALC_49 = File(
    id="F-CALC-49", setup_class="fleet_vehicles_table",
    basename="fleet-vehicles.xlsx", src=_src_fleet_vehicles,
)
F_CALC_50 = File(
    id="F-CALC-50", setup_class="employees_payroll_table",
    basename="employees-payroll.xlsx", src=_src_employees_payroll,
)
F_CALC_51 = File(
    id="F-CALC-51", setup_class="weather_log_table",
    basename="weather-log.xlsx", src=_src_weather_log,
)
F_CALC_52 = File(
    id="F-CALC-52", setup_class="restaurant_menu_table",
    basename="restaurant-menu.xlsx", src=_src_restaurant_menu,
)
F_CALC_53 = File(
    id="F-CALC-53", setup_class="donations_log_table",
    basename="donations-log.xlsx", src=_src_donations_log,
)
F_CALC_54 = File(
    id="F-CALC-54", setup_class="books_catalog_table",
    basename="books-catalog.xlsx", src=_src_books_catalog,
)
F_CALC_55 = File(
    id="F-CALC-55", setup_class="movie_box_office_table",
    basename="movie-box-office.xlsx", src=_src_movie_box_office,
)
F_CALC_56 = File(
    id="F-CALC-56", setup_class="real_estate_listings_table",
    basename="real-estate-listings.xlsx", src=_src_real_estate_listings,
)
F_CALC_57 = File(
    id="F-CALC-57", setup_class="purchase_orders_table",
    basename="purchase-orders.xlsx", src=_src_purchase_orders,
)
F_CALC_58 = File(
    id="F-CALC-58", setup_class="class_attendance_table",
    basename="class-attendance.xlsx", src=_src_class_attendance,
)
F_CALC_59 = File(
    id="F-CALC-59", setup_class="employee_skills_table",
    basename="employee-skills.xlsx", src=_src_employee_skills,
)
F_CALC_60 = File(
    id="F-CALC-60", setup_class="warehouse_orders_table",
    basename="warehouse-orders.xlsx", src=_src_warehouse_orders,
)
F_CALC_61 = File(
    id="F-CALC-61", setup_class="taxi_fares_table",
    basename="taxi-fares.xlsx", src=_src_taxi_fares,
)
F_CALC_62 = File(
    id="F-CALC-62", setup_class="workout_log_table",
    basename="workout-log.xlsx", src=_src_workout_log,
)
F_CALC_63 = File(
    id="F-CALC-63", setup_class="apartment_rents_table",
    basename="apartment-rents.xlsx", src=_src_apartment_rents,
)
F_CALC_64 = File(
    id="F-CALC-64", setup_class="blog_posts_table",
    basename="blog-posts.xlsx", src=_src_blog_posts,
)
F_CALC_65 = File(
    id="F-CALC-65", setup_class="classroom_quiz_table",
    basename="classroom-quiz.xlsx", src=_src_classroom_quiz,
)
F_CALC_66 = File(
    id="F-CALC-66", setup_class="film_festival_table",
    basename="film-festival.xlsx", src=_src_film_festival,
)
F_CALC_67 = File(
    id="F-CALC-67", setup_class="charity_pledges_table",
    basename="charity-pledges.xlsx", src=_src_charity_pledges,
)
F_CALC_68 = File(
    id="F-CALC-68", setup_class="software_bugs_table",
    basename="software-bugs.xlsx", src=_src_software_bugs,
)
F_CALC_69 = File(
    id="F-CALC-69", setup_class="hotel_bookings_table",
    basename="hotel-bookings.xlsx", src=_src_hotel_bookings,
)
F_CALC_70 = File(
    id="F-CALC-70", setup_class="app_downloads_table",
    basename="app-downloads.xlsx", src=_src_app_downloads,
)
F_CALC_71 = File(
    id="F-CALC-71", setup_class="country_capitals_table",
    basename="country-capitals.xlsx", src=_src_country_capitals,
)


# Final batch — F-CALC-72..F-CALC-90 (target: 90 files total).
F_CALC_72 = File(
    id="F-CALC-72", setup_class="gym_membership_table",
    basename="gym-membership.xlsx", src=_src_gym_membership,
)
F_CALC_73 = File(
    id="F-CALC-73", setup_class="garden_plants_table",
    basename="garden-plants.xlsx", src=_src_garden_plants,
)
F_CALC_74 = File(
    id="F-CALC-74", setup_class="warehouse_skus_table",
    basename="warehouse-skus.xlsx", src=_src_warehouse_skus,
)
F_CALC_75 = File(
    id="F-CALC-75", setup_class="streaming_subs_table",
    basename="streaming-subs.xlsx", src=_src_streaming_subs,
)
F_CALC_76 = File(
    id="F-CALC-76", setup_class="pet_clinic_table",
    basename="pet-clinic.xlsx", src=_src_pet_clinic,
)
F_CALC_77 = File(
    id="F-CALC-77", setup_class="concert_tickets_table",
    basename="concert-tickets.xlsx", src=_src_concert_tickets,
)
F_CALC_78 = File(
    id="F-CALC-78", setup_class="runners_log_table",
    basename="runners-log.xlsx", src=_src_runners_log,
)
F_CALC_79 = File(
    id="F-CALC-79", setup_class="recipe_ratings_table",
    basename="recipe-ratings.xlsx", src=_src_recipe_ratings,
)
F_CALC_80 = File(
    id="F-CALC-80", setup_class="invoice_aging_table",
    basename="invoice-aging.xlsx", src=_src_invoice_aging,
)
F_CALC_81 = File(
    id="F-CALC-81", setup_class="university_courses_table",
    basename="university-courses.xlsx", src=_src_university_courses,
)
F_CALC_82 = File(
    id="F-CALC-82", setup_class="smartphone_models_table",
    basename="smartphone-models.xlsx", src=_src_smartphone_models,
)
F_CALC_83 = File(
    id="F-CALC-83", setup_class="lab_results_table",
    basename="lab-results.xlsx", src=_src_lab_results,
)
F_CALC_84 = File(
    id="F-CALC-84", setup_class="freight_routes_table",
    basename="freight-routes.xlsx", src=_src_freight_routes,
)
F_CALC_85 = File(
    id="F-CALC-85", setup_class="internet_speeds_table",
    basename="internet-speeds.xlsx", src=_src_internet_speeds,
)
F_CALC_86 = File(
    id="F-CALC-86", setup_class="volunteer_hours_table",
    basename="volunteer-hours.xlsx", src=_src_volunteer_hours,
)
F_CALC_87 = File(
    id="F-CALC-87", setup_class="solar_panels_table",
    basename="solar-panels.xlsx", src=_src_solar_panels,
)
F_CALC_88 = File(
    id="F-CALC-88", setup_class="wine_inventory_table",
    basename="wine-inventory.xlsx", src=_src_wine_inventory,
)
F_CALC_89 = File(
    id="F-CALC-89", setup_class="pottery_orders_table",
    basename="pottery-orders.xlsx", src=_src_pottery_orders,
)
F_CALC_90 = File(
    id="F-CALC-90", setup_class="helpdesk_tickets_table",
    basename="helpdesk-tickets.xlsx", src=_src_helpdesk_tickets,
)


# Batch — eval-alignment files (P2-P6). Cap=2 tasks/file enforced
# by _emit_templates, so most new files host 1-2 FileTasks each.
F_CALC_91 = File(
    id="F-CALC-91", setup_class="sales_chart_data_table",
    basename="sales-chart-data.xlsx", src=_src_sales_chart_data,
)
F_CALC_92 = File(
    id="F-CALC-92", setup_class="customer_orders_5col_table",
    basename="customer-orders-5col.xlsx", src=_src_customer_orders_5col,
)
F_CALC_93 = File(
    id="F-CALC-93", setup_class="lars_two_sheet_table",
    basename="lars-resources.xlsx", src=_src_lars_two_sheet,
)
F_CALC_94 = File(
    id="F-CALC-94", setup_class="employee_birthday_table",
    basename="employee-birthday.xlsx", src=_src_employee_birthday,
)
F_CALC_95 = File(
    id="F-CALC-95", setup_class="old_id_padding_table",
    basename="old-id-padding.xlsx", src=_src_old_id_padding,
)
F_CALC_96 = File(
    id="F-CALC-96", setup_class="project_dates_table",
    basename="project-dates.xlsx", src=_src_project_dates,
)
F_CALC_97 = File(
    id="F-CALC-97", setup_class="company_financials_table",
    basename="company-financials.xlsx", src=_src_company_financials,
)
F_CALC_98 = File(
    id="F-CALC-98", setup_class="monthly_sales_growth_table",
    basename="monthly-sales-growth.xlsx", src=_src_monthly_sales_growth,
)
F_CALC_99 = File(
    id="F-CALC-99", setup_class="region_sales_table",
    basename="region-sales.xlsx", src=_src_region_sales,
)
F_CALC_100 = File(
    id="F-CALC-100", setup_class="expense_chart_data_table",
    basename="expense-chart-data.xlsx", src=_src_expense_chart_data,
)
# Cycle-N — eval-essence emulation files (see _src_weekend_calendar /
# _src_hourly_value docstrings).
F_CALC_101 = File(
    id="F-CALC-101", setup_class="weekend_calendar_table",
    basename="weekend-calendar.xlsx", src=_src_weekend_calendar,
)
F_CALC_102 = File(
    id="F-CALC-102", setup_class="hourly_value_table",
    basename="hourly-value.xlsx", src=_src_hourly_value,
)


# §I.d — Factory.

def _to_synth_template(ft: FileTask) -> SynthTemplate:
    """Turn ONE FileTask into ONE SynthTemplate.

    Per-seed: pick the i-th Param from ft.params (i = seed % len(params)),
    re-render the source xlsx (topic random per seed) and the gold xlsx
    (rotated gold_args). Eval rule rotates alongside.
    """
    pool = ft.params[:SYNTH_CAP_PARAMS_PER_TASK]
    template_id = f"{ft.file.id.lower().replace('-', '_')}__{ft.task_id}"

    def _params(seed: int) -> dict:
        variant = pool[seed % len(pool)]
        src_path = f"/home/user/Desktop/{ft.file.basename}"
        exp_path = f"/tmp/expected_calc_{template_id}_{seed:04d}.xlsx"
        pre = list(ft.file.pre_steps)
        pre.append(_py_step(ft.file.src(src_path, seed)))
        pre.append(_py_step(ft.gold(src_path, exp_path, **variant.gold_args)))
        # validation calc Save-As bridge (libreoffice_calc.md
        # row 1 of the quant snapshot). Synth previously launched LO Calc via
        # an `open_command` (= a bare `launch` step) but never opened the source
        # xlsx — agent saw blank Untitled1 → had to Save-As to the source path
        # → GTK Name-field Ctrl+A bug duplicated the basename. Eval's 44/46
        # rows use a `{"type": "open"}` step that both opens the file AND
        # launches LO. Switching synth to the same shape eliminates the
        # Save-As trap on 100% of calc rows (`measure_gap.calc_save_protocol`
        # → `open+ctrl_s`). The `common.py` auto-launch fallback already
        # treats `type=="open"` as "app already launched" so no `open_command`
        # is needed.
        pre.append({"type": "open", "parameters": {"path": src_path}})
        return {
            "instruction":   variant.instr,
            "rules":         variant.rules,
            "out_path":      src_path,
            "expected_path": exp_path,
            "pre_config_steps": pre,
            "post_open_config_steps": list(_LO_POSTLAUNCH_SETTLE),
            "oracle_after_postconfig": True,
        }

    return SynthTemplate(
        template_id=template_id,
        domain="libreoffice_calc",
        instruction_fn=lambda p: p["instruction"],
        evaluator_fn=lambda p: _eval_compare_table(
            p["out_path"], p["expected_path"], p["rules"]
        ),
        oracle_fn=lambda p: _cp_oracle(p["expected_path"], p["out_path"]),
        postconfig_fn=lambda _p: None,
        param_fn=_params,
        n_rows=len(pool),
        setup_class=ft.file.setup_class,
        eval_class=ft.eval_class,
    )


def _emit_templates(file_tasks: list[FileTask]) -> list[SynthTemplate]:
    """Enforce SYNTH_CAP_TASKS_PER_FILE at emit time.

    Tasks beyond the cap (per file) are kept in FILE_TASKS for ablation /
    review (headroom) but not emitted. Quality-rank tasks by listing the
    most eval-aligned first.
    """
    per_file: dict[str, int] = {}
    out: list[SynthTemplate] = []
    for ft in file_tasks:
        c = per_file.get(ft.file.id, 0)
        if c >= SYNTH_CAP_TASKS_PER_FILE:
            continue  # headroom for ablations
        per_file[ft.file.id] = c + 1
        out.append(_to_synth_template(ft))
    return out


# §I.e — FILE_TASKS: flat list. Each entry is one (file × task) pair.

FILE_TASKS: list[FileTask] = [
    # F-CALC-1 — PnL deck (Month / Revenue / Expenses, 24 months)
    FileTask(F_CALC_1, "copy_col_to_new_sheet", "multi_sheet_aggregate",
             _gold_sheet2_copy_col, params=[
        Param({"src_col_idx": 1, "new_sheet_name": "Forecast",
               "new_header": "Revenue"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Forecast"),
              "I'm preparing a forecast workbook from this P&L deck and need "
              "the Revenue numbers isolated on their own tab — please create "
              "a new sheet named 'Forecast' and copy the Revenue column with "
              "its header into it for the modelling team."),
        Param({"src_col_idx": 2, "new_sheet_name": "CostsOnly",
               "new_header": "Expenses"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("CostsOnly"),
              "Please help me prepare an expenses-only tab for the cost "
              "audit before the printing deadline — create a new sheet "
              "named 'CostsOnly' and copy the Expenses column with its "
              "header into it for the controller's review."),
    ]),
    FileTask(F_CALC_1, "aggregate_to_new_sheet", "multi_sheet_aggregate",
             _gold_sheet2_aggregate, params=[
        # validation calc P1: use the new `check_cell+sheet_data` rule combo on
        # this Param (eval has 2 check_cell rows; synth previously had 0). The
        # `_gold_sheet2_aggregate` builder writes the label_header ("Metric")
        # into Sheet 'Summary'!A1 — pinning that exact cell makes the agent's
        # header-text match the literal expected string, matching eval's
        # check_cell pattern on Sheet2 header cells.
        Param({"sum_cols": [(1, "Revenue"), (2, "Expenses")],
               "new_sheet_name": "Summary",
               "agg_label_header": "Metric", "agg_value_header": "Total"},
              _RULE_CHECK_CELL_NAMED_SHEET("Summary", "A1", "Metric"),
              "Could you help me build a summary tab before sharing this P&L "
              "with the leadership team? Add a new 'Summary' sheet with two "
              "columns 'Metric' and 'Total' — one row totalling Revenue across "
              "all months and one row totalling Expenses."),
        Param({"sum_cols": [(1, "Revenue")],
               "new_sheet_name": "RevTotal",
               "agg_label_header": "Metric", "agg_value_header": "Total"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("RevTotal"),
              "Add a new sheet 'RevTotal' with columns 'Metric' and "
              "'Total', containing a single row whose Metric is 'Revenue' "
              "and whose Total is the sum of the Revenue column."),
    ]),

    # F-CALC-2 — gradebook (Name / Score, 12 students)
    FileTask(F_CALC_2, "color_by_score", "conditional_format",
             _gold_two_color_by_predicate, params=[
        Param({"pred_a_py": "isinstance(row[1], (int,float)) and row[1] < 60",
               "argb_a": "FFFF0000",
               "pred_b_py": "isinstance(row[1], (int,float)) and row[1] >= 60",
               "argb_b": "FF00FF00", "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like to highlight pass/fail at a glance for parent-teacher "
              "night — in the gradebook, give every row whose Score is below "
              "60 a red fill and every row whose Score is 60 or above a "
              "green fill so the conferences run more smoothly."),
        Param({"pred_a_py": "isinstance(row[1], (int,float)) and row[1] >= 90",
               "argb_a": "FFFFD700",
               "pred_b_py": "isinstance(row[1], (int,float)) and row[1] < 60",
               "argb_b": "FFFF0000", "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Highlight honour-roll and failing rows: Score ≥ 90 in gold "
              "fill (gold = Custom Color #FFD700), Score < 60 in red fill. Leave the rest unstyled."),
    ]),
    FileTask(F_CALC_2, "filter_to_new_sheet", "multi_sheet_aggregate",
             _gold_sheet2_filter_score_threshold, params=[
        Param({"new_sheet_name": "Failing", "score_col_idx": 1,
               "predicate": "lt", "threshold": 60},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Failing"),
              "Please help me pull together a short list of students who need "
              "extra support this quarter — create a new sheet 'Failing' that "
              "holds the header row plus every student whose Score is strictly "
              "below 60 so I can share it with the counsellor."),
        Param({"new_sheet_name": "Honor Roll", "score_col_idx": 1,
               "predicate": "ge", "threshold": 90},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Honor Roll"),
              "Create a new sheet named 'Honor Roll' containing the header "
              "row plus every student whose Score is 90 or higher."),
    ]),

    # F-CALC-3 — orders (OrderID / Customer / Region / Amount, 24 orders)
    FileTask(F_CALC_3, "filter_by_region", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "North Orders", "filter_col_idx": 2,
               "filter_value": "North"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("North Orders"),
              "I'm working on the Northeast sales review and need a focused "
              "view — create a new sheet named 'North Orders' containing only "
              "the orders whose Region is North (copy the header too) so I "
              "can prep the regional manager's briefing."),
        Param({"new_sheet_name": "South Orders", "filter_col_idx": 2,
               "filter_value": "South"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("South Orders"),
              "I'm pulling a Southern-region order extract for tomorrow's "
              "sales standup — create a new sheet named 'South Orders' that "
              "holds the header row plus every order whose Region is South "
              "so the regional team has a focused view."),
    ]),
    FileTask(F_CALC_3, "groupby_region_totals", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "By Region", "key_col_idx": 2,
               "value_col_idx": 3, "agg": "sum",
               "key_header": "Region", "value_header": "TotalAmount"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Region"),
              "Could you help me roll up the orders by region for tomorrow's "
              "exec meeting? Add a new sheet 'By Region' summing the order "
              "Amount per Region with header row 'Region','TotalAmount' "
              "followed by one row per distinct region with its total."),
        Param({"new_sheet_name": "By Region Count", "key_col_idx": 2,
               "value_col_idx": 3, "agg": "count",
               "key_header": "Region", "value_header": "OrderCount"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Region Count"),
              "Add a new sheet 'By Region Count' tallying how many orders "
              "each Region has: header row 'Region','OrderCount' followed "
              "by one row per distinct region with its count."),
    ]),

    # F-CALC-4 — movies (Title / Year / Director / Rating / Genre, 30 rows)
    FileTask(F_CALC_4, "sort_by_year_asc", "sort_col",
             _gold_sort, params=[
        Param({"col_idx": 1, "reverse": False},
              _RULE_SHEET_DATA,
              "I'm putting together a film-history retrospective for the "
              "community library and need a chronological catalogue — please "
              "sort the movies by Year ascending (oldest first) and keep the "
              "header row at row 1 so the timeline reads naturally."),
        Param({"col_idx": 3, "reverse": True},
              _RULE_SHEET_DATA,
              "Sort the movies by Rating descending (highest first); keep "
              "the header row at row 1."),
    ]),
    FileTask(F_CALC_4, "filter_to_new_sheet_by_genre", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Sci-Fi", "filter_col_idx": 4,
               "filter_value": "Sci-Fi"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Sci-Fi"),
              "I'm curating a sci-fi marathon night and need a focused "
              "shortlist — create a new sheet 'Sci-Fi' containing only the "
              "movies whose Genre is Sci-Fi (copy the header too) so I can "
              "share it with the screening club."),
        Param({"new_sheet_name": "Drama", "filter_col_idx": 4,
               "filter_value": "Drama"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Drama"),
              "I'd like to share a drama-only film list with the community "
              "screening club next month — create a new sheet named 'Drama' "
              "that holds the header row plus every movie whose Genre is "
              "Drama so the curator can plan the lineup."),
    ]),

    # F-CALC-5 — household expenses (Month / Category / Amount / Notes, 24 rows)
    FileTask(F_CALC_5, "total_row_amount", "apply_formula",
             _gold_total_row, params=[
        Param({"sum_col_idxs": [2], "label": "Total", "label_col_idx": 1},
              _RULE_SHEET_DATA,
              "I'm closing the books on this month's household ledger and "
              "would like a single bottom-line figure — please append a row "
              "at the bottom with 'Total' in the Category column and the sum "
              "of the Amount column alongside it for our budget review."),
        Param({"sum_col_idxs": [2], "label": "Grand Total",
               "label_col_idx": 1},
              _RULE_SHEET_DATA,
              "Append a row at the bottom labelled 'Grand Total' under "
              "Category with the sum of the Amount column alongside."),
    ]),
    FileTask(F_CALC_5, "groupby_category_totals", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "By Category", "key_col_idx": 1,
               "value_col_idx": 2, "agg": "sum",
               "key_header": "Category", "value_header": "TotalAmount"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Category"),
              "I'd like to see where our spending actually goes before next "
              "month's budget meeting — add a new sheet 'By Category' summing "
              "the Amount per Category with header row 'Category','TotalAmount' "
              "then one row per distinct category with its total."),
        Param({"new_sheet_name": "Avg By Category", "key_col_idx": 1,
               "value_col_idx": 2, "agg": "avg",
               "key_header": "Category", "value_header": "AvgAmount"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Avg By Category"),
              "I'd like to compare the average spend per category for next "
              "year's budgeting — add a new sheet 'Avg By Category' with "
              "header 'Category','AvgAmount' followed by one row per "
              "distinct category showing its average for the planner."),
    ]),

    # F-CALC-6 — sales table (Month / Sales / COGS / Profit / Region / Channel)
    # Validation PARAM_REDUCIBLE: dropped the Margin (0.0%)
    # Param. The strict 0.0% number_format check was hostile (eval was
    # strict on number_format string). Kept the MarkupRatio Param whose
    # 0.00 format is a simpler / more discriminable target.
    FileTask(F_CALC_6, "derived_margin_col", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "MarkupRatio", "src_col_idx": 1,
               "expr": "float(ws.cell(r, 3).value) and "
                       "float(v) / float(ws.cell(r, 3).value)",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'MarkupRatio' = Sales / COGS formatted "
              "as 0.00 (two-decimal number)."),
    ]),
    FileTask(F_CALC_6, "filter_by_channel", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Online Sales", "filter_col_idx": 5,
               "filter_value": "Online"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Online Sales"),
              "I'm preparing the e-commerce channel report for the marketing "
              "team and need a clean slice — create a new sheet 'Online Sales' "
              "containing only the rows whose Channel is Online (copy the "
              "header too) so it's ready for the digital review."),
        Param({"new_sheet_name": "Retail Sales", "filter_col_idx": 5,
               "filter_value": "Retail"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Retail Sales"),
              "I'm preparing the retail-channel report for the brick-and-mortar "
              "team's monthly review — create a new sheet named 'Retail Sales' "
              "that holds the header row plus every sales row whose Channel "
              "is Retail so the store managers see a clean view."),
    ]),

    # F-CALC-7 — inventory (Product / Revenue / Margin)
    FileTask(F_CALC_7, "sort_by_revenue_desc", "sort_col",
             _gold_sort, params=[
        Param({"col_idx": 1, "reverse": True},
              _RULE_SHEET_DATA,
              "I'd like to highlight our top-revenue SKUs for the warehouse "
              "team's quarterly product review — please sort the inventory "
              "by Revenue descending (highest first) and keep the header "
              "row at row 1 so the leaders are easy to spot."),
        Param({"col_idx": 2, "reverse": True},
              _RULE_SHEET_DATA,
              "Sort the inventory by Margin descending (highest first); "
              "keep the header row at row 1."),
    ]),
    FileTask(F_CALC_7, "color_by_margin", "conditional_format",
             _gold_cell_color_by_predicate, params=[
        Param({"rules_py": "isinstance(row[2], (int,float)) and row[2] < 0.20",
               "apply_kind": "fill", "apply_argb": "FFFFC0C0"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'm running a low-margin investigation before the pricing "
              "committee meeting — give every row whose Margin is below 0.20 "
              "a pink fill so the at-risk SKUs jump out for the team to "
              "discuss in the printed handout (use Custom Color hex #FFC0C0 "
              "in the fill picker)."),
        Param({"rules_py": "isinstance(row[2], (int,float)) and row[2] >= 0.35",
               "apply_kind": "fill", "apply_argb": "FF90EE90"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like to highlight our most profitable SKUs ahead of the "
              "product strategy offsite — give every row whose Margin is "
              "0.35 or higher a light-green fill so the high-margin stars "
              "stand out on the printed handout (use Custom Color hex "
              "#90EE90 in the fill picker)."),
    ]),

    # F-CALC-8 — sales reps × quarter (Rep / Q1..Q6)
    FileTask(F_CALC_8, "summary_rollup", "multi_sheet_aggregate",
             _gold_sheet2_aggregate, params=[
        # validation calc P1: check_cell+sheet_data rule combo (eval-anchored
        # `check_cell` skill_class).
        Param({"sum_cols": [(1, "Q1"), (2, "Q2"), (3, "Q3"),
                            (4, "Q4"), (5, "Q5"), (6, "Q6")],
               "new_sheet_name": "Summary",
               "agg_label_header": "Quarter", "agg_value_header": "TotalSales"},
              _RULE_CHECK_CELL_NAMED_SHEET("Summary", "B1", "TotalSales"),
              "Please help me roll up the rep-by-quarter sales for the sales "
              "ops review — create a 'Summary' sheet with two columns "
              "'Quarter' and 'TotalSales' and one row per quarter Q1..Q6 "
              "totalling that column across all reps."),
        Param({"sum_cols": [(1, "Q1"), (2, "Q2"), (3, "Q3")],
               "new_sheet_name": "FirstHalf",
               "agg_label_header": "Quarter", "agg_value_header": "TotalSales"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("FirstHalf"),
              "I'm preparing a first-half summary for the sales ops review "
              "before next week's leadership meeting — create a 'FirstHalf' "
              "sheet with columns 'Quarter' and 'TotalSales' rolling up "
              "only Q1, Q2, Q3 across all reps."),
    ]),
    FileTask(F_CALC_8, "sort_by_q1_desc", "sort_col",
             _gold_sort, params=[
        Param({"col_idx": 1, "reverse": True},
              _RULE_SHEET_DATA,
              "Sort the reps by Q1 sales descending (highest first); keep "
              "the header row at row 1."),
        Param({"col_idx": 6, "reverse": True},
              _RULE_SHEET_DATA,
              "Sort the reps by Q6 sales descending (highest first); keep "
              "the header row at row 1."),
    ]),

    # F-CALC-9 — loan portfolio (LoanID / StartDate / Principal / Rate / Tenor)
    FileTask(F_CALC_9, "merge_header", "conditional_format",
             _gold_merge_header, params=[
        Param({"merge_range": "A1:E1",
               "merge_text": "Loan Portfolio Summary", "bold": True},
              _RULE_SHEET_DATA_AND_FONT,
              "I'm formatting the loan-book worksheet for the credit "
              "committee's printed handout — please insert a new row at "
              "the top of the sheet (so the existing data shifts down by "
              "one row), then merge cells A1:E1 in that new row and place "
              "the bold title 'Loan Portfolio Summary' into the merged "
              "banner so the report has a proper banner heading."),
        Param({"merge_range": "A1:E1",
               "merge_text": "Loan Book — 2026 Q1", "bold": True, "italic": True},
              _RULE_SHEET_DATA_AND_FONT,
              "Please help me finalise the Q1 loan-book worksheet for the "
              "credit committee's printed pack — insert a new row at the "
              "top of the sheet (so the existing data shifts down by one "
              "row), then merge cells A1:E1 in that new row and put the "
              "bold italic title 'Loan Book — 2026 Q1' into the merged "
              "banner so the report has a proper styled heading."),
    ]),
    FileTask(F_CALC_9, "color_by_rate", "conditional_format",
             _gold_cell_color_by_predicate, params=[
        Param({"rules_py": "isinstance(row[3], (int,float)) and row[3] > 0.06",
               "apply_kind": "fill", "apply_argb": "FFFFD0D0"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Give every loan row whose Rate is above 0.06 a pale-red "
              "fill (high-rate flag) (use Custom Color hex #FFD0D0 in the "
              "fill picker)."),
        Param({"rules_py": "isinstance(row[3], (int,float)) and row[3] < 0.05",
               "apply_kind": "fill", "apply_argb": "FFD0F0D0"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Give every loan row whose Rate is below 0.05 a pale-green "
              "fill (low-rate flag) (use Custom Color hex #D0F0D0 in the "
              "fill picker)."),
    ]),

    # F-CALC-10 — safety inspection (Item / Status, Pass/Fail)
    FileTask(F_CALC_10, "color_pass_fail", "conditional_format",
             _gold_two_color_by_predicate, params=[
        Param({"pred_a_py": "row[1] == 'Pass'",  "argb_a": "FF90EE90",
               "pred_b_py": "row[1] == 'Fail'",  "argb_b": "FFFF9090",
               "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like the safety-inspection results colour-coded before the "
              "compliance audit next week — in the inspection list, give "
              "every row whose Status is 'Pass' a green fill and every row "
              "whose Status is 'Fail' a red fill for the auditor's binder "
              "(Pass = Custom Color #90EE90, Fail = #FF9090 via Format → "
              "Cells → Background)."),
        Param({"pred_a_py": "row[1] == 'Pass'",  "argb_a": "FFA0D8EF",
               "pred_b_py": "row[1] == 'Fail'",  "argb_b": "FFFFFF00",
               "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like a colourblind-friendly inspection report for our "
              "factory-floor audit binder — give every row whose Status is "
              "'Pass' a light-blue fill and every row whose Status is 'Fail' "
              "a yellow fill for clearer print legibility (Pass = Custom "
              "Color #A0D8EF, Fail = #FFFF00 via Format → Cells → "
              "Background)."),
    ]),
    FileTask(F_CALC_10, "filter_failing_to_new_sheet", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Failing Items", "filter_col_idx": 1,
               "filter_value": "Fail"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Failing Items"),
              "Create a new sheet 'Failing Items' containing only the "
              "rows whose Status is 'Fail' (copy the header too)."),
        Param({"new_sheet_name": "Passing Items", "filter_col_idx": 1,
               "filter_value": "Pass"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Passing Items"),
              "I'm building a passing-items export for the safety officer's "
              "weekly compliance file — create a new sheet named 'Passing "
              "Items' that holds the header row plus every inspection row "
              "whose Status is 'Pass' for the archive."),
    ]),

    # ── F-CALC-11..F-CALC-16 ──────────────────────────────────────

    # F-CALC-11 — user_emails (UserId / Email / Clean) string-clean target
    FileTask(F_CALC_11, "string_clean_email", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 2, "op": "lower"},
              _RULE_SHEET_DATA,
              "I'm cleaning the user-email list before importing into our CRM "
              "and need the addresses normalised — please fill the Clean "
              "column with each Email lower-cased so the import job won't "
              "create duplicate records."),
        Param({"src_col_idx": 1, "dst_col_idx": 2, "op": "strip"},
              _RULE_SHEET_DATA,
              "I'd like to normalise the email addresses for the upcoming "
              "newsletter send — fill the Clean column with each Email "
              "with collapsed whitespace (single-space separation, case "
              "preserved) so the mailing list is clean for the campaign."),
    ]),
    FileTask(F_CALC_11, "sort_by_userid", "sort_col", _gold_sort, params=[
        Param({"col_idx": 0, "reverse": False},
              _RULE_SHEET_DATA,
              "Sort the user list by UserId ascending; keep the header at "
              "row 1."),
        Param({"col_idx": 0, "reverse": True},
              _RULE_SHEET_DATA,
              "Sort the user list by UserId descending; keep the header at "
              "row 1."),
    ]),

    # F-CALC-12 — product_codes (SKU / Code / Clean) string-clean target
    FileTask(F_CALC_12, "string_clean_proper", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 2, "op": "proper_strip"},
              _RULE_SHEET_DATA,
              "Could you help me tidy up the product-code list before sharing "
              "it with the catalog team? Fill the Clean column with each "
              "Code in title case — collapse extra spaces first, then "
              "title-case the result for consistency."),
        Param({"src_col_idx": 1, "dst_col_idx": 2, "op": "title"},
              _RULE_SHEET_DATA,
              "Fill the Clean column with each Code title-cased (capitalise "
              "the first letter of each word)."),
    ]),
    FileTask(F_CALC_12, "sort_by_sku", "sort_col", _gold_sort, params=[
        Param({"col_idx": 0, "reverse": False},
              _RULE_SHEET_DATA,
              "Sort the products by SKU ascending; keep the header at row 1."),
        Param({"col_idx": 1, "reverse": False},
              _RULE_SHEET_DATA,
              "Sort the products alphabetically by Code; keep the header at "
              "row 1."),
    ]),

    # F-CALC-13 — phonics titles (Idx / Raw / Year / Clean) — string-clean target
    FileTask(F_CALC_13, "clean_titles_to_proper", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 3, "op": "proper_strip"},
              _RULE_SHEET_DATA,
              "I'd like to clean up these mangled film titles before sending "
              "the catalogue to the printer — fill the Clean column with each "
              "Raw title normalised by collapsing extra whitespace and "
              "converting to Title Case for the typesetter."),
        Param({"src_col_idx": 1, "dst_col_idx": 3, "op": "lower"},
              _RULE_SHEET_DATA,
              "Fill the Clean column with each Raw title lower-cased."),
    ]),
    FileTask(F_CALC_13, "sort_titles_by_year", "sort_col", _gold_sort, params=[
        Param({"col_idx": 2, "reverse": False},
              _RULE_SHEET_DATA,
              "Sort the titles by Year ascending (oldest first); keep the "
              "header at row 1."),
        Param({"col_idx": 2, "reverse": True},
              _RULE_SHEET_DATA,
              "I'm preparing the most-recent-titles section for the gallery "
              "exhibition brochure — sort the titles by Year descending so "
              "the newest entries come first, and keep the header at row 1 "
              "for the printer's typesetter."),
    ]),

    # F-CALC-14 — questionnaire (Respondents / Sex / Civil Status / Educ)
    FileTask(F_CALC_14, "filter_by_status", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Married", "filter_col_idx": 2,
               "filter_value": "Married"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Married"),
              "I'm preparing a subgroup analysis from this survey for the "
              "research write-up — create a new sheet 'Married' containing "
              "only the respondents whose Civil Status is 'Married' (copy "
              "the header too) for the demographic breakdown."),
        Param({"new_sheet_name": "Single", "filter_col_idx": 2,
               "filter_value": "Single"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Single"),
              "Create a new sheet 'Single' that holds the header row plus "
              "every respondent whose Civil Status is 'Single'."),
    ]),
    FileTask(F_CALC_14, "groupby_educ_count", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "By Education", "key_col_idx": 3,
               "value_col_idx": 0, "agg": "count",
               "key_header": "Education", "value_header": "Count"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Education"),
              "Add a new sheet 'By Education' tallying the number of "
              "respondents per Education level: header row "
              "'Education','Count' followed by one row per level."),
        Param({"new_sheet_name": "By Sex", "key_col_idx": 1,
               "value_col_idx": 0, "agg": "count",
               "key_header": "Sex", "value_header": "Count"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Sex"),
              "I'd like a sex breakdown for the demographic analysis section "
              "of the research write-up — add a new sheet 'By Sex' tallying "
              "respondents per Sex with header row 'Sex','Count' followed "
              "by one row per distinct Sex value for the manuscript table."),
    ]),

    # F-CALC-15 — bus schedule (Route / Origin / Dest / Depart / Arrive / Duration)
    FileTask(F_CALC_15, "filter_by_origin", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Downtown Departures", "filter_col_idx": 1,
               "filter_value": "Downtown"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Downtown Departures"),
              "I'd like a Downtown-only schedule for the kiosk posters at "
              "the transit hub — create a new sheet 'Downtown Departures' "
              "containing only the rows whose Origin is 'Downtown' (copy "
              "the header too) so the printers have a clean source."),
        Param({"new_sheet_name": "Westside Departures", "filter_col_idx": 1,
               "filter_value": "Westside"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Westside Departures"),
              "Create a new sheet 'Westside Departures' that holds the "
              "header row plus every trip whose Origin is 'Westside'."),
    ]),
    FileTask(F_CALC_15, "sort_by_depart", "sort_col", _gold_sort, params=[
        Param({"col_idx": 3, "reverse": False},
              _RULE_SHEET_DATA,
              "Sort the schedule by Depart time ascending; keep the header "
              "at row 1."),
        Param({"col_idx": 4, "reverse": False},
              _RULE_SHEET_DATA,
              "I'd like an arrival-ordered schedule for the dispatch board "
              "before tomorrow's first shift — sort the schedule by Arrive "
              "time ascending and keep the header at row 1 so the operators "
              "can read the run-order at a glance."),
    ]),

    # F-CALC-16 — tournament bracket (Match / Team A / Team B / Score)
    FileTask(F_CALC_16, "filter_by_match_round", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Round QF1-4", "filter_col_idx": 0,
               "filter_value": "QF1"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Round QF1-4"),
              "I'm preparing the bracket pull-out for the tournament "
              "programme and need a focused round view — create a new sheet "
              "'Round QF1-4' containing only the QF1 row (copy the header "
              "too) so I can build the layout from there."),
        Param({"new_sheet_name": "Round QF5-8", "filter_col_idx": 0,
               "filter_value": "QF8"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Round QF5-8"),
              "Create a new sheet 'Round QF5-8' containing only the QF8 "
              "row (copy the header too)."),
    ]),
    FileTask(F_CALC_16, "sort_by_team_a", "sort_col", _gold_sort, params=[
        Param({"col_idx": 1, "reverse": False},
              _RULE_SHEET_DATA,
              "Sort the bracket alphabetically by Team A; keep the header "
              "at row 1."),
        Param({"col_idx": 2, "reverse": False},
              _RULE_SHEET_DATA,
              "I'd like an alphabetised Team B listing for the bracket "
              "lookup card we're handing out at the door — sort the bracket "
              "alphabetically by Team B and keep the header at row 1 for "
              "the printer's reference."),
    ]),

    # ── F-CALC-17..F-CALC-30 ─────────────────────────────────────

    FileTask(F_CALC_17, "filter_by_status", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Tardy", "filter_col_idx": 3, "filter_value": "Tardy"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Tardy"),
              "I'm preparing a tardiness report for next week's parent-teacher conferences — create a new sheet 'Tardy' containing only the attendance rows whose Status is 'Tardy' (copy the header too) so I can share it with the counsellor."),
        Param({"new_sheet_name": "Absent", "filter_col_idx": 3, "filter_value": "Absent"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Absent"),
              "Create a new sheet 'Absent' that holds the header row plus every row whose Status is 'Absent'."),
    ]),
    FileTask(F_CALC_17, "sort_by_tardy_min", "sort_col", _gold_sort, params=[
        Param({"col_idx": 4, "reverse": True}, _RULE_SHEET_DATA,
              "Sort the rows by Tardy_Min descending (largest delays first); keep the header at row 1."),
        Param({"col_idx": 1, "reverse": False}, _RULE_SHEET_DATA,
              "I'd like an alphabetised attendance roster for the homeroom teacher's morning check — sort the rows alphabetically by Student name and keep the header at row 1 so the printed roll-call is easy to scan."),
    ]),

    FileTask(F_CALC_18, "color_low_stock", "conditional_format",
             _gold_cell_color_by_predicate, params=[
        Param({"rules_py": "isinstance(row[3], (int,float)) and isinstance(row[4], (int,float)) and row[3] < row[4]",
               "apply_kind": "fill", "apply_argb": "FFFFC0C0"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like to flag low-stock items before tomorrow's purchasing meeting — give every row whose OnHand is below the ReorderPoint a pale-red fill so the buyer can prioritise reorders for the week ahead (use Custom Color hex #FFC0C0 in the fill picker)."),
        Param({"rules_py": "isinstance(row[3], (int,float)) and row[3] == 0",
               "apply_kind": "fill", "apply_argb": "FFFF0000"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like to flag the out-of-stock SKUs ahead of the emergency restock call with the supplier — give every row whose OnHand is zero a bright-red fill so the buyer can prioritise emergency orders today (use Custom Color hex #FF0000 in the fill picker)."),
    ]),
    FileTask(F_CALC_18, "sort_by_onhand_asc", "sort_col", _gold_sort, params=[
        Param({"col_idx": 3, "reverse": False}, _RULE_SHEET_DATA,
              "Sort the inventory by OnHand ascending (lowest stock first); keep the header at row 1."),
        Param({"col_idx": 0, "reverse": False}, _RULE_SHEET_DATA,
              "Sort the inventory alphabetically by SKU; keep the header at row 1."),
    ]),

    FileTask(F_CALC_19, "groupby_region_totals", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "By Region", "key_col_idx": 1, "value_col_idx": 2,
               "agg": "sum", "key_header": "Region", "value_header": "TotalRevenue"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Region"),
              "I'm putting together a regional revenue summary for the quarterly board review — add a new sheet 'By Region' summing the Revenue per Region with header row 'Region','TotalRevenue' followed by one row per distinct region so the board sees clean totals."),
        Param({"new_sheet_name": "By Region Units", "key_col_idx": 1, "value_col_idx": 3,
               "agg": "sum", "key_header": "Region", "value_header": "TotalUnits"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Region Units"),
              "Add a new sheet 'By Region Units' summing the UnitsSold per Region: header row 'Region','TotalUnits' followed by one row per distinct region."),
    ]),
    FileTask(F_CALC_19, "sort_by_revenue_desc", "sort_col", _gold_sort, params=[
        Param({"col_idx": 2, "reverse": True}, _RULE_SHEET_DATA,
              "Sort the rows by Revenue descending (highest first); keep the header at row 1."),
        Param({"col_idx": 3, "reverse": True}, _RULE_SHEET_DATA,
              "I'd like to surface the top-selling rows for the weekly sales-velocity discussion — sort the rows by UnitsSold descending so the highest-volume entries come first, and keep the header at row 1."),
    ]),

    FileTask(F_CALC_20, "color_high_step_days", "conditional_format",
             _gold_cell_color_by_predicate, params=[
        Param({"rules_py": "isinstance(row[1], (int,float)) and row[1] > 10000",
               "apply_kind": "fill", "apply_argb": "FF90EE90"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Could you help me celebrate my high-activity days in this fitness log? Give every row whose Steps is above 10,000 a light-green fill so the good days stand out when I share progress with my training partner (use Custom Color hex #90EE90 in the fill picker)."),
        Param({"rules_py": "isinstance(row[1], (int,float)) and row[1] < 5000",
               "apply_kind": "fill", "apply_argb": "FFFFC0C0"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like to flag the low-activity days in this fitness log before reviewing with my trainer this Saturday — give every row whose Steps is below 5,000 a pale-red fill so the recovery days are clearly marked (use Custom Color hex #FFC0C0 in the fill picker)."),
    ]),
    FileTask(F_CALC_20, "sort_by_steps_desc", "sort_col", _gold_sort, params=[
        Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              "Sort the rows by Steps descending (most active days first); keep the header at row 1."),
        Param({"col_idx": 2, "reverse": True}, _RULE_SHEET_DATA,
              "Sort the rows by Calories descending; keep the header at row 1."),
    ]),

    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_21, "sort_by_revenue", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 2, "reverse": True}, _RULE_SHEET_DATA,
              # "I'm reviewing tour performance for next year's booking strategy — please sort the shows by Revenue descending so the highest-grossing dates come first, and keep the header at row 1 so the schedule reads cleanly."),
        # Param({"col_idx": 0, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the shows alphabetically by Show name; keep the header at row 1."),
    # ]),
    FileTask(F_CALC_21, "groupby_venue_totals", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "By Venue", "key_col_idx": 1, "value_col_idx": 2,
               "agg": "sum", "key_header": "Venue", "value_header": "TotalRevenue"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Venue"),
              "Add a new sheet 'By Venue' summing the Revenue per Venue: header row 'Venue','TotalRevenue' followed by one row per distinct venue."),
        Param({"new_sheet_name": "Show Count", "key_col_idx": 1, "value_col_idx": 2,
               "agg": "count", "key_header": "Venue", "value_header": "ShowCount"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Show Count"),
              "I'm benchmarking venue performance for next year's tour routing — add a new sheet 'Show Count' tallying how many shows each Venue hosted with header row 'Venue','ShowCount' followed by one row per distinct venue for the booking review."),
    ]),

    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_22, "sort_by_share_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              # "I'd like to see the market leaders at the top of this share table before the strategy offsite — sort the brands by Share descending (largest share first) and keep the header at row 1 so the ranking is immediately visible."),
        # Param({"col_idx": 0, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the brands alphabetically by Brand name; keep the header at row 1."),
    # ]),
    FileTask(F_CALC_22, "color_top_share", "conditional_format",
             _gold_cell_color_by_predicate, params=[
        Param({"rules_py": "isinstance(row[1], (int,float)) and row[1] >= 0.20",
               "apply_kind": "fill", "apply_argb": "FFFFD700"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Give every row whose Share is 20% (0.20) or higher a gold fill (market-leader highlight) (use Custom Color hex #FFD700 in the fill picker)."),
        Param({"rules_py": "isinstance(row[1], (int,float)) and row[1] < 0.05",
               "apply_kind": "fill", "apply_argb": "FFD3D3D3"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like to fade out the long-tail brands for the market-leader-focused executive summary — give every row whose Share is below 5% (0.05) a light-grey fill so the long-tail recedes visually for the report (use Custom Color hex #D3D3D3 in the fill picker)."),
    ]),

    FileTask(F_CALC_23, "filter_by_status", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Unpaid", "filter_col_idx": 3, "filter_value": "Unpaid"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Unpaid"),
              "I'm chasing receivables before month-end close and need a focused worklist — create a new sheet 'Unpaid' containing only the invoices whose Status is 'Unpaid' (copy the header too) so the AR clerk can call through it."),
        Param({"new_sheet_name": "Overdue", "filter_col_idx": 3, "filter_value": "Overdue"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Overdue"),
              "Create a new sheet 'Overdue' that holds the header row plus every invoice whose Status is 'Overdue'."),
    ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_23, "sort_by_amount_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 2, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort the invoices by Amount descending (largest first); keep the header at row 1."),
        # Param({"col_idx": 1, "reverse": False}, _RULE_SHEET_DATA,
              # "I'd like the invoices grouped alphabetically by customer for the AR clerk's morning call sheet — sort the invoices alphabetically by Customer and keep the header at row 1 so the callbacks are organised."),
    # ]),

    FileTask(F_CALC_24, "color_low_uptime", "conditional_format",
             _gold_cell_color_by_predicate, params=[
        Param({"rules_py": "isinstance(row[1], (int,float)) and row[1] < 99.0",
               "apply_kind": "fill", "apply_argb": "FFFFC0C0"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Please help me flag the SLA-breach servers ahead of the ops post-mortem — give every server row whose UptimePct is below 99.0 a pale-red fill so the at-risk hosts are immediately visible in the report (use Custom Color hex #FFC0C0 in the fill picker)."),
        Param({"rules_py": "isinstance(row[2], (int,float)) and row[2] >= 5",
               "apply_kind": "fill", "apply_argb": "FFFFFF00"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Give every server row whose IncidentCount is 5 or more a yellow fill (high-incident flag) (use Custom Color hex #FFFF00 in the fill picker)."),
    ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_24, "sort_by_uptime_asc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 1, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the servers by UptimePct ascending (worst uptime first); keep the header at row 1."),
        # Param({"col_idx": 2, "reverse": True}, _RULE_SHEET_DATA,
              # "I'd like the highest-incident servers at the top of the on-call rotation review — sort the servers by IncidentCount descending so the most incident-prone come first, and keep the header at row 1 for the meeting."),
    # ]),

    # Validation: dropped paraphrase-only second Param — gold builder
    # is parameterless so two Params produced identical gold artifacts (PD 3b clone).
    FileTask(F_CALC_25, "vlookup_late_fee", "apply_formula", _gold_vlookup_late_fee, params=[
        Param({},
              [{"type": "sheet_data", "sheet_idx0": "RNAttendance", "sheet_idx1": "ENAttendance"}],
              "I'm computing the late-fee assessments for this month's attendance review and need the formula in place — fill the Fee column on the Attendance sheet by looking up each LateMin in the LateFeeScale sheet using VLOOKUP with TRUE / approximate match."),
    ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_25, "sort_attendance_by_latemin", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort the Attendance sheet by LateMin descending (worst tardiness first); keep the header at row 1."),
        # Param({"col_idx": 0, "reverse": False}, _RULE_SHEET_DATA,
              # "I'm preparing the alphabetised attendance roster for the homeroom binder — sort the Attendance sheet alphabetically by Name and keep the header at row 1 so the teacher can quickly find each student."),
    # ]),

    # Validation: dropped paraphrase-only second Param — gold builder
    # is parameterless so two Params produced identical gold artifacts (PD 3b clone).
    FileTask(F_CALC_26, "summary_rollup", "multi_sheet_aggregate",
             _gold_quarterly_summary, params=[
        Param({}, _RULE_SHEET_NAME_AND_DATA_NAMED("Summary"),
              "I'm rolling up the three-quarter regional results for the year-end board pack — create a new 'Summary' sheet aggregating Q1+Q2+Q3 by Region with columns Region / TotalRevenue / TotalExpenses, one row per region (Northeast, Midwest, South, West)."),
    ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_26, "sort_q1_by_revenue", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort the Q1 sheet by Revenue descending; keep the header at row 1."),
        # Param({"col_idx": 2, "reverse": True}, _RULE_SHEET_DATA,
              # "I'd like the highest-expense Q1 entries at the top of the cost review — sort the Q1 sheet by Expenses descending and keep the header at row 1 so the cost-control discussion focuses on the biggest items."),
    # ]),

    # F-CALC-27..30 — REAL CSV-staged files

    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_27, "sort_by_value_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              # "I'm preparing the US GDP trend chart for an economics class lecture — please sort the rows by GDP descending so the largest values come first, and keep the header at row 1 so the dataset reads cleanly for plotting."),
        # Param({"col_idx": 0, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort the rows by observation_date descending (most recent first); keep the header at row 1."),
    # ]),
    # validation fix (vacuous-predicate audit): retuned thresholds inside the
    # us-gdp first-32 range ($243B–$400B, 1947-1954) so both bands match a
    # non-empty, non-degenerate subset. Was: >20000/<5000 and >=25000/<1000
    # — all 0/32 matches.
    FileTask(F_CALC_27, "color_high_low", "conditional_format",
             _gold_two_color_by_predicate, params=[
        Param({"pred_a_py": "isinstance(row[1], (int,float)) and row[1] > 350", "argb_a": "FF90EE90",
               "pred_b_py": "isinstance(row[1], (int,float)) and row[1] < 270",  "argb_b": "FFFFC0C0",
               "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Give every row whose GDP is above 350 a light-green fill and every row whose GDP is below 270 a pale-red fill (light-green = Custom Color #90EE90, pale-red = #FFC0C0 via Format → Cells → Background)."),
        Param({"pred_a_py": "isinstance(row[1], (int,float)) and row[1] >= 380", "argb_a": "FFFFD700",
               "pred_b_py": "isinstance(row[1], (int,float)) and row[1] < 260",   "argb_b": "FFD3D3D3",
               "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like to highlight the boom and trough quarters for the macro briefing — give every row whose GDP is 380 or higher a gold fill and every row below 260 a light-grey fill for the presentation slides (gold = Custom Color #FFD700, light-grey = #D3D3D3 via Format → Cells → Background)."),
    ]),

    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_28, "sort_by_population_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              # "I'd like to highlight the largest states for a demographic atlas figure — sort the states by Population2020 descending so the most populous appear first, and keep the header at row 1 so the table is ready for the publisher."),
        # Param({"col_idx": 0, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the states alphabetically by State name; keep the header at row 1."),
    # ]),
    # validation B8 fix (vacuous-predicate retune): the source slice is
    # us-population-states first 30 rows (alphabetical: Alabama-NewHampshire).
    # Population2020 range across that slice: min = DC 689,545; max = California
    # 39,538,223. P[1] previously used `>= 20M` (only CA + FL = 2 matches —
    # fragile) and `< 600K` (0 matches — VACUOUS, DC at 690K is the smallest).
    # Retuned to `>= 7M` (7 matches: CA/FL/GA/IL/MI/AZ/MA) and `< 1.5M`
    # (4 matches: AK 733K, DE 990K, DC 690K, HI 1.46M). P[0] `>= 10M` (5
    # matches) and `< 1M` (3 matches) already pass the ≥3-per-band guard.
    FileTask(F_CALC_28, "color_large_small", "conditional_format",
             _gold_two_color_by_predicate, params=[
        Param({"pred_a_py": "isinstance(row[1], (int,float)) and row[1] >= 10000000", "argb_a": "FF90EE90",
               "pred_b_py": "isinstance(row[1], (int,float)) and row[1] < 1000000",   "argb_b": "FFFFC0C0",
               "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Give every state whose Population2020 is 10 million or more a light-green fill and every state below 1 million a pale-red fill (light-green = Custom Color #90EE90, pale-red = #FFC0C0 via Format → Cells → Background)."),
        Param({"pred_a_py": "isinstance(row[1], (int,float)) and row[1] >= 7000000", "argb_a": "FFFFD700",
               "pred_b_py": "isinstance(row[1], (int,float)) and row[1] < 1500000",  "argb_b": "FFD3D3D3",
               "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like to spotlight the extremes for the population-distribution chapter of the demographic textbook — highlight Population2020 of 7 million or more in gold and below 1.5 million in light grey so the giants and smallest states pop visually (gold = Custom Color #FFD700, light grey = #D3D3D3 via Format → Cells → Background)."),
    ]),

    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_29, "sort_by_unrate_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              # "I'm preparing a labour-market briefing for the policy team and want recession months at the top — sort the rows by UNRATE descending (highest unemployment first) and keep the header at row 1 so the worst periods are immediately visible."),
        # Param({"col_idx": 1, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the rows by UNRATE ascending (lowest unemployment first); keep the header at row 1."),
    # ]),
    # validation B9+B10 fix (vacuous-predicate retune): the source slice
    # is us-unemployment first 36 rows (1948-01 → 1950-12). UNRATE range over
    # that window: min 3.4 (1948-01), max 7.9 (1949-10) — the 1949 recession
    # peak. P[0] previously used `> 8.0` (0 matches — VACUOUS, max is 7.9)
    # and `< 4.0` (~14 matches). P[1] used `>= 10.0` (0 matches — VACUOUS,
    # max is 7.9) and `< 3.5` (1 match — fragile; only 1948-01 at 3.4).
    # Retuned to (all bands ≥3 matches over the 36-month slice):
    #   P[0]: `> 5.5` (12 matches, recession-cluster Q3-1949..Q3-1950) /
    #         `< 4.0` (10 matches, expansion 1948-01..1948-12)
    #   P[1]: `>= 6.5` (6 matches, 1949-07..1949-11 peak band) /
    #         `< 3.8` (5 matches, 1948-01,05,06,07,10 expansion trough)
    FileTask(F_CALC_29, "color_recession_band", "conditional_format",
             _gold_two_color_by_predicate, params=[
        Param({"pred_a_py": "isinstance(row[1], (int,float)) and row[1] > 5.5",  "argb_a": "FFFFC0C0",
               "pred_b_py": "isinstance(row[1], (int,float)) and row[1] < 4.0",  "argb_b": "FF90EE90",
               "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Give every row whose UNRATE is above 5.5 a pale-red fill (high-unemployment) and every row below 4.0 a light-green fill (low-unemployment) (high band = Custom Color #FFC0C0, low band = #90EE90 via Format → Cells → Background)."),
        Param({"pred_a_py": "isinstance(row[1], (int,float)) and row[1] >= 6.5", "argb_a": "FFFF0000",
               "pred_b_py": "isinstance(row[1], (int,float)) and row[1] < 3.8",  "argb_b": "FF00FF00",
               "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'd like to mark the labour-market extremes for next week's monetary-policy seminar — highlight UNRATE of 6.5 or higher in red and below 3.8 in bright green so the recession peaks and expansion troughs are immediately visible (high band = Custom Color #FF0000, low band = #00FF00 via Format → Cells → Background)."),
    ]),

    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_30, "sort_by_gdp_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 3, "reverse": True}, _RULE_SHEET_DATA,
              # "I'm building a global GDP league table for the international affairs course — sort the countries by GDP descending so the largest economies come first, and keep the header at row 1 so the printed handout reads naturally."),
        # Param({"col_idx": 1, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the countries alphabetically by CountryName; keep the header at row 1."),
    # ]),
    # validation fix (vacuous-predicate audit) + B2 follow-up retune:
    # Pre-B2 the slice was WB aggregates (`>=5e12` matched 17/30 aggregate
    # rows; `<5e11` matched 2/30). Post-B2 the slice is the REAL top-30
    # countries by GDP — range $509B (SGP) to $25.6T (USA). Re-tuned to:
    #   P[0]: `>= 2.5e12` → 7 matches (US, CN, JP, DE, IN, GB, FR)
    #   P[1]: `< 7e11`     → 10 matches (Switzerland-Singapore bottom tier)
    # both non-degenerate, both non-vacuous on the post-B2 country slice.
    FileTask(F_CALC_30, "color_top_economies", "conditional_format",
             _gold_cell_color_by_predicate, params=[
        Param({"rules_py": "isinstance(row[3], (int,float)) and row[3] >= 2500000000000",
               "apply_kind": "fill", "apply_argb": "FFFFD700"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Give every country whose GDP is 2.5 trillion (2.5e12) or more a gold fill (top-tier economy) (use Custom Color hex #FFD700 in the fill picker)."),
        Param({"rules_py": "isinstance(row[3], (int,float)) and row[3] < 700000000000",
               "apply_kind": "fill", "apply_argb": "FFD3D3D3"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Give every country whose GDP is below 700 billion (7e11) a light-grey fill (smaller-economy callout) (use Custom Color hex #D3D3D3 in the fill picker)."),
    ]),

    # ── F-CALC-31..F-CALC-50 — real CSV-staged + authored xlsx
    # Skill mix tilted toward `apply_formula` + `text_manipulation` to match the
    # eval taxonomy weights (apply_formula=12, text_manipulation=4 vs current
    # over-share on sort_col / multi_sheet_aggregate).

    # F-CALC-31 — oil WTI daily (date / DCOILWTICO)
    FileTask(F_CALC_31, "derived_log_price", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "LogPrice", "src_col_idx": 1,
               "expr": "round(__import__('math').log(float(v)), 4) "
                       "if isinstance(v, (int, float)) and v > 0 else None",
               "number_format": "0.0000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm preparing the oil-price series for an econometrics homework set and need a log-transformed column — add a derived column 'LogPrice' = ln(DCOILWTICO) rounded to 4 decimals and format the new column as 0.0000 for the regression input."),
        Param({"new_header": "PriceUSD2", "src_col_idx": 1,
               "expr": "round(float(v), 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm preparing the oil-price series for the daily trading desk hand-off — add a derived column 'PriceUSD2' that copies DCOILWTICO rounded to 2 decimals and format the new column as $#,##0.00 so the rows match the standard desk template."),
    ]),
    # validation fix (vacuous-predicate audit): retuned both Params inside
    # the WTI 2021-05-11..2021-06-21 window ($61.95–$73.64). Was Param[0]
    # >100/<30 and Param[1] >80/<40 — entire template was vacuous (0/29 for
    # all four bands).
    FileTask(F_CALC_31, "color_oil_extremes", "conditional_format",
             _gold_two_color_by_predicate, params=[
        Param({"pred_a_py": "isinstance(row[1], (int,float)) and row[1] > 70", "argb_a": "FFFF6347",
               "pred_b_py": "isinstance(row[1], (int,float)) and row[1] < 64", "argb_b": "FFADD8E6",
               "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Highlight oil-price extremes: DCOILWTICO above 70 in tomato red and below 64 in light blue (high band = Custom Color #FF6347, low band = #ADD8E6 via Format → Cells → Background)."),
        Param({"pred_a_py": "isinstance(row[1], (int,float)) and row[1] > 72", "argb_a": "FFFFA500",
               "pred_b_py": "isinstance(row[1], (int,float)) and row[1] < 65", "argb_b": "FF87CEEB",
               "apply_kind": "fill"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Flag DCOILWTICO above 72 in orange and below 65 in sky-blue (high band = Custom Color #FFA500, low band = #87CEEB via Format → Cells → Background)."),
    ]),

    # F-CALC-32 — fed-funds (date / FEDFUNDS)
    FileTask(F_CALC_32, "derived_basis_points", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "BasisPoints", "src_col_idx": 1,
               "expr": "int(round(float(v) * 100)) if isinstance(v, (int,float)) else None",
               "number_format": "#,##0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the FEDFUNDS series re-expressed in basis points for the rates desk note — add a derived column 'BasisPoints' = FEDFUNDS × 100 rounded to integer and format the new column as #,##0 for the daily briefing."),
        Param({"new_header": "RateDecimal", "src_col_idx": 1,
               "expr": "round(float(v) / 100.0, 4) if isinstance(v, (int,float)) else None",
               "number_format": "0.00%"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the fed-funds rate re-expressed as a decimal fraction for the macro-spreadsheet model — add a derived column 'RateDecimal' equal to FEDFUNDS / 100 and format the new column as 0.00% (two-decimal percent) for the calculations."),
    ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_32, "sort_by_rate_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort by FEDFUNDS descending (highest-rate months first); keep the header at row 1."),
        # Param({"col_idx": 0, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the rows chronologically by observation_date ascending; keep the header at row 1."),
    # ]),

    # F-CALC-33 — housing starts (date / HOUST)
    FileTask(F_CALC_33, "derived_thousand_units", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "MillionUnits", "src_col_idx": 1,
               "expr": "round(float(v) / 1000.0, 3) if isinstance(v, (int,float)) else None",
               "number_format": "0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm preparing a housing-starts chart for the macro report and need the series rescaled to millions — add a derived column 'MillionUnits' = HOUST / 1000 rounded to 3 decimals and format as 0.000 for cleaner labels."),
        Param({"new_header": "DoubledUnits", "src_col_idx": 1,
               "expr": "float(v) * 2 if isinstance(v, (int,float)) else None",
               "number_format": "#,##0.0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a what-if doubled-output column for the housing capacity scenario in the construction-industry report — add a derived column 'DoubledUnits' = HOUST × 2 and format as #,##0.0 for the scenario tab."),
    ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_33, "color_low_housing", "conditional_format",
             # _gold_cell_color_by_predicate, params=[
        # Param({"rules_py": "isinstance(row[1], (int,float)) and row[1] < 1000",
               # "apply_kind": "fill", "apply_argb": "FFFFC0C0"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every row whose HOUST is below 1000 a pale-red fill (slow-construction flag)."),
        # Param({"rules_py": "isinstance(row[1], (int,float)) and row[1] > 2000",
               # "apply_kind": "fill", "apply_argb": "FF90EE90"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every row whose HOUST is above 2000 a light-green fill (boom-period highlight)."),
    # ]),

    # F-CALC-34 — CPI (date / CPIAUCSL)
    FileTask(F_CALC_34, "derived_cpi_index", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "Index100", "src_col_idx": 1,
               "expr": "round(float(v), 1) if isinstance(v, (int,float)) else None",
               "number_format": "0.0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm cleaning up the CPI series for an inflation-tracking dashboard — add a derived column 'Index100' that copies CPIAUCSL rounded to 1 decimal and format as 0.0 for consistent display across charts."),
        Param({"new_header": "HalfIndex", "src_col_idx": 1,
               "expr": "round(float(v) / 2.0, 2) if isinstance(v, (int,float)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a half-CPI normalised column for the textbook chapter exercise on price-index transformations — add a derived column 'HalfIndex' = CPIAUCSL / 2 rounded to 2 decimals and format as 0.00 for the student handout."),
    ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_34, "sort_by_cpi_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort by CPIAUCSL descending (highest CPI rows first); keep the header at row 1."),
        # Param({"col_idx": 0, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort by observation_date descending (most recent first); keep the header at row 1."),
    # ]),

    # F-CALC-35 — mortgage 30yr (date / MORTGAGE30US)
    FileTask(F_CALC_35, "derived_monthly_rate", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "MonthlyRate", "src_col_idx": 1,
               "expr": "round(float(v) / 12.0, 4) if isinstance(v, (int,float)) else None",
               "number_format": "0.0000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Could you help me convert the 30-year mortgage series into monthly equivalents for the home-buying calculator? Add a derived column 'MonthlyRate' = MORTGAGE30US / 12 rounded to 4 decimals and format as 0.0000."),
        Param({"new_header": "RateBand", "src_col_idx": 1,
               "expr": "'High' if isinstance(v,(int,float)) and v >= 7 else "
                       "('Low' if isinstance(v,(int,float)) and v < 4 else 'Mid')",
               "number_format": None},
              _RULE_SHEET_DATA,
              "I'd like a rate-band classifier on the mortgage series for the home-affordability dashboard — add a derived column 'RateBand' that classifies MORTGAGE30US as 'High' if ≥ 7, 'Low' if < 4, and 'Mid' otherwise."),
    ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_35, "color_low_mortgage", "conditional_format",
             # _gold_cell_color_by_predicate, params=[
        # Param({"rules_py": "isinstance(row[1], (int,float)) and row[1] < 3.5",
               # "apply_kind": "fill", "apply_argb": "FF98FB98"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every week whose MORTGAGE30US is below 3.5% a pale-green fill (low-rate window)."),
        # Param({"rules_py": "isinstance(row[1], (int,float)) and row[1] > 7.5",
               # "apply_kind": "fill", "apply_argb": "FFFFB6C1"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every week whose MORTGAGE30US is above 7.5% a pink fill (high-rate flag)."),
    # ]),

    # F-CALC-36 — state median income (State / MedianHouseholdIncome_USD / StateFIPS)
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_36, "sort_by_income_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              # "I'm building a state-by-state income comparison for a policy brief and want the wealthiest states up top — sort the states by MedianHouseholdIncome_USD descending (richest first) and keep the header at row 1."),
        # Param({"col_idx": 0, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the states alphabetically by State name; keep the header at row 1."),
    # ]),
    FileTask(F_CALC_36, "derived_income_thousands", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "IncomeK", "src_col_idx": 1,
               "expr": "round(float(v) / 1000.0, 1) if isinstance(v, (int,float)) else None",
               "number_format": "0.0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'IncomeK' = MedianHouseholdIncome_USD / 1000 rounded to 1 decimal; format as 0.0."),
        Param({"new_header": "IncomeBand", "src_col_idx": 1,
               "expr": "'High' if isinstance(v,(int,float)) and v >= 80000 else "
                       "('Low' if isinstance(v,(int,float)) and v < 55000 else 'Mid')",
               "number_format": None},
              _RULE_SHEET_DATA,
              "I'd like an income-band classifier on the state-income table for the regional inequality study — add a derived column 'IncomeBand' classifying MedianHouseholdIncome_USD as 'High' if ≥ 80000, 'Low' if < 55000, and 'Mid' otherwise for the comparison."),
    ]),

    # F-CALC-37 — world population (CountryCode / CountryName / Year / Population)
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_37, "sort_by_population_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 3, "reverse": True}, _RULE_SHEET_DATA,
              # "I'd like a global ranking by population for the international affairs course — sort by Population descending so the most populous countries appear first, and keep the header at row 1 so the printed handout reads naturally."),
        # Param({"col_idx": 1, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort alphabetically by CountryName; keep the header at row 1."),
    # ]),
    FileTask(F_CALC_37, "derived_pop_millions", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "PopMillion", "src_col_idx": 3,
               "expr": "round(float(v) / 1_000_000.0, 2) if isinstance(v, (int,float)) else None",
               "number_format": "#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'PopMillion' = Population / 1,000,000 rounded to 2 decimals; format as #,##0.00."),
        Param({"new_header": "PopBillion", "src_col_idx": 3,
               "expr": "round(float(v) / 1_000_000_000.0, 4) if isinstance(v, (int,float)) else None",
               "number_format": "0.0000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the world-population values expressed in billions for the global-trends infographic — add a derived column 'PopBillion' = Population / 1,000,000,000 rounded to 4 decimals and format as 0.0000 for the design team."),
    ]),

    # F-CALC-38 — bank transactions (Date / Description / Amount / Balance)
    # task_id renamed from 'fill_running_balance' to 'derived_abs_or_direction'
    # to match the actual Param instructions (AbsAmount + Direction derived
    # columns) rather than the misleading original name.
    FileTask(F_CALC_38, "derived_abs_or_direction", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "AbsAmount", "src_col_idx": 2,
               "expr": "abs(float(v)) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm preparing a transaction-magnitude column for the bank-statement reconciliation — please add a derived column 'AbsAmount' equal to |Amount| and format it as $#,##0.00 so the analysts can sort by size."),
        Param({"new_header": "Direction", "src_col_idx": 2,
               "expr": "'Credit' if isinstance(v,(int,float)) and v >= 0 else 'Debit'",
               "number_format": None},
              _RULE_SHEET_DATA,
              "I'd like a Credit/Debit direction tag on this bank statement for the bookkeeping import to QuickBooks — add a derived column 'Direction' that says 'Credit' when Amount ≥ 0 and 'Debit' otherwise so the import job can route correctly."),
    ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_38, "color_overdraft", "conditional_format",
             # _gold_cell_color_by_predicate, params=[
        # Param({"rules_py": "isinstance(row[2], (int,float)) and row[2] < -100",
               # "apply_kind": "fill", "apply_argb": "FFFFC0C0"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every transaction whose Amount is below -100 (large outflow) a pale-red fill."),
        # Param({"rules_py": "isinstance(row[2], (int,float)) and row[2] >= 1000",
               # "apply_kind": "fill", "apply_argb": "FF90EE90"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every transaction whose Amount is 1000 or higher (large inflow) a light-green fill."),
    # ]),

    # F-CALC-39 — event schedule (EventID / Title / Date / Status)
    FileTask(F_CALC_39, "filter_confirmed_to_sheet", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Confirmed", "filter_col_idx": 3, "filter_value": "Confirmed"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Confirmed"),
              "I'd like to share only the confirmed events with the venue coordinator — create a new sheet 'Confirmed' containing only the events whose Status is 'Confirmed' (copy the header too) so the schedule is clean before the planning call."),
        Param({"new_sheet_name": "Cancelled", "filter_col_idx": 3, "filter_value": "Cancelled"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Cancelled"),
              "Create a new sheet 'Cancelled' that holds the header row plus every event whose Status is 'Cancelled'."),
    ]),
    FileTask(F_CALC_39, "groupby_status_count", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "Status Counts", "key_col_idx": 3, "value_col_idx": 0,
               "agg": "count", "key_header": "Status", "value_header": "EventCount"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Status Counts"),
              "Add a new sheet 'Status Counts' tallying events per Status: header row 'Status','EventCount' followed by one row per distinct Status value."),
        Param({"new_sheet_name": "Status Totals", "key_col_idx": 3, "value_col_idx": 0,
               "agg": "count", "key_header": "Status", "value_header": "Total"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Status Totals"),
              "I'd like a status-totals overview for the event-planning dashboard — build a sheet 'Status Totals' with header 'Status','Total' that counts events per Status so the planner sees the pipeline distribution at a glance for the upcoming season."),
    ]),

    # F-CALC-40 — product catalog (SKU / Product / Price / Category)
    FileTask(F_CALC_40, "derived_price_with_tax", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "PriceWithTax", "src_col_idx": 2,
               "expr": "round(float(v) * 1.0875, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm updating the product catalog with tax-inclusive prices ahead of the holiday catalog drop — add a derived column 'PriceWithTax' = Price × 1.0875 (8.75% sales tax) rounded to 2 decimals and format as $#,##0.00 for the printed sleeves."),
        Param({"new_header": "DiscountedPrice", "src_col_idx": 2,
               "expr": "round(float(v) * 0.85, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm preparing the holiday-sale catalog and need discounted prices computed — add a derived column 'DiscountedPrice' = Price × 0.85 (15% off) rounded to 2 decimals and format as $#,##0.00 so the catalog can go to print tomorrow morning."),
    ]),
    FileTask(F_CALC_40, "filter_by_category", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Tools", "filter_col_idx": 3, "filter_value": "Tools"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Tools"),
              "Create a new sheet 'Tools' with the header plus every product whose Category is 'Tools'."),
        Param({"new_sheet_name": "Kitchen", "filter_col_idx": 3, "filter_value": "Kitchen"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Kitchen"),
              "I'd like a kitchen-category extract for the cooking-section flyer we're sending out before Black Friday — create a new sheet 'Kitchen' with the header row plus every product whose Category is 'Kitchen' for the design team."),
    ]),

    # F-CALC-41 — clinic visits (PatientID / Age / VisitType / Department / Cost)
    FileTask(F_CALC_41, "groupby_dept_cost", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "By Department", "key_col_idx": 3, "value_col_idx": 4,
               "agg": "sum", "key_header": "Department", "value_header": "TotalCost"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Department"),
              "I'm preparing the cost-by-department rollup for next month's clinic operations review — add a new sheet 'By Department' summing Cost per Department with header row 'Department','TotalCost' followed by one row per distinct department for the admin team."),
        Param({"new_sheet_name": "Avg By Dept", "key_col_idx": 3, "value_col_idx": 4,
               "agg": "avg", "key_header": "Department", "value_header": "AvgCost"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Avg By Dept"),
              "I'd like to compare average cost per department for the clinic-operations efficiency review — add a new sheet 'Avg By Dept' with header 'Department','AvgCost' followed by one row per distinct department for the practice manager."),
    ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_41, "color_emergency_visits", "conditional_format",
             # _gold_cell_color_by_predicate, params=[
        # Param({"rules_py": "row[2] == 'Emergency'",
               # "apply_kind": "fill", "apply_argb": "FFFF6347"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every visit whose VisitType is 'Emergency' a tomato-red fill."),
        # Param({"rules_py": "row[2] == 'Routine'",
               # "apply_kind": "fill", "apply_argb": "FFADD8E6"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every visit whose VisitType is 'Routine' a light-blue fill."),
    # ]),

    # F-CALC-42 — survey (RespondentID / Region / Rating / Complete)
    FileTask(F_CALC_42, "groupby_region_avg_rating", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "Region Avg", "key_col_idx": 1, "value_col_idx": 2,
               "agg": "avg", "key_header": "Region", "value_header": "AvgRating"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Region Avg"),
              "I'd like a regional satisfaction summary for the customer-experience deck — add a new sheet 'Region Avg' with header 'Region','AvgRating' showing the average Rating per Region so the CX team has a clean comparison."),
        Param({"new_sheet_name": "Region Sum", "key_col_idx": 1, "value_col_idx": 2,
               "agg": "sum", "key_header": "Region", "value_header": "RatingSum"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Region Sum"),
              "I'd like a regional rating-sum view for the customer-feedback report's appendix — add a new sheet 'Region Sum' with header 'Region','RatingSum' summing Rating per Region so the CX team has a clean aggregate for the slide."),
    ]),
    FileTask(F_CALC_42, "filter_complete_yes", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Completed", "filter_col_idx": 3, "filter_value": "Yes"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Completed"),
              "Create a new sheet 'Completed' containing only the survey rows whose Complete is 'Yes' (copy the header too)."),
        Param({"new_sheet_name": "Incomplete", "filter_col_idx": 3, "filter_value": "No"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Incomplete"),
              "Create a new sheet 'Incomplete' that holds the header row plus every survey row whose Complete is 'No'."),
    ]),

    # F-CALC-43 — temperature log (Timestamp / Station / TempC)
    FileTask(F_CALC_43, "derived_tempf", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "TempF", "src_col_idx": 2,
               "expr": "round(float(v) * 9/5 + 32, 2) if isinstance(v, (int,float)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm preparing the weather log for sharing with our US partner station — please add a derived column 'TempF' converting TempC to Fahrenheit (TempC × 9/5 + 32) rounded to 2 decimals and format as 0.00 for their reporting conventions."),
        Param({"new_header": "TempK", "src_col_idx": 2,
               "expr": "round(float(v) + 273.15, 2) if isinstance(v, (int,float)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the temperature log in Kelvin for the physics-class lab handout — add a derived column 'TempK' converting TempC to Kelvin (TempC + 273.15) rounded to 2 decimals and format as 0.00 so the students see the absolute scale."),
    ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_43, "color_temp_extremes", "conditional_format",
             # _gold_two_color_by_predicate, params=[
        # Param({"pred_a_py": "isinstance(row[2], (int,float)) and row[2] >= 20", "argb_a": "FFFF8C00",
               # "pred_b_py": "isinstance(row[2], (int,float)) and row[2] < 9",   "argb_b": "FFB0E0E6",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every row whose TempC is 20 or higher a dark-orange fill and every row below 9 a powder-blue fill."),
        # Param({"pred_a_py": "isinstance(row[2], (int,float)) and row[2] >= 18", "argb_a": "FFFFA500",
               # "pred_b_py": "isinstance(row[2], (int,float)) and row[2] < 10",  "argb_b": "FFADD8E6",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Highlight TempC ≥ 18 in orange and TempC < 10 in light-blue."),
    # ]),

    # F-CALC-44 — commute log (Date / Mode / Miles / Minutes)
    FileTask(F_CALC_44, "derived_speed", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "Mph", "src_col_idx": 2,
               "expr": "round(float(v) / (float(ws.cell(r, 4).value) / 60.0), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) "
                       "and ws.cell(r,4).value > 0 else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm analysing my commute speeds over the past month for a transit blog post — add a derived column 'Mph' = Miles / (Minutes / 60) rounded to 2 decimals and format as 0.00 so the trip-by-trip pace is easy to scan."),
        Param({"new_header": "MinPerMile", "src_col_idx": 3,
               "expr": "round(float(v) / float(ws.cell(r, 3).value), 3) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) "
                       "and ws.cell(r,3).value > 0 else None",
               "number_format": "0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'MinPerMile' = Minutes / Miles rounded to 3 decimals; format as 0.000."),
    ]),
    FileTask(F_CALC_44, "groupby_mode_totals", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "By Mode", "key_col_idx": 1, "value_col_idx": 2,
               "agg": "sum", "key_header": "Mode", "value_header": "TotalMiles"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Mode"),
              "Add a new sheet 'By Mode' summing Miles per Mode: header row 'Mode','TotalMiles' followed by one row per distinct mode."),
        Param({"new_sheet_name": "Mode Avg", "key_col_idx": 1, "value_col_idx": 3,
               "agg": "avg", "key_header": "Mode", "value_header": "AvgMinutes"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Mode Avg"),
              "I'm comparing average commute times across transit modes for a personal-finance blog post — add a new sheet 'Mode Avg' with header 'Mode','AvgMinutes' followed by one row per distinct mode for the article's comparison table."),
    ]),

    # F-CALC-45 — subscription data (SubID / StartDate / Plan / Status)
    FileTask(F_CALC_45, "filter_active_subs", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "Active", "filter_col_idx": 3, "filter_value": "Active"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Active"),
              "I'm preparing an MRR snapshot before sharing with the investor and need a focused list of paying subscribers — create a new sheet 'Active' containing only the subscriptions whose Status is 'Active' (copy the header too) for the data room."),
        Param({"new_sheet_name": "Cancelled", "filter_col_idx": 3, "filter_value": "Cancelled"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Cancelled"),
              "Create a new sheet 'Cancelled' containing only the subscriptions whose Status is 'Cancelled' (copy the header too)."),
    ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_45, "color_status_bands", "conditional_format",
             # _gold_two_color_by_predicate, params=[
        # Param({"pred_a_py": "row[3] == 'Active'",    "argb_a": "FF90EE90",
               # "pred_b_py": "row[3] == 'Cancelled'", "argb_b": "FFFFC0C0",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every Active subscription a light-green fill and every Cancelled subscription a pale-red fill."),
        # Param({"pred_a_py": "row[2] == 'Premium'", "argb_a": "FFFFD700",
               # "pred_b_py": "row[2] == 'Basic'",   "argb_b": "FFD3D3D3",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "I'd like to visually segment the subscription tiers for the customer-success team's playbook — highlight Premium rows in gold and Basic rows in light grey so the tier-specific outreach is immediately clear."),
    # ]),

    # F-CALC-46 — recipe ingredients (Recipe / Ingredient / Qty / UnitCost)
    FileTask(F_CALC_46, "derived_line_cost", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "LineCost", "src_col_idx": 2,
               "expr": "round(float(v) * float(ws.cell(r, 4).value), 4) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) "
                       "else None",
               "number_format": "0.0000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like to compute line-level ingredient costs for the menu engineering spreadsheet — add a derived column 'LineCost' = Qty × UnitCost rounded to 4 decimals and format as 0.0000 so the chef can compare recipes."),
        Param({"new_header": "UnitCostUSD", "src_col_idx": 3,
               "expr": "round(float(v), 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the unit costs cleaned up for the menu-engineering spreadsheet I'm sharing with the chef — add a derived column 'UnitCostUSD' that copies UnitCost rounded to 2 decimals and format as $#,##0.00 for consistency across all recipes."),
    ]),
    FileTask(F_CALC_46, "groupby_recipe_qty", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "By Recipe", "key_col_idx": 0, "value_col_idx": 2,
               "agg": "sum", "key_header": "Recipe", "value_header": "TotalQty"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Recipe"),
              "Add a new sheet 'By Recipe' summing Qty per Recipe: header row 'Recipe','TotalQty' followed by one row per distinct recipe."),
        Param({"new_sheet_name": "Recipe Lines", "key_col_idx": 0, "value_col_idx": 2,
               "agg": "count", "key_header": "Recipe", "value_header": "LineCount"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Recipe Lines"),
              "Add a new sheet 'Recipe Lines' tallying ingredient lines per Recipe: header 'Recipe','LineCount' then one row per distinct recipe."),
    ]),

    # F-CALC-47 — student grades multi-subject (StudentID / Name / Math / Science / English)
    FileTask(F_CALC_47, "derived_average", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "Average", "src_col_idx": 2,
               "expr": "round((float(v) + float(ws.cell(r,4).value) + float(ws.cell(r,5).value)) / 3.0, 2) "
                       "if all(isinstance(ws.cell(r,c).value,(int,float)) for c in (3,4,5)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm finalising grades for the term-end report cards and need a composite per student — add a derived column 'Average' = (Math + Science + English) / 3 rounded to 2 decimals and format as 0.00 for the report-card mail merge."),
        Param({"new_header": "TotalScore", "src_col_idx": 2,
               "expr": "(float(v) + float(ws.cell(r,4).value) + float(ws.cell(r,5).value)) "
                       "if all(isinstance(ws.cell(r,c).value,(int,float)) for c in (3,4,5)) else None",
               "number_format": "0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a combined-score column for the honor-roll cutoff decision next week — add a derived column 'TotalScore' = Math + Science + English and format as 0 so the principal can see each student's total at a glance."),
    ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_47, "color_low_scores", "conditional_format",
             # _gold_cell_color_by_predicate, params=[
        # Param({"rules_py": "isinstance(row[2], (int,float)) and row[2] < 60",
               # "apply_kind": "fill", "apply_argb": "FFFFC0C0"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every student whose Math score is below 60 a pale-red fill (failing-Math flag)."),
        # Param({"rules_py": "isinstance(row[4], (int,float)) and row[4] < 65",
               # "apply_kind": "fill", "apply_argb": "FFFFE4B5"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every student whose English score is below 65 a moccasin-orange fill (English-attention flag)."),
    # ]),

    # F-CALC-48 — library loans (LoanID / Book / Patron / DueDate / Returned)
    FileTask(F_CALC_48, "filter_unreturned", "multi_sheet_aggregate",
             _gold_sheet2_filter, params=[
        Param({"new_sheet_name": "NotReturned", "filter_col_idx": 4, "filter_value": "No"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("NotReturned"),
              "I'm running an outstanding-loans recall round before the holiday closure — create a new sheet 'NotReturned' containing only the loans whose Returned is 'No' (copy the header too) so the librarian can email patrons."),
        Param({"new_sheet_name": "Returned", "filter_col_idx": 4, "filter_value": "Yes"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Returned"),
              "Create a new sheet 'Returned' containing only the loans whose Returned is 'Yes' (copy the header too)."),
    ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_48, "color_returned_state", "conditional_format",
             # _gold_two_color_by_predicate, params=[
        # Param({"pred_a_py": "row[4] == 'Yes'", "argb_a": "FF98FB98",
               # "pred_b_py": "row[4] == 'No'",  "argb_b": "FFFFB6C1",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every Returned='Yes' row a pale-green fill and every Returned='No' row a pink fill."),
        # Param({"pred_a_py": "row[4] == 'Yes'", "argb_a": "FFADD8E6",
               # "pred_b_py": "row[4] == 'No'",  "argb_b": "FFFFFFE0",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "I'd like a colourblind-friendly returned-state palette for the library audit printout — give every Returned='Yes' row a light-blue fill and every Returned='No' row a light-yellow fill so the audit binder is legible to all staff."),
    # ]),

    # F-CALC-49 — fleet vehicles (Plate / Make / Year / Mileage / Status)
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_49, "sort_by_mileage_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 3, "reverse": True}, _RULE_SHEET_DATA,
              # "I'd like to identify the highest-mileage vehicles before next month's fleet maintenance review — sort the fleet by Mileage descending so the most-driven vehicles come first, and keep the header at row 1 for the workshop's worklist."),
        # Param({"col_idx": 2, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the fleet by Year ascending (oldest first); keep the header at row 1."),
    # ]),
    FileTask(F_CALC_49, "groupby_make_count", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "By Make", "key_col_idx": 1, "value_col_idx": 3,
               "agg": "count", "key_header": "Make", "value_header": "VehicleCount"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Make"),
              "Add a new sheet 'By Make' tallying vehicles per Make: header row 'Make','VehicleCount' followed by one row per distinct make."),
        Param({"new_sheet_name": "Mileage By Make", "key_col_idx": 1, "value_col_idx": 3,
               "agg": "sum", "key_header": "Make", "value_header": "TotalMileage"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Mileage By Make"),
              "I'd like to roll up total fleet mileage by manufacturer for the procurement renewal discussion — add a new sheet 'Mileage By Make' summing Mileage per Make with header 'Make','TotalMileage' followed by one row per distinct make for the contract negotiation."),
    ]),

    # F-CALC-50 — employees payroll (EmpID / Name / Dept / Salary / Bonus)
    FileTask(F_CALC_50, "derived_total_comp", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "TotalComp", "src_col_idx": 3,
               "expr": "(float(v) + float(ws.cell(r, 5).value)) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,5).value,(int,float)) else None",
               "number_format": "$#,##0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm preparing the year-end compensation summary for HR before the budgeting cycle — add a derived column 'TotalComp' = Salary + Bonus and format as $#,##0 so total comp is visible at a glance for each employee."),
        Param({"new_header": "BonusPct", "src_col_idx": 4,
               "expr": "round(float(v) / float(ws.cell(r, 4).value), 4) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) "
                       "and ws.cell(r,4).value > 0 else None",
               "number_format": "0.00%"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'BonusPct' = Bonus / Salary; format as 0.00% (two-decimal percent)."),
    ]),
    FileTask(F_CALC_50, "groupby_dept_salary", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "Dept Total", "key_col_idx": 2, "value_col_idx": 3,
               "agg": "sum", "key_header": "Dept", "value_header": "TotalSalary"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Dept Total"),
              "Add a new sheet 'Dept Total' summing Salary per Dept: header row 'Dept','TotalSalary' followed by one row per distinct department."),
        Param({"new_sheet_name": "Dept Avg", "key_col_idx": 2, "value_col_idx": 3,
               "agg": "avg", "key_header": "Dept", "value_header": "AvgSalary"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("Dept Avg"),
              "I'd like to benchmark average department salaries for the compensation review with the HR director — add a new sheet 'Dept Avg' with header 'Dept','AvgSalary' followed by one row per distinct department for the discussion."),
    ]),

    # ── F-CALC-51..F-CALC-90 — final 80 FileTasks ─────────────────
    # Distribution leans apply_formula + text_manipulation + multi_sheet_aggregate
    # so a finite `TARGET` in catalog.py still keeps every template active.

    # F-CALC-51 — weather log (Date / HighC / LowC / Precip_mm)
    FileTask(F_CALC_51, "derived_temp_range", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "RangeC", "src_col_idx": 1,
               "expr": "round(float(v) - float(ws.cell(r, 3).value), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm preparing a daily temperature-swing column for the climate study handout — add a derived column 'RangeC' = HighC − LowC rounded to 2 decimals and format as 0.00 so each day's swing is easy to read in the printed report."),
        Param({"new_header": "AvgC", "src_col_idx": 1,
               "expr": "round((float(v) + float(ws.cell(r, 3).value)) / 2.0, 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'AvgC' = (HighC + LowC) / 2 rounded to 2 decimals; format as 0.00."),
    ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_51, "color_rainy_days", "conditional_format",
             # _gold_cell_color_by_predicate, params=[
        # Param({"rules_py": "isinstance(row[3], (int,float)) and row[3] >= 5",
               # "apply_kind": "fill", "apply_argb": "FFB0C4DE"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every day whose Precip_mm is 5 or higher a steel-blue fill (rainy-day flag)."),
        # Param({"rules_py": "isinstance(row[1], (int,float)) and row[1] >= 22",
               # "apply_kind": "fill", "apply_argb": "FFFFA07A"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "I'd like to flag the warm days for the gardener's planting calendar review — give every day whose HighC is 22 or higher a salmon fill so the warm-weather windows are immediately visible on the printed sheet."),
    # ]),

    # F-CALC-52 — restaurant menu (Item / Category / Price / SoldCount)
    FileTask(F_CALC_52, "derived_revenue", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "Revenue", "src_col_idx": 2,
               "expr": "round(float(v) * float(ws.cell(r, 4).value), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like to see per-item revenue to identify the menu best-sellers ahead of the rebranding meeting — add a derived column 'Revenue' = Price × SoldCount rounded to 2 decimals and format as $#,##0.00 so the chef can plan accordingly."),
        Param({"new_header": "PriceCents", "src_col_idx": 2,
               "expr": "int(round(float(v) * 100)) if isinstance(v, (int,float)) else None",
               "number_format": "#,##0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the menu prices expressed in cents for the POS-system data import — add a derived column 'PriceCents' = Price × 100 rounded to integer and format as #,##0 so the integer payload matches the POS schema."),
    ]),
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_52, "filter_by_category", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "Mains", "filter_col_idx": 1, "filter_value": "Main"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Mains"),
              # "Create a new sheet 'Mains' containing only the items whose Category is 'Main' (copy the header too)."),
        # Param({"new_sheet_name": "Desserts", "filter_col_idx": 1, "filter_value": "Dessert"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Desserts"),
              # "Create a new sheet 'Desserts' that holds the header row plus every item whose Category is 'Dessert'."),
    # ]),

    # F-CALC-53 — donations (Donor / Date / Amount / Campaign)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_53, "groupby_campaign_totals", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Campaign", "key_col_idx": 3, "value_col_idx": 2,
               # "agg": "sum", "key_header": "Campaign", "value_header": "TotalAmount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Campaign"),
              # "I'm putting together the year-end fundraising recap for the board — add a new sheet 'By Campaign' summing Amount per Campaign with header 'Campaign','TotalAmount' followed by one row per distinct campaign so the development director can compare."),
        # Param({"new_sheet_name": "Avg Campaign", "key_col_idx": 3, "value_col_idx": 2,
               # "agg": "avg", "key_header": "Campaign", "value_header": "AvgGift"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Avg Campaign"),
              # "Add a new sheet 'Avg Campaign' showing average Amount per Campaign: header 'Campaign','AvgGift' then one row per distinct campaign."),
    # ]),
    FileTask(F_CALC_53, "derived_amount_with_match", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "MatchedAmount", "src_col_idx": 2,
               "expr": "round(float(v) * 2.0, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'MatchedAmount' = Amount × 2 (corporate match) rounded to 2 decimals; format as $#,##0.00."),
        Param({"new_header": "AmountTier", "src_col_idx": 2,
               "expr": "'Major' if isinstance(v,(int,float)) and v >= 200 else "
                       "('Sustaining' if isinstance(v,(int,float)) and v >= 75 else 'Standard')",
               "number_format": None},
              _RULE_SHEET_DATA,
              "I'd like a donor-tier classifier to streamline the year-end stewardship outreach campaign — add a derived column 'AmountTier' classifying gifts as 'Major' if Amount ≥ 200, 'Sustaining' if ≥ 75, and 'Standard' otherwise for the development team."),
    ]),

    # F-CALC-54 — books (ISBN / Title / Author / Year / Pages)
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_54, "sort_by_year_asc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 3, "reverse": False}, _RULE_SHEET_DATA,
              # "I'm preparing the library catalogue for the rare-books exhibition opening next week — please sort the books by Year ascending so the oldest volumes appear first, and keep the header at row 1 for the printed display labels."),
        # Param({"col_idx": 2, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the books alphabetically by Author; keep the header at row 1."),
    # ]),
    FileTask(F_CALC_54, "derived_decade", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "Decade", "src_col_idx": 3,
               "expr": "int(int(v) // 10 * 10) if isinstance(v, (int,float)) else None",
               "number_format": "0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'Decade' = floor(Year / 10) × 10; format as 0."),
        Param({"new_header": "PagesK", "src_col_idx": 4,
               "expr": "round(float(v) / 1000.0, 3) if isinstance(v, (int,float)) else None",
               "number_format": "0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the book lengths expressed in thousands of pages for the long-reads list compared on the library website — add a derived column 'PagesK' = Pages / 1000 rounded to 3 decimals and format as 0.000 for the comparison table."),
    ]),

    # F-CALC-55 — movie box office (Title / OpeningWeek / DomesticTotal / IntlTotal)
    FileTask(F_CALC_55, "derived_world_total", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "WorldTotal", "src_col_idx": 2,
               "expr": "(float(v) + float(ws.cell(r, 4).value)) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) else None",
               "number_format": "$#,##0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm building the global box-office summary for the studio's quarterly performance review — add a derived column 'WorldTotal' = DomesticTotal + IntlTotal and format as $#,##0 for the investor-relations slides."),
        Param({"new_header": "OpeningPctOfDomestic", "src_col_idx": 1,
               "expr": "round(float(v) / float(ws.cell(r, 3).value), 4) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) "
                       "and ws.cell(r,3).value > 0 else None",
               "number_format": "0.00%"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like an opening-week share metric for the box-office trade-press article — add a derived column 'OpeningPctOfDomestic' = OpeningWeek / DomesticTotal and format as 0.00% (two-decimal percent) so the article's table shows each film's front-loading."),
    ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_55, "sort_by_domestic_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 2, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort the films by DomesticTotal descending (highest grossing first); keep the header at row 1."),
        # Param({"col_idx": 0, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the films alphabetically by Title; keep the header at row 1."),
    # ]),

    # F-CALC-56 — real estate listings (ListingID / City / Beds / SqFt / Price)
    FileTask(F_CALC_56, "derived_price_per_sqft", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "PricePerSqFt", "src_col_idx": 4,
               "expr": "round(float(v) / float(ws.cell(r, 4).value), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) "
                       "and ws.cell(r,4).value > 0 else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a normalised price metric on this MLS extract before sharing with prospective buyers — add a derived column 'PricePerSqFt' = Price / SqFt rounded to 2 decimals and format as $#,##0.00 so comps are easy to compare."),
        Param({"new_header": "PriceK", "src_col_idx": 4,
               "expr": "round(float(v) / 1000.0, 1) if isinstance(v, (int,float)) else None",
               "number_format": "0.0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the listing prices expressed in thousands of dollars for the buyer-friendly comparison page — add a derived column 'PriceK' = Price / 1000 rounded to 1 decimal and format as 0.0 so the prices fit cleanly on the listing card."),
    ]),
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_56, "filter_by_city", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "Portland", "filter_col_idx": 1, "filter_value": "Portland"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Portland"),
              # "Create a new sheet 'Portland' containing only the listings whose City is 'Portland' (copy the header too)."),
        # Param({"new_sheet_name": "Seattle", "filter_col_idx": 1, "filter_value": "Seattle"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Seattle"),
              # "Create a new sheet 'Seattle' that holds the header row plus every listing whose City is 'Seattle'."),
    # ]),

    # F-CALC-57 — purchase orders (PO# / Vendor / Amount / Status)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_57, "groupby_vendor_totals", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "Vendor Totals", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "sum", "key_header": "Vendor", "value_header": "TotalAmount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Vendor Totals"),
              # "I'm preparing the vendor-spend report for the procurement review — add a new sheet 'Vendor Totals' summing Amount per Vendor with header 'Vendor','TotalAmount' followed by one row per distinct vendor so the buyer can renegotiate larger contracts."),
        # Param({"new_sheet_name": "Vendor Counts", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "count", "key_header": "Vendor", "value_header": "POCount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Vendor Counts"),
              # "I'd like to count POs per vendor for the procurement review with the finance director next Monday — add a new sheet 'Vendor Counts' tallying POs per Vendor with header 'Vendor','POCount' followed by one row per distinct vendor for the review."),
    # ]),
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_57, "filter_approved", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "Approved", "filter_col_idx": 3, "filter_value": "Approved"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Approved"),
              # "Create a new sheet 'Approved' containing only the POs whose Status is 'Approved' (copy the header too)."),
        # Param({"new_sheet_name": "Pending", "filter_col_idx": 3, "filter_value": "Pending"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Pending"),
              # "Create a new sheet 'Pending' that holds the header row plus every PO whose Status is 'Pending'."),
    # ]),

    # F-CALC-58 — class attendance (Date / Course / Section / Headcount)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_58, "groupby_course_avg", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "Avg By Course", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "avg", "key_header": "Course", "value_header": "AvgHeadcount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Avg By Course"),
              # "I'd like to compute average classroom headcounts ahead of the registrar's enrolment review — add a new sheet 'Avg By Course' with the average Headcount per Course using header 'Course','AvgHeadcount' then one row per distinct course for the dean."),
        # Param({"new_sheet_name": "Total By Course", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "sum", "key_header": "Course", "value_header": "TotalSeats"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Total By Course"),
              # "I'd like total seat-counts by course for the budget allocation discussion with the dean next month — add a new sheet 'Total By Course' summing Headcount per Course with header 'Course','TotalSeats' followed by one row per distinct course."),
    # ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_58, "color_low_attendance", "conditional_format",
             # _gold_cell_color_by_predicate, params=[
        # Param({"rules_py": "isinstance(row[3], (int,float)) and row[3] < 25",
               # "apply_kind": "fill", "apply_argb": "FFFFC0C0"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every row whose Headcount is below 25 a pale-red fill (low-attendance flag)."),
        # Param({"rules_py": "isinstance(row[3], (int,float)) and row[3] >= 30",
               # "apply_kind": "fill", "apply_argb": "FF98FB98"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every row whose Headcount is 30 or higher a pale-green fill (full-class flag)."),
    # ]),

    # F-CALC-59 — employee skills (EmpID / Name / Skill / Level)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_59, "filter_by_skill", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "Python", "filter_col_idx": 2, "filter_value": "Python"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Python"),
              # "I'm staffing a Python project and need a focused list of qualified engineers — create a new sheet 'Python' containing only the rows whose Skill is 'Python' (copy the header too) so I can share it with the project manager."),
        # Param({"new_sheet_name": "SQL", "filter_col_idx": 2, "filter_value": "SQL"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("SQL"),
              # "Create a new sheet 'SQL' that holds the header row plus every employee whose Skill is 'SQL'."),
    # ]),
    FileTask(F_CALC_59, "string_clean_name_upper", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "upper"},
              _RULE_SHEET_DATA,
              "Replace each Name with its UPPERCASE form (overwrite the Name column)."),
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "lower"},
              _RULE_SHEET_DATA,
              "I'd like the employee names normalised to lowercase for the legacy HRIS import — replace each Name with its lowercase form (overwrite the Name column) so the import job won't reject the records on casing mismatch."),
    ]),

    # F-CALC-60 — warehouse outbound (OrderID / Warehouse / Items / Weight_kg)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_60, "groupby_warehouse_weight", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Warehouse", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "sum", "key_header": "Warehouse", "value_header": "TotalWeight"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Warehouse"),
              # "I'm preparing the warehouse throughput summary for the logistics steering committee — add a new sheet 'By Warehouse' summing Weight_kg per Warehouse with header 'Warehouse','TotalWeight' followed by one row per distinct warehouse for the comparison."),
        # Param({"new_sheet_name": "Item Counts", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "sum", "key_header": "Warehouse", "value_header": "TotalItems"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Item Counts"),
              # "Add a new sheet 'Item Counts' summing Items per Warehouse: header 'Warehouse','TotalItems' then one row per distinct warehouse."),
    # ]),
    FileTask(F_CALC_60, "derived_weight_per_item", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "WeightPerItem", "src_col_idx": 3,
               "expr": "round(float(v) / float(ws.cell(r, 3).value), 3) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) "
                       "and ws.cell(r,3).value > 0 else None",
               "number_format": "0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'WeightPerItem' = Weight_kg / Items rounded to 3 decimals; format as 0.000."),
        Param({"new_header": "WeightLb", "src_col_idx": 3,
               "expr": "round(float(v) * 2.20462, 2) if isinstance(v, (int,float)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the shipment weights expressed in pounds for the US-distribution partner's reference sheet — add a derived column 'WeightLb' = Weight_kg × 2.20462 rounded to 2 decimals and format as 0.00 so the US partners read familiar units."),
    ]),

    # F-CALC-61 — taxi fares (TripID / Distance_km / Duration_min / Fare_USD)
    FileTask(F_CALC_61, "derived_fare_per_km", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "FarePerKm", "src_col_idx": 3,
               "expr": "round(float(v) / float(ws.cell(r, 2).value), 3) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,2).value,(int,float)) "
                       "and ws.cell(r,2).value > 0 else None",
               "number_format": "0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like to compute a unit fare for our taxi-fleet pricing analysis — add a derived column 'FarePerKm' = Fare_USD / Distance_km rounded to 3 decimals and format as 0.000 so we can compare rate efficiency across trips."),
        Param({"new_header": "AvgSpeed", "src_col_idx": 1,
               "expr": "round(float(v) / (float(ws.cell(r, 3).value) / 60.0), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) "
                       "and ws.cell(r,3).value > 0 else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'AvgSpeed' = Distance_km / (Duration_min / 60) (km/h) rounded to 2 decimals; format as 0.00."),
    ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_61, "sort_by_fare_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 3, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort the trips by Fare_USD descending (most expensive first); keep the header at row 1."),
        # Param({"col_idx": 1, "reverse": True}, _RULE_SHEET_DATA,
              # "I'd like the longest trips at the top of the fare-distance correlation analysis — sort the trips by Distance_km descending so the longest journeys appear first, and keep the header at row 1 for the analyst's review."),
    # ]),

    # F-CALC-62 — workout log (Date / Activity / Duration_min / CaloriesBurned)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_62, "groupby_activity_calories", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Activity", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "sum", "key_header": "Activity", "value_header": "TotalCalories"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Activity"),
              # "I'm reviewing my workout balance with my personal trainer this week — add a new sheet 'By Activity' summing CaloriesBurned per Activity with header 'Activity','TotalCalories' followed by one row per distinct activity for the coaching session."),
        # Param({"new_sheet_name": "Avg Activity", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "avg", "key_header": "Activity", "value_header": "AvgDuration"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Avg Activity"),
              # "Add a new sheet 'Avg Activity' showing average Duration_min per Activity: header 'Activity','AvgDuration' then one row per distinct activity."),
    # ]),
    FileTask(F_CALC_62, "derived_calories_per_min", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "CaloriesPerMin", "src_col_idx": 3,
               "expr": "round(float(v) / float(ws.cell(r, 3).value), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) "
                       "and ws.cell(r,3).value > 0 else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'CaloriesPerMin' = CaloriesBurned / Duration_min rounded to 2 decimals; format as 0.00."),
        Param({"new_header": "DurationHours", "src_col_idx": 2,
               "expr": "round(float(v) / 60.0, 3) if isinstance(v, (int,float)) else None",
               "number_format": "0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the workout durations expressed in hours for the training-load tracker my coach uses — add a derived column 'DurationHours' = Duration_min / 60 rounded to 3 decimals and format as 0.000 so the weekly totals are easy to interpret."),
    ]),

    # F-CALC-63 — apartment rents (Building / Unit / Beds / MonthlyRent)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_63, "groupby_building_avg_rent", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "Bldg Avg", "key_col_idx": 0, "value_col_idx": 3,
               # "agg": "avg", "key_header": "Building", "value_header": "AvgRent"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Bldg Avg"),
              # "I'm benchmarking rents across our portfolio for the asset manager's review — add a new sheet 'Bldg Avg' showing the average MonthlyRent per Building with header 'Building','AvgRent' then one row per distinct building for the rent-roll discussion."),
        # Param({"new_sheet_name": "Bldg Total", "key_col_idx": 0, "value_col_idx": 3,
               # "agg": "sum", "key_header": "Building", "value_header": "TotalRent"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Bldg Total"),
              # "Add a new sheet 'Bldg Total' summing MonthlyRent per Building: header 'Building','TotalRent' then one row per distinct building."),
    # ]),
    FileTask(F_CALC_63, "derived_annual_rent", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "AnnualRent", "src_col_idx": 3,
               "expr": "int(float(v) * 12) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'AnnualRent' = MonthlyRent × 12; format as $#,##0."),
        Param({"new_header": "RentPerBed", "src_col_idx": 3,
               "expr": "round(float(v) / float(ws.cell(r, 3).value), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) "
                       "and ws.cell(r,3).value > 0 else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a per-bedroom rent comparison for the multifamily-asset performance review next quarter — add a derived column 'RentPerBed' = MonthlyRent / Beds rounded to 2 decimals and format as $#,##0.00 for the asset-manager's deep-dive."),
    ]),

    # F-CALC-64 — blog posts (PostID / Title / Author / ViewCount)
    FileTask(F_CALC_64, "string_clean_title_lower", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "lower"},
              _RULE_SHEET_DATA,
              "I'm normalising the blog-post titles before pushing the feed to social media — replace each Title with its lowercase form (overwrite the Title column) so the casing is consistent across our channels."),
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "upper"},
              _RULE_SHEET_DATA,
              "Replace each Title with its UPPERCASE form (overwrite the Title column)."),
    ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_64, "sort_by_views_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 3, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort the posts by ViewCount descending (most viewed first); keep the header at row 1."),
        # Param({"col_idx": 2, "reverse": False}, _RULE_SHEET_DATA,
              # "I'd like the blog posts grouped alphabetically by author for the editor's monthly content audit — sort the posts alphabetically by Author and keep the header at row 1 so the audit list is easy to navigate."),
    # ]),

    # F-CALC-65 — classroom quiz (QuestionID / Topic / Difficulty / CorrectPct)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_65, "groupby_topic_avg", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Topic", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "avg", "key_header": "Topic", "value_header": "AvgCorrectPct"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Topic"),
              # "I'd like to identify weak-topic areas for next week's review session — add a new sheet 'By Topic' showing the average CorrectPct per Topic with header 'Topic','AvgCorrectPct' followed by one row per distinct topic for the lesson plan."),
        # Param({"new_sheet_name": "By Difficulty", "key_col_idx": 2, "value_col_idx": 3,
               # "agg": "avg", "key_header": "Difficulty", "value_header": "AvgCorrectPct"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Difficulty"),
              # "Add a new sheet 'By Difficulty' showing average CorrectPct per Difficulty: header 'Difficulty','AvgCorrectPct' then one row per distinct level."),
    # ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_65, "color_hard_questions", "conditional_format",
             # _gold_two_color_by_predicate, params=[
        # Param({"pred_a_py": "row[2] == 'Hard'",  "argb_a": "FFFFB6C1",
               # "pred_b_py": "row[2] == 'Easy'",  "argb_b": "FF98FB98",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every Hard-difficulty question a pink fill and every Easy-difficulty question a pale-green fill."),
        # Param({"pred_a_py": "isinstance(row[3], (int,float)) and row[3] < 40", "argb_a": "FFFF6347",
               # "pred_b_py": "isinstance(row[3], (int,float)) and row[3] >= 80", "argb_b": "FF90EE90",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "I'd like to flag the troublesome and easy quiz questions for next semester's curriculum rewrite — give every question whose CorrectPct is below 40 a tomato fill and every question whose CorrectPct is 80 or higher a light-green fill for the review."),
    # ]),

    # F-CALC-66 — film festival (Title / Country / RuntimeMin / Rating)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_66, "groupby_country_count", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Country", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "count", "key_header": "Country", "value_header": "FilmCount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Country"),
              # "I'm preparing the country-distribution summary for the film-festival programme — add a new sheet 'By Country' tallying films per Country with header 'Country','FilmCount' followed by one row per distinct country for the printed brochure."),
        # Param({"new_sheet_name": "Country Avg", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "avg", "key_header": "Country", "value_header": "AvgRating"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Country Avg"),
              # "Add a new sheet 'Country Avg' showing average Rating per Country: header 'Country','AvgRating' then one row per distinct country."),
    # ]),
    FileTask(F_CALC_66, "derived_runtime_hours", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "RuntimeHours", "src_col_idx": 2,
               "expr": "round(float(v) / 60.0, 3) if isinstance(v, (int,float)) else None",
               "number_format": "0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'RuntimeHours' = RuntimeMin / 60 rounded to 3 decimals; format as 0.000."),
        Param({"new_header": "Rating10", "src_col_idx": 3,
               "expr": "round(float(v) * 10, 0) if isinstance(v, (int,float)) else None",
               "number_format": "0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a 1-100 ratings scale for the film festival's public voting site — add a derived column 'Rating10' = Rating × 10 rounded to integer and format as 0 so the audience-facing display feels natural."),
    ]),

    # F-CALC-67 — charity pledges (PledgeID / Donor / Amount / Type)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_67, "filter_monthly", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "Monthly", "filter_col_idx": 3, "filter_value": "Monthly"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Monthly"),
              # "I'm preparing the recurring-donor list for the nonprofit's annual stewardship campaign — create a new sheet 'Monthly' containing only the pledges whose Type is 'Monthly' (copy the header too) so the donor-relations team can send thank-you notes."),
        # Param({"new_sheet_name": "OneTime", "filter_col_idx": 3, "filter_value": "OneTime"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("OneTime"),
              # "Create a new sheet 'OneTime' that holds the header row plus every pledge whose Type is 'OneTime'."),
    # ]),
    FileTask(F_CALC_67, "derived_annual_pledge", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "AnnualValue", "src_col_idx": 2,
               "expr": "(float(v) * 12 if ws.cell(r, 4).value == 'Monthly' else float(v)) "
                       "if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'AnnualValue' that equals Amount × 12 when Type is 'Monthly' and just Amount otherwise; format as $#,##0.00."),
        Param({"new_header": "AmountTier", "src_col_idx": 2,
               "expr": "'Major' if isinstance(v,(int,float)) and v >= 250 else "
                       "('Standard' if isinstance(v,(int,float)) and v >= 100 else 'Small')",
               "number_format": None},
              _RULE_SHEET_DATA,
              "I'd like a donor-tier classifier on this pledge log to plan the upcoming stewardship calls — add a derived column 'AmountTier' classifying Amount as 'Major' if ≥ 250, 'Standard' if ≥ 100, and 'Small' otherwise for the development team's outreach."),
    ]),

    # F-CALC-68 — software bugs (BugID / Severity / Status / DaysOpen)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_68, "filter_critical_open", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "Critical", "filter_col_idx": 1, "filter_value": "Critical"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Critical"),
              # "I'm preparing the critical-bug triage list for tomorrow's incident-review meeting — create a new sheet 'Critical' containing only the bugs whose Severity is 'Critical' (copy the header too) so the on-call team can prioritise fixes."),
        # Param({"new_sheet_name": "OpenBugs", "filter_col_idx": 2, "filter_value": "Open"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("OpenBugs"),
              # "Create a new sheet 'OpenBugs' that holds the header row plus every bug whose Status is 'Open'."),
    # ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_68, "color_severity_band", "conditional_format",
             # _gold_two_color_by_predicate, params=[
        # Param({"pred_a_py": "row[1] == 'Critical'", "argb_a": "FFFF0000",
               # "pred_b_py": "row[1] == 'Major'",    "argb_b": "FFFFA500",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every Critical bug a red fill and every Major bug an orange fill."),
        # Param({"pred_a_py": "isinstance(row[3], (int,float)) and row[3] >= 30", "argb_a": "FFFF6347",
               # "pred_b_py": "isinstance(row[3], (int,float)) and row[3] < 7",   "argb_b": "FF98FB98",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "I'd like to age the bug backlog visually for the next sprint planning ceremony — give every bug open ≥ 30 days a tomato fill and every bug open < 7 days a pale-green fill so the team prioritises the stale issues first."),
    # ]),

    # F-CALC-69 — hotel bookings (BookingID / Guest / Room / Nights / TotalUSD)
    FileTask(F_CALC_69, "derived_nightly_rate", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "NightlyRate", "src_col_idx": 4,
               "expr": "round(float(v) / float(ws.cell(r, 4).value), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) "
                       "and ws.cell(r,4).value > 0 else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like to derive nightly rates for the hotel revenue-management dashboard — add a derived column 'NightlyRate' = TotalUSD / Nights rounded to 2 decimals and format as $#,##0.00 so the revenue manager can compare RevPAR trends."),
        Param({"new_header": "WithTaxTotal", "src_col_idx": 4,
               "expr": "round(float(v) * 1.12, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a tax-inclusive booking total for the hotel revenue audit binder — add a derived column 'WithTaxTotal' = TotalUSD × 1.12 (12% tax) rounded to 2 decimals and format as $#,##0.00 so the audit team has the tax-loaded figures ready."),
    ]),
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_69, "groupby_room_total", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Room", "key_col_idx": 2, "value_col_idx": 4,
               # "agg": "sum", "key_header": "Room", "value_header": "TotalRevenue"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Room"),
              # "Add a new sheet 'By Room' summing TotalUSD per Room: header 'Room','TotalRevenue' followed by one row per distinct room type."),
        # Param({"new_sheet_name": "Room Avg", "key_col_idx": 2, "value_col_idx": 3,
               # "agg": "avg", "key_header": "Room", "value_header": "AvgNights"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Room Avg"),
              # "Add a new sheet 'Room Avg' showing average Nights per Room: header 'Room','AvgNights' then one row per distinct room type."),
    # ]),

    # F-CALC-70 — app downloads (AppID / Platform / DownloadCount / RatingAvg)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_70, "groupby_platform_downloads", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Platform", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "sum", "key_header": "Platform", "value_header": "TotalDownloads"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Platform"),
              # "I'm preparing the app-store performance summary for the marketing all-hands — add a new sheet 'By Platform' summing DownloadCount per Platform with header 'Platform','TotalDownloads' followed by one row per distinct platform for the slide deck."),
        # Param({"new_sheet_name": "Platform Avg", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "avg", "key_header": "Platform", "value_header": "AvgRating"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Platform Avg"),
              # "Add a new sheet 'Platform Avg' showing average RatingAvg per Platform: header 'Platform','AvgRating' then one row per distinct platform."),
    # ]),
    FileTask(F_CALC_70, "derived_downloads_k", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "DownloadsK", "src_col_idx": 2,
               "expr": "round(float(v) / 1000.0, 2) if isinstance(v, (int,float)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'DownloadsK' = DownloadCount / 1000 rounded to 2 decimals; format as 0.00."),
        Param({"new_header": "RatingScaled", "src_col_idx": 3,
               "expr": "round(float(v) * 20, 1) if isinstance(v, (int,float)) else None",
               "number_format": "0.0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a 0-100 scale rating for the app-store comparison featured on the tech-review site — add a derived column 'RatingScaled' = RatingAvg × 20 rounded to 1 decimal and format as 0.0 so readers see familiar percentages."),
    ]),

    # F-CALC-71 — country capitals (Country / Capital / Continent / Pop_M)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_71, "groupby_continent_pop", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Continent", "key_col_idx": 2, "value_col_idx": 3,
               # "agg": "sum", "key_header": "Continent", "value_header": "TotalPopM"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Continent"),
              # "I'm building a population-by-continent overview for the geography curriculum — add a new sheet 'By Continent' summing Pop_M per Continent with header 'Continent','TotalPopM' followed by one row per distinct continent for the classroom handout."),
        # Param({"new_sheet_name": "Continent Count", "key_col_idx": 2, "value_col_idx": 3,
               # "agg": "count", "key_header": "Continent", "value_header": "CapitalCount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Continent Count"),
              # "Add a new sheet 'Continent Count' tallying capitals per Continent: header 'Continent','CapitalCount' then one row per distinct continent."),
    # ]),
    FileTask(F_CALC_71, "string_clean_country_upper", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 0, "dst_col_idx": 0, "op": "upper"},
              _RULE_SHEET_DATA,
              "Replace each Country with its UPPERCASE form (overwrite the Country column)."),
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "upper"},
              _RULE_SHEET_DATA,
              "I'd like the capital names in UPPERCASE for the printed atlas index — replace each Capital with its UPPERCASE form (overwrite the Capital column) so the atlas typesetter has the canonical formatting."),
    ]),

    # F-CALC-72 — gym membership (MemberID / Tier / MonthlyFee / Status)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_72, "filter_active_members", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "Active", "filter_col_idx": 3, "filter_value": "Active"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Active"),
              # "I'd like to prepare the active-member roster for the front-desk welcome packets — create a new sheet 'Active' containing only the members whose Status is 'Active' (copy the header too) so the desk team has a clean list."),
        # Param({"new_sheet_name": "Cancelled", "filter_col_idx": 3, "filter_value": "Cancelled"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Cancelled"),
              # "Create a new sheet 'Cancelled' that holds the header row plus every member whose Status is 'Cancelled'."),
    # ]),
    FileTask(F_CALC_72, "derived_annual_fee", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "AnnualFee", "src_col_idx": 2,
               "expr": "round(float(v) * 12, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'AnnualFee' = MonthlyFee × 12 rounded to 2 decimals; format as $#,##0.00."),
        Param({"new_header": "FeeQuarterly", "src_col_idx": 2,
               "expr": "round(float(v) * 3, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a quarterly-fee column for the bursar's billing reconciliation — add a derived column 'FeeQuarterly' = MonthlyFee × 3 rounded to 2 decimals and format as $#,##0.00 so the quarterly billing run matches the contract terms."),
    ]),

    # F-CALC-73 — garden plants (PlantID / Species / SunNeeded / WaterNeeded)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_73, "filter_full_sun", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "FullSun", "filter_col_idx": 2, "filter_value": "Full"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("FullSun"),
              # "I'm planning the sunny side of the garden for spring planting — create a new sheet 'FullSun' containing only the plants whose SunNeeded is 'Full' (copy the header too) so I can take it to the nursery for shopping."),
        # Param({"new_sheet_name": "Shade", "filter_col_idx": 2, "filter_value": "Shade"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Shade"),
              # "Create a new sheet 'Shade' that holds the header row plus every plant whose SunNeeded is 'Shade'."),
    # ]),
    FileTask(F_CALC_73, "string_clean_species_lower", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "lower"},
              _RULE_SHEET_DATA,
              "Replace each Species with its lowercase form (overwrite the Species column)."),
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "title"},
              _RULE_SHEET_DATA,
              "I'd like the plant species names in Title Case for the garden-club newsletter feature this month — replace each Species with its Title Case form (overwrite the Species column) so the species names read cleanly in the article."),
    ]),

    # F-CALC-74 — warehouse SKUs (SKU / Item / OnHand / Bin) — leading zeros preserved
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_74, "color_low_stock", "conditional_format",
             # _gold_cell_color_by_predicate, params=[
        # Param({"rules_py": "isinstance(row[2], (int,float)) and row[2] < 100",
               # "apply_kind": "fill", "apply_argb": "FFFFC0C0"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "I'd like to flag low-stock SKUs ahead of the warehouse manager's reorder review next week — give every SKU whose OnHand is below 100 a pale-red fill so the at-risk items jump out on the printed worklist."),
        # Param({"rules_py": "isinstance(row[2], (int,float)) and row[2] >= 500",
               # "apply_kind": "fill", "apply_argb": "FF98FB98"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every SKU whose OnHand is 500 or higher a pale-green fill (well-stocked highlight)."),
    # ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_74, "sort_by_onhand_asc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 2, "reverse": False}, _RULE_SHEET_DATA,
              # "Sort the SKUs by OnHand ascending (lowest stock first); keep the header at row 1."),
        # Param({"col_idx": 3, "reverse": False}, _RULE_SHEET_DATA,
              # "I'd like the SKUs grouped by warehouse bin for the warehouse picker's morning route — sort the SKUs alphabetically by Bin location and keep the header at row 1 so the pick list follows the aisle layout."),
    # ]),

    # F-CALC-75 — streaming subs (SubID / Plan / DeviceCount / MonthlyFee)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_75, "groupby_plan_revenue", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Plan", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "sum", "key_header": "Plan", "value_header": "TotalMonthlyFee"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Plan"),
              # "I'm preparing a subscription-tier revenue breakdown for the investor data room — add a new sheet 'By Plan' summing MonthlyFee per Plan with header 'Plan','TotalMonthlyFee' followed by one row per distinct plan for the upcoming pitch."),
        # Param({"new_sheet_name": "Plan Devices", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "sum", "key_header": "Plan", "value_header": "TotalDevices"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Plan Devices"),
              # "Add a new sheet 'Plan Devices' summing DeviceCount per Plan: header 'Plan','TotalDevices' then one row per distinct plan."),
    # ]),
    FileTask(F_CALC_75, "derived_per_device_fee", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "FeePerDevice", "src_col_idx": 3,
               "expr": "round(float(v) / float(ws.cell(r, 3).value), 4) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) "
                       "and ws.cell(r,3).value > 0 else None",
               "number_format": "$0.0000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'FeePerDevice' = MonthlyFee / DeviceCount rounded to 4 decimals; format as $0.0000."),
        Param({"new_header": "AnnualFee", "src_col_idx": 3,
               "expr": "round(float(v) * 12, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the annual streaming-subscription fees computed for the investor data room — add a derived column 'AnnualFee' = MonthlyFee × 12 rounded to 2 decimals and format as $#,##0.00 so the LTV slide reads in clean annual figures."),
    ]),

    # F-CALC-76 — pet clinic (VisitID / Pet / Species / Cost)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_76, "groupby_species_cost", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Species", "key_col_idx": 2, "value_col_idx": 3,
               # "agg": "sum", "key_header": "Species", "value_header": "TotalCost"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Species"),
              # "I'd like to break down clinic spending by species for the practice owner's monthly P&L review — add a new sheet 'By Species' summing Cost per Species with header 'Species','TotalCost' followed by one row per distinct species for the print-out."),
        # Param({"new_sheet_name": "Species Avg", "key_col_idx": 2, "value_col_idx": 3,
               # "agg": "avg", "key_header": "Species", "value_header": "AvgCost"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Species Avg"),
              # "Add a new sheet 'Species Avg' showing average Cost per Species: header 'Species','AvgCost' then one row per distinct species."),
    # ]),
    FileTask(F_CALC_76, "derived_cost_with_tax", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "CostWithTax", "src_col_idx": 3,
               "expr": "round(float(v) * 1.085, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'CostWithTax' = Cost × 1.085 (8.5% tax) rounded to 2 decimals; format as $#,##0.00."),
        Param({"new_header": "CostBand", "src_col_idx": 3,
               "expr": "'High' if isinstance(v,(int,float)) and v >= 200 else "
                       "('Low' if isinstance(v,(int,float)) and v < 100 else 'Mid')",
               "number_format": None},
              _RULE_SHEET_DATA,
              "I'd like a cost-band classifier on the pet-clinic visits for the practice manager's pricing review — add a derived column 'CostBand' classifying Cost as 'High' if ≥ 200, 'Low' if < 100, and 'Mid' otherwise for the segmentation analysis."),
    ]),

    # F-CALC-77 — concert tickets (TicketID / Section / Price / Sold)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_77, "groupby_section_revenue", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Section", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "sum", "key_header": "Section", "value_header": "TotalPrice"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Section"),
              # "I'm preparing a section-by-section revenue summary for the venue's settlement statement — add a new sheet 'By Section' summing Price per Section with header 'Section','TotalPrice' followed by one row per distinct section for the promoter."),
        # Param({"new_sheet_name": "Section Count", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "count", "key_header": "Section", "value_header": "TicketCount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Section Count"),
              # "Add a new sheet 'Section Count' tallying tickets per Section: header 'Section','TicketCount' then one row per distinct section."),
    # ]),
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_77, "filter_sold", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "Sold", "filter_col_idx": 3, "filter_value": "Yes"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Sold"),
              # "Create a new sheet 'Sold' containing only the tickets whose Sold is 'Yes' (copy the header too)."),
        # Param({"new_sheet_name": "Available", "filter_col_idx": 3, "filter_value": "No"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Available"),
              # "I'd like an unsold-tickets report ahead of tonight's promoter pricing call — create a new sheet 'Available' that holds the header row plus every ticket whose Sold is 'No' so the promoter can decide on a last-minute discount."),
    # ]),

    # F-CALC-78 — runners log (RunnerID / Name / Distance_km / TimeMin)
    FileTask(F_CALC_78, "derived_pace", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "PaceMinPerKm", "src_col_idx": 3,
               "expr": "round(float(v) / float(ws.cell(r, 3).value), 3) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) "
                       "and ws.cell(r,3).value > 0 else None",
               "number_format": "0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm comparing pace data across the team ahead of the next marathon block — add a derived column 'PaceMinPerKm' = TimeMin / Distance_km rounded to 3 decimals and format as 0.000 so the coach can review every runner's tempo."),
        Param({"new_header": "SpeedKmh", "src_col_idx": 2,
               "expr": "round(float(v) / (float(ws.cell(r, 4).value) / 60.0), 3) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) "
                       "and ws.cell(r,4).value > 0 else None",
               "number_format": "0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'SpeedKmh' = Distance_km / (TimeMin / 60) rounded to 3 decimals; format as 0.000."),
    ]),
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_78, "groupby_distance_avg", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Distance", "key_col_idx": 2, "value_col_idx": 3,
               # "agg": "avg", "key_header": "Distance_km", "value_header": "AvgTimeMin"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Distance"),
              # "Add a new sheet 'By Distance' showing average TimeMin per Distance_km: header 'Distance_km','AvgTimeMin' then one row per distance."),
        # Param({"new_sheet_name": "Count By Distance", "key_col_idx": 2, "value_col_idx": 3,
               # "agg": "count", "key_header": "Distance_km", "value_header": "RunnerCount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Count By Distance"),
              # "I'd like to tally the runners by race distance for the upcoming marathon-club newsletter — add a new sheet 'Count By Distance' with header 'Distance_km','RunnerCount' followed by one row per distance for the participation summary."),
    # ]),

    # F-CALC-79 — recipe ratings (RecipeID / Cuisine / Rating / ReviewCount)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_79, "groupby_cuisine_avg", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "Cuisine Avg", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "avg", "key_header": "Cuisine", "value_header": "AvgRating"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Cuisine Avg"),
              # "I'd like to compare cuisine ratings before publishing the food-blog roundup post — add a new sheet 'Cuisine Avg' showing the average Rating per Cuisine with header 'Cuisine','AvgRating' then one row per distinct cuisine for the article."),
        # Param({"new_sheet_name": "Cuisine Reviews", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "sum", "key_header": "Cuisine", "value_header": "TotalReviews"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Cuisine Reviews"),
              # "Add a new sheet 'Cuisine Reviews' summing ReviewCount per Cuisine: header 'Cuisine','TotalReviews' then one row per distinct cuisine."),
    # ]),
    FileTask(F_CALC_79, "derived_weighted_score", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "WeightedScore", "src_col_idx": 2,
               "expr": "round(float(v) * float(ws.cell(r, 4).value), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'WeightedScore' = Rating × ReviewCount rounded to 2 decimals; format as 0.00."),
        Param({"new_header": "RatingPct", "src_col_idx": 2,
               "expr": "round(float(v) / 5.0, 4) if isinstance(v, (int,float)) else None",
               "number_format": "0.00%"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the recipe ratings expressed as a percentage for the food blog's review card UI — add a derived column 'RatingPct' = Rating / 5 and format as 0.00% (two-decimal percent) so readers see familiar percent scores."),
    ]),

    # F-CALC-80 — invoice aging (InvoiceID / Customer / DaysOverdue / Amount)
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_80, "color_overdue_buckets", "conditional_format",
             # _gold_two_color_by_predicate, params=[
        # Param({"pred_a_py": "isinstance(row[2], (int,float)) and row[2] >= 60", "argb_a": "FFFF6347",
               # "pred_b_py": "isinstance(row[2], (int,float)) and row[2] == 0",  "argb_b": "FF98FB98",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "I'd like to flag the most-overdue invoices ahead of the AR director's collections meeting next Monday — give every invoice with DaysOverdue ≥ 60 a tomato fill and every invoice with DaysOverdue = 0 a pale-green fill for the discussion."),
        # Param({"pred_a_py": "isinstance(row[2], (int,float)) and row[2] > 30", "argb_a": "FFFFB6C1",
               # "pred_b_py": "isinstance(row[2], (int,float)) and row[2] <= 7", "argb_b": "FFADD8E6",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Highlight aging buckets: DaysOverdue > 30 in pink, DaysOverdue ≤ 7 in light-blue."),
    # ]),
    FileTask(F_CALC_80, "derived_late_fee", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "LateFee", "src_col_idx": 3,
               "expr": "round(float(v) * 0.015, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'LateFee' = Amount × 0.015 (1.5% late fee) rounded to 2 decimals; format as $#,##0.00."),
        Param({"new_header": "TotalDue", "src_col_idx": 3,
               "expr": "round(float(v) * 1.015, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a Total-Due column with the late fee baked in for tomorrow's collections-call sheet — add a derived column 'TotalDue' = Amount × 1.015 (with 1.5% late fee) rounded to 2 decimals and format as $#,##0.00 for the AR clerks."),
    ]),

    # F-CALC-81 — university courses (CourseCode / Title / Credits / Enrolled)
    FileTask(F_CALC_81, "derived_credit_hours", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "CreditHours", "src_col_idx": 2,
               "expr": "(float(v) * float(ws.cell(r, 4).value)) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) else None",
               "number_format": "#,##0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm computing aggregate credit-hour load for the registrar's enrolment report — add a derived column 'CreditHours' = Credits × Enrolled and format as #,##0 so the dean can see total student-credit-hours per course for the budget."),
        Param({"new_header": "TuitionUSD", "src_col_idx": 2,
               "expr": "round(float(v) * 850, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a per-credit tuition column on the course catalog for the bursar's billing-rate update — add a derived column 'TuitionUSD' = Credits × 850 rounded to 2 decimals (per-credit tuition) and format as $#,##0.00 for the catalog."),
    ]),
    FileTask(F_CALC_81, "string_clean_title_lower", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "lower"},
              _RULE_SHEET_DATA,
              "Replace each Title with its lowercase form (overwrite the Title column)."),
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "title"},
              _RULE_SHEET_DATA,
              "Replace each Title with its Title Case form (overwrite the Title column)."),
    ]),

    # F-CALC-82 — smartphone models (ModelID / Brand / StorageGB / Price)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_82, "groupby_brand_avg_price", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "Brand Avg", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "avg", "key_header": "Brand", "value_header": "AvgPrice"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Brand Avg"),
              # "I'd like to benchmark brand pricing for the smartphone-market research piece — add a new sheet 'Brand Avg' showing the average Price per Brand with header 'Brand','AvgPrice' followed by one row per distinct brand for the article tables."),
        # Param({"new_sheet_name": "Brand Models", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "count", "key_header": "Brand", "value_header": "ModelCount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Brand Models"),
              # "Add a new sheet 'Brand Models' tallying models per Brand: header 'Brand','ModelCount' then one row per distinct brand."),
    # ]),
    FileTask(F_CALC_82, "derived_price_per_gb", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "PricePerGB", "src_col_idx": 3,
               "expr": "round(float(v) / float(ws.cell(r, 3).value), 3) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) "
                       "and ws.cell(r,3).value > 0 else None",
               "number_format": "$0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'PricePerGB' = Price / StorageGB rounded to 3 decimals; format as $0.000."),
        Param({"new_header": "PriceEUR", "src_col_idx": 3,
               "expr": "round(float(v) * 0.92, 2) if isinstance(v, (int,float)) else None",
               "number_format": "€#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like Euro-converted prices for the European-market launch announcement next month — add a derived column 'PriceEUR' = Price × 0.92 rounded to 2 decimals and format as €#,##0.00 for the EU pricing tables."),
    ]),

    # F-CALC-83 — lab results (SampleID / TestType / Result / FlagStatus)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_83, "filter_high_flag", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "HighFlag", "filter_col_idx": 3, "filter_value": "High"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("HighFlag"),
              # "I'm preparing the abnormal-results worklist for the lab director's morning review — create a new sheet 'HighFlag' containing only the lab results whose FlagStatus is 'High' (copy the header too) so the physician follow-up is straightforward."),
        # Param({"new_sheet_name": "Normal", "filter_col_idx": 3, "filter_value": "Normal"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Normal"),
              # "Create a new sheet 'Normal' that holds the header row plus every row whose FlagStatus is 'Normal'."),
    # ]),
    # Pruned (calc rebalance, eval_class=conditional_format OVER):
    # FileTask(F_CALC_83, "color_flags", "conditional_format",
             # _gold_two_color_by_predicate, params=[
        # Param({"pred_a_py": "row[3] == 'High'", "argb_a": "FFFF6347",
               # "pred_b_py": "row[3] == 'Low'",  "argb_b": "FFADD8E6",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "Give every High-flag row a tomato fill and every Low-flag row a light-blue fill."),
        # Param({"pred_a_py": "row[3] == 'Normal'", "argb_a": "FF98FB98",
               # "pred_b_py": "row[3] == 'High'",   "argb_b": "FFFFB6C1",
               # "apply_kind": "fill"},
              # _RULE_SHEET_DATA_AND_STYLE,
              # "I'd like the lab results colour-coded for the morning physician review meeting — give every Normal-flag row a pale-green fill and every High-flag row a pink fill so the abnormal results are immediately apparent."),
    # ]),

    # F-CALC-84 — freight routes (RouteID / Origin / Destination / Distance_km)
    FileTask(F_CALC_84, "derived_miles", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "DistanceMi", "src_col_idx": 3,
               "expr": "round(float(v) * 0.621371, 2) if isinstance(v, (int,float)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like to convert the freight routes to miles for the US dispatch team's reference sheet — add a derived column 'DistanceMi' = Distance_km × 0.621371 rounded to 2 decimals and format as 0.00 so the US partners read the routes in their preferred units."),
        Param({"new_header": "FuelGallons", "src_col_idx": 3,
               "expr": "round(float(v) / 12.0, 2) if isinstance(v, (int,float)) else None",
               "number_format": "0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like an estimated-fuel-usage column for the next dispatch cost review — add a derived column 'FuelGallons' = Distance_km / 12 (km per gallon) rounded to 2 decimals and format as 0.00 so the operations team can plan fuel budgets."),
    ]),
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_84, "groupby_origin_count", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Origin", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "count", "key_header": "Origin", "value_header": "RouteCount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Origin"),
              # "Add a new sheet 'By Origin' tallying routes per Origin: header 'Origin','RouteCount' followed by one row per distinct origin."),
        # Param({"new_sheet_name": "Origin Distance", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "sum", "key_header": "Origin", "value_header": "TotalDistance"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Origin Distance"),
              # "Add a new sheet 'Origin Distance' summing Distance_km per Origin: header 'Origin','TotalDistance' then one row per distinct origin."),
    # ]),

    # F-CALC-85 — internet speeds (TestID / ISP / DownloadMbps / UploadMbps)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_85, "groupby_isp_avg", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "ISP Avg Down", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "avg", "key_header": "ISP", "value_header": "AvgDownload"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("ISP Avg Down"),
              # "I'm preparing an ISP comparison for the residential broadband review article — add a new sheet 'ISP Avg Down' showing the average DownloadMbps per ISP with header 'ISP','AvgDownload' followed by one row per distinct ISP for the readers."),
        # Param({"new_sheet_name": "ISP Avg Up", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "avg", "key_header": "ISP", "value_header": "AvgUpload"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("ISP Avg Up"),
              # "Add a new sheet 'ISP Avg Up' showing average UploadMbps per ISP: header 'ISP','AvgUpload' then one row per distinct ISP."),
    # ]),
    FileTask(F_CALC_85, "derived_ratio", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "DownUpRatio", "src_col_idx": 2,
               "expr": "round(float(v) / float(ws.cell(r, 4).value), 3) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) "
                       "and ws.cell(r,4).value > 0 else None",
               "number_format": "0.000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'DownUpRatio' = DownloadMbps / UploadMbps rounded to 3 decimals; format as 0.000."),
        Param({"new_header": "DownGbps", "src_col_idx": 2,
               "expr": "round(float(v) / 1000.0, 4) if isinstance(v, (int,float)) else None",
               "number_format": "0.0000"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the download speeds expressed in gigabits per second for the next-gen-broadband article — add a derived column 'DownGbps' = DownloadMbps / 1000 rounded to 4 decimals and format as 0.0000 so the high-end ISPs read in familiar units."),
    ]),

    # F-CALC-86 — volunteer hours (VolunteerID / Project / HoursLogged / Status)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_86, "groupby_project_hours", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Project", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "sum", "key_header": "Project", "value_header": "TotalHours"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Project"),
              # "I'd like to roll up volunteer commitment for the nonprofit's annual impact report — add a new sheet 'By Project' summing HoursLogged per Project with header 'Project','TotalHours' followed by one row per distinct project for the development team."),
        # Param({"new_sheet_name": "Project Avg", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "avg", "key_header": "Project", "value_header": "AvgHours"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Project Avg"),
              # "Add a new sheet 'Project Avg' showing average HoursLogged per Project: header 'Project','AvgHours' then one row per distinct project."),
    # ]),
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_86, "filter_pending", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "Pending", "filter_col_idx": 3, "filter_value": "Pending"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Pending"),
              # "Create a new sheet 'Pending' containing only the rows whose Status is 'Pending' (copy the header too)."),
        # Param({"new_sheet_name": "Approved", "filter_col_idx": 3, "filter_value": "Approved"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Approved"),
              # "I'd like the approved-volunteer-hours list for the nonprofit's verified-hours report submitted to the funder — create a new sheet 'Approved' that holds the header row plus every row whose Status is 'Approved' for the grant report."),
    # ]),

    # F-CALC-87 — solar panels (InstallID / Panels / KwCapacity / CostUSD)
    FileTask(F_CALC_87, "derived_cost_per_kw", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "CostPerKw", "src_col_idx": 3,
               "expr": "round(float(v) / float(ws.cell(r, 3).value), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,3).value,(int,float)) "
                       "and ws.cell(r,3).value > 0 else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm preparing a cost-efficiency comparison for the solar-installation case studies — add a derived column 'CostPerKw' = CostUSD / KwCapacity rounded to 2 decimals (apply the formula to every data row, not just the first) and format as $#,##0.00 so prospective customers can compare installs."),
        Param({"new_header": "AnnualKWh", "src_col_idx": 2,
               "expr": "round(float(v) * 1500, 0) if isinstance(v, (int,float)) else None",
               "number_format": "#,##0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like an annual-energy-production column for the solar-payback marketing brochure — add a derived column 'AnnualKWh' = KwCapacity × 1500 (annual production estimate) rounded to integer and format as #,##0 for the brochure's case studies."),
    ]),
    # Pruned (calc rebalance, eval_class=sort_col OVER):
    # FileTask(F_CALC_87, "sort_by_capacity_desc", "sort_col", _gold_sort, params=[
        # Param({"col_idx": 2, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort the installs by KwCapacity descending (largest first); keep the header at row 1."),
        # Param({"col_idx": 3, "reverse": True}, _RULE_SHEET_DATA,
              # "Sort the installs by CostUSD descending (most expensive first); keep the header at row 1."),
    # ]),

    # F-CALC-88 — wine inventory (BottleID / Varietal / Vintage / BottleCount)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_88, "groupby_varietal_count", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Varietal", "key_col_idx": 1, "value_col_idx": 3,
               # "agg": "sum", "key_header": "Varietal", "value_header": "TotalBottles"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Varietal"),
              # "I'd like to inventory the wine collection by varietal before the sommelier's tasting next month — add a new sheet 'By Varietal' summing BottleCount per Varietal with header 'Varietal','TotalBottles' followed by one row per distinct varietal for the cellar audit."),
        # Param({"new_sheet_name": "Vintage Avg", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "avg", "key_header": "Varietal", "value_header": "AvgVintage"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Vintage Avg"),
              # "Add a new sheet 'Vintage Avg' showing average Vintage per Varietal: header 'Varietal','AvgVintage' then one row per distinct varietal."),
    # ]),
    FileTask(F_CALC_88, "derived_age", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "AgeYears", "src_col_idx": 2,
               "expr": "(2026 - int(v)) if isinstance(v, (int,float)) else None",
               "number_format": "0"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Add a derived column 'AgeYears' = 2026 - Vintage; format as 0."),
        Param({"new_header": "BottleValueUSD", "src_col_idx": 3,
               "expr": "round(float(v) * 35.0, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like an estimated-value column for the cellar's insurance appraisal renewal — add a derived column 'BottleValueUSD' = BottleCount × 35 (per-bottle estimated value) rounded to 2 decimals and format as $#,##0.00 for the appraiser's worksheet."),
    ]),

    # F-CALC-89 — pottery orders (OrderID / Item / Quantity / UnitPrice)
    FileTask(F_CALC_89, "derived_line_total", "apply_formula",
             _gold_derived_col, params=[
        Param({"new_header": "LineTotal", "src_col_idx": 2,
               "expr": "round(float(v) * float(ws.cell(r, 4).value), 2) "
                       "if isinstance(v,(int,float)) and isinstance(ws.cell(r,4).value,(int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm computing per-line totals on this pottery order sheet for invoicing my craft-fair customers — add a derived column 'LineTotal' = Quantity × UnitPrice rounded to 2 decimals and format as $#,##0.00 for the bills."),
        Param({"new_header": "WithDiscount", "src_col_idx": 3,
               "expr": "round(float(v) * 0.9, 2) if isinstance(v, (int,float)) else None",
               "number_format": "$#,##0.00"},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like a discounted-price column for the holiday craft-fair flyer next weekend — add a derived column 'WithDiscount' = UnitPrice × 0.9 (10% off) rounded to 2 decimals and format as $#,##0.00 so the flyer shows the festive price."),
    ]),
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_89, "groupby_item_qty", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Item", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "sum", "key_header": "Item", "value_header": "TotalQty"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Item"),
              # "Add a new sheet 'By Item' summing Quantity per Item: header 'Item','TotalQty' followed by one row per distinct item."),
        # Param({"new_sheet_name": "Item Orders", "key_col_idx": 1, "value_col_idx": 2,
               # "agg": "count", "key_header": "Item", "value_header": "OrderCount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Item Orders"),
              # "Add a new sheet 'Item Orders' tallying orders per Item: header 'Item','OrderCount' then one row per distinct item."),
    # ]),

    # F-CALC-90 — helpdesk tickets (TicketID / Priority / Category / Status)
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_90, "filter_open_tickets", "multi_sheet_aggregate",
             # _gold_sheet2_filter, params=[
        # Param({"new_sheet_name": "Open", "filter_col_idx": 3, "filter_value": "Open"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Open"),
              # "I'd like a focused worklist of open tickets ahead of the helpdesk team's daily stand-up — create a new sheet 'Open' containing only the tickets whose Status is 'Open' (copy the header too) so the on-call agent has a clean queue to triage."),
        # Param({"new_sheet_name": "Resolved", "filter_col_idx": 3, "filter_value": "Resolved"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("Resolved"),
              # "I'd like an archive of resolved tickets for the quarterly KPI presentation to the steering committee — create a new sheet 'Resolved' that holds the header row plus every ticket whose Status is 'Resolved' for the throughput slide."),
    # ]),
    # Pruned (calc rebalance, eval_class=multi_sheet_aggregate OVER):
    # FileTask(F_CALC_90, "groupby_category_count", "multi_sheet_aggregate",
             # _gold_sheet2_groupby_sum, params=[
        # Param({"new_sheet_name": "By Category", "key_col_idx": 2, "value_col_idx": 0,
               # "agg": "count", "key_header": "Category", "value_header": "TicketCount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Category"),
              # "Add a new sheet 'By Category' tallying tickets per Category: header 'Category','TicketCount' followed by one row per distinct category."),
        # Param({"new_sheet_name": "By Priority", "key_col_idx": 1, "value_col_idx": 0,
               # "agg": "count", "key_header": "Priority", "value_header": "TicketCount"},
              # _RULE_SHEET_NAME_AND_DATA_NAMED("By Priority"),
              # "Add a new sheet 'By Priority' tallying tickets per Priority: header 'Priority','TicketCount' then one row per distinct priority."),
    # ]),

    # ── Eval-alignment FileTasks (P2-P6). ────────────────────────
    # Each new FileTask emits 2 Param variants → 2 templates each.
    #
    # P2 (charts): compare_table verifies the source-data sheet is intact AND
    # that the target chart-host sheet contains a chart of the expected type.
    # Added an explicit `chart` rule on the target
    # sheet (RN<chart_sheet_name>/EN<chart_sheet_name>) so chart presence is
    # enforced. Without it the eval was vacuous (gold had an empty target
    # sheet, agent's no-op workbook also had 0 charts there → false-pass at
    # `--max-turns 0`). Now the gold places a real BarChart and the rule
    # checks `chart_props=['type']` via compare_calc_chart_type.
    #
    # P2 — F-CALC-91: clustered column chart in Sheet2.
    FileTask(F_CALC_91, "chart_clustered_column", "chart_create",
             _gold_chart_data_passthrough, params=[
        Param({"chart_sheet_name": "Sheet2"},
              [{"type": "sheet_name"},
               {"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
               {"type": "chart", "sheet_idx0": "RNSheet2", "sheet_idx1": "ENSheet2",
                "chart_props": ["type"]}],
              "I'm preparing the weekly sales overview for the steering "
              "committee — please create a clustered column chart in a new "
              "sheet named 'Sheet2' showing the Sales and COGS data for each "
              "week from Sheet1 so the trend is easy to read."),
        Param({"chart_sheet_name": "ChartTab"},
              [{"type": "sheet_name"},
               {"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
               {"type": "chart", "sheet_idx0": "RNChartTab", "sheet_idx1": "ENChartTab",
                "chart_props": ["type"]}],
              "Create a clustered column chart on a new sheet named 'ChartTab' "
              "comparing the Sales and COGS columns across all 12 weeks."),
    ]),

    # P2 — F-CALC-100: chart variant (revenue vs expenses by month).
    FileTask(F_CALC_100, "chart_revenue_expenses", "chart_create",
             _gold_chart_data_passthrough, params=[
        Param({"chart_sheet_name": "Sheet2"},
              [{"type": "sheet_name"},
               {"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
               {"type": "chart", "sheet_idx0": "RNSheet2", "sheet_idx1": "ENSheet2",
                "chart_props": ["type"]}],
              "I'd like a visual side-by-side of revenue versus expenses for "
              "the board pack — please add a clustered column chart on a "
              "new sheet named 'Sheet2' charting Revenue and Expenses across "
              "all twelve months from Sheet1."),
        Param({"chart_sheet_name": "MonthlyChart"},
              [{"type": "sheet_name"},
               {"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
               {"type": "chart", "sheet_idx0": "RNMonthlyChart", "sheet_idx1": "ENMonthlyChart",
                "chart_props": ["type"]}],
              "Create a clustered column chart on a new sheet named 'MonthlyChart' "
              "comparing the Revenue and Expenses columns across all months."),
    ]),

    # P3 — F-CALC-92: column reorder.
    FileTask(F_CALC_92, "column_reorder", "column_reorder",
             _gold_reorder_columns, params=[
        Param({"new_order": [4, 2, 3, 0, 1]},
              _RULE_SHEET_DATA,
              "I'm preparing the customer-orders extract for the CRM import "
              "and need the columns in a specific order — please reorder the "
              "columns to be Date, First Name, Last Name, Order ID, Sales so "
              "the import template matches."),
        Param({"new_order": [2, 3, 4, 0, 1]},
              _RULE_SHEET_DATA,
              "Reorder the columns so they appear as First Name, Last Name, "
              "Date, Order ID, Sales (left to right)."),
    ]),

    # P3 — F-CALC-93: sheet rename + copy with suffix.
    FileTask(F_CALC_93, "sheet_rename_and_copy", "sheet_management",
             _gold_rename_and_copy_sheet, params=[
        Param({"old_name": "Sheet1", "new_name": "LARS Resources",
               "copy_suffix": " (Backup)"},
              [{"type": "sheet_name"}],
              "I'd like to safeguard this resources workbook before sending "
              "it for review — please rename Sheet1 to 'LARS Resources', "
              "then copy it and place the copy before Sheet2 with the suffix "
              "' (Backup)' on its name."),
        Param({"old_name": "Sheet1", "new_name": "Resources",
               "copy_suffix": " (Copy)"},
              [{"type": "sheet_name"}],
              "Rename Sheet1 to 'Resources' and create a duplicate placed "
              "before Sheet2 named 'Resources (Copy)'."),
    ]),

    # P3 — F-CALC-93 variant: sheet rename + delete (uses same File).
    FileTask(F_CALC_93, "sheet_rename_and_delete", "sheet_management",
             _gold_rename_and_delete_sheet, params=[
        Param({"old_name": "Sheet1", "new_name": "LARS Resources",
               "delete_name": "Sheet2"},
              [{"type": "sheet_name"}],
              "I'm tidying up this workbook before archiving it for the "
              "compliance team — rename Sheet1 to 'LARS Resources' and then "
              "delete Sheet2 entirely so only the renamed sheet remains."),
        Param({"old_name": "Sheet1", "new_name": "Active Resources",
               "delete_name": "Sheet2"},
              [{"type": "sheet_name"}],
              "Rename Sheet1 to 'Active Resources' and delete Sheet2 to keep "
              "only the active list."),
    ]),

    # P4 — F-CALC-94: date age computation.
    FileTask(F_CALC_94, "age_from_birthday", "date_arithmetic",
             _gold_age_from_birthday, params=[
        Param({"birthday_col_idx": 3, "age_header": "Age",
               "reference_year": 2026},
              _RULE_SHEET_DATA,
              "I'm preparing the employee directory for the HR review and "
              "would like to add an Age column for each staff member — "
              "calculate each employee's age from their Birthday column "
              "(using 2026 as the reference year) and put it in a new "
              "column headed 'Age'."),
        Param({"birthday_col_idx": 3, "age_header": "AgeYears",
               "reference_year": 2026},
              _RULE_SHEET_DATA,
              "Add a new column headed 'AgeYears' computing each employee's "
              "current age from their Birthday (use 2026 as today's year)."),
    ]),

    # P4 — F-CALC-95: pad zeros to 7 digits.
    FileTask(F_CALC_95, "pad_zeros_seven_digits", "text_manipulation",
             _gold_pad_zeros, params=[
        Param({"src_col_idx": 0, "dst_col_idx": 1, "width": 7},
              _RULE_SHEET_DATA,
              "I'd like to migrate our legacy customer numbers to a standard "
              "seven-digit format for the new billing system — please copy "
              "all the numbers in the 'Old ID' column to the 'New 7 Digit Id' "
              "column and pad them with zeros in front to fill them up to "
              "seven digits."),
        Param({"src_col_idx": 0, "dst_col_idx": 1, "width": 8},
              _RULE_SHEET_DATA,
              "Copy each number in the 'Old ID' column into the 'New 7 Digit "
              "Id' column padded with leading zeros to 8 digits."),
    ]),

    # P4 — F-CALC-96: date duration.
    FileTask(F_CALC_96, "date_duration_days", "date_arithmetic",
             _gold_date_duration, params=[
        Param({"start_col_idx": 2, "end_col_idx": 3,
               "duration_header": "DurationDays"},
              _RULE_SHEET_DATA,
              "I'm putting together a project-duration summary for the PMO's "
              "year-end retrospective — please add a new column 'DurationDays' "
              "computing each project's length in days from StartDate to "
              "EndDate so the PMs can compare timelines."),
        Param({"start_col_idx": 2, "end_col_idx": 3,
               "duration_header": "DaysActive"},
              _RULE_SHEET_DATA,
              "Add a new column headed 'DaysActive' giving the number of "
              "days each project ran (EndDate − StartDate)."),
    ]),

    # P5 — F-CALC-97: custom numfmt M / B (millions / billions).
    FileTask(F_CALC_97, "numfmt_millions_billions", "number_format",
             _gold_numfmt_scaled, params=[
        Param({"col_idx": 1, "number_format": '0.0,,"M"'},
              [{"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'd like the financial figures to read more cleanly for the "
              "investor briefing — please format the Revenue column to "
              "display values in millions with one decimal and an 'M' "
              "suffix (use the custom format 0.0,,\"M\")."),
        Param({"col_idx": 2, "number_format": '0.0,,,"B"'},
              [{"type": "sheet_data", "sheet_idx0": 0, "sheet_idx1": "EI0"},
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Format the MarketCap column to display values in billions "
              "with one decimal and a 'B' suffix (custom format 0.0,,,\"B\")."),
    ]),

    # P6 — F-CALC-98: total row + month-on-month growth column.
    FileTask(F_CALC_98, "total_and_mom_growth", "compound_multi_step",
             _gold_total_and_growth, params=[
        Param({"value_col_idx": 1, "label_col_idx": 0,
               "total_label": "Total", "growth_header": "Growth%"},
              _RULE_SHEET_DATA,
              "I'm wrapping up the year-end financial summary for the "
              "leadership team — please calculate the Total of the Sales "
              "column appended as a new bottom row labelled 'Total', AND "
              "add a 'Growth%' column showing the month-on-month percentage "
              "growth of Sales for each row."),
        Param({"value_col_idx": 1, "label_col_idx": 0,
               "total_label": "Grand Total", "growth_header": "MoMChange"},
              _RULE_SHEET_DATA,
              "Append a 'Grand Total' row summing the Sales column and add "
              "a 'MoMChange' column showing each month's percentage change "
              "vs the previous month."),
    ]),

    # P6 — F-CALC-99: filter+aggregate compound.
    FileTask(F_CALC_99, "filter_region_with_total", "compound_multi_step",
             _gold_filter_to_sheet_with_total, params=[
        Param({"new_sheet_name": "North Sales", "filter_col_idx": 0,
               "filter_value": "North", "sum_col_idx": 2,
               "total_label_col_idx": 0, "total_label": "Total"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("North Sales"),
              "I'd like a regional rollup of the North-region performance "
              "before the regional manager's review — create a new sheet "
              "'North Sales' containing only the rows whose Region is 'North' "
              "with the header copied, and append a bottom row labelled "
              "'Total' summing the Sales column."),
        Param({"new_sheet_name": "South Sales", "filter_col_idx": 0,
               "filter_value": "South", "sum_col_idx": 2,
               "total_label_col_idx": 0, "total_label": "Total"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("South Sales"),
              "Create a new sheet 'South Sales' with the header plus every "
              "South-region row, then append a 'Total' row summing the "
              "Sales column."),
    ]),

    # =====================================================================
    # Additions — fill eval-tested skill gaps after the validation 71-task
    # cut. Each FileTask reuses an existing File whose content thematically
    # fits the op; cap-2×2 maintained (the 2nd active task per File).
    # =====================================================================

    # ---- chart_create (7 adds) — uses real _gold_real_chart so eval's
    # `chart` rule (chart_props=['type']) reads a real LineChart / BarChart.
    # Series ref pins the file's natural numeric column.

    # F-CALC-21 concert-tour (Show / Venue / Revenue) — bar chart over Revenue.
    FileTask(F_CALC_21, "chart_revenue_bar", "chart_create",
             _gold_real_chart, params=[
        Param({"chart_type": "bar",
               "series_ref": "Tour!$C$2:$C$13",
               "cat_ref": "Tour!$B$2:$B$13"},
              _RULE_CHART_TYPE_ACTIVE,
              "I'm wrapping up the tour-revenue board pack and need a "
              "quick visual of receipts per stop — please insert a bar "
              "chart on the Tour sheet showing the Revenue column across "
              "all twelve venues so the leadership can scan it at a glance."),
    ]),
    # F-CALC-27 us-gdp (CSV-backed, date / GDP) — line chart on GDP.
    FileTask(F_CALC_27, "chart_gdp_line", "chart_create",
             _gold_real_chart, params=[
        Param({"chart_type": "line",
               "series_ref": "GDP!$B$2:$B$33",
               "cat_ref": "GDP!$A$2:$A$33"},
              _RULE_CHART_TYPE_ACTIVE,
              "Could you help me prepare the macro briefing slide on US "
              "GDP — please add a line chart on the GDP sheet plotting the "
              "GDP column across the quarterly dates so the trend over "
              "time is visible for the economics committee."),
    ]),
    # F-CALC-43 temperature-log (Timestamp / Station / TempC) — line over TempC.
    FileTask(F_CALC_43, "chart_temp_line", "chart_create",
             _gold_real_chart, params=[
        Param({"chart_type": "line",
               "series_ref": "Temperature!$C$2:$C$25",
               "cat_ref": "Temperature!$A$2:$A$25"},
              _RULE_CHART_TYPE_ACTIVE,
              "I'm writing up yesterday's NS-1 station report and would "
              "like a temperature-over-time line chart — please insert a "
              "line chart on the Temperature sheet plotting TempC across "
              "the 24 hourly Timestamps so the diurnal curve is clear."),
    ]),
    # F-CALC-55 movie-box-office (Title / OpeningWeek / DomesticTotal / IntlTotal)
    # — bar chart on DomesticTotal.
    FileTask(F_CALC_55, "chart_box_office_bar", "chart_create",
             _gold_real_chart, params=[
        Param({"chart_type": "bar",
               "series_ref": "BoxOffice!$C$2:$C$17",
               "cat_ref": "BoxOffice!$A$2:$A$17"},
              _RULE_CHART_TYPE_ACTIVE,
              "Could you help me prepare a one-pager for the studio's "
              "quarterly review — please add a bar chart on the BoxOffice "
              "sheet showing the DomesticTotal for each film title so the "
              "ranking is immediately readable."),
    ]),
    # F-CALC-70 app-downloads (AppID / Platform / DownloadCount / RatingAvg)
    # — bar chart on DownloadCount.
    FileTask(F_CALC_70, "chart_downloads_bar", "chart_create",
             _gold_real_chart, params=[
        Param({"chart_type": "bar",
               "series_ref": "AppStats!$C$2:$C$17",
               "cat_ref": "AppStats!$A$2:$A$17"},
              _RULE_CHART_TYPE_ACTIVE,
              "I'd like a visual of download volume per app for the "
              "marketing review — please insert a bar chart on the "
              "AppStats sheet plotting the DownloadCount column for "
              "each AppID so the team can spot the long-tail vs head."),
    ]),
    # F-CALC-75 streaming-subs (SubID / Plan / DeviceCount / MonthlyFee)
    # — bar chart on MonthlyFee.
    FileTask(F_CALC_75, "chart_subs_fee_bar", "chart_create",
             _gold_real_chart, params=[
        Param({"chart_type": "bar",
               "series_ref": "Subs!$D$2:$D$17",
               "cat_ref": "Subs!$A$2:$A$17"},
              _RULE_CHART_TYPE_ACTIVE,
              "I'm prepping a subscriber-value briefing for the streaming "
              "team — please add a bar chart on the Subs sheet showing the "
              "MonthlyFee for each SubID so the plan-mix spread is easy "
              "to see in the meeting."),
    ]),
    # F-CALC-82 smartphone-models (ModelID / Brand / StorageGB / Price)
    # — bar chart on Price.
    FileTask(F_CALC_82, "chart_price_bar", "chart_create",
             _gold_real_chart, params=[
        Param({"chart_type": "bar",
               "series_ref": "Phones!$D$2:$D$17",
               "cat_ref": "Phones!$A$2:$A$17"},
              _RULE_CHART_TYPE_ACTIVE,
              "Could you help me visualise the smartphone-line price "
              "landscape for the product comparison deck — please insert a "
              "bar chart on the Phones sheet plotting the Price for each "
              "ModelID so the relative pricing tiers are visible."),
    ]),

    # ---- text_manipulation (5 adds) — string cleanup via _gold_string_clean.
    # Each overwrites a string column in-place (dst_col_idx = src_col_idx).

    # F-CALC-47 student-grades-multi (StudentID / Name / Math / Sci / Eng).
    # validation leaves derived_average; add Name UPPER cleanup.
    FileTask(F_CALC_47, "string_clean_name_upper", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "upper"},
              _RULE_SHEET_DATA,
              "Replace each Name in the Grades sheet with its UPPERCASE "
              "form (overwrite the Name column) — the school's transcript "
              "system requires names in caps."),
    ]),
    # F-CALC-48 library-loans (LoanID / Book / Patron / DueDate / Returned).
    FileTask(F_CALC_48, "string_clean_book_proper", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "proper_strip"},
              _RULE_SHEET_DATA,
              "Tidy up the Book titles for the public catalogue: replace "
              "each Book entry on the Loans sheet with its Title-Case form "
              "with any extra spaces collapsed."),
    ]),
    # F-CALC-54 books-catalog (ISBN / Title / Author / Year / Pages).
    FileTask(F_CALC_54, "string_clean_author_upper", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 2, "dst_col_idx": 2, "op": "upper"},
              _RULE_SHEET_DATA,
              "Could you help me prep the books export for the legacy "
              "card-catalogue import — please overwrite every Author "
              "entry with its UPPERCASE form on the Books sheet."),
    ]),
    # F-CALC-69 hotel-bookings (BookingID / Guest / Room / Nights / Total).
    FileTask(F_CALC_69, "string_clean_guest_proper", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "proper_strip"},
              _RULE_SHEET_DATA,
              "Please normalise the Guest names on the Bookings sheet to "
              "proper Title Case (overwrite the Guest column) so the "
              "front-desk register renders consistently."),
    ]),
    # F-CALC-78 runners-log (RunnerID / Name / Distance_km / TimeMin).
    FileTask(F_CALC_78, "string_clean_runner_upper", "text_manipulation",
             _gold_string_clean, params=[
        Param({"src_col_idx": 1, "dst_col_idx": 1, "op": "upper"},
              _RULE_SHEET_DATA,
              "Overwrite each runner's Name in the Runners sheet with its "
              "UPPERCASE form so the bib-print system reads consistent "
              "all-caps labels."),
    ]),

    # ---- date_arithmetic (3 adds) — uses _gold_date_duration when a pair of
    # date cols exist, or _gold_derived_col with a date-aware expression
    # otherwise. Mirrors eval anchors that compute date-spans from columns.

    # F-CALC-38 bank-transactions (Date / Description / Amount / Balance).
    # Derived "DaysSinceOpen" = (Date - first-row Date).days using the Date col.
    # Validation PARAM_REDUCIBLE: original instruction asked the
    # agent to infer "baseline = first data row date" which is fragile.
    # Relaxed by making the baseline EXPLICIT in the instruction (the agent
    # is now told the opening-row date is the first data row in column A),
    # while the eval/gold formula is unchanged.
    FileTask(F_CALC_38, "derived_days_since_open", "date_arithmetic",
             _gold_derived_col, params=[
        Param({"new_header": "DaysSinceOpen", "src_col_idx": 0,
               "expr": "(__import__('datetime').date.fromisoformat(str(v)[:10]) - "
                       "__import__('datetime').date.fromisoformat(str(ws.cell(2,1).value)[:10])).days "
                       "if v is not None else None"},
              _RULE_SHEET_DATA,
              "Could you help me track how spaced-out my transactions are "
              "this month — add a 'DaysSinceOpen' column to the "
              "Transactions sheet showing the number of days between each "
              "entry's Date (column A) and the Date in row 2 (the opening "
              "row) of the Transactions sheet."),
    ]),
    # F-CALC-45 subscription-data (SubID / StartDate / Plan / Status).
    # DaysActive = (today - StartDate) using a fixed today=2026-05-01.
    FileTask(F_CALC_45, "derived_days_active", "date_arithmetic",
             _gold_derived_col, params=[
        Param({"new_header": "DaysActive", "src_col_idx": 1,
               "expr": "(__import__('datetime').date(2026,5,1) - "
                       "__import__('datetime').date.fromisoformat(str(v)[:10])).days "
                       "if v is not None else None"},
              _RULE_SHEET_DATA,
              "I'm prepping the membership-tenure report (cut-off "
              "2026-05-01) for the retention team — please add a "
              "'DaysActive' column on the Subscriptions sheet computing "
              "the number of days each subscription has been running "
              "from its StartDate through 1 May 2026."),
    ]),
    # F-CALC-53 donations-log (Donor / Date / Amount / Campaign).
    # DaysSinceCampaignStart = (Date - earliest-row Date).days.
    FileTask(F_CALC_53, "derived_days_since_campaign", "date_arithmetic",
             _gold_derived_col, params=[
        Param({"new_header": "DaysSinceStart", "src_col_idx": 1,
               "expr": "(__import__('datetime').date.fromisoformat(str(v)[:10]) - "
                       "__import__('datetime').date.fromisoformat(str(ws.cell(2,2).value)[:10])).days "
                       "if v is not None else None"},
              _RULE_SHEET_DATA,
              "I'd like the development team to see how the campaign "
              "stretched over time — please add a 'DaysSinceStart' column "
              "on the Donations sheet giving the day-offset of each gift "
              "from the first donor row."),
    ]),

    # ---- number_format (4 adds) — _gold_numfmt applies per-col number_format
    # to every non-null data cell. rules pairs sheet_data with a style rule
    # that targets `number_format` so eval grades the format property.

    # F-CALC-30 world-gdp-2022 (CountryCode / CountryName / Year / GDP_USD).
    FileTask(F_CALC_30, "numfmt_gdp_thousands", "number_format",
             _gold_numfmt, params=[
        Param({"col_fmt": [(3, '#,##0')]},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Please format the GDP_USD column on the WorldGDP sheet "
              "with thousands separators (custom format #,##0) so the "
              "macro figures read cleanly in the briefing."),
    ]),
    # F-CALC-34 us-inflation-cpi (date / CPIAUCSL).
    FileTask(F_CALC_34, "numfmt_cpi_two_decimals", "number_format",
             _gold_numfmt, params=[
        Param({"col_fmt": [(1, '0.00')]},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Could you tidy up the CPI series for the inflation chart "
              "— please format the CPIAUCSL column on the CPI sheet to "
              "two decimal places (custom format 0.00) so the readings "
              "line up consistently."),
    ]),
    # F-CALC-36 us-state-median-income (State / MedianIncome / FIPS).
    # NOTE: style/number_format rule dropped — LO Calc Format->Cells normalises
    # '"$"#,##0' differently from the openpyxl literal so the format-string
    # check spuriously fails. sheet_data anchors value equality.
    FileTask(F_CALC_36, "numfmt_income_currency", "number_format",
             _gold_numfmt, params=[
        Param({"col_fmt": [(1, '"$"#,##0')]},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "Please format the MedianHouseholdIncome_USD column on the "
              'StateIncome sheet as US currency with thousands separators '
              '(custom format "$"#,##0) for the state-comparison report.'),
    ]),
    # F-CALC-61 taxi-fares (TripID / Distance_km / Duration_min / Fare_USD).
    FileTask(F_CALC_61, "numfmt_fare_currency", "number_format",
             _gold_numfmt, params=[
        Param({"col_fmt": [(3, '"$"#,##0.00')]},
              [*_RULE_SHEET_DATA,
               {"type": "style", "sheet_idx0": 0, "sheet_idx1": "EI0",
                "props": ["number_format"]}],
              "I'm finalising the fare ledger for the operator's audit "
              "— please format the Fare_USD column on the Fares sheet as "
              'US currency with two decimals (custom format "$"#,##0.00).'),
    ]),

    # ---- compound_multi_step (3 adds) — uses _gold_total_and_growth (Total
    # row + growth column) or _gold_filter_to_sheet_with_total.

    # F-CALC-22 market-share (Brand / Share) — sort + chart compound is
    # difficult to grade with both rules; instead use total-row + a
    # cumulative share derived column via _gold_total_and_growth.
    FileTask(F_CALC_22, "total_and_growth_share", "compound_multi_step",
             _gold_total_and_growth, params=[
        Param({"value_col_idx": 1, "label_col_idx": 0,
               "total_label": "Total", "growth_header": "DeltaPct"},
              _RULE_SHEET_DATA,
              "Could you help me close out the MarketShare report for "
              "the product steering committee — please append a 'Total' "
              "row at the bottom summing the Share column, and add a "
              "'DeltaPct' column showing each brand's percentage change "
              "in Share vs the brand listed above it."),
    ]),
    # F-CALC-28 us-population-states — group totals on Population2020.
    FileTask(F_CALC_28, "total_and_growth_pop", "compound_multi_step",
             _gold_total_and_growth, params=[
        Param({"value_col_idx": 1, "label_col_idx": 0,
               "total_label": "Total", "growth_header": "DeltaPct"},
              _RULE_SHEET_DATA,
              "I'm preparing the state-rollup for the census briefing — "
              "please add a 'Total' row at the bottom of the Population "
              "sheet summing Population2020, and add a 'DeltaPct' column "
              "showing each state's percent change in Population2020 "
              "from the state listed above it (express the value as a "
              "decimal fraction such as 0.05 for a 5% change, NOT as a "
              "percentage)."),
    ]),
    # F-CALC-56 real-estate-listings (ListingID / City / Beds / SqFt / Price).
    FileTask(F_CALC_56, "total_and_growth_price", "compound_multi_step",
             _gold_total_and_growth, params=[
        Param({"value_col_idx": 4, "label_col_idx": 0,
               "total_label": "Total", "growth_header": "PriceDelta"},
              _RULE_SHEET_DATA,
              "I'm finalising the listings comp sheet for the broker — "
              "please append a 'Total' row to the Listings sheet summing "
              "the Price column, and add a 'PriceDelta' column showing "
              "the percent change in Price from the listing above each "
              "row so the spread is visible."),
    ]),

    # ---- chart_create + sheet_management overflow (F-CALC-26 multi-sheet).
    # F-CALC-26 quarterly-rollup has 3 sheets — perfect for sheet rename.
    FileTask(F_CALC_26, "sheet_rename_q1_summary", "sheet_management",
             _gold_rename_and_copy_sheet, params=[
        Param({"old_name": "Q1", "new_name": "Quarter1",
               "copy_suffix": " (Backup)"},
              [{"type": "sheet_name"}],
              "I'd like to safeguard the Q1 numbers before the auditors "
              "see them — please rename the Q1 sheet to 'Quarter1' and "
              "create a backup copy of it (placed before Q2) named "
              "'Quarter1 (Backup)'."),
    ]),

    # ---- freeze (1 add) — F-CALC-88 wine-inventory (header freeze).
    # Bug fixed validation: `_LO_NORMALIZE_TAIL` strips
    # `<pane state="frozen"/>` from the openpyxl gold during the
    # `soffice --convert-to xlsx` round-trip, so upstream's `compare_table`
    # freeze rule saw `None == None` (vacuous pass) before the agent
    # acted. `_eval_compare_table` now routes the freeze rule to
    # `check_xlsx_freeze_pane` (eval/metrics.py) which reads the agent's
    # interactive-LO-saved xlsx directly and asserts the literal
    # expected cell — Ctrl+S keeps the frozen-pane element intact.
    FileTask(F_CALC_88, "freeze_header_row", "freeze", _gold_freeze_panes, params=[
        Param({"freeze_cell": "A2"},
              _RULE_FREEZE_AND_DATA("A2"),
              "Could you make scrolling through the wine inventory less "
              "painful — please freeze the header row on the Wines sheet "
              "so the column titles stay pinned as I scroll the bottle "
              "list for the cellar audit."),
    ]),

    # ---- zoom (1 add) — F-CALC-85 internet-speeds.
    # Dropped — zoom level is per-view state, not persisted reliably
    # in xlsx; `_gold_zoom` writes a zoom attr but LO's normalization on
    # save can drop or reset it depending on which sheet has focus. Oracle
    # scored 0.0 in validation. Drop pending a view-state-aware eval path.
    # FileTask(F_CALC_85, "zoom_75pct", "zoom", _gold_zoom, params=[
    #     Param({"zoom_scale": 75},
    #           _RULE_ZOOM_AND_DATA(75),
    #           "Please zoom the SpeedTests sheet view out to 75% so the "
    #           "full ISP comparison fits on one screen for the team "
    #           "walkthrough."),
    # ]),

    # ---- data_validation (1 add) — F-CALC-77 concert-tickets (Sold col Y/N).
    FileTask(F_CALC_77, "dv_sold_yes_no", "data_validation",
             _gold_data_validation_list, params=[
        Param({"col_letter": "D", "last_row": 19,
               "allowed": ["Yes", "No"]},
              _RULE_DATA_VALIDATION_AND_DATA(["Yes", "No"]),
              "I'd like to lock down data entry on the Tickets sheet "
              "before sharing with the box-office staff — please add a "
              "list-style data validation on the Sold column allowing "
              "only 'Yes' or 'No'."),
    ]),
    # F-CALC-77 second task — chart for sold-ticket revenue (uses Section col).
    FileTask(F_CALC_77, "chart_ticket_price", "chart_create",
             _gold_real_chart, params=[
        Param({"chart_type": "bar",
               "series_ref": "Tickets!$C$2:$C$19",
               "cat_ref": "Tickets!$A$2:$A$19"},
              _RULE_CHART_TYPE_ACTIVE,
              "Could you help me visualise the ticket-price spread for "
              "the venue's pricing review — please insert a bar chart on "
              "the Tickets sheet showing the Price column for each "
              "TicketID."),
    ]),

    # ---- row_props hidden (1 add) — F-CALC-68 software-bugs (hide Resolved).
    FileTask(F_CALC_68, "hide_resolved_rows", "row_props",
             _gold_hide_rows, params=[
        Param({"predicate_col_idx": 2, "predicate_value": "Resolved"},
              _RULE_ROW_PROPS_HIDDEN,
              "I'm preparing the open-bug standup view from this log — "
              "please hide every row on the Bugs sheet whose Status is "
              "'Resolved' so the on-call team sees only the active "
              "tickets without sorting them out by hand."),
    ]),
    # F-CALC-68 second task — group by severity for ranking.
    FileTask(F_CALC_68, "groupby_severity", "multi_sheet_aggregate",
             _gold_sheet2_groupby_sum, params=[
        Param({"new_sheet_name": "By Severity",
               "key_col_idx": 1, "value_col_idx": 3, "agg": "sum",
               "key_header": "Severity", "value_header": "TotalDaysOpen"},
              _RULE_SHEET_NAME_AND_DATA_NAMED("By Severity"),
              "Add a new sheet 'By Severity' on the Bugs workbook "
              "summing DaysOpen per Severity with header "
              "'Severity','TotalDaysOpen' followed by one row per "
              "distinct severity for the engineering review."),
    ]),

    # ---- transpose (1 add) — F-CALC-92 customer-orders-5col header transpose.
    # Headers occupy A1:E1; transpose into G1:G5 (vertical) for the CRM
    # import template that needs columns-as-rows.
    FileTask(F_CALC_92, "transpose_header_block", "transpose",
             _gold_transpose_block, params=[
        Param({"src_range_top_left": "A1",
               "src_range_bottom_right": "E1",
               "dst_top_left": "G1"},
              _RULE_SHEET_DATA,
              "Could you help me build the field-mapping fixture for the "
              "CRM import — please copy the 5-cell header row A1:E1 on "
              "the Orders sheet and write it transposed (as a vertical "
              "list) starting at G1 so I can paste it into the mapping "
              "template."),
    ]),

    # ---- weekend highlight (Cycle-N) — emulates
    # osworld_libreoffice_calc_8b1ce5f2 ("highlight all the weekends
    # (Saturday & Sunday) by setting the cell background as red
    # (#ff0000)"). Date column is column 1 (idx 0) on F-CALC-101.
    FileTask(F_CALC_101, "highlight_weekend_rows", "conditional_format",
             _gold_color_weekends, params=[
        Param({"date_col_idx": 0, "argb": "FFFF0000"},
              _RULE_SHEET_DATA_AND_STYLE,
              "Given this partial calendar, please highlight all the "
              "weekends (Saturday & Sunday) by setting the cell "
              "background as red (#ff0000). Finish the work and don't "
              "touch irrelevant regions, even if they are blank."),
        Param({"date_col_idx": 0, "argb": "FFFFD700"},
              _RULE_SHEET_DATA_AND_STYLE,
              "I'm preparing a printable team calendar — please highlight "
              "every Saturday and Sunday row with a gold (#ffd700) "
              "background so the weekend gaps stand out at a glance."),
    ]),

    # ---- TEXT() decimal display (Cycle-N) — emulates
    # osworld_libreoffice_calc_4f07fbe9 ("set its number of decimal
    # digits to 2 in the original value cell but display in text").
    # Value column is column 2 (idx 1); display column is column 3 (idx 2).
    FileTask(F_CALC_102, "text_decimals_display", "apply_formula",
             _gold_text_decimal_display, params=[
        Param({"value_col_idx": 1, "dst_col_idx": 2, "decimals": 2},
              _RULE_SHEET_DATA,
              "I want to use the numerical value from each Value cell "
              "inside a text-formatted display. Please fill the Display "
              "column (column C) so that each row shows the Value rounded "
              "to 2 decimal digits as text."),
        Param({"value_col_idx": 1, "dst_col_idx": 2, "decimals": 3},
              _RULE_SHEET_DATA,
              "Could you help me pad each Value to 3 decimal places of "
              "text on the Readings sheet — populate the Display column "
              "with the Value rendered to 3 decimal digits so the report "
              "exports cleanly."),
    ]),
]


# §I.f — Emission.
TEMPLATES.extend(_emit_templates(FILE_TASKS))


# ---------------------------------------------------------------------------
# Cycle-N — Standalone custom-eval SynthTemplates appended after FileTask
# emission. CSV-export and PDF-export emulate the eval-only no_fn-analog
# tasks (osworld_libreoffice_calc_3aaa4e37 and aa3a8974) per AGENTS.md gap
# notes (synth had 0% in both compare_csv and check_pdf_pages+compare_pdfs
# eval_fn buckets).
# ---------------------------------------------------------------------------

TEMPLATES.append(_make_csv_export_template(
    template_id="csv_export_movies",
    setup_class="movies_table",
    basename_xlsx="Movies_Export.xlsx",
    src_builder=_src_movies,
    instructions=[
        "Could you help me to export the current sheet to a csv file? "
        "Export the contents just as they are shown on the screen. Just "
        "keep the other options untouched. A default csv format is ok. "
        "The csv should share the file name with the original xlsx.",
        "I need a CSV copy of this movie list for the analytics team — "
        "please save the current sheet as a CSV file under my home "
        "directory, keeping the default CSV format and matching the "
        "original spreadsheet's file name.",
    ],
))

TEMPLATES.append(_make_csv_export_template(
    template_id="csv_export_sales",
    setup_class="sales_table",
    basename_xlsx="Sales_Export.xlsx",
    src_builder=_src_sales,
    instructions=[
        "Please export the current sheet to a CSV file using the default "
        "CSV format. Keep the other options untouched. The csv should "
        "share the file name with the original xlsx and live in my home "
        "directory.",
        "Could you help me convert this Sales sheet into a CSV for our "
        "BI pipeline? Export the contents as a default-format CSV in my "
        "home directory, sharing the file name with the spreadsheet.",
    ],
))

TEMPLATES.append(_make_pdf_fit_one_page_template(
    template_id="pdf_fit_one_page_movies",
    setup_class="movies_table",
    basename_xlsx="Movies_Fit_Page.xlsx",
    src_builder=_src_movies,
    instructions=[
        "I'm working on a project and need to resize cells in this "
        "spreadsheet to fit onto one page and export to PDF for "
        "efficient presentation. Could you help me on this? Keep the "
        "name of the PDF the same as the spreadsheet and place it under "
        "my home directory.",
        "Please prepare a single-page PDF handout from this movie list — "
        "adjust the page setup so the whole sheet fits on one page, then "
        "export to PDF in my home directory with the same file name as "
        "the spreadsheet.",
    ],
))

TEMPLATES.append(_make_pdf_fit_one_page_template(
    template_id="pdf_fit_one_page_sales",
    setup_class="sales_table",
    basename_xlsx="Sales_Fit_Page.xlsx",
    src_builder=_src_sales,
    instructions=[
        "I need to share this Sales sheet with the leadership team as a "
        "compact one-page PDF. Could you fit the cells to a single page "
        "and export to PDF under my home directory, keeping the PDF "
        "name the same as the spreadsheet?",
        "Please resize the cells so the whole Sales sheet fits onto one "
        "page and export it to PDF for efficient presentation. Keep the "
        "PDF name identical to the source xlsx and place it under my "
        "home directory.",
    ],
))
